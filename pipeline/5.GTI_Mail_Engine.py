# -*- coding: utf-8 -*-
"""
GTI STEP5 Mail Engine - report quality form v2

Report form
-----------
1. 총평
2. Top3 Deep Analysis
3. Regulation
4. 주요뉴스

This step does not reselect STEP4 results. It keeps all selected regulation/news
items, then rewrites weak STEP4 text into an executive report style.
"""

from __future__ import annotations

import argparse
import html
import os
import re
import smtplib
import ssl
from datetime import datetime
from email.message import EmailMessage
from email.utils import formataddr
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


REGULATION_INPUT_FILE = Path(os.getenv("GTI_REGULATION_INPUT_FILE", r"C:\Temp\4-1.regulation_ai_summary.xlsx"))
NEWS_INPUT_FILE = Path(os.getenv("GTI_NEWS_INPUT_FILE", r"C:\Temp\4-2.news_ai_summary.xlsx"))
OUTPUT_DIR = Path(os.getenv("GTI_OUTPUT_DIR", r"C:\Temp\12345\c_type_outputs"))
RUN_DATE = os.getenv("GTI_RUN_DATE", datetime.now().strftime("%Y-%m-%d"))

NEWS_MAX_ROWS = int(os.getenv("GTI_NEWS_MAX_ROWS", "0"))  # 0 = no cap
SEND_EMAIL = os.getenv("GTI_SEND_EMAIL", "Y").strip().upper() in {"Y", "YES", "TRUE", "1"}
SMTP_HOST = os.getenv("GTI_SMTP_HOST", "smtp.naver.com")
SMTP_PORT = int(os.getenv("GTI_SMTP_PORT", "465"))
SMTP_USER = os.getenv("GTI_SMTP_USER", "kch8872@naver.com").strip()
SMTP_PASS = (os.getenv("GTI_SMTP_PASS") or os.getenv("GTI_MAIL_PW") or "").strip()
MAIL_TO = os.getenv("GTI_MAIL_TO", "").strip()
MAIL_FROM_NAME = os.getenv("GTI_MAIL_FROM_NAME", "GTI Radar").strip()
RECIPIENT_FILE = Path(os.getenv("GTI_RECIPIENT_FILE", r"C:\Temp\00.xlsx"))


OUTPUT_COLUMNS = [
    "No", "Content Type", "Mail Group", "Samsung Impact", "Affected Subsidiary", "Impact Reason",
    "Date", "Headline", "Major Changes", "Summary", "AI Analysis", "Action Plan", "Country", "Agency",
    "Risk", "Importance Score", "Priority Group", "Issue", "Cluster", "URL", "Source", "Source File",
]

GROUP_REGULATION = "Regulation"
GROUP_NEWS = "주요뉴스"


def output_paths() -> dict[str, Path]:
    return {
        "analysis": OUTPUT_DIR / "4.news_ai_analysis.xlsx",
        "mail_xlsx": OUTPUT_DIR / f"[GTI Radar] Global Trade Intelligence({RUN_DATE}).xlsx",
        "mail_html": OUTPUT_DIR / f"[GTI Radar] Global Trade Intelligence({RUN_DATE}).html",
        "cumulative": OUTPUT_DIR / "gti_news_cumulative.xlsx",
    }


def clean(value) -> str:
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except Exception:
        pass
    return re.sub(r"\s+", " ", str(value)).strip()


def pick_col(df: pd.DataFrame, candidates: list[str]) -> str | None:
    lookup = {str(c).strip().lower(): c for c in df.columns}
    for cand in candidates:
        if cand.lower() in lookup:
            return lookup[cand.lower()]
    return None


def safe_num(value) -> float:
    try:
        if value is None or pd.isna(value):
            return 0.0
    except Exception:
        pass
    text = re.sub(r"[^0-9.\-]", "", str(value))
    try:
        return float(text) if text else 0.0
    except Exception:
        return 0.0


def normalize_risk(value) -> str:
    raw = clean(value)
    low = raw.lower()
    if raw in {"상", "중", "하"}:
        return raw
    if low in {"high", "h", "red"}:
        return "상"
    if low in {"medium", "med", "m", "orange"}:
        return "중"
    if low in {"low", "l", "blue"}:
        return "하"
    return "중"


def risk_weight(value) -> int:
    return {"상": 300, "중": 150, "하": 0}.get(normalize_risk(value), 0)


def priority_weight(value) -> int:
    p = clean(value).upper()
    return {"CORE": 1000, "POLICY_WATCH": 850, "USABLE": 650, "REFERENCE": 300, "WATCH": 250}.get(p, 200)


def parse_date(value):
    dt = pd.to_datetime(value, errors="coerce")
    return pd.Timestamp.min if pd.isna(dt) else dt


def display_date(value) -> str:
    dt = pd.to_datetime(value, errors="coerce")
    if pd.isna(dt):
        return clean(value)[:16]
    if dt.hour == 0 and dt.minute == 0:
        return dt.strftime("%Y-%m-%d")
    return dt.strftime("%Y-%m-%d %H:%M")


def best_url_from_values(values) -> str:
    invalid = {
        "", "nan", "none", "null", "new", "https://new", "http://new",
        "https://news", "http://news", "https://news.google.com", "https://news.google.com/",
    }
    candidates: list[str] = []
    for value in values:
        text = clean(value)
        if not text:
            continue
        for item in [text] + re.findall(r"https?://[^'\"),\s]+", text):
            url = html.unescape(item).strip().strip("<>'\"").rstrip(".,);]}")
            if url.lower() in invalid:
                continue
            if re.match(r"^https?://", url, re.I) and url not in candidates:
                candidates.append(url)
    for url in candidates:
        low = url.lower()
        if "news.google.com/rss/articles/" not in low and "news.google.com/articles/" not in low:
            return url
    return candidates[0] if candidates else ""


def non_empty_hint(value: str) -> str:
    text = clean(value)
    if not text or text in {"본문에서 확인 불가", "nan", "None"}:
        return ""
    return text


def normalize_input(df: pd.DataFrame, content_type: str, source_file: Path) -> pd.DataFrame:
    df = df.copy()
    df.columns = [str(c).strip() for c in df.columns]

    col_date = pick_col(df, ["Date", "date"])
    col_headline = pick_col(df, ["Headline", "Title", "headline"])
    col_country = pick_col(df, ["Country", "country"])
    col_agency = pick_col(df, ["Agency", "Publisher", "agency", "source"])
    col_risk = pick_col(df, ["Risk", "risk"])
    col_score = pick_col(df, ["Importance Score", "final_score", "samsung_score", "Score", "Importance"])
    col_priority = pick_col(df, ["Priority Group", "priority_group", "mail_section", "Tier"])
    col_issue = pick_col(df, ["Issue", "issue_type", "topic_keyword", "topic", "IssueKey"])
    col_cluster = pick_col(df, ["Cluster", "cluster_key", "ClusterHeadlines"])
    col_summary = pick_col(df, ["Summary", "summary", "ExecutiveMessage"])
    col_analysis = pick_col(df, ["AI Analysis", "analysis", "samsung_reason"])
    col_action = pick_col(df, ["Action Plan", "RequiredAction", "action"])
    col_source = pick_col(df, ["Source", "SourceFile", "source"])
    col_impact = pick_col(df, ["Samsung Impact", "samsung_impact"])
    col_subs = pick_col(df, ["Affected Subsidiary", "affected_subsidiary", "affected_subsidiaries"])
    col_reason = pick_col(df, ["Impact Reason", "subsidiary_reason", "samsung_reason", "SelectReason"])

    hint_cols = {
        "effective_date_hint": pick_col(df, ["effective_date_hint"]),
        "change_detail_hint": pick_col(df, ["change_detail_hint"]),
        "hs_hint": pick_col(df, ["hs_hint"]),
        "tariff_rate_hint": pick_col(df, ["tariff_rate_hint"]),
        "KeywordMatches": pick_col(df, ["KeywordMatches"]),
        "affected_products": pick_col(df, ["affected_products", "impact_products", "subsidiary_products"]),
        "fta_impact": pick_col(df, ["fta_impact"]),
        "export_control_impact": pick_col(df, ["export_control_impact"]),
        "hs_impact": pick_col(df, ["hs_impact"]),
        "tariff_impact": pick_col(df, ["tariff_impact"]),
    }

    url_cols = [
        pick_col(df, ["BestLinkURL"]),
        pick_col(df, ["OriginalURLCandidate"]),
        pick_col(df, ["original_url"]),
        pick_col(df, ["URL", "url", "Link"]),
        pick_col(df, ["GoogleURL"]),
        col_source,
    ]
    url_cols = [c for c in url_cols if c]

    out = pd.DataFrame()
    out["Date"] = df[col_date].apply(display_date) if col_date else ""
    out["_sort_date"] = df[col_date].apply(parse_date) if col_date else pd.Timestamp.min
    out["Headline"] = df[col_headline].apply(clean) if col_headline else ""
    out["Country"] = df[col_country].apply(clean) if col_country else ""
    out["Agency"] = df[col_agency].apply(clean) if col_agency else ""
    out["Risk"] = df[col_risk].apply(normalize_risk) if col_risk else "중"
    out["Importance Score"] = df[col_score].apply(safe_num) if col_score else 0
    out["Priority Group"] = df[col_priority].apply(lambda v: clean(v).upper()) if col_priority else ("CORE" if content_type == "Regulation" else "USABLE")
    out["Issue"] = df[col_issue].apply(clean) if col_issue else ""
    out["Issue"] = out["Issue"].replace({
        "TARIFF": "관세정책", "SECTION_301_232": "관세정책",
        "CUSTOMS": "통관", "CUSTOMS_CLEARANCE": "통관",
        "ORIGIN_FTA": "FTA/원산지", "CBAM_CARBON": "CBAM",
        "HS_CLASSIFICATION": "HS/품목분류", "AD_CVD": "AD/CVD",
        "EXPORT_CONTROL": "수출통제",
    })
    out["Cluster"] = df[col_cluster].apply(clean) if col_cluster else ""
    out["Summary"] = df[col_summary].apply(clean) if col_summary else ""
    out["AI Analysis"] = df[col_analysis].apply(clean) if col_analysis else ""
    out["Action Plan"] = df[col_action].apply(clean) if col_action else ""
    out["Samsung Impact"] = df[col_impact].apply(lambda v: clean(v).title() if clean(v).lower() in {"direct", "indirect", "watch"} else clean(v)) if col_impact else "Watch"
    out["Samsung Impact"] = out["Samsung Impact"].replace({"": "Watch", "직접": "Direct", "간접": "Indirect", "모니터링": "Watch"})
    out["Affected Subsidiary"] = df[col_subs].apply(clean) if col_subs else ""
    out["Impact Reason"] = df[col_reason].apply(clean) if col_reason else ""
    out["Source"] = df[col_source].apply(clean) if col_source else ""
    out["Source File"] = str(source_file)
    out["Content Type"] = content_type
    out["Mail Group"] = GROUP_REGULATION if content_type == "Regulation" else GROUP_NEWS
    out["URL"] = df.apply(lambda r: best_url_from_values([r.get(c, "") for c in url_cols]), axis=1) if len(df) else ""

    for out_col, src_col in hint_cols.items():
        out[out_col] = df[src_col].apply(clean) if src_col else ""

    out = out[out["Headline"].astype(str).str.strip().ne("")]
    return out.reset_index(drop=True)


def read_step4_results() -> pd.DataFrame:
    frames = []
    if REGULATION_INPUT_FILE.exists():
        frames.append(normalize_input(pd.read_excel(REGULATION_INPUT_FILE), "Regulation", REGULATION_INPUT_FILE))
    if NEWS_INPUT_FILE.exists():
        news = normalize_input(pd.read_excel(NEWS_INPUT_FILE), "News", NEWS_INPUT_FILE)
        if NEWS_MAX_ROWS > 0:
            news = news.head(NEWS_MAX_ROWS)
        frames.append(news)
    if not frames:
        raise FileNotFoundError(f"STEP4 outputs not found: {REGULATION_INPUT_FILE}, {NEWS_INPUT_FILE}")

    rows = pd.concat(frames, ignore_index=True)
    rows["_dedup_key"] = rows.apply(
        lambda r: clean(r.get("URL")) or (clean(r.get("Headline"))[:160] + "|" + clean(r.get("Agency")) + "|" + clean(r.get("Date"))),
        axis=1,
    )
    rows = rows.drop_duplicates(subset=["_dedup_key"], keep="first").drop(columns=["_dedup_key"], errors="ignore")
    rows["_integrated_score"] = rows.apply(
        lambda r: priority_weight(r["Priority Group"]) + risk_weight(r["Risk"]) + (180 if r["Content Type"] == "Regulation" else 0) + safe_num(r["Importance Score"]),
        axis=1,
    )
    return rows.reset_index(drop=True)


def dedup_report_rows(rows: pd.DataFrame) -> pd.DataFrame:
    """Remove near-duplicate report items after STEP4 merge.

    This intentionally catches cases where the same policy appears through two
    official pages, such as CBAM certificate price "to be published" and
    "now available", or the same bonded warehouse notice from two boards.
    """
    if rows.empty:
        return rows
    rows = rows.copy()
    rows["_report_dedup_key"] = rows.apply(report_dedup_key, axis=1)
    rows["_dedup_rank"] = rows.apply(dedup_rank, axis=1)
    rows = rows.sort_values(["_dedup_rank", "_integrated_score", "_sort_date"], ascending=[False, False, False])
    rows = rows.drop_duplicates(subset=["_report_dedup_key"], keep="first")
    return rows.drop(columns=["_report_dedup_key", "_dedup_rank"], errors="ignore").reset_index(drop=True)


def report_dedup_key(row: pd.Series) -> str:
    issue = clean(row.get("Issue")) or issue_for(row)
    title = clean(row.get("Headline")).lower()
    source = clean(row.get("Agency")).lower()
    normalized = re.sub(r"\([^)]*\)|\[[^]]*\]", " ", title)
    normalized = re.sub(r"제출기한[:：]?\s*\d{4}[.\-/]\d{1,2}[.\-/]\d{1,2}.*", " ", normalized)
    normalized = re.sub(r"\b(to be published|now available|published|available|first)\b", " ", normalized)
    normalized = re.sub(r"[^0-9a-z가-힣]+", " ", normalized)
    normalized = re.sub(r"\s+", " ", normalized).strip()

    if "cbam" in normalized and "certificate price" in normalized:
        return "REG:CBAM_CERTIFICATE_PRICE"
    if "보세창고" in normalized and "특허" in normalized and "운영" in normalized:
        return "REG:BONDED_WAREHOUSE_LICENSE_OPERATION"
    if "환전영업자" in normalized and "관리" in normalized:
        return "REG:FX_BUSINESS_OPERATOR_MANAGEMENT"
    if "수입신고" in normalized and "가산세" in normalized:
        return "REG:IMPORT_DECLARATION_DELAY_SURCHARGE"

    tokens = [t for t in normalized.split() if len(t) >= 2]
    return f"{clean(row.get('Content Type'))}:{issue}:{' '.join(tokens[:9])}:{source[:24]}"


def dedup_rank(row: pd.Series) -> float:
    title = clean(row.get("Headline")).lower()
    rank = safe_num(row.get("Importance Score")) + priority_weight(row.get("Priority Group")) + risk_weight(row.get("Risk"))
    if "now available" in title or "successfully entered into force" in title:
        rank += 150
    if "to be published" in title or "reminder" in title:
        rank -= 80
    if "제출기한" in title:
        rank += 60
    if clean(row.get("Agency")).startswith("Korea Customs") or "관세청" in clean(row.get("Agency")):
        rank += 40
    return rank


def issue_for(row) -> str:
    issue = clean(row.get("Issue"))
    if issue and issue.lower() not in {"watch", "policy_watch", "usable", "core"}:
        return issue
    text = " ".join(clean(row.get(c)) for c in ["Headline", "Summary", "AI Analysis", "Action Plan", "KeywordMatches"]).lower()
    if any(k in text for k in ["section 301", "section 232", "tariff", "quota", "duty", "관세", "쿼터"]):
        return "관세정책"
    if any(k in text for k in ["anti-dumping", "antidumping", "countervailing", "ad/cvd", "반덤핑", "상계관세"]):
        return "AD/CVD"
    if any(k in text for k in ["cbam", "carbon border"]):
        return "CBAM"
    if any(k in text for k in ["fta", "cepa", "origin", "원산지"]):
        return "FTA/원산지"
    if any(k in text for k in ["export control", "entity list", "uflpa", "forced labor", "수출통제"]):
        return "수출통제"
    if any(k in text for k in ["hs code", "classification", "품목분류"]):
        return "HS/품목분류"
    if clean(row.get("Content Type")) == "Regulation":
        return "법규"
    return "Watch"


def prepare_rows(rows: pd.DataFrame) -> pd.DataFrame:
    rows = rows.copy()
    rows["Issue"] = rows.apply(issue_for, axis=1)
    rows = dedup_report_rows(rows)
    rows["Mail Group"] = rows["Content Type"].map({"Regulation": GROUP_REGULATION}).fillna(GROUP_NEWS)
    rows["Major Changes"] = rows.apply(major_changes, axis=1)
    rows["Summary"] = rows.apply(report_summary, axis=1)
    rows["AI Analysis"] = rows.apply(report_impact, axis=1)
    rows["Action Plan"] = rows.apply(report_action, axis=1)
    rows["_report_score"] = rows.apply(report_score, axis=1)
    rows = rows.sort_values(["_integrated_score", "_sort_date"], ascending=[False, False]).reset_index(drop=True)
    rows["No"] = range(1, len(rows) + 1)
    return rows


def choose_top3(rows: pd.DataFrame) -> pd.DataFrame:
    pool = rows.copy()
    pool["_top3_score"] = pool.apply(top3_deep_score, axis=1)
    pool = pool.sort_values(["_top3_score", "_sort_date"], ascending=[False, False])
    selected = []
    used_issues = set()
    for _, row in pool.iterrows():
        issue = clean(row.get("Issue"))
        if issue in used_issues and len(selected) < 3:
            continue
        selected.append(row)
        used_issues.add(issue)
        if len(selected) == 3:
            break
    if len(selected) < 3:
        for _, row in pool.iterrows():
            if any(clean(row.get("Headline")) == clean(x.get("Headline")) for x in selected):
                continue
            selected.append(row)
            if len(selected) == 3:
                break
    out = pd.DataFrame(selected).reset_index(drop=True)
    if not out.empty:
        out["No"] = range(1, len(out) + 1)
    return out


def top3_deep_score(row: pd.Series) -> float:
    text = " ".join(clean(row.get(c)) for c in [
        "Headline", "Major Changes", "Summary", "AI Analysis", "Action Plan", "Issue"
    ]).lower()
    score = report_score(row)

    high_terms = [
        "anti-dumping", "antidumping", "countervailing", "ad/cvd", "반덤핑", "상계관세",
        "cbam", "carbon border", "탄소국경", "탄소세",
        "export control", "entity list", "forced labor", "uflpa", "수출통제", "강제노동",
        "section 301", "section 232", "tariff quota", "duty-free quota", "customs duty",
        "관세", "무관세", "쿼터", "원산지", "rules of origin", "hs code",
    ]
    medium_terms = [
        "fta", "cepa", "usmca", "통상협정", "fta", "통관", "보세", "신고", "classification",
    ]
    low_terms = [
        "수출 85.9", "수출입 현황", "잠정치", "재정적자", "refunds", "customs revenue",
        "브랜드", "주식", "전략회의", "칼럼", "market outlook",
    ]

    if any(t in text for t in high_terms):
        score += 1200
    if any(t in text for t in medium_terms):
        score += 450
    if clean(row.get("Content Type")) == "Regulation":
        score += 300
    if any(t in text for t in low_terms):
        score -= 900
    return score


def report_score(row: pd.Series) -> float:
    impact_weight = {"Direct": 2200, "Indirect": 900, "Watch": 0}.get(clean(row.get("Samsung Impact")), 0)
    type_weight = 350 if clean(row.get("Content Type")) == "Regulation" else 0
    issue_weight = {
        "관세정책": 500,
        "AD/CVD": 500,
        "반덤핑/상계관세": 500,
        "CBAM": 450,
        "수출통제": 450,
        "FTA/원산지": 350,
        "통관": 300,
        "통관/세관": 300,
        "HS/품목분류": 300,
    }.get(clean(row.get("Issue")), 150)
    return safe_num(row.get("Importance Score")) + priority_weight(row.get("Priority Group")) + risk_weight(row.get("Risk")) + impact_weight + type_weight + issue_weight


def hint_line(label: str, value: str) -> str:
    value = non_empty_hint(value)
    return f"{label}: {value}" if value else ""


def compact_parts(parts: list[str], fallback: str) -> str:
    parts = [p for p in parts if clean(p)]
    return "; ".join(parts) if parts else fallback


def major_changes(row: pd.Series) -> str:
    issue = clean(row.get("Issue"))
    headline = clean(row.get("Headline"))
    title_l = headline.lower()

    if "보세창고" in headline and "특허" in headline:
        return (
            "개정 사유: 자가용보세창고 특허요건 완화 및 불명확한 규정 보완 필요. "
            "주요 개정 내용: 자가용보세창고 반입 대상에 국제무역선·기 적재 자가화물 외 수리용 예비부분품 및 부속품 장치를 허용하고, "
            "관세법 제178조상 물품반입 정지기간을 오해 없이 적용할 수 있도록 규정을 명확화하는 내용입니다."
        )
    if "환전영업자" in headline and "관리" in headline:
        return (
            "주요 내용: 환전영업자의 등록·관리, 보고·자료제출, 영업장 운영 및 관세청 관리 기준과 관련된 고시입니다. "
            "해외출장·주재원·외환거래 지원 프로세스와 연결될 수 있어 실제 법인 업무 해당 여부 확인이 필요합니다."
        )
    if "cbam" in title_l and "certificate price" in title_l:
        return (
            "주요 내용: EU CBAM 인증서 가격이 공표되었거나 공표 일정이 확정된 사안입니다. "
            "EU 수입품의 내재배출량 신고, 인증서 구매 비용, 공급사 배출량 자료 확보 체계에 영향을 줄 수 있습니다."
        )
    if "customs enforcement" in title_l and "executive order" in title_l:
        return (
            "주요 내용: 미국 세관 집행 강화 행정명령 관련 사안입니다. "
            "수입신고 정확성, 저가신고·우회수입·전자상거래 물품 관리 및 CBP 심사 강화 가능성을 확인해야 합니다."
        )
    if "수입신고" in headline and "가산세" in headline:
        return (
            "주요 내용: 수입신고 지연 가산세 부과 대상이 되는 매점매석 금지 품목의 적용기간 연장 공고입니다. "
            "해당 품목 수입 시 신고 지연, 재고 운영, 통관 일정 관리 기준을 확인해야 합니다."
        )

    parts = [
        hint_line("시행/적용일", row.get("effective_date_hint")),
        hint_line("변경 내용", row.get("change_detail_hint")),
        hint_line("대상 HS", row.get("hs_hint")),
        hint_line("관세율/쿼터", row.get("tariff_rate_hint")),
        hint_line("키워드", row.get("KeywordMatches")),
    ]
    if any(parts):
        return compact_parts(parts, "")

    if issue == "관세정책":
        return "관세율, 쿼터, 면세/환급 또는 Section 301/232 등 관세 비용에 영향을 줄 수 있는 정책 변화입니다."
    if issue in {"AD/CVD", "반덤핑/상계관세"}:
        return "반덤핑 또는 상계관세 조사·판정·연장 가능성이 있는 사안입니다. 공급국, 대상 품목, 조사 기간과 관세율 확인이 필요합니다."
    if issue == "CBAM":
        return "CBAM 신고, 인증서 가격, 배출량 자료 또는 EU 수입통관 절차와 연결되는 탄소국경조정 변화입니다."
    if issue == "FTA/원산지":
        return "FTA/CEPA 협정, 원산지 기준, CO 발급 또는 특혜관세 적용 가능성에 영향을 주는 변화입니다."
    if issue == "수출통제":
        return "Entity List, ECCN, UFLPA, forced labor 또는 전략물자·제재 스크리닝 관련 변화입니다."
    if issue == "통관":
        return "보세, 통관, 신고, 세관 심사 또는 행정절차 기준에 영향을 줄 수 있는 공식 공지입니다."
    if issue == "HS/품목분류":
        return "HS 분류 기준 또는 품목 해석이 달라질 수 있어 품목 마스터와 신고 기준 점검이 필요한 사안입니다."
    return f"{headline} 관련 관세·통상 모니터링 사안입니다."


def report_summary(row: pd.Series) -> str:
    issue = clean(row.get("Issue"))
    country = clean(row.get("Country")) or "관련 국가"
    agency = clean(row.get("Agency")) or "관련 기관"
    change = major_changes(row)
    if clean(row.get("Content Type")) == "Regulation":
        return f"{agency}의 공식 법규/공지입니다. 핵심은 {change} 원문 기준으로 시행일, 적용 품목, HS, 세율 또는 신고 절차를 확인해야 합니다."
    return f"{country}에서 포착된 {issue} 뉴스입니다. 핵심은 {change} 삼성전자 관련 법인·품목에 직접 적용되는지 확인할 필요가 있습니다."


def report_impact(row: pd.Series) -> str:
    issue = clean(row.get("Issue"))
    impact = clean(row.get("Samsung Impact")) or "Watch"
    subs = clean(row.get("Affected Subsidiary")) or "SEC/HQ"
    products = non_empty_hint(row.get("affected_products"))
    product_txt = f" 대상 제품 후보는 {products}입니다." if products else ""
    if issue == "관세정책":
        return f"{subs} 기준 수입가격, 관세환급, 할당관세/쿼터, 공급국 선택에 영향을 줄 수 있습니다.{product_txt} Impact는 {impact}로 분류됩니다."
    if issue in {"AD/CVD", "반덤핑/상계관세"}:
        return f"{subs}의 철강·부품·원재료 조달에서 AD/CVD 추가관세 또는 조사 대응 자료 부담이 생길 수 있습니다.{product_txt} 공급국과 HS별 노출도를 확인해야 합니다."
    if issue == "CBAM":
        return f"{subs}의 EU향 판매·공급망에서 CBAM 신고자료, 배출량 증빙, 인증서 비용 관리가 필요할 수 있습니다.{product_txt}"
    if issue == "FTA/원산지":
        return f"{subs}의 FTA 활용, 원산지 판정, CO 발급, BOM 원산지 증빙 체계에 영향을 줄 수 있습니다.{product_txt}"
    if issue == "수출통제":
        return f"{subs}의 거래처 스크리닝, ECCN/전략물자 분류, 우회수출 통제와 연결될 수 있습니다.{product_txt}"
    if issue == "통관":
        return f"{subs}의 수입신고, 보세창고, 통관 심사, 세관 제출자료 운영 기준에 반영 여부를 확인해야 합니다.{product_txt}"
    if issue == "HS/품목분류":
        return f"{subs}의 HS 마스터, 품목 설명, 관세율 산정 및 신고 정확성에 영향을 줄 수 있습니다.{product_txt}"
    return f"{subs} 기준 관세·통상 리스크 모니터링 가치가 있습니다. Impact는 {impact}입니다."


def report_action(row: pd.Series) -> str:
    issue = clean(row.get("Issue"))
    subs = clean(row.get("Affected Subsidiary")) or "SEC/HQ"
    if issue == "관세정책":
        return f"{subs}: 대상 HS·공급국·거래금액을 매핑하고 세율/쿼터/환급 가능성을 산출해 관세비용 영향표에 반영하십시오."
    if issue in {"AD/CVD", "반덤핑/상계관세"}:
        return f"{subs}: 대상 품목·공급국·벤더를 확인하고 조사대응 자료, 원산지 증빙, 가격자료 보관 필요성을 점검하십시오."
    if issue == "CBAM":
        return f"{subs}: EU향 품목, 공급사 배출량 자료, CBAM 신고·인증서 비용 반영 여부를 ESG/구매/통관 담당과 확인하십시오."
    if issue == "FTA/원산지":
        return f"{subs}: BOM 원산지, CO 발급, 직접운송, 누적기준, 특혜세율 적용 가능성을 FTA 마스터와 대조하십시오."
    if issue == "수출통제":
        return f"{subs}: ECCN/전략물자 분류, 거래처·최종사용자 스크리닝, 제재국 우회거래 가능성을 재점검하십시오."
    if issue == "통관":
        return f"{subs}: 통관 SOP, 보세/신고 체크리스트, 관세사 안내문, 세관 제출자료 양식을 업데이트하십시오."
    if issue == "HS/품목분류":
        return f"{subs}: 관련 제품의 HS 설명, 판정 근거, 해외법인 신고코드와 한국 본사 마스터 간 차이를 점검하십시오."
    return f"{subs}: 원문 기준으로 대상 국가, 품목, 시행일, 담당 부서를 확인하고 후속 모니터링하십시오."


def html_link(title: str, url: str) -> str:
    title_e = html.escape(clean(title))
    url = best_url_from_values([url])
    if not url:
        return title_e
    return f'<a href="{html.escape(url)}" target="_blank">{title_e}</a>'


def risk_color(risk: str) -> str:
    return {"상": "#C00000", "중": "#C55A11", "하": "#4472C4"}.get(normalize_risk(risk), "#555")


def short_text(value, fallback: str, limit: int = 360) -> str:
    text = clean(value) or fallback
    return text[:limit] + ("..." if len(text) > limit else "")


def one_line(row: pd.Series) -> str:
    return f"{clean(row.get('Issue'))} / {clean(row.get('Country')) or '-'} / {clean(row.get('Samsung Impact'))}: {short_text(row.get('Major Changes'), '주요 변경내역 확인 필요', 130)}"


def top3_summary_sentence(row: pd.Series) -> str:
    issue = clean(row.get("Issue"))
    if issue in {"AD/CVD", "반덤핑/상계관세"}:
        return "반덤핑·상계관세 조사는 대상 품목의 HS, 원산지, 가격자료 방어체계를 우선 점검해야 합니다."
    if issue == "FTA/원산지":
        return "FTA·원산지 변경은 관련 법인의 CO 발급요건과 특혜세율 적용 가능성을 재검토해야 합니다."
    if issue in {"관세정책", "통관", "통관/세관", "HS/품목분류"}:
        return "관세·통상 정책 변화는 관련 품목의 HS, 원산지, 관세율 영향을 확인해야 합니다."
    if issue == "CBAM":
        return "CBAM 변화는 EU향 품목의 배출량 자료, 인증서 비용, 신고의무 반영 여부를 우선 점검해야 합니다."
    if issue == "수출통제":
        return "수출통제 변화는 ECCN·전략물자 분류와 거래처·최종사용자 스크리닝 체계를 우선 확인해야 합니다."
    return short_text(row.get("Major Changes"), "관세·통상 영향 여부를 원문 기준으로 확인해야 합니다.", 150)


def top3_summary_rows(rows: pd.DataFrame, top3: pd.DataFrame) -> pd.DataFrame:
    """Issue-level summary rows for the executive summary.

    The executive summary should be thematic and practical, not just repeat the
    three article titles. Prefer AD/CVD, FTA/origin, tariff/customs if present.
    """
    preferred_groups = [
        {"AD/CVD", "반덤핑/상계관세"},
        {"FTA/원산지", "ORIGIN", "원산지"},
        {"관세정책", "통관", "통관/세관", "HS/품목분류"},
        {"CBAM"},
        {"수출통제"},
    ]
    selected = []
    used = set()
    for group in preferred_groups:
        cand = rows[rows["Issue"].astype(str).isin(group)].copy()
        if cand.empty:
            continue
        cand = cand.sort_values(["_report_score", "_sort_date"], ascending=[False, False])
        row = cand.iloc[0]
        key = clean(row.get("Issue"))
        if key in used:
            continue
        selected.append(row)
        used.add(key)
        if len(selected) >= 3:
            break
    if len(selected) < 3:
        for _, row in top3.iterrows():
            key = clean(row.get("Issue"))
            if key in used:
                continue
            selected.append(row)
            used.add(key)
            if len(selected) >= 3:
                break
    return pd.DataFrame(selected)


def overall_html(rows: pd.DataFrame, top3: pd.DataFrame) -> str:
    reg = rows[rows["Content Type"].eq("Regulation")]
    news = rows[rows["Content Type"].eq("News")]
    direct = rows[rows["Samsung Impact"].eq("Direct")]
    indirect = rows[rows["Samsung Impact"].eq("Indirect")]
    watch = rows[rows["Samsung Impact"].eq("Watch")]
    issues = rows["Issue"].value_counts().head(6)
    issue_txt = ", ".join(f"{k} {v}건" for k, v in issues.items())
    summary_rows = top3_summary_rows(rows, top3)
    top_lines = "".join(f"<li>{html.escape(top3_summary_sentence(r))}</li>" for _, r in summary_rows.iterrows())
    return f"""
    <div style="padding:15px;background:#F4F6F8;border-left:6px solid #1F4E78;margin-bottom:18px;">
      <div style="font-size:14px;color:#555;margin-bottom:8px;">
        금일 선별 결과: 법규 {len(reg)}건, 주요뉴스 {len(news)}건 | Direct {len(direct)}건, Indirect {len(indirect)}건, Watch {len(watch)}건
      </div>
      <div style="font-size:15px;font-weight:bold;line-height:1.8;margin-bottom:8px;">
        금일 GTI Radar는 {html.escape(issue_txt)} 중심으로 관세·통상 변화가 포착되었습니다. 법규는 시행일·HS·세율·신고절차 반영 여부를, 뉴스는 실제 비용·원산지·수출통제 영향 가능성을 우선 확인해야 합니다.
      </div>
      <div style="margin-top:8px;"><b>Top3 요약</b><ol style="margin-top:6px;">{top_lines}</ol></div>
    </div>
    """


def top3_html(top3: pd.DataFrame) -> str:
    blocks = []
    for idx, row in top3.iterrows():
        blocks.append(f"""
        <div style="margin:14px 0 18px 0;padding:15px;border-left:5px solid #C00000;background:#FFF7F7;">
          <div style="font-size:15px;font-weight:bold;margin-bottom:6px;">Top {idx + 1}. {html_link(row.get('Headline'), row.get('URL'))}</div>
          <div style="font-size:12px;color:#555;margin-bottom:9px;">
            Type: {html.escape(clean(row.get('Content Type')))} | Topic: {html.escape(clean(row.get('Issue')))} |
            Samsung Impact: <b>{html.escape(clean(row.get('Samsung Impact')))}</b> |
            Subsidiary: {html.escape(clean(row.get('Affected Subsidiary')) or 'SEC/HQ')} |
            Agency: {html.escape(clean(row.get('Agency')))} | Publish Date: {html.escape(clean(row.get('Date')))} |
            Country: {html.escape(clean(row.get('Country')))} |
            Risk: <span style="color:{risk_color(row.get('Risk'))};font-weight:bold;">{html.escape(clean(row.get('Risk')))}</span> |
            Score: {safe_num(row.get('Importance Score')):.0f}
          </div>
          <div style="margin-top:8px;"><b>Executive Impact</b><br>{html.escape(one_line(row))}</div>
          <div style="margin-top:8px;"><b>주요 변경내역</b><br>{html.escape(short_text(row.get('Major Changes'), '주요 변경내역 확인 필요', 520))}</div>
          <div style="margin-top:8px;"><b>삼성 영향</b><br>{html.escape(short_text(row.get('AI Analysis'), '삼성 영향 검토 필요', 520))}</div>
          <div style="margin-top:8px;"><b>Action</b><br>{html.escape(short_text(row.get('Action Plan'), '담당 부서 확인 필요', 520))}</div>
        </div>
        """)
    return "".join(blocks)


def table_html(title: str, rows: pd.DataFrame, color: str) -> str:
    if rows.empty:
        return f"<h3 style='color:{color};'>{html.escape(title)} (0건)</h3>"
    trs = []
    for _, row in rows.iterrows():
        trs.append(f"""
        <tr>
          <td style="padding:7px;border:1px solid #ddd;text-align:center;vertical-align:top;">{html.escape(str(row.get('No')))}</td>
          <td style="padding:7px;border:1px solid #ddd;text-align:center;vertical-align:top;">{html.escape(clean(row.get('Issue')))}</td>
          <td style="padding:7px;border:1px solid #ddd;vertical-align:top;">{html_link(row.get('Headline'), row.get('URL'))}</td>
          <td style="padding:7px;border:1px solid #ddd;vertical-align:top;">{html.escape(short_text(row.get('Major Changes'), '주요 변경내역 확인 필요', 260))}</td>
          <td style="padding:7px;border:1px solid #ddd;vertical-align:top;">{html.escape(short_text(row.get('AI Analysis'), '영향 검토 필요', 260))}</td>
          <td style="padding:7px;border:1px solid #ddd;vertical-align:top;">{html.escape(short_text(row.get('Action Plan'), '담당 부서 확인 필요', 260))}</td>
          <td style="padding:7px;border:1px solid #ddd;text-align:center;vertical-align:top;">{html.escape(clean(row.get('Country')))}</td>
          <td style="padding:7px;border:1px solid #ddd;text-align:center;vertical-align:top;color:{risk_color(row.get('Risk'))};font-weight:bold;">{html.escape(clean(row.get('Risk')))}</td>
          <td style="padding:7px;border:1px solid #ddd;text-align:center;vertical-align:top;">{html.escape(clean(row.get('Samsung Impact')))}</td>
        </tr>
        """)
    return f"""
    <h3 style="margin-top:24px;color:{color};">{html.escape(title)} ({len(rows)}건)</h3>
    <table style="border-collapse:collapse;width:100%;font-size:12px;table-layout:fixed;">
      <colgroup>
        <col style="width:3%;"><col style="width:8%;"><col style="width:22%;">
        <col style="width:22%;"><col style="width:18%;"><col style="width:18%;">
        <col style="width:5%;"><col style="width:4%;"><col style="width:5%;">
      </colgroup>
      <thead>
        <tr style="background:{color};color:white;">
          <th style="padding:7px;border:1px solid #ddd;">No</th>
          <th style="padding:7px;border:1px solid #ddd;">Issue</th>
          <th style="padding:7px;border:1px solid #ddd;">Headline</th>
          <th style="padding:7px;border:1px solid #ddd;">주요 변경내역</th>
          <th style="padding:7px;border:1px solid #ddd;">삼성 영향</th>
          <th style="padding:7px;border:1px solid #ddd;">Action</th>
          <th style="padding:7px;border:1px solid #ddd;">Country</th>
          <th style="padding:7px;border:1px solid #ddd;">Risk</th>
          <th style="padding:7px;border:1px solid #ddd;">Impact</th>
        </tr>
      </thead>
      <tbody>{''.join(trs)}</tbody>
    </table>
    """


def build_html(rows: pd.DataFrame, top3: pd.DataFrame) -> str:
    subject = f"[GTI Radar] Global Trade Intelligence | {RUN_DATE}"
    regulation = rows[rows["Content Type"].eq("Regulation")]
    news = rows[rows["Content Type"].eq("News")]
    return f"""<!DOCTYPE html>
<html lang="ko">
<head><meta charset="utf-8"><title>{html.escape(subject)}</title></head>
<body style="font-family:Arial,'Malgun Gothic',sans-serif;font-size:13px;color:#222;line-height:1.55;">
  <div style="max-width:1320px;margin:0 auto;">
    <h2 style="margin-bottom:3px;color:#1F4E78;">[GTI Radar] Global Trade Intelligence</h2>
    <div style="font-size:13px;color:#555;margin-bottom:16px;">{RUN_DATE} | Samsung Electronics Customs & Trade Intelligence</div>

    <h3 style="margin-top:18px;margin-bottom:6px;">1. 총평</h3>
    {overall_html(rows, top3)}

    <h3 style="margin-top:22px;color:#C00000;">2. Top3 Deep Analysis</h3>
    {top3_html(top3)}

    {table_html('3. Regulation', regulation, '#1F4E78')}
    {table_html('4. 주요뉴스', news, '#548235')}

    <p style="margin-top:18px;color:#666;font-size:12px;">첨부 Excel에는 전체 선별 결과와 원문 링크가 포함되어 있습니다.</p>
  </div>
</body>
</html>"""


def style_sheet(ws) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border
    widths = {
        "A": 5, "B": 13, "C": 13, "D": 13, "E": 18, "F": 30, "G": 15, "H": 48,
        "I": 46, "J": 46, "K": 46, "L": 46, "M": 14, "N": 18, "O": 8, "P": 12,
        "Q": 16, "R": 16, "S": 24, "T": 36, "U": 24, "V": 28,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A2"


def append_output_row(ws, row: pd.Series) -> None:
    ws.append([row.get(c, "") for c in OUTPUT_COLUMNS])
    headline_col = OUTPUT_COLUMNS.index("Headline") + 1
    cell = ws.cell(row=ws.max_row, column=headline_col)
    url = best_url_from_values([row.get("URL")])
    if url:
        cell.hyperlink = url
        cell.font = Font(color="0563C1", underline="single", bold=True)


def save_excel(rows: pd.DataFrame, top3: pd.DataFrame, paths: dict[str, Path]) -> None:
    wb = Workbook()
    sheets = [
        ("GTI Radar", rows),
        ("Top3 Deep Analysis", top3),
        ("Regulation", rows[rows["Content Type"].eq("Regulation")]),
        ("주요뉴스", rows[rows["Content Type"].eq("News")]),
    ]
    first = True
    for name, frame in sheets:
        ws = wb.active if first else wb.create_sheet(name[:31])
        first = False
        ws.title = name[:31]
        ws.append(OUTPUT_COLUMNS)
        for _, row in frame.iterrows():
            append_output_row(ws, row)
        style_sheet(ws)

    runlog = wb.create_sheet("Run Log")
    runlog.append(["item", "value"])
    runlog.append(["regulation_input", str(REGULATION_INPUT_FILE)])
    runlog.append(["news_input", str(NEWS_INPUT_FILE)])
    runlog.append(["run_date", RUN_DATE])
    runlog.append(["total_rows", len(rows)])
    runlog.append(["regulation_rows", int(rows["Content Type"].eq("Regulation").sum())])
    runlog.append(["news_rows", int(rows["Content Type"].eq("News").sum())])
    runlog.append(["direct_rows", int(rows["Samsung Impact"].eq("Direct").sum())])
    runlog.append(["indirect_rows", int(rows["Samsung Impact"].eq("Indirect").sum())])
    runlog.append(["watch_rows", int(rows["Samsung Impact"].eq("Watch").sum())])
    style_sheet(runlog)

    wb.save(paths["mail_xlsx"])
    wb.save(paths["analysis"])
    rows[OUTPUT_COLUMNS].to_excel(paths["cumulative"], index=False)


def read_recipients() -> list[str]:
    recipients = []
    if MAIL_TO:
        recipients.extend([x.strip() for x in re.split(r"[;,]", MAIL_TO) if x.strip()])
    if RECIPIENT_FILE.exists():
        try:
            df = pd.read_excel(RECIPIENT_FILE)
            for col in df.columns:
                for value in df[col].dropna().astype(str):
                    text = clean(value)
                    if re.match(r"^[^@\s]+@[^@\s]+\.[^@\s]+$", text):
                        recipients.append(text)
        except Exception:
            pass
    seen, out = set(), []
    for email in recipients:
        low = email.lower()
        if low not in seen:
            seen.add(low)
            out.append(email)
    return out


def send_email(html_body: str, attachment: Path) -> None:
    if not SEND_EMAIL:
        print("[MAIL SKIP] GTI_SEND_EMAIL=N or --no-email")
        return
    recipients = read_recipients()
    if not recipients:
        print("[MAIL SKIP] recipients missing")
        return
    if not SMTP_USER or not SMTP_PASS:
        print("[MAIL SKIP] SMTP credential missing")
        return

    msg = EmailMessage()
    msg["Subject"] = f"[GTI Radar] Global Trade Intelligence({RUN_DATE})"
    msg["From"] = formataddr((MAIL_FROM_NAME, SMTP_USER))
    msg["To"] = ", ".join(recipients)
    msg.set_content("GTI Radar report is attached. HTML mail requires an HTML-capable client.")
    msg.add_alternative(html_body, subtype="html")
    data = attachment.read_bytes()
    msg.add_attachment(
        data,
        maintype="application",
        subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=attachment.name,
    )

    context = ssl.create_default_context()
    with smtplib.SMTP_SSL(SMTP_HOST, SMTP_PORT, context=context, timeout=30) as server:
        server.login(SMTP_USER, SMTP_PASS)
        server.send_message(msg)
    print(f"[MAIL SENT] {len(recipients)} recipients")


def main() -> None:
    paths = output_paths()
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    rows = prepare_rows(read_step4_results())
    top3 = choose_top3(rows)
    html_body = build_html(rows, top3)
    save_excel(rows, top3, paths)
    paths["mail_html"].write_text(html_body, encoding="utf-8")
    send_email(html_body, paths["mail_xlsx"])

    reg_n = int(rows["Content Type"].eq("Regulation").sum())
    news_n = int(rows["Content Type"].eq("News").sum())
    print(f"[DONE] HTML: {paths['mail_html']}")
    print(f"[DONE] XLSX: {paths['mail_xlsx']}")
    print(
        f"[ROWS] total={len(rows)}, regulation={reg_n}, news={news_n}, "
        f"direct={(rows['Samsung Impact'] == 'Direct').sum()}, "
        f"indirect={(rows['Samsung Impact'] == 'Indirect').sum()}, "
        f"watch={(rows['Samsung Impact'] == 'Watch').sum()}"
    )


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--date", default=None)
    parser.add_argument("--regulation-input", default=None)
    parser.add_argument("--news-input", default=None)
    parser.add_argument("--output-dir", default=None)
    parser.add_argument("--no-email", action="store_true")
    args = parser.parse_args()
    if args.date:
        RUN_DATE = args.date
    if args.regulation_input:
        REGULATION_INPUT_FILE = Path(args.regulation_input)
    if args.news_input:
        NEWS_INPUT_FILE = Path(args.news_input)
    if args.output_dir:
        OUTPUT_DIR = Path(args.output_dir)
    if args.no_email:
        SEND_EMAIL = False
    main()
