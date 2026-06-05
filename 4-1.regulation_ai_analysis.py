# =========================================================
# GTI STEP4-1 Regulation AI Analysis - Regulation Dedicated Prompt v4.1
# INPUT  : C:/Temp/3-1.regulation_article_summary.xlsx preferred; fallback C:/Temp/3-1.regulation_summary.xlsx
# DAILY  : C:/Temp/4-1.regulation_ai_summary.xlsx
# CUMUL  : C:/Temp/4-1.regulation_ai_cumulative.xlsx
#
# Operating rule
# - 3-1 remains the official regulation source.
# - 4-1 is regulation-only: it reads 3-1 output only and never imports 3-2 news.
# - HR/recruitment/personnel, statistics/history, drug/smuggling enforcement,
#   and generic administrative sanctions are excluded unless there is direct
#   customs/trade-law substance.
# - Regulation output has no Top-N cap.
# - Cumulative is re-cleaned by the current rules before saving.
# =========================================================

import json
import os
import re
from datetime import datetime
from urllib.parse import urlparse

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


REG_INPUT = os.getenv("GTI_REGULATION_INPUT", r"C:/Temp/3-1.regulation_summary.xlsx")
REG_ARTICLE_INPUT = os.getenv("GTI_REGULATION_ARTICLE_INPUT", r"C:/Temp/3-1.regulation_article_summary.xlsx")
OUTPUT_DAILY = os.getenv("GTI_REGULATION_OUTPUT", r"C:/Temp/4-1.regulation_ai_summary.xlsx")
OUTPUT_CUMUL = os.getenv("GTI_REGULATION_CUMULATIVE", r"C:/Temp/4-1.regulation_ai_cumulative.xlsx")
OUTPUT_EXCLUDED = os.getenv("GTI_REGULATION_EXCLUDED", r"C:/Temp/4-1.regulation_ai_excluded.xlsx")

GEMINI_API_KEY = os.getenv("GEMINI_API_KEY", "").strip()
GEMINI_MODEL = os.getenv("GEMINI_MODEL", "gemini-2.0-flash")
USE_AI = bool(GEMINI_API_KEY)
MISSING_TEXT = "본문에서 확인 불가"

client = None
if USE_AI:
    try:
        from google import genai

        client = genai.Client(api_key=GEMINI_API_KEY)
    except Exception:
        USE_AI = False
        client = None


CUSTOMS_LAW_KEYWORDS = [
    "관세", "관세율", "할당관세", "덤핑방지관세", "반덤핑", "상계관세",
    "통관", "수입신고", "수출신고", "보세", "관세환급", "세관장확인",
    "품목분류", "HS", "HS CODE", "원산지", "FTA", "CEPA", "EPA",
    "수출통제", "전략물자", "제재", "CBAM", "탄소국경", "Section 301",
    "Section 232", "USTR", "CBP", "customs", "tariff", "duty",
    "import duty", "anti-dumping", "antidumping", "countervailing",
    "drawback", "rules of origin", "export control", "sanctions",
]

REGULATION_ACTION_KEYWORDS = [
    "법률", "법령", "시행령", "시행규칙", "행정규칙", "대통령령", "부령",
    "규칙", "고시", "공고", "관보", "입법예고", "행정예고", "개정",
    "일부개정", "시행", "공포", "제정", "폐지", "시행일", "적용",
    "등록 시스템", "가산세", "부과 대상", "notice", "regulation",
    "rule", "law", "decree", "ordinance", "amendment", "final rule",
    "proposed rule", "effective date", "implementation",
]

HR_EXCLUDE = [
    "채용", "인턴", "합격자", "경력경쟁", "공모직위", "공개모집", "임용",
    "인사", "후보", "승진", "면접", "서류전형", "시험공고", "recruit",
    "recruitment", "hiring", "vacancy", "career", "personnel",
    "appointment", "intern", "job opening",
]

ADMIN_EXCLUDE = [
    "공시송달", "제재처분 통지", "연구개발과제", "부정당업자", "입찰참가자격",
    "과징금", "과태료", "행정처분", "처분사전통지", "송달불능",
]

LOW_VALUE_NOTICE_EXCLUDE = [
    "마약", "밀수", "코카인", "펜타닐", "narcotic", "drug", "smuggling",
    "집중검사", "적발", "recruitment event", "beginnings to today",
    "history", "trade statistics", "provisional", "통계", "워크숍", "선발",
    "행사", "교육", "홍보",
]

OFFICIAL_NOTICE_KEEP_SIGNALS = [
    "수입신고", "가산세", "할당관세", "개정 관세법", "전자상거래업자 등록",
    "등록 시스템", "원산지", "품목분류 고시", "수출통제", "전략물자",
    "반덤핑", "상계관세", "덤핑방지관세", "관세환급", "세관장확인",
    "시행규칙", "시행령", "고시", "공고", "입법예고", "행정예고",
]

KEEP_KEYWORDS = [
    "덤핑방지관세", "반덤핑", "상계관세", "AD/CVD", "anti-dumping",
    "antidumping", "countervailing", "세관장확인", "통합공고",
    "할당관세", "FTA", "원산지", "품목분류", "HS", "관세환급",
    "수입신고", "수출신고", "수출통제", "전략물자",
]

ADMIN_KEEP_IF_ALSO = [
    "관세", "통관", "수입신고", "수출신고", "원산지", "FTA", "CEPA", "EPA",
    "품목분류", "HS", "수출통제", "전략물자", "CBAM", "반덤핑", "상계관세",
    "덤핑방지관세", "관세환급", "할당관세", "세관장확인",
]

SAMSUNG_KEYWORDS = [
    "samsung", "삼성", "semiconductor", "chip", "반도체", "hbm", "dram",
    "nand", "smartphone", "mobile", "galaxy", "display", "battery",
    "electronics", "network", "server", "vietnam", "베트남", "india",
    "인도", "mexico", "멕시코", "china", "중국", "korea", "한국",
]

OUTPUT_COLUMNS = [
    "Rank", "Date", "Headline", "URL", "Country", "Agency", "ResponsibleAgency",
    "source_type", "regulation_related", "regulation_transfer_type",
    "priority_group", "selected", "must_include", "exclude_reason", "Risk",
    "Score", "regulation_type", "issue_type", "cluster_key", "impact_area",
    "samsung_relevance", "samsung_score", "samsung_reason", "Summary",
    "AI Analysis", "Action Plan",
    "effective_date", "hs_code", "old_tariff_rate", "new_tariff_rate", "tariff_rate", "target_country", "affected_products",
    "effective_date_hint", "hs_hint", "tariff_rate_hint", "change_detail_hint",
    "compliance_tasks", "KeywordMatches", "FilterReason", "Source",
    "last_checked",
]


def safe_str(value):
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except Exception:
        pass
    return str(value).strip()


def norm(value):
    text = safe_str(value).lower()
    text = re.sub(r"https?://\S+", " ", text)
    text = re.sub(r"[^0-9a-z가-힣/\-\s]", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def contains_any(text, keywords):
    n = norm(text)
    return any(norm(k) in n for k in keywords)


def trim(value, limit=700):
    text = safe_str(value).replace("**", "").replace("##", "")
    return re.sub(r"\s+", " ", text).strip()[:limit].strip()


def first_nonempty(*values):
    for value in values:
        text = safe_str(value)
        if text:
            return text
    return ""


def extract_hs_codes(*texts):
    joined = " ".join(safe_str(t) for t in texts)
    candidates = []

    # Prefer explicit HS context to avoid collecting phone numbers, law numbers, or addresses.
    for m in re.finditer(r"(HS|H\.S\.|품목분류|세번|관세번호|tariff\s*classification|tariff\s*line|HTS|CN\s*code).{0,40}?(?<![\d-])(\d{4}(?:\.\d{2}){0,3}|\d{6}|\d{8}|\d{10})(?![\d-])", joined, flags=re.I | re.S):
        candidates.append(m.group(2))

    # Also accept tariff-style dotted codes even without an HS prefix.
    candidates.extend(re.findall(r"(?<!\d)(\d{4}\.\d{2}(?:\.\d{2}){0,2})(?!\d)", joined))

    cleaned = []
    for code in candidates:
        digits = re.sub(r"\D", "", code)
        if 4 <= len(digits) <= 10 and digits not in cleaned:
            cleaned.append(digits)
    return "; ".join(cleaned[:10])


def extract_tariff_rates(*texts):
    joined = " ".join(safe_str(t) for t in texts)
    found = re.findall(r"\b\d+(?:\.\d+)?\s*%", joined)
    return "; ".join(dict.fromkeys(found).keys())


def extract_old_new_tariff_rates(*texts):
    joined = " ".join(safe_str(t) for t in texts)
    old_rate = ""
    new_rate = ""

    patterns = [
        (r"(?:기존|현행|종전|previous|current)[^0-9%]{0,30}(\d+(?:\.\d+)?\s*%)", "old"),
        (r"(?:변경|개정|신설|인하|인상|적용|new|revised|changed)[^0-9%]{0,30}(\d+(?:\.\d+)?\s*%)", "new"),
    ]
    for pat, kind in patterns:
        m = re.search(pat, joined, flags=re.I)
        if m and kind == "old":
            old_rate = m.group(1)
        if m and kind == "new":
            new_rate = m.group(1)

    arrow = re.search(r"(\d+(?:\.\d+)?\s*%)\s*(?:→|->|에서|에서\s*)\s*(\d+(?:\.\d+)?\s*%)", joined)
    if arrow:
        old_rate = old_rate or arrow.group(1)
        new_rate = new_rate or arrow.group(2)

    rates = [r.strip() for r in extract_tariff_rates(joined).split(";") if r.strip()]
    if len(rates) >= 2:
        old_rate = old_rate or rates[0]
        new_rate = new_rate or rates[1]
    elif len(rates) == 1:
        new_rate = new_rate or rates[0]

    return old_rate, new_rate


def normalize_blank(value):
    text = safe_str(value)
    if not text:
        return ""
    lowered = text.lower().strip()
    if lowered in {"n/a", "na", "none", "null", "-", "없음", "해당없음", "미상", "불명", "확인필요", "원문 확인 필요", "본문에서 확인 불가"}:
        return ""
    return text


def clean_regulation_text(value, fallback=""):
    text = safe_str(value) or safe_str(fallback)
    banned_patterns = [
        r"원문\s*확인\s*필요(?:합니다|함|하십시오| 필요)?",
        r"원문을\s*확인(?:해야|하십시오| 필요)",
        r"담당\s*부서\s*확인\s*필요(?:합니다|함)?",
        r"추가\s*확인\s*필요(?:합니다|함)?",
    ]
    for pat in banned_patterns:
        text = re.sub(pat, MISSING_TEXT, text, flags=re.I)
    text = re.sub(r"\s+", " ", text).strip(" ;,./")
    return text or MISSING_TEXT


def missing_if_blank(value):
    return normalize_blank(value) or MISSING_TEXT


def compact_join(items, sep="; "):
    cleaned = []
    for item in items:
        text = normalize_blank(item)
        if text and text not in cleaned:
            cleaned.append(text)
    return sep.join(cleaned)


def derive_target_country(row):
    title = safe_str(row.get("headline"))
    country = safe_str(row.get("country"))
    agency = safe_str(row.get("agency"))
    hints = safe_str(row.get("change_detail_hint"))
    body = safe_str(row.get("article_body"))[:2000]
    text = " ".join([title, country, agency, hints, body])

    country_aliases = [
        ("사우디아라비아", ["사우디아라비아", "사우디", "Saudi Arabia", "Saudi"]),
        ("중국", ["중국", "China", "Chinese"]),
        ("미국", ["미국", "United States", "U.S.", "USA", "US "]),
        ("EU", ["EU", "European Union", "유럽연합"]),
        ("인도", ["인도", "India"]),
        ("베트남", ["베트남", "Vietnam", "Viet Nam"]),
        ("멕시코", ["멕시코", "Mexico"]),
        ("튀르키예", ["튀르키예", "터키", "Türkiye", "Turkey"]),
        ("네덜란드", ["네덜란드", "Netherlands"]),
        ("한국", ["한국", "대한민국", "Korea", "Republic of Korea"]),
    ]
    found = []
    for canonical, aliases in country_aliases:
        if any(alias in text for alias in aliases):
            found.append(canonical)
    if found:
        return "; ".join(dict.fromkeys(found).keys())
    return country


def derive_affected_products(row):
    hints = hint_context(row)
    title = safe_str(row.get("headline"))
    candidates = []
    if hints["change_detail_hint"] and not contains_any(hints["change_detail_hint"], ["할당관세의 적용에 관한 규정", "할당관세의적용", "관세법 제71조", "대통령령", "시행령", "시행규칙"]):
        candidates.append(hints["change_detail_hint"])
    # Common Korean legal title pattern: "OO에 대한 ..."
    m = re.search(r"(.{2,80}?)에 대한", title)
    if m:
        product = m.group(1)
        product = re.sub(r"^(규칙|고시|공고|대통령령|기획재정부령|관세청고시)\s*", "", product).strip()
        if product and not contains_any(product, ["관세법 제71조", "시행령", "시행규칙", "할당관세의 적용", "할당관세", "규정", "대통령령"]):
            candidates.append(product)
    return trim(compact_join(candidates), 300)


def infer_regulation_summary(row, issue):
    title = safe_str(row.get("headline"))
    effective = normalize_blank(row.get("effective_date")) or safe_str(row.get("effective_date_hint"))
    hs = normalize_blank(row.get("hs_code")) or safe_str(row.get("hs_hint"))
    rate = normalize_blank(row.get("tariff_rate")) or safe_str(row.get("tariff_rate_hint"))
    country = normalize_blank(row.get("target_country")) or derive_target_country(row)
    products = normalize_blank(row.get("affected_products")) or derive_affected_products(row)

    parts = []
    if effective:
        parts.append(f"시행일 {effective}")
    if hs:
        parts.append(f"HS {hs}")
    if rate:
        parts.append(f"관세율 {rate}")
    if country:
        parts.append(f"대상국 {country}")
    if products:
        parts.append(f"대상품목 {products}")
    detail = "; ".join(parts) if parts else "시행일·HS·관세율·대상국·대상품목 구조화 대상"
    return f"{title} 건은 {issue} 법규 변경사항입니다. {detail}."


def build_regulation_analysis(row):
    issue = classify_issue(combined(row))
    effective = normalize_blank(row.get("effective_date")) or safe_str(row.get("effective_date_hint"))
    hs = normalize_blank(row.get("hs_code")) or safe_str(row.get("hs_hint"))
    rate = normalize_blank(row.get("tariff_rate")) or safe_str(row.get("tariff_rate_hint"))
    country = normalize_blank(row.get("target_country")) or derive_target_country(row)
    products = normalize_blank(row.get("affected_products")) or derive_affected_products(row)

    if issue == "TARIFF_DUTY":
        base = "할당관세·기본관세·잠정세율 등 수입세율 적용 기준이 변경되는 사안입니다."
    elif issue == "AD_CVD":
        base = "반덤핑·상계관세 부과대상, 부과기간, 공급자별 세율 또는 적용품목이 변경되는 사안입니다."
    elif issue == "HS_CLASSIFICATION":
        base = "품목분류 기준 또는 HS 적용 해석이 변경되어 신고 HS Master와 품목별 세율 판단에 영향을 줄 수 있습니다."
    elif issue == "FTA_ORIGIN":
        base = "원산지 판정, 협정세율 적용 또는 증빙관리 기준이 변경되는 사안입니다."
    else:
        base = "관세·통관 컴플라이언스 운영 기준이 변경되는 공식 법규/고시 사안입니다."

    facts = compact_join([
        f"시행일={effective}" if effective else "",
        f"HS={hs}" if hs else "",
        f"관세율={rate}" if rate else "",
        f"대상국={country}" if country else "",
        f"대상품목={products}" if products else "",
    ])
    return f"{base} {facts}".strip()


def build_regulation_action_plan(row):
    issue = classify_issue(combined(row))
    tasks = []
    if normalize_blank(row.get("effective_date")) or safe_str(row.get("effective_date_hint")):
        tasks.append("시행일 기준으로 적용 시작일·선적일·수입신고일 기준을 캘린더화")
    if normalize_blank(row.get("hs_code")) or safe_str(row.get("hs_hint")):
        tasks.append("HS Master와 글로벌 수입신고 이력에서 해당 HS 사용 법인 추출")
    if normalize_blank(row.get("tariff_rate")) or safe_str(row.get("tariff_rate_hint")):
        tasks.append("변경 관세율을 관세비용 산정 로직과 신고 검증표에 반영")
    if issue == "AD_CVD":
        tasks.append("공급자·원산국·품목 기준 AD/CVD 적용대상 거래를 분리 점검")
    elif issue == "TARIFF_DUTY":
        tasks.append("할당관세 적용 가능 품목은 수입계획·Quota 잔량·세율 적용 조건을 점검")
    elif issue == "HS_CLASSIFICATION":
        tasks.append("품목분류 사전심사·국가별 HS 불일치 후보를 재점검")
    else:
        tasks.append("대상 법인·관세사에 변경사항을 배포하고 신고 체크리스트를 업데이트")
    return "; ".join(tasks[:5])


def hint_context(row):
    return {
        "effective_date_hint": safe_str(row.get("effective_date_hint")),
        "hs_hint": safe_str(row.get("hs_hint")),
        "tariff_rate_hint": safe_str(row.get("tariff_rate_hint")),
        "change_detail_hint": safe_str(row.get("change_detail_hint")),
        "article_body": trim(row.get("article_body"), 3000),
    }


def parse_json(text):
    text = safe_str(text)
    start = text.find("{")
    end = text.rfind("}") + 1
    if start < 0 or end <= start:
        return {}
    try:
        return json.loads(text[start:end])
    except Exception:
        return {}


def pick(df, *names):
    lookup = {str(c).strip().lower(): c for c in df.columns}
    for name in names:
        if name.lower() in lookup:
            return lookup[name.lower()]
    return None


def recover_headline_url(row):
    headline = safe_str(row.get("headline"))
    url = safe_str(row.get("url"))
    m = re.search(r'=HYPERLINK\("([^"]+)","([^"]+)"\)', headline, re.I)
    if m:
        url = url or m.group(1)
        headline = m.group(2)
    if not headline and url.startswith("http"):
        last = urlparse(url).path.strip("/").split("/")[-1]
        headline = last[:100] if last else url[:100]
    return pd.Series({"headline": headline, "url": url})


def normalize_regulation(df):
    mapping = {
        "date": pick(df, "Date", "date", "published"),
        "headline": pick(df, "Headline", "headline", "title"),
        "url": pick(df, "URL", "url", "link"),
        "country": pick(df, "Country", "country"),
        "agency": pick(df, "Agency", "agency", "source", "publisher"),
        "responsible_agency": pick(df, "ResponsibleAgency", "responsible_agency"),
        "score": pick(df, "Score", "score"),
        "keyword_matches": pick(df, "KeywordMatches", "keyword_matches", "keyword"),
        "filter_reason": pick(df, "FilterReason", "filter_reason", "SelectReason"),
        "source": pick(df, "Source", "source", "SourceFile", "source_file"),
        "original_url": pick(df, "original_url", "OriginalURL", "canonical_url"),
        "article_body": pick(df, "article_body", "ArticleBody", "body", "content"),
        "article_extract_status": pick(df, "article_extract_status", "ArticleExtractStatus"),
        "article_source_type": pick(df, "article_source_type", "ArticleSourceType"),
        "effective_date_hint": pick(df, "effective_date_hint", "EffectiveDateHint"),
        "change_detail_hint": pick(df, "change_detail_hint", "ChangeDetailHint"),
        "hs_hint": pick(df, "hs_hint", "HsHint", "HS_Hint"),
        "tariff_rate_hint": pick(df, "tariff_rate_hint", "TariffRateHint"),
    }
    out = pd.DataFrame()
    for target, source in mapping.items():
        out[target] = df[source] if source in df.columns else ""
    recovered = out.apply(recover_headline_url, axis=1)
    out["headline"], out["url"] = recovered["headline"], recovered["url"]
    out["source_type"] = "REGULATION_ORIGINAL"
    out["regulation_related"] = "Y"
    out["regulation_transfer_type"] = "Original"
    return out


def combined(row):
    return " ".join(
        safe_str(row.get(c))
        for c in [
            "headline", "country", "agency", "responsible_agency",
            "keyword_matches", "filter_reason", "source",
            "effective_date_hint", "hs_hint", "tariff_rate_hint", "change_detail_hint", "article_body",
        ]
    )


def exclusion_reason(row):
    text = combined(row)

    # GTI Core Regulation must be kept even if STEP3 body extraction is weak.
    if contains_any(text, KEEP_KEYWORDS):
        return ""

    if contains_any(text, HR_EXCLUDE):
        return "HR_RECRUITMENT_PERSONNEL"
    if contains_any(text, ADMIN_EXCLUDE) and not contains_any(text, ADMIN_KEEP_IF_ALSO):
        return "GENERIC_ADMIN_SANCTION"
    if not contains_any(text, CUSTOMS_LAW_KEYWORDS):
        return "NOT_CUSTOMS_TRADE_REGULATION"
    return ""


def must_include(row):
    return exclusion_reason(row) == ""


def classify_issue(text):
    checks = [
        ("FTA_ORIGIN", ["fta", "cepa", "epa", "원산지", "rules of origin", "origin"]),
        ("EXPORT_CONTROL", ["수출통제", "전략물자", "export control", "ear", "sanction", "sanctions"]),
        ("AD_CVD", ["반덤핑", "덤핑방지관세", "상계관세", "anti-dumping", "antidumping", "countervailing"]),
        ("CBAM", ["cbam", "탄소국경", "carbon border"]),
        ("SECTION_301_232", ["section 301", "301조", "section 232", "232조", "ustr"]),
        ("CUSTOMS_CLEARANCE", ["통관", "수입신고", "세관장확인", "customs clearance", "customs audit", "customs valuation", "cbp"]),
        ("DUTY_REFUND", ["관세환급", "drawback", "duty refund"]),
        ("HS_CLASSIFICATION", ["품목분류", "hs code", "classification"]),
        ("TARIFF_DUTY", ["관세", "관세율", "할당관세", "tariff", "duty", "import duty"]),
    ]
    for label, keys in checks:
        if contains_any(text, keys):
            return label
    return "TRADE_REGULATION"


def make_fact_key(row):
    """
    Cluster relaxation key.
    Same legal title can stay separate when effective date / HS / tariff /
    change detail is different. This prevents Article 71 quota-tariff notices
    from being collapsed into a single row when they contain different
    implementation facts.
    """
    facts = []
    for col in [
        "effective_date_hint", "hs_hint", "tariff_rate_hint", "change_detail_hint",
        "effective_date", "hs_code", "tariff_rate", "old_tariff_rate",
        "new_tariff_rate", "affected_products",
    ]:
        val = safe_str(row.get(col))
        if val:
            facts.append(val)
    fact_key = norm(" ".join(facts))
    return fact_key[:140]


def regulation_title_key(row):
    title = norm(row.get("headline"))
    # Keep enough title detail to distinguish: decree, amendment, notice, rule etc.
    title = re.sub(r"\s+", " ", title).strip()
    return title[:120]


def cluster_key(row):
    text = norm(combined(row))
    issue = classify_issue(text)
    title_key = regulation_title_key(row)
    fact_key = make_fact_key(row)

    # Do NOT collapse all Article 71 quota-tariff rows into one.
    # Keep by issue + title + structured facts.
    if "관세법" in text and "71" in text and "할당관세" in text:
        return f"KR_QUOTA_TARIFF_ARTICLE_71:{title_key}:{fact_key}"

    if "부틸글리콜에테르" in text or "butyl glycol ether" in text:
        return f"KR_AD_SAUDI_BUTYL_GLYCOL_ETHER:{title_key}:{fact_key}"

    if "에너지이용 합리화법" in text and "세관장확인" in text:
        return f"KR_ENERGY_USE_CUSTOMS_CONFIRMATION:{title_key}:{fact_key}"

    if "전자상거래업자" in text and "등록" in text:
        return f"KR_ECOMMERCE_OPERATOR_REGISTRATION:{title_key}:{fact_key}"

    if "수입신고" in text and "가산세" in text and "할당관세" in text:
        return f"KR_IMPORT_DECLARATION_SURCHARGE_QUOTA_TARIFF:{title_key}:{fact_key}"

    return f"{issue}:{title_key}:{fact_key}"


def base_score(row):
    raw = pd.to_numeric(pd.Series([row.get("score")]), errors="coerce").fillna(0).iloc[0]
    text = combined(row)
    issue = classify_issue(text)
    score = int(raw)
    if issue == "AD_CVD":
        score += 70
    elif issue == "CUSTOMS_CLEARANCE":
        score += 60
    elif issue == "TARIFF_DUTY":
        score += 55
    elif issue in {"FTA_ORIGIN", "HS_CLASSIFICATION", "EXPORT_CONTROL", "CBAM", "SECTION_301_232"}:
        score += 50
    elif issue == "DUTY_REFUND":
        score += 40
    if contains_any(text, REGULATION_ACTION_KEYWORDS):
        score += 15
    if safe_str(row.get("source_type")) == "NEWS_OFFICIAL_NOTICE":
        score += 10
    if contains_any(text, SAMSUNG_KEYWORDS):
        score += 10
    return max(score, 60)


def samsung_info(row):
    text = combined(row)
    score = 0
    reasons = []
    if contains_any(text, SAMSUNG_KEYWORDS):
        score += 45
        reasons.append("삼성 생산/제품 키워드 직접 관련")
    if classify_issue(text) in {"TARIFF_DUTY", "AD_CVD", "EXPORT_CONTROL", "FTA_ORIGIN", "CUSTOMS_CLEARANCE"}:
        score += 25
        reasons.append("관세·통상 컴플라이언스 영향")
    if contains_any(text, ["반도체", "semiconductor", "chip", "display", "battery"]):
        score += 20
        reasons.append("전자·부품 공급망 관련")
    score = min(score, 100)
    if score >= 70:
        relevance = "상"
    elif score >= 40:
        relevance = "중"
    else:
        relevance = "하"
    return relevance, score, "; ".join(reasons) if reasons else "직접 영향 제한적"


def risk_from_score(score):
    if score >= 95:
        return "상"
    if score >= 75:
        return "중"
    return "하"


def make_fallback(row):
    headline = safe_str(row.get("headline"))
    issue = classify_issue(combined(row))
    hints = hint_context(row)
    hs_code = first_nonempty(hints["hs_hint"], extract_hs_codes(headline, hints["change_detail_hint"], hints["article_body"]))
    tariff_rate = first_nonempty(hints["tariff_rate_hint"], extract_tariff_rates(headline, hints["change_detail_hint"], hints["article_body"]))
    old_tariff_rate, new_tariff_rate = extract_old_new_tariff_rates(headline, hints["change_detail_hint"], hints["tariff_rate_hint"], hints["article_body"])
    effective_date = hints["effective_date_hint"]
    target_country = derive_target_country(row)
    affected_products = derive_affected_products(row)

    temp_row = row.copy() if hasattr(row, "copy") else dict(row)
    temp_row["effective_date"] = effective_date
    temp_row["hs_code"] = hs_code
    temp_row["old_tariff_rate"] = old_tariff_rate
    temp_row["new_tariff_rate"] = new_tariff_rate
    temp_row["tariff_rate"] = tariff_rate or compact_join([old_tariff_rate, new_tariff_rate], " -> ")
    temp_row["target_country"] = target_country
    temp_row["affected_products"] = affected_products

    return {
        "summary": infer_regulation_summary(temp_row, issue),
        "analysis": build_regulation_analysis(temp_row),
        "action_plan": build_regulation_action_plan(temp_row),
        "effective_date": missing_if_blank(effective_date),
        "hs_code": missing_if_blank(hs_code),
        "old_tariff_rate": missing_if_blank(old_tariff_rate),
        "new_tariff_rate": missing_if_blank(new_tariff_rate),
        "tariff_rate": missing_if_blank(tariff_rate or compact_join([old_tariff_rate, new_tariff_rate], " -> ")),
        "target_country": missing_if_blank(target_country),
        "affected_products": missing_if_blank(affected_products),
        "compliance_tasks": build_regulation_action_plan(temp_row),
    }


def regulation_prompt(row):
    """STEP4-1 dedicated regulation prompt: STEP3 hints -> legal impact analysis."""
    return f"""
너는 삼성전자 본사 관세 담당자다.
아래 법규/공고 원문 및 STEP3 추출 힌트를 바탕으로 반드시 다음 항목을 추출하라.

1. 시행일
2. 변경 HS Code
3. 기존 세율
4. 변경 세율
5. 영향 품목
6. 삼성전자 영향
7. 즉시 조치사항

중요 원칙:
- STEP3는 원문에서 단서를 뽑고, STEP4는 그 단서를 근거로 법규 영향분석을 완성하는 단계다.
- 법규/공고/관보/고시/시행령/시행규칙은 뉴스처럼 해설하지 말고, 시행일·HS·세율·대상품목 중심으로 구조화하라.
- 정보가 없으면 "원문 확인 필요"라고 쓰지 말고 반드시 "본문에서 확인 불가"라고 표기하라.
- "담당 부서 확인 필요", "추가 확인 필요" 같은 일반 문구도 쓰지 말라.
- 반드시 한국어로 작성하라.

Return JSON only with these exact keys:
summary, analysis, action_plan,
effective_date, hs_code, old_tariff_rate, new_tariff_rate, tariff_rate,
target_country, affected_products, compliance_tasks.

필드 작성 기준:
- summary: 법규상 변경 내용을 시행일/HS/세율/대상국/대상품목 중심으로 2~3문장 요약.
- analysis: 삼성전자 영향. HS Master, 관세율 Master, 수입신고, FTA/원산지, 관세비용, 해외법인/관세사 신고로 연결해서 작성.
- action_plan: 즉시 조치사항. 마스터 변경, 대상 거래 추출, 법인/관세사 배포, 시행일 기준 신고 로직 반영 등 실행형으로 작성.
- effective_date: 시행일/적용일/발효일. 없으면 "본문에서 확인 불가".
- hs_code: 변경 HS Code/세번/품목분류번호. 여러 개면 세미콜론으로 구분. 없으면 "본문에서 확인 불가".
- old_tariff_rate: 기존/현행/종전 세율. 없으면 "본문에서 확인 불가".
- new_tariff_rate: 변경/개정/적용 세율. 없으면 "본문에서 확인 불가".
- tariff_rate: 세율 변경 요약. 가능하면 "기존세율 -> 변경세율" 형식. 없으면 "본문에서 확인 불가".
- target_country: 대상국/원산국/수입국/수출국. 없으면 "본문에서 확인 불가".
- affected_products: 영향 품목/대상품목/물품명. 없으면 "본문에서 확인 불가".
- compliance_tasks: 내부통제 관점 즉시 조치사항을 세미콜론으로 구분.

입력 메타데이터:
Title: {safe_str(row.get('headline'))}
Country: {safe_str(row.get('country'))}
Agency: {safe_str(row.get('agency'))}
ResponsibleAgency: {safe_str(row.get('responsible_agency'))}
Issue: {classify_issue(combined(row))}
URL: {safe_str(row.get('url'))}

STEP3 Article Extractor 힌트:
effective_date_hint: {safe_str(row.get('effective_date_hint'))}
hs_hint: {safe_str(row.get('hs_hint'))}
tariff_rate_hint: {safe_str(row.get('tariff_rate_hint'))}
change_detail_hint: {safe_str(row.get('change_detail_hint'))}

법규/공고 본문 발췌:
{trim(row.get('article_body'), 4500)}
"""


def ai_enrich(row):
    fallback = make_fallback(row)
    if not USE_AI or client is None:
        return fallback
    prompt = regulation_prompt(row)
    try:
        resp = client.models.generate_content(model=GEMINI_MODEL, contents=prompt)
        data = parse_json(getattr(resp, "text", ""))
        result = {
            "summary": clean_regulation_text(data.get("summary"), fallback["summary"]),
            "analysis": clean_regulation_text(data.get("analysis"), fallback["analysis"]),
            "action_plan": clean_regulation_text(data.get("action_plan"), fallback["action_plan"]),
            "effective_date": trim(normalize_blank(data.get("effective_date")) or fallback["effective_date"], 120),
            "hs_code": trim(normalize_blank(data.get("hs_code")) or fallback["hs_code"], 250),
            "old_tariff_rate": trim(normalize_blank(data.get("old_tariff_rate")) or fallback["old_tariff_rate"], 250),
            "new_tariff_rate": trim(normalize_blank(data.get("new_tariff_rate")) or fallback["new_tariff_rate"], 250),
            "tariff_rate": trim(normalize_blank(data.get("tariff_rate")) or fallback["tariff_rate"], 250),
            "target_country": trim(normalize_blank(data.get("target_country")) or fallback["target_country"], 250),
            "affected_products": trim(normalize_blank(data.get("affected_products")) or fallback["affected_products"], 300),
            "compliance_tasks": clean_regulation_text(data.get("compliance_tasks"), fallback["compliance_tasks"]),
        }

        # If AI summarized generically, rebuild the key sentences from structured fields.
        temp = row.copy() if hasattr(row, "copy") else dict(row)
        for k in ["effective_date", "hs_code", "old_tariff_rate", "new_tariff_rate", "tariff_rate", "target_country", "affected_products"]:
            temp[k] = result.get(k, "")
        if not result["summary"] or len(result["summary"]) < 20:
            result["summary"] = infer_regulation_summary(temp, classify_issue(combined(row)))
        if not result["analysis"] or "확인 필요" in result["analysis"]:
            result["analysis"] = build_regulation_analysis(temp)
        if not result["action_plan"] or "확인 필요" in result["action_plan"]:
            result["action_plan"] = build_regulation_action_plan(temp)
        if not result["compliance_tasks"] or "확인 필요" in result["compliance_tasks"]:
            result["compliance_tasks"] = build_regulation_action_plan(temp)
        return {k: trim(v, 700 if k in {"summary", "analysis", "action_plan", "compliance_tasks"} else 300) for k, v in result.items()}
    except Exception:
        return fallback


def read_inputs():
    frames = []
    reg_path = REG_ARTICLE_INPUT if os.path.exists(REG_ARTICLE_INPUT) else REG_INPUT
    if os.path.exists(reg_path):
        frames.append(normalize_regulation(pd.read_excel(reg_path)))
        print(f"[INPUT] Regulation source: {reg_path}")
    if not frames:
        return pd.DataFrame()
    return pd.concat(frames, ignore_index=True)

def choose_representative(group):
    # Only choose one row inside a truly identical cluster. Since cluster_key now
    # includes title + effective/HS/tariff/change facts, this no longer removes
    # distinct legal changes under the same regulation family.
    g = group.copy()
    g["_len"] = g["headline"].astype(str).str.len()
    g["_body_len"] = g.get("article_body", "").astype(str).str.len() if "article_body" in g.columns else 0
    g["_hint_len"] = (
        g.get("effective_date_hint", "").astype(str).str.len()
        + g.get("hs_hint", "").astype(str).str.len()
        + g.get("tariff_rate_hint", "").astype(str).str.len()
        + g.get("change_detail_hint", "").astype(str).str.len()
    )
    g["_source_rank"] = g["source_type"].map({"REGULATION_ORIGINAL": 2, "NEWS_OFFICIAL_NOTICE": 1}).fillna(0)
    g["_score"] = g.apply(base_score, axis=1)
    g = g.sort_values(
        ["_score", "_source_rank", "_hint_len", "_body_len", "_len"],
        ascending=[False, False, False, False, False],
    )
    return g.iloc[0]


def build_output(df):
    if df.empty:
        return pd.DataFrame(columns=OUTPUT_COLUMNS)
    df = df.copy()
    df["exclude_reason"] = df.apply(exclusion_reason, axis=1)
    selected = df[df["exclude_reason"].eq("")].copy()
    if selected.empty:
        return pd.DataFrame(columns=OUTPUT_COLUMNS)
    selected["cluster_key"] = selected.apply(cluster_key, axis=1)
    rows = []
    for _, group in selected.groupby("cluster_key", dropna=False):
        rows.append(choose_representative(group))
    final = pd.DataFrame(rows)
    final["Score"] = final.apply(base_score, axis=1)
    final["issue_type"] = final.apply(lambda r: classify_issue(combined(r)), axis=1)
    final["regulation_type"] = final["issue_type"].map(
        {
            "AD_CVD": "관세조치",
            "TARIFF_DUTY": "관세율/할당관세",
            "CUSTOMS_CLEARANCE": "통관/신고",
            "FTA_ORIGIN": "FTA/원산지",
            "EXPORT_CONTROL": "수출통제",
            "CBAM": "탄소국경",
            "HS_CLASSIFICATION": "품목분류",
            "DUTY_REFUND": "관세환급",
        }
    ).fillna("통상·관세 법규")
    final["Risk"] = final.apply(lambda r: "상" if r["issue_type"] in {"AD_CVD", "CUSTOMS_CLEARANCE", "TARIFF_DUTY", "FTA_ORIGIN", "HS_CLASSIFICATION", "EXPORT_CONTROL"} else risk_from_score(r["Score"]), axis=1)
    final["priority_group"] = final.apply(lambda r: "CORE" if r["issue_type"] in {"AD_CVD", "CUSTOMS_CLEARANCE", "TARIFF_DUTY", "FTA_ORIGIN", "HS_CLASSIFICATION", "EXPORT_CONTROL"} or r["Score"] >= 90 else "USABLE", axis=1)
    final["selected"] = "Y"
    final["must_include"] = "Y"
    final["last_checked"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    enriched = final.apply(ai_enrich, axis=1, result_type="expand")
    final["Summary"] = enriched["summary"]
    final["AI Analysis"] = enriched["analysis"]
    final["Action Plan"] = enriched["action_plan"]
    final["effective_date"] = enriched["effective_date"]
    final["hs_code"] = enriched["hs_code"]
    final["old_tariff_rate"] = enriched["old_tariff_rate"]
    final["new_tariff_rate"] = enriched["new_tariff_rate"]
    final["tariff_rate"] = enriched["tariff_rate"]
    final["target_country"] = enriched["target_country"]
    final["affected_products"] = enriched["affected_products"]
    final["effective_date_hint"] = final.get("effective_date_hint", "")
    final["hs_hint"] = final.get("hs_hint", "")
    final["tariff_rate_hint"] = final.get("tariff_rate_hint", "")
    final["change_detail_hint"] = final.get("change_detail_hint", "")
    final["compliance_tasks"] = enriched["compliance_tasks"]
    sams = final.apply(samsung_info, axis=1, result_type="expand")
    final["samsung_relevance"] = sams[0]
    final["samsung_score"] = sams[1]
    final["samsung_reason"] = sams[2]
    final["impact_area"] = final["issue_type"].map(
        {
            "AD_CVD": "수입관세/가격",
            "TARIFF_DUTY": "관세율/수입비용",
            "CUSTOMS_CLEARANCE": "통관절차/신고",
            "FTA_ORIGIN": "원산지/협정세율",
            "EXPORT_CONTROL": "수출허가/제재",
            "CBAM": "탄소규제/비용",
            "HS_CLASSIFICATION": "품목분류",
            "DUTY_REFUND": "환급/사후관리",
        }
    ).fillna("통상 컴플라이언스")
    final = final.sort_values(["Score", "date"], ascending=[False, False], kind="stable").reset_index(drop=True)
    final["Rank"] = range(1, len(final) + 1)
    rename_map = {
        "date": "Date",
        "headline": "Headline",
        "url": "URL",
        "country": "Country",
        "agency": "Agency",
        "responsible_agency": "ResponsibleAgency",
        "keyword_matches": "KeywordMatches",
        "filter_reason": "FilterReason",
        "source": "Source",
    }
    final = final.rename(columns=rename_map)
    for col in OUTPUT_COLUMNS:
        if col not in final.columns:
            final[col] = ""
    return final[OUTPUT_COLUMNS]


def clean_cumulative(existing, daily):
    frames = []
    if existing is not None and not existing.empty:
        existing = existing.copy()
        for col in OUTPUT_COLUMNS:
            if col not in existing.columns:
                existing[col] = ""

        # Drop rows produced by older STEP4-1 versions.  They used blank
        # source_type, boolean selected values, and a different score scale,
        # so cluster_key-based de-duplication cannot reliably catch them.
        valid_source_type = existing["source_type"].astype(str).str.strip().isin(
            ["REGULATION_ORIGINAL", "NEWS_OFFICIAL_NOTICE"]
        )
        valid_selected = existing["selected"].astype(str).str.strip().eq("Y")
        valid_score = pd.to_numeric(existing["Score"], errors="coerce").fillna(0).le(180)
        existing = existing[valid_source_type & valid_selected & valid_score].copy()
        frames.append(existing)
    if daily is not None and not daily.empty:
        frames.append(daily)
    if not frames:
        return pd.DataFrame(columns=OUTPUT_COLUMNS)
    combined_df = pd.concat(frames, ignore_index=True)
    combined_df = combined_df[combined_df["Headline"].astype(str).str.strip().ne("")]
    combined_df = combined_df[~combined_df["Headline"].astype(str).apply(lambda x: contains_any(x, HR_EXCLUDE))]
    combined_df["_headline_key"] = combined_df["Headline"].astype(str).apply(norm)
    combined_df = combined_df.drop_duplicates(subset=["cluster_key"], keep="first")
    # Do not drop by headline only: same legal title may have different effective date/HS/rate.
    combined_df = combined_df.sort_values(["Score", "Date"], ascending=[False, False], kind="stable").reset_index(drop=True)
    combined_df["Rank"] = range(1, len(combined_df) + 1)
    combined_df = combined_df.drop(columns=["_headline_key"], errors="ignore")
    for col in OUTPUT_COLUMNS:
        if col not in combined_df.columns:
            combined_df[col] = ""
    return combined_df[OUTPUT_COLUMNS]


def format_excel(path):
    if not os.path.exists(path):
        return
    wb = load_workbook(path)
    ws = wb.active
    ws.freeze_panes = "A2"
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in ws[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    widths = {
        "A": 8, "B": 16, "C": 56, "D": 32, "E": 14, "F": 22, "G": 22,
        "H": 20, "K": 16, "M": 12, "N": 12, "O": 12, "P": 10, "Q": 20,
        "R": 20, "S": 20, "T": 18, "U": 12, "V": 36, "W": 54, "X": 54,
        "Y": 54, "Z": 18, "AA": 18, "AB": 18, "AC": 18, "AD": 32,
        "AE": 22, "AF": 22, "AG": 22, "AH": 34,
    }
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].width = widths.get(letter, 18)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    wb.save(path)



def build_excluded(df):
    if df.empty:
        return pd.DataFrame(columns=OUTPUT_COLUMNS)
    ex = df.copy()
    ex["exclude_reason"] = ex.apply(exclusion_reason, axis=1)
    ex = ex[ex["exclude_reason"].ne("")].copy()
    if ex.empty:
        return pd.DataFrame(columns=OUTPUT_COLUMNS)
    ex["cluster_key"] = ex.apply(cluster_key, axis=1)
    ex["Score"] = ex.apply(base_score, axis=1)
    ex["issue_type"] = ex.apply(lambda r: classify_issue(combined(r)), axis=1)
    ex["Risk"] = ex["Score"].apply(risk_from_score)
    ex["priority_group"] = "EXCLUDED"
    ex["selected"] = "N"
    ex["must_include"] = "N"
    ex["last_checked"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ex = ex.rename(columns={
        "date": "Date", "headline": "Headline", "url": "URL",
        "country": "Country", "agency": "Agency",
        "responsible_agency": "ResponsibleAgency",
        "keyword_matches": "KeywordMatches", "filter_reason": "FilterReason",
        "source": "Source",
    })
    for col in OUTPUT_COLUMNS:
        if col not in ex.columns:
            ex[col] = ""
    ex["Rank"] = range(1, len(ex) + 1)
    return ex[OUTPUT_COLUMNS]


def main():
    print("[STEP4-1] Regulation analysis start")
    print("[AI]", "ON" if USE_AI else "OFF")
    df = read_inputs()
    daily = build_output(df)
    excluded = build_excluded(df)
    daily.to_excel(OUTPUT_DAILY, index=False)
    excluded.to_excel(OUTPUT_EXCLUDED, index=False)

    existing = pd.read_excel(OUTPUT_CUMUL) if os.path.exists(OUTPUT_CUMUL) else pd.DataFrame(columns=OUTPUT_COLUMNS)
    cumulative = clean_cumulative(existing, daily)
    cumulative.to_excel(OUTPUT_CUMUL, index=False)

    format_excel(OUTPUT_DAILY)
    format_excel(OUTPUT_CUMUL)
    format_excel(OUTPUT_EXCLUDED)
    print(f"[DONE] Daily: {OUTPUT_DAILY}")
    print(f"[DONE] Cumulative: {OUTPUT_CUMUL}")
    print(f"[DONE] Excluded: {OUTPUT_EXCLUDED}")
    print(f"[ROWS] daily={len(daily)}, cumulative={len(cumulative)}, excluded={len(excluded)}")


# =========================================================
# GTI Regulation-first relaxed analysis override
# Keep immediately before __main__.
# Official regulation rows are analyzed even when article_body is short, using
# title/agency/date/url/source metadata and STEP3 fallback body.
# =========================================================

def _u4(s0: str) -> str:
    return s0.encode("ascii").decode("unicode_escape")


K_TARIFF = _u4("\\uad00\\uc138")
K_TARIFF_RATE = _u4("\\uad00\\uc138\\uc728")
K_CUSTOMS = _u4("\\ud1b5\\uad00")
K_IMPORT = _u4("\\uc218\\uc785")
K_EXPORT = _u4("\\uc218\\ucd9c")
K_ORIGIN = _u4("\\uc6d0\\uc0b0\\uc9c0")
K_HS = _u4("\\ud488\\ubaa9\\ubd84\\ub958")
K_AD = _u4("\\ubc18\\ub364\\ud551")
K_CVD = _u4("\\uc0c1\\uacc4\\uad00\\uc138")
K_EXPORT_CONTROL = _u4("\\uc218\\ucd9c\\ud1b5\\uc81c")
K_NOTICE = _u4("\\uace0\\uc2dc")
K_ANNOUNCE = _u4("\\uacf5\\uace0")
K_GAZETTE = _u4("\\uad00\\ubcf4")
K_AMEND = _u4("\\uac1c\\uc815")
K_ENACT = _u4("\\uc81c\\uc815")
K_EFFECTIVE = _u4("\\uc2dc\\ud589")
K_LAW = _u4("\\ubc95\\ub839")
K_RULE = _u4("\\uaddc\\uce59")

REG_SAFE_OFFICIAL_DOMAINS = [
    "law.go.kr", "gwanbo.go.kr", "customs.go.kr", "unipass.customs.go.kr",
    "motir.go.kr", "federalregister.gov", "ustr.gov", "cbp.gov", "usitc.gov",
    "eur-lex.europa.eu", "taxation-customs.ec.europa.eu", "wto.org",
]

REG_SAFE_TRADE_TERMS = [
    K_TARIFF, K_TARIFF_RATE, K_CUSTOMS, K_IMPORT, K_EXPORT, K_ORIGIN, K_HS,
    K_AD, K_CVD, K_EXPORT_CONTROL,
    "customs", "tariff", "duty", "import", "export", "rules of origin",
    "origin", "fta", "cepa", "epa", "hs code", "classification",
    "anti-dumping", "antidumping", "countervailing", "safeguard",
    "section 301", "section 232", "export control", "entity list", "cbam",
]

REG_SAFE_OFFICIAL_TERMS = [
    K_NOTICE, K_ANNOUNCE, K_GAZETTE, K_AMEND, K_ENACT, K_EFFECTIVE, K_LAW, K_RULE,
    "notice", "regulation", "rule", "law", "decree", "ordinance",
    "amendment", "final rule", "proposed rule", "effective date",
]


def _safe_contains_text(text: str, terms: list[str]) -> bool:
    low = safe_str(text).lower()
    return any(str(t).lower() in low for t in terms if safe_str(t))


def _is_official_regulation(row) -> bool:
    text = combined(row).lower()
    if safe_str(row.get("official_regulation_flag")).upper() == "Y":
        return True
    if safe_str(row.get("article_source_type")).upper() == "OFFICIAL":
        return True
    if any(d in text for d in REG_SAFE_OFFICIAL_DOMAINS):
        return True
    return _safe_contains_text(text, REG_SAFE_OFFICIAL_TERMS)


def _has_trade_signal(row) -> bool:
    return _safe_contains_text(combined(row), REG_SAFE_TRADE_TERMS)


def _reg_doc_strength(row) -> int:
    score = 0
    text = combined(row)
    if _is_official_regulation(row):
        score += 45
    if _has_trade_signal(row):
        score += 35
    try:
        score += int(float(row.get("official_regulation_score") or 0)) // 4
    except Exception:
        pass
    if safe_str(row.get("article_body_ok")).upper() == "Y":
        score += 10
    if _safe_contains_text(text, [K_NOTICE, K_ANNOUNCE, K_GAZETTE, "federal register", "official journal"]):
        score += 10
    return min(score, 100)


def normalize_regulation(df):
    mapping = {
        "date": pick(df, "Date", "date", "published"),
        "headline": pick(df, "Headline", "headline", "title"),
        "url": pick(df, "URL", "url", "link"),
        "country": pick(df, "Country", "country"),
        "agency": pick(df, "Agency", "agency", "source", "publisher"),
        "responsible_agency": pick(df, "ResponsibleAgency", "responsible_agency"),
        "score": pick(df, "Score", "score", "official_regulation_score", "policy_score"),
        "keyword_matches": pick(df, "KeywordMatches", "keyword_matches", "keyword", "matched_policy_terms"),
        "filter_reason": pick(df, "FilterReason", "filter_reason", "SelectReason", "official_regulation_reason"),
        "source": pick(df, "Source", "source", "SourceFile", "source_file"),
        "original_url": pick(df, "original_url", "OriginalURL", "canonical_url"),
        "article_body": pick(df, "article_body", "ArticleBody", "body", "content", "regulation_fallback_body"),
        "article_extract_status": pick(df, "article_extract_status", "ArticleExtractStatus"),
        "article_source_type": pick(df, "article_source_type", "ArticleSourceType"),
        "effective_date_hint": pick(df, "effective_date_hint", "EffectiveDateHint"),
        "change_detail_hint": pick(df, "change_detail_hint", "ChangeDetailHint"),
        "hs_hint": pick(df, "hs_hint", "HsHint", "HS_Hint"),
        "tariff_rate_hint": pick(df, "tariff_rate_hint", "TariffRateHint"),
        "official_regulation_flag": pick(df, "official_regulation_flag"),
        "official_regulation_score": pick(df, "official_regulation_score"),
        "official_regulation_type": pick(df, "official_regulation_type"),
        "official_regulation_reason": pick(df, "official_regulation_reason"),
    }
    out = pd.DataFrame()
    for target, source in mapping.items():
        out[target] = df[source] if source in df.columns else ""
    recovered = out.apply(recover_headline_url, axis=1)
    out["headline"], out["url"] = recovered["headline"], recovered["url"]
    out["source_type"] = "REGULATION_ORIGINAL"
    out["regulation_related"] = "Y"
    out["regulation_transfer_type"] = "Original"
    # Build metadata body if STEP3 body is still short.
    weak = out["article_body"].fillna("").astype(str).str.len().lt(80)
    out.loc[weak, "article_body"] = out[weak].apply(
        lambda r: " | ".join([
            f"Title: {safe_str(r.get('headline'))}",
            f"Date: {safe_str(r.get('date'))}",
            f"Agency: {safe_str(r.get('agency'))}",
            f"URL: {safe_str(r.get('url'))}",
            f"Source: {safe_str(r.get('source'))}",
            f"Signals: {safe_str(r.get('official_regulation_reason')) or safe_str(r.get('keyword_matches'))}",
        ]),
        axis=1,
    )
    return out


def exclusion_reason(row):
    text = combined(row)
    if _is_official_regulation(row) and _has_trade_signal(row):
        return ""
    if _is_official_regulation(row) and _reg_doc_strength(row) >= 55:
        # Official regulation with weak trade signal is retained as USABLE so
        # headquarters can see newly issued legal changes and mark no impact.
        return ""
    if contains_any(text, HR_EXCLUDE):
        return "HR_RECRUITMENT_PERSONNEL"
    if contains_any(text, ADMIN_EXCLUDE) and not contains_any(text, ADMIN_KEEP_IF_ALSO):
        return "GENERIC_ADMIN_SANCTION"
    if not _has_trade_signal(row) and not contains_any(text, CUSTOMS_LAW_KEYWORDS):
        return "NOT_CUSTOMS_TRADE_REGULATION"
    return ""


def classify_issue(text):
    if _safe_contains_text(text, [K_AD, K_CVD, "anti-dumping", "antidumping", "countervailing", "ad/cvd"]):
        return "AD_CVD"
    if _safe_contains_text(text, [K_EXPORT_CONTROL, "export control", "entity list", "ear", "sanction", "sanctions"]):
        return "EXPORT_CONTROL"
    if _safe_contains_text(text, ["cbam", "carbon border"]):
        return "CBAM"
    if _safe_contains_text(text, ["section 301", "section 232", "ustr"]):
        return "SECTION_301_232"
    if _safe_contains_text(text, [K_ORIGIN, "fta", "cepa", "epa", "rules of origin", "origin"]):
        return "FTA_ORIGIN"
    if _safe_contains_text(text, [K_HS, "hs code", "classification"]):
        return "HS_CLASSIFICATION"
    if _safe_contains_text(text, [K_CUSTOMS, "customs clearance", "customs audit", "customs valuation", "cbp"]):
        return "CUSTOMS_CLEARANCE"
    if _safe_contains_text(text, [K_TARIFF, K_TARIFF_RATE, "tariff", "duty", "import duty"]):
        return "TARIFF_DUTY"
    return "TRADE_REGULATION"


def base_score(row):
    raw = pd.to_numeric(pd.Series([row.get("score")]), errors="coerce").fillna(0).iloc[0]
    issue = classify_issue(combined(row))
    score = int(raw) + _reg_doc_strength(row)
    if issue == "AD_CVD":
        score += 70
    elif issue == "CUSTOMS_CLEARANCE":
        score += 60
    elif issue == "TARIFF_DUTY":
        score += 55
    elif issue in {"FTA_ORIGIN", "HS_CLASSIFICATION", "EXPORT_CONTROL", "CBAM", "SECTION_301_232"}:
        score += 50
    elif issue == "TRADE_REGULATION" and _is_official_regulation(row):
        score += 20
    if contains_any(combined(row), SAMSUNG_KEYWORDS):
        score += 10
    return max(score, 60)


def make_fallback(row):
    headline = safe_str(row.get("headline"))
    issue = classify_issue(combined(row))
    hints = hint_context(row)
    hs_code = first_nonempty(hints["hs_hint"], extract_hs_codes(headline, hints["change_detail_hint"], hints["article_body"]))
    tariff_rate = first_nonempty(hints["tariff_rate_hint"], extract_tariff_rates(headline, hints["change_detail_hint"], hints["article_body"]))
    old_tariff_rate, new_tariff_rate = extract_old_new_tariff_rates(headline, hints["change_detail_hint"], hints["tariff_rate_hint"], hints["article_body"])
    effective_date = hints["effective_date_hint"]
    target_country = derive_target_country(row)
    affected_products = derive_affected_products(row)
    official_note = "Official regulation metadata fallback used; verify original URL for full legal text." if _is_official_regulation(row) else ""

    temp_row = row.copy() if hasattr(row, "copy") else dict(row)
    temp_row["effective_date"] = effective_date
    temp_row["hs_code"] = hs_code
    temp_row["old_tariff_rate"] = old_tariff_rate
    temp_row["new_tariff_rate"] = new_tariff_rate
    temp_row["tariff_rate"] = tariff_rate or compact_join([old_tariff_rate, new_tariff_rate], " -> ")
    temp_row["target_country"] = target_country
    temp_row["affected_products"] = affected_products

    summary = infer_regulation_summary(temp_row, issue)
    analysis = build_regulation_analysis(temp_row)
    action = build_regulation_action_plan(temp_row)
    if issue == "TRADE_REGULATION" and _is_official_regulation(row):
        summary = f"Official regulation/notice detected: {headline}. Customs/trade impact is not explicit in the extracted text; keep for legal monitoring."
        analysis = f"{official_note} Samsung impact should be reviewed by HQ customs if the regulation affects HS, tariff rate, import/export declaration, origin, export control, or CBAM processes."
        action = "Open original URL; confirm whether HS/tariff/origin/export-control terms are present; mark no impact if unrelated; update regulation source master if recurring."

    return {
        "summary": summary,
        "analysis": analysis,
        "action_plan": action,
        "effective_date": missing_if_blank(effective_date),
        "hs_code": missing_if_blank(hs_code),
        "old_tariff_rate": missing_if_blank(old_tariff_rate),
        "new_tariff_rate": missing_if_blank(new_tariff_rate),
        "tariff_rate": missing_if_blank(tariff_rate or compact_join([old_tariff_rate, new_tariff_rate], " -> ")),
        "target_country": missing_if_blank(target_country),
        "affected_products": missing_if_blank(affected_products),
        "compliance_tasks": action,
    }


if __name__ == "__main__":
    main()
