# -*- coding: utf-8 -*-
"""
GTI STEP5 Mail Engine - EXECUTIVE SAMSUNG IMPACT FINAL

Input
-----
    C:/Temp/4-1.regulation_ai_summary.xlsx
    C:/Temp/4-2.news_ai_summary.xlsx

Output
------
    C:/Temp/12345/c_type_outputs/[GTI Radar] Global Trade Intelligence(YYYY-MM-DD).html
    C:/Temp/12345/c_type_outputs/[GTI Radar] Global Trade Intelligence(YYYY-MM-DD).xlsx
    C:/Temp/12345/c_type_outputs/4.news_ai_analysis.xlsx

Operating rule
--------------
    STEP5 does not reselect STEP4 results.
    Top3 is selected across Regulation + News by integrated importance.
    Mail body is split into:
    HTML Form: Topic / Headline / Summary / Impact / Action / Country / Agency / Risk / Publish Date
      Section 1: Regulation
      Section 2-1: News CORE
      Section 2-2: News USABLE / Reference
"""

from __future__ import annotations

import argparse
import html
import json

import os
import re
import smtplib
import ssl
import urllib.parse
import urllib.request
from datetime import datetime
from email.message import EmailMessage
from email.utils import formataddr
from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


REGULATION_INPUT_FILE = Path(os.getenv("GTI_REGULATION_INPUT_FILE", r"C:\Temp\4-1.regulation_ai_summary.xlsx"))
NEWS_INPUT_FILE = Path(os.getenv("GTI_NEWS_INPUT_FILE", r"C:\Temp\4-2.news_ai_summary.xlsx"))
OUTPUT_DIR = Path(os.getenv("GTI_OUTPUT_DIR", r"C:\Temp\12345\c_type_outputs"))
RUN_DATE = os.getenv("GTI_RUN_DATE", datetime.now().strftime("%Y-%m-%d"))
NEWS_MAX_ROWS = int(os.getenv("GTI_NEWS_MAX_ROWS", "30"))

SEND_EMAIL = os.getenv("GTI_SEND_EMAIL", "Y").strip().upper() in {"Y", "YES", "TRUE", "1"}
SMTP_HOST = os.getenv("GTI_SMTP_HOST", "smtp.naver.com")
SMTP_PORT = int(os.getenv("GTI_SMTP_PORT", "465"))
SMTP_USER = os.getenv("GTI_SMTP_USER", "kch8872@naver.com").strip()
SMTP_PASS = (os.getenv("GTI_SMTP_PASS") or os.getenv("GTI_MAIL_PW") or "").strip()
MAIL_TO = os.getenv("GTI_MAIL_TO", "").strip()
MAIL_FROM_NAME = os.getenv("GTI_MAIL_FROM_NAME", "GTI Radar").strip()
RECIPIENT_FILE = Path(os.getenv("GTI_RECIPIENT_FILE", r"C:\Temp\00.xlsx"))


OUTPUT_COLUMNS = [
    "No", "Content Type", "Mail Group", "Samsung Impact", "Affected Subsidiary", "Impact Reason",
    "Date", "Headline", "Summary", "AI Analysis", "Action Plan", "Country", "Agency", "Risk",
    "Importance Score", "Priority Group", "Issue", "Cluster", "URL", "Source", "Source File",
]


def output_paths() -> dict[str, Path]:
    return {
        "analysis": OUTPUT_DIR / "4.news_ai_analysis.xlsx",
        "mail_xlsx": OUTPUT_DIR / f"[GTI Radar] Global Trade Intelligence({RUN_DATE}).xlsx",
        "mail_html": OUTPUT_DIR / f"[GTI Radar] Global Trade Intelligence({RUN_DATE}).html",
        "cumulative": OUTPUT_DIR / "gti_news_cumulative.xlsx",
    }


def clean(value) -> str:
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except Exception:
        pass
    return re.sub(r"\s+", " ", str(value)).strip()



# -----------------------------------------------------------------------------
# URL hygiene
# -----------------------------------------------------------------------------
_GOOGLE_NEWS_RESOLVE_CACHE: dict[str, str] = {}


def is_google_news_rss_url(url: str) -> bool:
    u = clean(url).lower()
    return "news.google.com/rss/articles/" in u or "news.google.com/articles/" in u


def is_valid_http_url(url: str) -> bool:
    """클릭 가능한 정상 URL인지 강하게 검증합니다.

    이전 버전에서 Google News 해제 과정 중 "https://new" 같은 조각 문자열이
    정상 URL로 오인되어 메일 링크에 들어가는 문제가 있었습니다.
    임원 보고용 메일에서는 잘못된 링크가 1개라도 있으면 신뢰도가 크게 떨어지므로
    scheme, host, dot, 차단어를 모두 확인합니다.
    """
    u = clean(url)
    if not re.match(r"^https?://", u, flags=re.I):
        return False
    low = u.lower().strip()
    invalid_exact = {
        "http://", "https://", "https://new", "http://new",
        "https://news", "http://news", "nan", "none", "null", "new"
    }
    if low in invalid_exact:
        return False
    try:
        parsed = urllib.parse.urlparse(u)
        host = (parsed.netloc or "").lower()
        if not host or "." not in host:
            return False
        if host in {"new", "news"}:
            return False
    except Exception:
        return False
    return True


def is_preferred_article_url(url: str) -> bool:
    """사용자 클릭용 원문 URL 후보인지 판단합니다.

    Google News RSS 링크는 브라우저에서 리디렉션 오류를 일으킬 수 있으므로
    최종 링크로 쓰지 않습니다. 또한 https://new 같은 조각 URL도 차단합니다.
    """
    u = clean(url)
    if not is_valid_http_url(u):
        return False
    low = u.lower()
    bad_fragments = [
        "news.google.com/",
        "accounts.google.",
        "policies.google.",
        "support.google.",
        "consent.google.",
        "google.com/search",
        "google.com/amp/s/",
    ]
    return not any(x in low for x in bad_fragments)


def _clean_url_candidate(cand: str) -> str:
    cand = html.unescape(urllib.parse.unquote(clean(cand)))
    cand = cand.strip().strip('"\'<>')
    cand = cand.rstrip(".,);]}")
    # Google HTML/JS 조각에서 https://new 같은 조각이 잡히는 것을 차단
    if not is_valid_http_url(cand):
        return ""
    return cand



def _google_news_article_id(url: str) -> str:
    """Google News RSS/article URL에서 article id(CBMi...)를 추출합니다."""
    try:
        parsed = urllib.parse.urlparse(clean(url))
        parts = [p for p in parsed.path.split("/") if p]
        if not parts:
            return ""
        return parts[-1]
    except Exception:
        return ""


def _extract_preferred_url_from_text(text: str) -> str:
    """Google News HTML/JS 응답에서 원문 URL 후보만 추출합니다."""
    if not text:
        return ""
    # escaped unicode/slash를 풀기 위한 1차 정규화
    variants = [text]
    try:
        variants.append(text.encode("utf-8", errors="ignore").decode("unicode_escape", errors="ignore"))
    except Exception:
        pass
    for t in variants:
        for pat in [
            r"data-n-au=[\"'](https?://[^\"']+)[\"']",
            r"data-url=[\"'](https?://[^\"']+)[\"']",
            r"href=[\"'](https?://[^\"']+)[\"']",
            r"url=(https?%3A%2F%2F[^&\"'<>]+)",
            r"(https?:\\/\\/[^\"'<>\\]+)",
            r"(https?://[^\"'<>\s]+)",
        ]:
            for m in re.finditer(pat, t, flags=re.I):
                cand = m.group(1).replace('\\/', '/')
                cand = _clean_url_candidate(cand)
                if is_preferred_article_url(cand):
                    return cand
    return ""


def _decode_google_news_batchexecute(article_id: str, page_text: str, timeout: int = 8) -> str:
    """Google News 신형 article id를 batchexecute로 원문 URL 해제합니다.

    Google News RSS 링크는 단순 GET 시 https://news.google.com/ 으로 끝나는 경우가 있어
    data-n-a-sg(signature), data-n-a-ts(timestamp)를 읽어 Fbv4je RPC를 호출합니다.
    실패하면 빈 문자열을 반환해 잘못된 Google 링크를 메일에 노출하지 않습니다.
    """
    if not article_id or not page_text:
        return ""
    sg = ""
    ts = ""
    m = re.search(r'data-n-a-sg=["\']([^"\']+)["\']', page_text)
    if m:
        sg = html.unescape(m.group(1))
    m = re.search(r'data-n-a-ts=["\']([^"\']+)["\']', page_text)
    if m:
        ts = html.unescape(m.group(1))
    if not sg or not ts:
        return ""
    try:
        # Google News가 내부적으로 사용하는 garturlreq RPC payload.
        req_obj = [
            "garturlreq",
            [["en-US", "US", ["FINANCE_TOP_INDICES", "WEB_TEST_1_0_0"], None, None, 1, 1, "US:en", None, 180, None, None, None, None, None, 0, None, None, [int(ts), 0]],
             "en-US", "US", 1, [2, 3, 4, 8], 1, 0, "655000234", 0, 0, None, 0],
            article_id,
            int(ts),
            sg,
        ]
        f_req = [[["Fbv4je", json.dumps(req_obj, ensure_ascii=False, separators=(",", ":")), None, "generic"]]]
        data = ("f.req=" + urllib.parse.quote(json.dumps(f_req, ensure_ascii=False, separators=(",", ":")))).encode("utf-8")
        req = urllib.request.Request(
            "https://news.google.com/_/DotsSplashUi/data/batchexecute?rpcids=Fbv4je",
            data=data,
            headers={
                "Content-Type": "application/x-www-form-urlencoded;charset=UTF-8",
                "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/120 Safari/537.36",
                "Accept-Language": "ko-KR,ko;q=0.9,en-US;q=0.8,en;q=0.7",
                "Referer": "https://news.google.com/",
            },
            method="POST",
        )
        with urllib.request.urlopen(req, timeout=timeout) as resp:
            text = resp.read(500000).decode("utf-8", errors="ignore")
        cand = _extract_preferred_url_from_text(text)
        if is_preferred_article_url(cand):
            return cand
        # batchexecute 응답 내부 JSON 문자열 안에 URL이 직접 들어있는 경우
        try:
            unescaped = text.encode("utf-8", errors="ignore").decode("unicode_escape", errors="ignore")
            cand = _extract_preferred_url_from_text(unescaped)
            if is_preferred_article_url(cand):
                return cand
        except Exception:
            pass
    except Exception:
        return ""
    return ""

def resolve_google_news_url(url: str, timeout: int = 8) -> str:
    """Google News RSS article URL을 원문 URL로 해제합니다.

    성공하면 원문 URL을 반환하고, 실패하면 빈 문자열을 반환합니다.
    절대 https://news.google.com/ 또는 Google RSS 링크를 최종 링크로 반환하지 않습니다.
    """
    u = clean(url)
    if not is_google_news_rss_url(u):
        return u if is_preferred_article_url(u) else ""
    if u in _GOOGLE_NEWS_RESOLVE_CACHE:
        return _GOOGLE_NEWS_RESOLVE_CACHE[u]

    resolved = ""
    page_text = ""
    article_id = _google_news_article_id(u)
    try:
        req = urllib.request.Request(
            u,
            headers={
                "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/120 Safari/537.36",
                "Accept-Language": "ko-KR,ko;q=0.9,en-US;q=0.8,en;q=0.7",
            },
        )
        with urllib.request.urlopen(req, timeout=timeout) as resp:
            final_url = _clean_url_candidate(resp.geturl())
            if is_preferred_article_url(final_url):
                resolved = final_url
            page_text = resp.read(500000).decode("utf-8", errors="ignore")

        if not resolved:
            resolved = _extract_preferred_url_from_text(page_text)

        if not resolved:
            resolved = _decode_google_news_batchexecute(article_id, page_text, timeout=timeout)
    except Exception:
        resolved = ""

    if not is_preferred_article_url(resolved):
        resolved = ""
    _GOOGLE_NEWS_RESOLVE_CACHE[u] = resolved
    return resolved


def best_url_from_values(values: list[str]) -> str:
    """URL 우선순위: resolved/article/original/Representative/URL.

    정상 원문 URL만 반환합니다. Google RSS만 있고 해제에 실패하면 빈 문자열을 반환하여
    HTML/Excel에서 잘못된 링크가 생성되지 않도록 합니다.
    """
    invalid_tokens = {"", "nan", "none", "null", "new", "https://new", "http://new", "https://news", "http://news"}
    cleaned = []
    for v in values:
        vv = clean(v)
        if vv.lower() in invalid_tokens:
            continue
        cleaned.append(vv)
        for found in re.findall(r"https?://[^'\"),\s]+", vv):
            if found not in cleaned:
                cleaned.append(found)
        for found in re.findall(r"https?://[^'\"),\s]+", vv):
            if found not in cleaned:
                cleaned.append(found)
        for found in re.findall(r"https?://[^'\"),\s]+", vv):
            if found not in cleaned:
                cleaned.append(found)

    # 1) 이미 원문 URL인 후보 우선
    for v in cleaned:
        if is_preferred_article_url(v):
            return v

    # 2) Google News RSS는 해제 성공 시에만 사용
    for v in cleaned:
        if is_google_news_rss_url(v):
            resolved = resolve_google_news_url(v)
            if is_preferred_article_url(resolved):
                return resolved

    # 3) 실패 시 빈 값 반환. 나쁜 링크를 넣지 않음.
    return ""

def pick_col(df: pd.DataFrame, candidates: list[str]) -> str | None:
    lookup = {str(c).strip().lower(): c for c in df.columns}
    for cand in candidates:
        key = cand.lower()
        if key in lookup:
            return lookup[key]
    return None


def safe_num(value) -> int:
    try:
        return int(float(value))
    except Exception:
        return 0


def normalize_risk(value) -> str:
    raw = clean(value)
    low = raw.lower()
    if raw in {"상", "중", "하"}:
        return raw
    if low in {"high", "h", "red"}:
        return "상"
    if low in {"medium", "med", "m", "orange"}:
        return "중"
    if low in {"low", "l", "blue"}:
        return "하"
    return "중"


def risk_weight(risk: str) -> int:
    return {"상": 300, "중": 150, "하": 0}.get(normalize_risk(risk), 0)


def priority_weight(priority: str) -> int:
    p = clean(priority).upper()
    return {"CORE": 1000, "USABLE": 500, "REFERENCE": 200}.get(p, 0)


def type_weight(content_type: str) -> int:
    return 150 if clean(content_type).lower().startswith("reg") else 0


def parse_date_for_sort(value):
    dt = pd.to_datetime(value, errors="coerce")
    if pd.isna(dt):
        return pd.Timestamp.min
    return dt


def display_date(value) -> str:
    dt = pd.to_datetime(value, errors="coerce")
    if pd.isna(dt):
        return clean(value)[:16]
    if dt.hour == 0 and dt.minute == 0:
        return dt.strftime("%Y-%m-%d")
    return dt.strftime("%Y-%m-%d %H:%M")


def normalize_input(df: pd.DataFrame, content_type: str, source_file: Path) -> pd.DataFrame:
    df = df.copy()
    df.columns = [str(c).strip() for c in df.columns]

    col_date = pick_col(df, ["Date", "date"])
    col_headline = pick_col(df, ["Headline", "Title", "headline"])
    col_url = pick_col(df, ["URL", "Link", "url"])
    col_original_url = pick_col(df, ["resolved_url", "article_url", "original_url", "RepresentativeURL", "CanonicalURL", "FinalURL", "SourceURL"])
    col_rep_url = pick_col(df, ["RepresentativeURL", "rep_url", "cluster_url"])
    col_country = pick_col(df, ["Country", "country"])
    col_agency = pick_col(df, ["Agency", "Publisher", "agency"])
    col_risk = pick_col(df, ["Risk", "risk"])
    col_score = pick_col(df, ["final_score", "Score", "samsung_score", "Importance"])
    col_priority = pick_col(df, ["priority_group", "Priority Group", "Tier"])
    col_issue = pick_col(df, ["issue_type", "Issue", "IssueKey"])
    col_cluster = pick_col(df, ["cluster_key", "Cluster"])
    col_summary = pick_col(df, ["Summary", "summary"])
    col_analysis = pick_col(df, ["AI Analysis", "analysis"])
    col_action = pick_col(df, ["Action Plan", "action"])
    col_source = pick_col(df, ["Source", "SourceFile", "source"])

    out = pd.DataFrame()
    out["Date"] = df[col_date].apply(display_date) if col_date else ""
    out["_sort_date"] = df[col_date].apply(parse_date_for_sort) if col_date else pd.Timestamp.min
    out["Headline"] = df[col_headline].apply(clean) if col_headline else ""

    def _row_best_url(src_row: pd.Series) -> str:
        vals = []
        for c in [col_original_url, col_rep_url, col_url, col_source]:
            if c is not None and c in src_row.index:
                vals.append(src_row.get(c))
        return best_url_from_values(vals)

    out["URL"] = df.apply(_row_best_url, axis=1) if len(df) else ""
    out["Country"] = df[col_country].apply(clean) if col_country else ""
    out["Agency"] = df[col_agency].apply(clean) if col_agency else ""
    out["Risk"] = df[col_risk].apply(normalize_risk) if col_risk else "중"
    out["Importance Score"] = df[col_score].apply(safe_num) if col_score else 0
    out["Priority Group"] = df[col_priority].apply(lambda v: clean(v).upper()) if col_priority else "USABLE"
    out["Issue"] = df[col_issue].apply(clean) if col_issue else ""
    out["Cluster"] = df[col_cluster].apply(clean) if col_cluster else ""
    out["Summary"] = df[col_summary].apply(clean) if col_summary else ""
    out["AI Analysis"] = df[col_analysis].apply(clean) if col_analysis else ""
    out["Action Plan"] = df[col_action].apply(clean) if col_action else ""
    out["Source"] = df[col_source].apply(clean) if col_source else ""
    out["Source File"] = str(source_file)
    out["Content Type"] = content_type
    blank_url = out["URL"].astype(str).str.strip().isin(["", "nan", "None"])
    source_url = out["Source"].astype(str).str.strip().str.startswith("http")
    # Source가 실제 원문 URL인 경우에만 보완. Google RSS는 클릭용 URL로 사용하지 않습니다.
    out.loc[blank_url & source_url, "URL"] = out.loc[blank_url & source_url, "Source"].apply(
        lambda v: best_url_from_values([v])
    )
    # URL이 없는 기사도 임원 보고 대상이면 유지합니다.
    # 단, HTML/Excel에서는 잘못된 Google RSS 또는 https://new 링크를 절대 표시하지 않습니다.
    out = out[out["Headline"].astype(str).str.strip().ne("")]
    return out.reset_index(drop=True)


def read_step4_results() -> pd.DataFrame:
    frames = []
    if REGULATION_INPUT_FILE.exists():
        frames.append(normalize_input(pd.read_excel(REGULATION_INPUT_FILE), "Regulation", REGULATION_INPUT_FILE))
    if NEWS_INPUT_FILE.exists():
        news = normalize_input(pd.read_excel(NEWS_INPUT_FILE), "News", NEWS_INPUT_FILE)
        news = news.head(NEWS_MAX_ROWS)
        frames.append(news)
    if not frames:
        raise FileNotFoundError(f"STEP4 outputs not found: {REGULATION_INPUT_FILE}, {NEWS_INPUT_FILE}")
    rows = pd.concat(frames, ignore_index=True)
    rows = rows[
        rows["Headline"].astype(str).str.strip().ne("")
        & rows["URL"].astype(str).str.strip().ne("")
    ].copy()
    rows = rows.drop_duplicates(subset=["URL", "Headline"], keep="first")
    rows["_integrated_score"] = rows.apply(
        lambda r: priority_weight(r["Priority Group"]) + risk_weight(r["Risk"]) + type_weight(r["Content Type"]) + safe_num(r["Importance Score"]),
        axis=1,
    )
    return rows.reset_index(drop=True)


def assign_mail_groups(rows: pd.DataFrame) -> pd.DataFrame:
    rows = rows.copy()
    rows["Mail Group"] = "News - 주요/참고"
    rows.loc[rows["Content Type"].eq("Regulation"), "Mail Group"] = "Regulation"
    rows.loc[
        rows["Content Type"].eq("News") & rows["Priority Group"].eq("CORE"),
        "Mail Group",
    ] = "News - 핵심"
    rows = apply_samsung_impact(rows)
    return rows


def choose_top3(rows: pd.DataFrame) -> pd.DataFrame:
    pool = rows.copy()
    pool = pool.sort_values(["_integrated_score", "_sort_date"], ascending=[False, False])
    selected = []
    used_clusters = set()
    used_issues = set()
    for _, row in pool.iterrows():
        cluster = clean(row.get("Cluster")) or clean(row.get("Headline"))[:80]
        issue = clean(row.get("Issue")) or clean(row.get("Priority Group"))
        if cluster in used_clusters:
            continue
        if issue in used_issues and len(used_issues) < 3:
            continue
        selected.append(row)
        used_clusters.add(cluster)
        used_issues.add(issue)
        if len(selected) == 3:
            break
    if len(selected) < 3:
        for _, row in pool.iterrows():
            if any(clean(row["Headline"]) == clean(x["Headline"]) for x in selected):
                continue
            selected.append(row)
            if len(selected) == 3:
                break
    return pd.DataFrame(selected).reset_index(drop=True)


def final_order(rows: pd.DataFrame, top3: pd.DataFrame) -> pd.DataFrame:
    rows = assign_mail_groups(rows)
    top_keys = set(top3["Headline"].astype(str))
    top = rows[rows["Headline"].astype(str).isin(top_keys)].copy()
    rest = rows[~rows["Headline"].astype(str).isin(top_keys)].copy()
    group_order = {"Regulation": 0, "News - 핵심": 1, "News - 주요/참고": 2}
    rest["_group_order"] = rest["Mail Group"].map(group_order).fillna(9)
    rest = rest.sort_values(["_group_order", "_integrated_score", "_sort_date"], ascending=[True, False, False])
    top = top.sort_values(["_integrated_score", "_sort_date"], ascending=[False, False])
    out = pd.concat([top, rest], ignore_index=True)
    out["No"] = range(1, len(out) + 1)
    return out


def risk_color(risk: str) -> str:
    return {"상": "#C00000", "중": "#C55A11", "하": "#2F5597"}.get(normalize_risk(risk), "#444")


def html_link(title: str, url: str) -> str:
    title = html.escape(clean(title))
    url = best_url_from_values([url])
    if is_preferred_article_url(url):
        return f'<a href="{html.escape(url)}" target="_blank">{title}</a>'
    # Google News RSS 해제 실패 시 잘못된 리디렉션 링크를 노출하지 않습니다.
    return title


def compact_text(row: pd.Series, field: str, fallback: str) -> str:
    text = clean(row.get(field))
    return text or fallback



def is_english_dominant(text: str) -> bool:
    """HTML 메일 표시용: 영어/JSON 원문이 그대로 노출되는 경우를 감지합니다."""
    t = clean(text)
    if not t:
        return False
    letters = re.findall(r"[A-Za-z]", t)
    korean = re.findall(r"[가-힣]", t)
    # JSON/list/dict 형태가 그대로 나온 경우도 사용자 가독성이 낮으므로 정리 대상
    json_like = ("{'" in t or "[{'" in t or '"step"' in t or "responsibility" in t or "timeline" in t)
    return json_like or (len(letters) >= 35 and len(letters) > len(korean) * 1.8)


def clean_json_like_text(text: str) -> str:
    """Gemini 결과가 Python list/dict처럼 들어온 경우 사람이 읽는 문장으로 최대한 정리합니다."""
    t = clean(text)
    if not t:
        return ""
    # dict/list 표시 제거 및 흔한 key 제거
    replacements = {
        "{": "", "}": "", "[": "", "]": "", "'": "", '"': "",
        "step:": "", "details:": "", "description:": "", "responsibility:": "담당:", "timeline:": "일정:",
        "Step:": "", "Details:": "", "Description:": "", "Responsibility:": "담당:", "Timeline:": "일정:",
    }
    for a, b in replacements.items():
        t = t.replace(a, b)
    t = re.sub(r"\s*,\s*", " / ", t)
    t = re.sub(r"\s*:\s*", ": ", t)
    t = re.sub(r"\s+", " ", t).strip(" /;")
    return t



def row_text_for_issue(row: pd.Series, include_analysis: bool = False) -> str:
    fields = ["Issue", "Cluster", "Headline", "Summary"]
    if include_analysis:
        fields += ["AI Analysis", "Action Plan"]
    return " ".join(clean(row.get(f)) for f in fields).lower()


def infer_topic_ko(row: pd.Series) -> str:
    """토픽 판정은 STEP4 Issue/Headline 중심으로 수행합니다.
    기존 오류 원인: AI Analysis 안의 일반 문구(FTA/원산지)를 먼저 잡아 모든 Top3가 같은 문장으로 출력됨.
    """
    issue = clean(row.get("Issue")).upper()
    headline = clean(row.get("Headline")).lower()
    cluster = clean(row.get("Cluster")).lower()
    base = f"{issue.lower()} {cluster} {headline}"

    exact_map = {
        "FTA_CEPA": "FTA/CEPA",
        "ORIGIN": "원산지/CO",
        "CBAM": "CBAM",
        "AD_CVD": "AD/CVD",
        "EXPORT_CONTROL": "수출통제/제재",
        "ENTITY_LIST": "Entity List",
        "SECTION_301_232": "미국 301/232 관세",
        "SEMICONDUCTOR_TARIFF": "반도체 관세",
        "TARIFF_DUTY": "관세율/할당관세",
        "DUTY_REFUND": "관세환급",
        "TRADE_NEGOTIATION": "통상협상",
        "TARIFF": "관세율",
    }
    if issue in exact_map:
        return exact_map[issue]

    # 명시 토픽 우선순위: 더 구체적인 규제부터 판정
    checks = [
        (["entity list", "entity_list", "우려거래자", "제재명단"], "Entity List"),
        (["export control", "export_control", "sanction", "제재", "수출통제", "iran", "이란"], "수출통제/제재"),
        (["section_301", "section_232", "301", "232", "usmca"], "미국 301/232 관세"),
        (["ad_cvd", "anti-dumping", "antidumping", "countervailing", "반덤핑", "상계관세"], "AD/CVD"),
        (["cbam", "탄소국경"], "CBAM"),
        (["semiconductor", "반도체", "chip"], "반도체 관세"),
        (["duty_refund", "refund", "환급"], "관세환급"),
        (["origin", "원산지", "certificate of origin", "co 발급", "coo"], "원산지/CO"),
        (["fta", "cepa", "협정", "무역협정"], "FTA/CEPA"),
        (["customs", "세관", "통관", "수입신고", "신고"], "통관/세관심사"),
        (["tariff", "duty", "관세", "세율", "할당관세"], "관세율"),
    ]
    for keys, label in checks:
        if any(k in base for k in keys):
            return label
    return clean(row.get("Issue")) or "관세·통상"


SUBSIDIARY_MASTER = {
    # country keyword: 대표 법인/거점
    "UNITED STATES": ["SEA", "SAS", "SSI", "SRA"],
    "USA": ["SEA", "SAS", "SSI", "SRA"],
    "US": ["SEA", "SAS", "SSI", "SRA"],
    "미국": ["SEA", "SAS", "SSI", "SRA"],
    "KOREA": ["SEC"],
    "한국": ["SEC"],
    "INDIA": ["SIEL"],
    "인도": ["SIEL"],
    "VIETNAM": ["SEV", "SEVT", "SEHC"],
    "베트남": ["SEV", "SEVT", "SEHC"],
    "CHINA": ["SSEC", "SSCX", "SEHZ"],
    "중국": ["SSEC", "SSCX", "SEHZ"],
    "MEXICO": ["SEM", "SAMEX"],
    "멕시코": ["SEM", "SAMEX"],
    "BRAZIL": ["SEDA", "SEDB"],
    "브라질": ["SEDA", "SEDB"],
    "EU": ["SEUK", "SEG", "SEF", "SEI", "SEPOL"],
    "EUROPE": ["SEUK", "SEG", "SEF", "SEI", "SEPOL"],
    "GERMANY": ["SEG"],
    "독일": ["SEG"],
    "UK": ["SEUK"],
    "UNITED KINGDOM": ["SEUK"],
    "영국": ["SEUK"],
    "POLAND": ["SEPOL"],
    "폴란드": ["SEPOL"],
    "HUNGARY": ["SEH"],
    "헝가리": ["SEH"],
    "TURKIYE": ["SETK"],
    "TURKEY": ["SETK"],
    "튀르키예": ["SETK"],
    "THAILAND": ["TSE"],
    "태국": ["TSE"],
    "MALAYSIA": ["SEMA"],
    "말레이시아": ["SEMA"],
    "INDONESIA": ["SEIN"],
    "인도네시아": ["SEIN"],
    "CANADA": ["SECA"],
    "캐나다": ["SECA"],
    "UAE": ["SGE"],
    "OMAN": ["SGE"],
}


DIRECT_TOPIC_KEYWORDS = {
    "ENTITY LIST", "EXPORT_CONTROL", "수출통제/제재", "Entity List",
    "SEMICONDUCTOR_TARIFF", "반도체 관세",
    "SECTION_301_232", "미국 301/232 관세",
    "AD_CVD", "AD/CVD",
    "CBAM", "FTA_CEPA", "FTA/CEPA", "ORIGIN", "원산지/CO",
    "DUTY_REFUND", "관세환급", "TARIFF_DUTY", "관세율/할당관세",
}


LOW_RELEVANCE_PATTERNS = [
    "우유", "유제품", "참치", "여행 산업", "공항 세관", "저작권 침해",
    "shot glasses", "sombrero", "마쓰다", "mazda",
]


def split_country_tokens(country: str) -> list[str]:
    raw = clean(country)
    if not raw:
        return []
    parts = re.split(r"[,;/|·&]+|\band\b", raw, flags=re.I)
    return [p.strip() for p in parts if p.strip()]


def matched_subsidiaries(row: pd.Series) -> list[str]:
    country = clean(row.get("Country"))
    text = f"{country} {clean(row.get('Summary'))} {clean(row.get('AI Analysis'))} {clean(row.get('Action Plan'))}"
    text_u = text.upper()
    found = []
    for key, subs in SUBSIDIARY_MASTER.items():
        if key.upper() in text_u:
            found.extend(subs)
    # STEP4 분석문에 이미 SIEL/SEA 등 법인이 직접 기재된 경우 보존
    known_codes = sorted({c for codes in SUBSIDIARY_MASTER.values() for c in codes}, key=len, reverse=True)
    for code in known_codes:
        if re.search(rf"\b{re.escape(code)}\b", text_u):
            found.append(code)
    return list(dict.fromkeys(found))


def is_low_relevance(row: pd.Series) -> bool:
    text = f"{clean(row.get('Headline'))} {clean(row.get('Summary'))} {clean(row.get('AI Analysis'))}".lower()
    return any(p.lower() in text for p in LOW_RELEVANCE_PATTERNS)


def determine_samsung_impact(row: pd.Series) -> tuple[str, str, str]:
    topic = infer_topic_ko(row)
    issue = clean(row.get("Issue")).upper()
    subs = matched_subsidiaries(row)
    country = clean(row.get("Country")) or "국가 미상"
    summary = clean(row.get("Summary"))
    analysis = clean(row.get("AI Analysis"))
    text = f"{summary} {analysis}".lower()

    if is_low_relevance(row) and not subs:
        return "Indirect", "", "삼성전자 제품·법인 직접 연결성이 낮은 참고성 이슈"

    if subs:
        return "Direct", ", ".join(subs), f"Country/본문 법인 매칭: {country} → {', '.join(subs)}"

    if issue in DIRECT_TOPIC_KEYWORDS or topic in DIRECT_TOPIC_KEYWORDS:
        if any(k in text for k in ["삼성전자", "반도체", "휴대폰", "가전", "네트워크", "서버", "수출", "수입", "hs", "관세", "원산지"]):
            return "Direct", "", f"직접영향 토픽({topic})이며 삼성전자 수출입·HS·관세 영향 문구 포함"

    if clean(row.get("Content Type")).lower().startswith("reg"):
        return "Direct", "SEC/HQ", f"법규/공고성 이슈로 본사 관세 Master·신고 프로세스 점검 필요"

    return "Indirect", "", f"국가({country})가 Subsidiary_Master와 직접 매칭되지 않거나 삼성전자 거래 연결성 확인 필요"


def apply_samsung_impact(rows: pd.DataFrame) -> pd.DataFrame:
    rows = rows.copy()
    impacts = rows.apply(determine_samsung_impact, axis=1)
    rows["Samsung Impact"] = [x[0] for x in impacts]
    rows["Affected Subsidiary"] = [x[1] for x in impacts]
    rows["Impact Reason"] = [x[2] for x in impacts]
    return rows


def force_korean_text(row: pd.Series, field: str, fallback: str) -> str:
    """HTML 메일에는 한국어 중심 문장만 표시합니다.
    영어 또는 JSON 형식이 강하게 감지되면 삼성전자 관세 실무 관점의 한국어 문장으로 대체합니다.
    """
    raw = clean(row.get(field))
    if not raw:
        return fallback

    cleaned = clean_json_like_text(raw)
    if not is_english_dominant(raw):
        return cleaned

    topic = infer_topic_ko(row)
    country = clean(row.get("Country")) or "관련 국가"
    headline = clean(row.get("Headline"))

    if field == "Summary":
        return f"{headline} 관련 {topic} 이슈입니다. 원문 기준으로 대상 국가, 적용 품목, 시행일 및 관세·통관 영향 여부 확인이 필요합니다."

    if field == "AI Analysis":
        if topic == "AD/CVD":
            return f"{country}의 AD/CVD 이슈는 대상 HS, 공급국, 과세가격 및 원산지 증빙에 영향을 줄 수 있습니다. 삼성전자 관련 원재료·부품·완제품과의 직접 매칭 여부를 우선 확인해야 합니다."
        if topic == "반도체 관세":
            return f"{country}의 반도체 관세 이슈는 삼성전자 반도체 및 관련 부품의 수입원가, 대미 수출가격, 공급망 전략에 직접 영향을 줄 수 있습니다. 대상 HS와 적용 시점, 예외·환급 가능성을 별도 확인해야 합니다."
        if topic in ["FTA/CEPA", "원산지/CO"]:
            return f"{country}의 {topic} 이슈는 생산거점별 원산지 판정, CO 발급, 특혜세율 적용 가능성에 영향을 줄 수 있습니다. 관련 제품군의 원산지 기준 충족 여부와 증빙 체계를 재점검해야 합니다."
        if topic == "CBAM":
            return f"{country}의 CBAM 이슈는 알루미늄·철강 등 원재료 조달, 탄소자료 확보, EU향 수입신고 증빙 체계에 영향을 줄 수 있습니다. 관련 법인의 공급망 및 탄소데이터 확보 수준을 점검해야 합니다."
        if topic in ["수출통제/제재", "Entity List"]:
            return f"{country}의 {topic} 이슈는 고객·거래처 스크리닝, ECCN/전략물자 분류, 제재국 우회거래 점검에 직접 영향을 줄 수 있습니다."
        if topic == "통관/세관심사":
            return f"{country}의 통관·세관심사 이슈는 신고 지연, HS 오류, 과세가격 및 증빙 보관 리스크로 연결될 수 있습니다. 법인별 신고 프로세스와 관세사 제출자료의 정합성 점검이 필요합니다."
        return f"{country}의 {topic} 이슈는 삼성전자 주요 생산거점·판매법인의 관세원가, HS 분류, 원산지 및 통관 내부통제에 영향을 줄 수 있습니다. 대상 품목과 법인별 실제 거래 여부를 우선 확인해야 합니다."

    if field == "Action Plan":
        if topic == "AD/CVD":
            return "① 규제 대상 HS·품목 확인 ② 삼성전자 수입·수출 품목과 매칭 ③ 공급국·공급업체별 거래 여부 점검 ④ 가격자료·원산지 증빙 및 방어자료 준비"
        if topic == "반도체 관세":
            return "① 반도체 관련 대상 HS 및 세율 확인 ② 생산거점·판매법인별 관세원가 영향 산출 ③ 예외·환급·유예 가능성 검토 ④ 사업부 및 법무·관세사와 대응 시나리오 공유"
        if topic in ["FTA/CEPA", "원산지/CO"]:
            return "① 협정문·원산지 기준 확인 ② 제품별 BOM/공정 기준 충족 여부 점검 ③ CO 발급·보관 증빙 재확인 ④ 특혜세율 적용 가능성과 가격 영향 산출"
        if topic == "CBAM":
            return "① 대상 원재료·HS 확인 ② EU향 제품 사용 여부 점검 ③ 공급사 탄소자료 확보 ④ CBAM 신고·증빙 체계와 구매전략 연계"
        if topic in ["수출통제/제재", "Entity List"]:
            return "① 거래처·고객 스크리닝 ② ECCN/전략물자 분류 재점검 ③ 제재국 우회거래 여부 확인 ④ Hold/Release 의사결정 기록화"
        if topic == "통관/세관심사":
            return "① 적용 품목·시행일 확인 ② HS·과세가격·신고시점 점검 ③ 관세사 업무지침 및 증빙 체크리스트 개정 ④ 법인별 오류 발생 가능성 점검"
        return "① 원문 공고·법령 확인 ② 대상 국가·HS·제품군 매핑 ③ 법인별 수출입 신고 영향 검토 ④ 필요 시 관세사·법무·사업부 공동 대응"

    return cleaned or fallback


def top3_one_line_summary(row: pd.Series) -> str:
    """총평 하단 Top3 1줄 요약용 문장. 반드시 헤드라인별로 다르게 생성."""
    topic = infer_topic_ko(row)
    country = clean(row.get("Country")) or "주요국"
    title = clean(row.get("Headline"))
    title_short = re.sub(r"\s+", " ", title)
    if len(title_short) > 42:
        title_short = title_short[:42].rstrip() + "…"
    impact = clean(row.get("Samsung Impact")) or "Indirect"
    subs = clean(row.get("Affected Subsidiary"))
    prefix = f"{topic} / {country}"
    if impact == "Direct":
        prefix += f" / 직접영향{('(' + subs + ')') if subs else ''}"

    if topic == "AD/CVD":
        action = "대상 HS·공급국·가격자료 및 원산지 방어자료 확인"
    elif topic == "반도체 관세":
        action = "대상 HS·세율·적용시점 확인 및 반도체 공급망 원가 영향 산출"
    elif topic in ["FTA/CEPA", "원산지/CO"]:
        action = "협정세율·원산지 기준·CO 발급 가능성 및 법인별 활용 여부 재점검"
    elif topic == "CBAM":
        action = "알루미늄·철강 등 대상 원재료와 탄소자료 확보 체계 점검"
    elif topic in ["수출통제/제재", "Entity List"]:
        action = "거래처 스크리닝·ECCN·제재국 우회거래 여부 즉시 확인"
    elif topic == "관세환급":
        action = "환급 프로세스·증빙 자율발급·시스템 반영 여부 확인"
    elif topic == "미국 301/232 관세":
        action = "미국향 수출입 HS·원산지·추가관세 노출도 점검"
    else:
        action = "대상 국가·HS·제품군 기준으로 삼성전자 관세·통상 영향 확인"
    return f"{prefix} → {title_short} | {action}"



def build_overall_review_html(rows: pd.DataFrame, top3: pd.DataFrame) -> str:
    regulation = rows[rows["Mail Group"].eq("Regulation")]
    news_core = rows[rows["Mail Group"].eq("News - 핵심")]
    news_usable = rows[rows["Mail Group"].eq("News - 주요/참고")]
    direct_rows = rows[rows["Samsung Impact"].eq("Direct")]
    indirect_rows = rows[rows["Samsung Impact"].ne("Direct")]

    issues = []
    for _, r in top3.iterrows():
        issue = infer_topic_ko(r)
        if issue not in issues:
            issues.append(issue)
    issue_text = "·".join(issues[:3]) if issues else "관세·통상"

    top_direct = direct_rows.sort_values(["_integrated_score", "_sort_date"], ascending=[False, False]).head(3)
    top_indirect = indirect_rows.sort_values(["_integrated_score", "_sort_date"], ascending=[False, False]).head(5)

    total_line = (
        f"금일 GTI Radar는 {issue_text} 이슈를 중심으로 선별되었으며, "
        f"삼성전자 직접영향 {len(direct_rows)}건 / 간접영향 {len(indirect_rows)}건으로 구분됩니다. "
        f"직접영향 건은 해당 법인의 HS·원산지·관세율·수출통제·환급 프로세스 반영 여부를 우선 확인해야 합니다."
    )

    bullets = []
    seen = set()
    for idx, (_, r) in enumerate(top3.iterrows(), 1):
        line = top3_one_line_summary(r)
        key = re.sub(r"[^가-힣A-Za-z0-9]+", "", line[:80]).lower()
        if key in seen:
            line = f"{line} / 중복 이슈 아님: {clean(r.get('Headline'))[:36]}"
        seen.add(key)
        bullets.append(
            "<div style='margin-top:8px;line-height:1.8;'>"
            f"• Top{idx}. {html.escape(line)}"
            "</div>"
        )

    def mini_list(title: str, frame: pd.DataFrame, color: str) -> str:
        if frame.empty:
            return f"<div style='margin-top:8px;color:#777;'>{html.escape(title)}: 해당 없음</div>"
        items = []
        for _, r in frame.iterrows():
            subs = clean(r.get("Affected Subsidiary")) or "-"
            headline = clean(r.get("Headline"))
            if len(headline) > 58:
                headline = headline[:58].rstrip() + "…"
            items.append(
                f"<li style='margin:3px 0;'><b>{html.escape(infer_topic_ko(r))}</b> "
                f"[{html.escape(clean(r.get('Country')))} / {html.escape(subs)}] "
                f"{html.escape(headline)}</li>"
            )
        return (
            f"<div style='margin-top:12px;'><b style='color:{color};'>{html.escape(title)}</b>"
            f"<ol style='margin-top:6px;margin-bottom:4px;padding-left:22px;'>{''.join(items)}</ol></div>"
        )

    counts = (
        f"법규 {len(regulation)}건, 핵심 뉴스 {len(news_core)}건, 주요/참고 뉴스 {len(news_usable)}건 "
        f"| 직접영향 {len(direct_rows)}건, 간접영향 {len(indirect_rows)}건"
    )
    return f"""
    <div style="margin-top:10px;margin-bottom:20px;padding:16px;background:#F4F6F8;border-left:6px solid #1F4E78;color:#222;">
      <div style="font-size:14px;color:#555;margin-bottom:8px;">금일 선별 결과: {html.escape(counts)}</div>
      <div style="font-size:15px;font-weight:bold;line-height:1.8;margin-bottom:12px;">{html.escape(total_line)}</div>
      {''.join(bullets)}
      {mini_list('직접영향 Top 3', top_direct, '#C00000')}
      {mini_list('간접영향 Top 5', top_indirect, '#7F7F7F')}
    </div>
    """

def display_topic(row: pd.Series) -> str:
    """HTML 메일용 Topic 표시값.
    STEP4의 Issue 컬럼을 우선 사용하고, 비어 있으면 Cluster/Priority Group/Mail Group 순으로 보완합니다.
    """
    for field in ["Issue", "Cluster", "Priority Group", "Mail Group"]:
        value = clean(row.get(field))
        if value:
            return value
    return "-"


def build_table(title: str, rows: pd.DataFrame, color: str) -> str:
    """메일 본문 섹션 테이블.
    기존 Form 기준으로 Topic / Agency / Publish Date 컬럼을 복원합니다.
    """
    if rows.empty:
        return ""
    trs = []
    for _, row in rows.iterrows():
        trs.append(f"""
        <tr>
          <td style="padding:7px;border:1px solid #d9d9d9;text-align:center;">{html.escape(str(row['No']))}</td>
          <td style="padding:7px;border:1px solid #d9d9d9;text-align:center;">{html.escape(display_topic(row))}</td>
          <td style="padding:7px;border:1px solid #d9d9d9;text-align:center;font-weight:bold;color:{'#C00000' if clean(row.get('Samsung Impact')) == 'Direct' else '#666'};">{html.escape(clean(row.get('Samsung Impact')))}</td>
          <td style="padding:7px;border:1px solid #d9d9d9;text-align:center;">{html.escape(clean(row.get('Affected Subsidiary')) or '-')}</td>
          <td style="padding:7px;border:1px solid #d9d9d9;">{html_link(row['Headline'], row['URL'])}</td>
          <td style="padding:7px;border:1px solid #d9d9d9;">{html.escape(force_korean_text(row, 'Summary', '요약 정보 확인 필요'))}</td>
          <td style="padding:7px;border:1px solid #d9d9d9;">{html.escape(force_korean_text(row, 'AI Analysis', '영향 검토 필요'))}</td>
          <td style="padding:7px;border:1px solid #d9d9d9;">{html.escape(force_korean_text(row, 'Action Plan', '담당 부서 확인 필요'))}</td>
          <td style="padding:7px;border:1px solid #d9d9d9;text-align:center;">{html.escape(clean(row['Country']))}</td>
          <td style="padding:7px;border:1px solid #d9d9d9;text-align:center;">{html.escape(clean(row['Agency']))}</td>
          <td style="padding:7px;border:1px solid #d9d9d9;text-align:center;color:{risk_color(row['Risk'])};font-weight:bold;">{html.escape(clean(row['Risk']))}</td>
          <td style="padding:7px;border:1px solid #d9d9d9;text-align:center;">{html.escape(clean(row['Date']))}</td>
        </tr>
        """)
    return f"""
    <h3 style="margin-top:24px;color:{color};">{html.escape(title)} ({len(rows)}건)</h3>
    <table style="border-collapse:collapse;width:100%;font-size:12px;">
      <thead>
        <tr style="background:{color};color:white;">
          <th style="padding:7px;border:1px solid #d9d9d9;width:42px;">No</th>
          <th style="padding:7px;border:1px solid #d9d9d9;width:110px;">Topic</th>
          <th style="padding:7px;border:1px solid #d9d9d9;width:80px;">Impact</th>
          <th style="padding:7px;border:1px solid #d9d9d9;width:100px;">Subsidiary</th>
          <th style="padding:7px;border:1px solid #d9d9d9;width:23%;">Headline</th>
          <th style="padding:7px;border:1px solid #d9d9d9;">Summary</th>
          <th style="padding:7px;border:1px solid #d9d9d9;">Impact</th>
          <th style="padding:7px;border:1px solid #d9d9d9;">Action</th>
          <th style="padding:7px;border:1px solid #d9d9d9;width:90px;">Country</th>
          <th style="padding:7px;border:1px solid #d9d9d9;width:130px;">Agency</th>
          <th style="padding:7px;border:1px solid #d9d9d9;width:52px;">Risk</th>
          <th style="padding:7px;border:1px solid #d9d9d9;width:95px;">Publish Date</th>
        </tr>
      </thead>
      <tbody>{''.join(trs)}</tbody>
    </table>
    """


def build_html(rows: pd.DataFrame, top3: pd.DataFrame) -> str:
    subject = f"[GTI Radar] Global Trade Intelligence | {RUN_DATE}"
    top_blocks = []
    for idx, row in top3.iterrows():
        top_blocks.append(f"""
        <div style="margin:14px 0 16px 0;padding:14px;border-left:5px solid #C00000;background:#FFF7F7;">
          <div style="font-size:15px;font-weight:bold;margin-bottom:6px;">Top {idx + 1}. {html_link(row['Headline'], row['URL'])}</div>
          <div style="font-size:12px;color:#555;margin-bottom:8px;">Type: {html.escape(clean(row['Content Type']))} | Group: {html.escape(clean(row['Mail Group']))} | Topic: {html.escape(display_topic(row))} | Samsung Impact: <b>{html.escape(clean(row.get('Samsung Impact')))}</b> | Subsidiary: {html.escape(clean(row.get('Affected Subsidiary')) or '-')} | Agency: {html.escape(clean(row['Agency']))} | Publish Date: {html.escape(clean(row['Date']))} | Country: {html.escape(clean(row['Country']))} | Risk: <span style="color:{risk_color(row['Risk'])};font-weight:bold;">{html.escape(clean(row['Risk']))}</span> | Score: {safe_num(row['Importance Score'])}</div>
          <div style="margin-top:7px;"><b>요약</b><br>{html.escape(force_korean_text(row, 'Summary', '요약 정보 확인 필요'))}</div>
          <div style="margin-top:7px;"><b>영향</b><br>{html.escape(force_korean_text(row, 'AI Analysis', '영향 검토 필요'))}</div>
          <div style="margin-top:7px;"><b>대응조치</b><br>{html.escape(force_korean_text(row, 'Action Plan', '담당 부서 확인 필요'))}</div>
        </div>
        """)

    regulation = rows[rows["Mail Group"].eq("Regulation")]
    news_core = rows[rows["Mail Group"].eq("News - 핵심")]
    news_usable = rows[rows["Mail Group"].eq("News - 주요/참고")]
    return f"""<!DOCTYPE html>
<html lang="ko">
<head><meta charset="utf-8"><title>{html.escape(subject)}</title></head>
<body style="font-family:Arial,'Malgun Gothic',sans-serif;font-size:13px;color:#222;line-height:1.5;">
  <div style="max-width:1320px;margin:0 auto;">
    <h2 style="margin-bottom:4px;color:#1F4E78;">[GTI Radar] Global Trade Intelligence</h2>
    <div style="font-size:14px;margin-bottom:4px;"><b>Date:</b> {RUN_DATE}</div>
    <div style="font-size:12px;color:#555;margin-bottom:16px;">Focus: Samsung Electronics Customs & Trade Intelligence</div>

    <h3 style="margin-top:18px;margin-bottom:6px;">총평</h3>
    {build_overall_review_html(rows, top3)}

    <h3 style="color:#C00000;margin-top:22px;">Top 3 통합 중요 이슈</h3>
    {''.join(top_blocks)}

    {build_table('Section 1: Regulation', regulation, '#1F4E78')}
    {build_table('Section 2-1: News CORE', news_core, '#548235')}
    {build_table('Section 2-2: News USABLE / Reference', news_usable, '#7F7F7F')}
    <p style="margin-top:18px;color:#666;font-size:12px;">첨부 Excel에는 STEP4 선별 결과와 메일 그룹이 함께 포함되어 있습니다.</p>
  </div>
</body>
</html>"""


def style_sheet(ws) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border
        ws.row_dimensions[row[0].row].height = 76
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def append_row(ws, row: pd.Series) -> None:
    ws.append([row.get(c, "") for c in OUTPUT_COLUMNS])
    headline_cell = ws.cell(row=ws.max_row, column=OUTPUT_COLUMNS.index("Headline") + 1)
    url = best_url_from_values([row.get("URL")])
    if is_preferred_article_url(url):
        headline_cell.hyperlink = url
        headline_cell.font = Font(color="0563C1", underline="single", bold=True)


def save_excel(rows: pd.DataFrame, top3: pd.DataFrame, paths: dict[str, Path]) -> None:
    wb = Workbook()
    sheets = [
        ("GTI Radar", rows),
        ("Top3", top3),
        ("Regulation", rows[rows["Mail Group"].eq("Regulation")]),
        ("News CORE", rows[rows["Mail Group"].eq("News - 핵심")]),
        ("News USABLE", rows[rows["Mail Group"].eq("News - 주요/참고")]),
    ]
    first = True
    for name, frame in sheets:
        ws = wb.active if first else wb.create_sheet(name)
        first = False
        ws.title = name
        ws.append(OUTPUT_COLUMNS)
        for _, row in frame.iterrows():
            append_row(ws, row)
        widths = {
            "A": 6, "B": 14, "C": 18, "D": 14, "E": 18, "F": 42,
            "G": 14, "H": 48, "I": 52, "J": 52, "K": 68, "L": 16,
            "M": 22, "N": 8, "O": 14, "P": 14, "Q": 20, "R": 28,
            "S": 36, "T": 30, "U": 38,
        }
        for col, width in widths.items():
            ws.column_dimensions[col].width = width
        style_sheet(ws)

    runlog = wb.create_sheet("Run Log")
    runlog.append(["item", "value"])
    runlog.append(["regulation_input", str(REGULATION_INPUT_FILE)])
    runlog.append(["news_input", str(NEWS_INPUT_FILE)])
    runlog.append(["run_date", RUN_DATE])
    runlog.append(["news_max_rows", NEWS_MAX_ROWS])
    runlog.append(["selected_total", len(rows)])
    runlog.append(["regulation_rows", int(rows["Mail Group"].eq("Regulation").sum())])
    runlog.append(["news_core_rows", int(rows["Mail Group"].eq("News - 핵심").sum())])
    runlog.append(["news_usable_rows", int(rows["Mail Group"].eq("News - 주요/참고").sum())])
    style_sheet(runlog)

    wb.save(paths["mail_xlsx"])
    wb.save(paths["analysis"])
    rows[OUTPUT_COLUMNS].to_excel(paths["cumulative"], index=False)


def load_recipients() -> list[str]:
    recipients = [
        x.strip()
        for x in re.split(r"[;,]", MAIL_TO)
        if re.fullmatch(r"[A-Za-z0-9._%+\-]+@[A-Za-z0-9.\-]+\.[A-Za-z]{2,}", x.strip())
    ]
    if recipients:
        return list(dict.fromkeys(recipients))
    for fp in [RECIPIENT_FILE, Path(r"C:\Temp\00.xlsx"), Path(r"C:\Temp\mail.xlsx")]:
        if not fp.exists():
            continue
        try:
            df = pd.read_excel(fp)
            values = df.astype(str).values.ravel().tolist()
            found = [
                v.strip()
                for v in values
                if re.fullmatch(r"[A-Za-z0-9._%+\-]+@[A-Za-z0-9.\-]+\.[A-Za-z]{2,}", v.strip())
            ]
            if found:
                return list(dict.fromkeys(found))
        except Exception:
            continue
    return []


def send_email(html_body: str, attachment: Path) -> None:
    if not SEND_EMAIL:
        print("[MAIL] skipped: GTI_SEND_EMAIL=N")
        return
    recipients = load_recipients()
    if not recipients:
        print("[MAIL] skipped: no recipients")
        return
    if not SMTP_PASS:
        print("[MAIL] skipped: SMTP password not set")
        return
    msg = EmailMessage()
    msg["Subject"] = f"[GTI Radar] Global Trade Intelligence | {RUN_DATE}"
    msg["From"] = formataddr((MAIL_FROM_NAME, SMTP_USER))
    msg["To"] = ", ".join(recipients)
    msg.set_content("HTML 메일을 확인해 주세요.")
    msg.add_alternative(html_body, subtype="html")
    data = attachment.read_bytes()
    msg.add_attachment(
        data,
        maintype="application",
        subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=attachment.name,
    )
    context = ssl.create_default_context()
    with smtplib.SMTP_SSL(SMTP_HOST, SMTP_PORT, context=context) as server:
        server.login(SMTP_USER, SMTP_PASS)
        server.send_message(msg)
    print(f"[MAIL SENT] {len(recipients)} recipients")



# =========================================================
# V3.0 EXECUTIVE INTELLIGENCE OVERRIDES
# - Executive Summary: article list -> Samsung impact narrative
# - Today's Required Actions: owner/action driven
# - Residual noise filter in STEP5
# - Direct impact shrink and cluster representative display
# =========================================================

RESIDUAL_NOISE_PATTERNS = [
    "농정", "농민", "농업", "농가", "축산", "축산업", "돼지고기", "양돈", "한우",
    "우유", "낙농", "유제품", "치즈", "버터", "영양제", "supplement", "vitamin",
    "시장 후보", "당진시장", "선거", "지방선거", "후보", "시민사회", "정책협약",
    "블라우스", "의류", "패션", "화장품", "k-beauty", "mask pack", "마스크팩",
    "일자리", "채용", "노조", "성과급", "수상", "관세대상", "출입국", "현금과 금",
    "개인", "관광", "여행", "공항", "연료", "휘발유", "디젤", "petrol", "diesel",
]

EXECUTIVE_OWNER_MAP = {
    "CBAM": "구매/ESG/관세",
    "원산지/CO": "FTA 운영/생산법인",
    "FTA/CEPA": "FTA 운영/생산법인",
    "관세율": "통관운영/사업부",
    "관세율/할당관세": "통관운영/세무",
    "반도체 관세": "DS사업부/통관기획",
    "수출통제/제재": "수출통제/법무",
    "Entity List": "수출통제/법무",
    "AD/CVD": "통상대응/구매",
    "통관/세관심사": "통관운영/관세사",
    "관세환급": "관세환급/세무",
}



# ---- V3.0.1 compatibility helpers: required by V3.0 override block ----
BROAD_COUNTRY_VALUES = {"", "GLOBAL", "WORLD", "WORLDWIDE", "INTERNATIONAL", "MULTIPLE", "VARIOUS", "글로벌", "전세계", "세계", "-"}

SAMSUNG_PRODUCT_PATTERNS_V3 = [
    "삼성전자", "samsung", "반도체", "semiconductor", "chip", "memory", "메모리", "hbm",
    "display", "디스플레이", "mobile", "phone", "smartphone", "휴대폰", "스마트폰",
    "tv", "television", "가전", "appliance", "electronics", "전자", "server", "서버", "network", "네트워크",
    "battery", "배터리", "pcb", "camera module", "module", "component", "부품", "완제품",
    "aluminium", "aluminum", "알루미늄", "steel", "철강", "steel sheet", "copper", "구리",
    "customs", "tariff", "origin", "fta", "cepa", "cbam", "export control", "uflpa", "hs",
    "관세", "원산지", "수출통제", "품목분류", "공급망", "제조", "생산", "수입", "수출"
]

def country_based_subsidiaries(country: str) -> list[str]:
    """Country 컬럼만 기준으로 삼성 법인을 매칭합니다. Global/World는 제외합니다."""
    found = []
    for token in split_country_tokens(country):
        token_u = clean(token).upper()
        if token_u in BROAD_COUNTRY_VALUES:
            continue
        # normalize common variants
        if token_u in {"UNITED STATES OF AMERICA", "AMERICA"}:
            token_u = "UNITED STATES"
        if token_u in {"SOUTH KOREA", "REPUBLIC OF KOREA"}:
            token_u = "KOREA"
        if token_u in {"BRITAIN", "GREAT BRITAIN"}:
            token_u = "UNITED KINGDOM"
        for key, subs in SUBSIDIARY_MASTER.items():
            key_u = clean(key).upper()
            if token_u == key_u:
                found.extend(subs)
    return list(dict.fromkeys(found))

def explicit_subsidiaries_in_text(row: pd.Series) -> list[str]:
    """SEA/SIEL 등 법인 코드가 본문에 명시된 경우에만 보조 매칭합니다."""
    text_u = " ".join(clean(row.get(f)) for f in ["Headline", "Summary", "AI Analysis", "Action Plan", "Affected Subsidiary"]).upper()
    known_codes = sorted({c for codes in SUBSIDIARY_MASTER.values() for c in codes}, key=len, reverse=True)
    found = []
    for code in known_codes:
        if re.search(rf"\b{re.escape(code)}\b", text_u):
            found.append(code)
    return list(dict.fromkeys(found))

def has_samsung_product_signal(row: pd.Series) -> bool:
    text = f"{clean(row.get('Headline'))} {clean(row.get('Summary'))} {clean(row.get('AI Analysis'))} {clean(row.get('Action Plan'))}".lower()
    return any(p.lower() in text for p in SAMSUNG_PRODUCT_PATTERNS_V3)

def is_global_country(row: pd.Series) -> bool:
    country = clean(row.get("Country")).upper()
    return country in BROAD_COUNTRY_VALUES


def is_residual_noise_v3(row: pd.Series) -> bool:
    text = f"{clean(row.get('Headline'))} {clean(row.get('Summary'))} {clean(row.get('AI Analysis'))} {clean(row.get('Action Plan'))}".lower()
    if any(p.lower() in text for p in RESIDUAL_NOISE_PATTERNS):
        # 단, 실제 삼성/반도체/수출통제/CBAM 직접 키워드가 강한 경우는 살림
        strong = any(k in text for k in ["samsung", "삼성전자", "semiconductor", "반도체", "cbam", "uflpa", "entity list", "수출통제"])
        return not strong
    return False


def cluster_key_v3(row: pd.Series) -> str:
    topic = infer_topic_ko(row)
    country = clean(row.get("Country")).upper()
    title = clean(row.get("Headline")).lower()

    # Known issue clusters
    if "india" in country and "united kingdom" in country and any(k in title for k in ["uk", "britain", "fta", "trade pact", "steel", "cbam", "carbon"]):
        return "INDIA_UK_FTA_CBAM_STEEL"
    if ("india" in country and "oman" in title and "cepa" in title):
        return "INDIA_OMAN_CEPA_ORIGIN"
    if "vietnam" in country and any(k in title for k in ["asean", "china", "acfta", "free trade"]):
        return "VIETNAM_ACFTA_ORIGIN"
    if "korea" in country and "india" in country and "cepa" in title:
        return "KOREA_INDIA_CEPA_UPGRADE"
    if topic == "CBAM" and ("vietnam" in country or "베트남" in country):
        return "KR_VN_CBAM_SUPPLY_CHAIN"
    if topic == "반도체 관세":
        return "SEMICONDUCTOR_TARIFF_GENERAL"
    if topic in {"원산지/CO", "FTA/CEPA"}:
        # use country pair + topic
        return f"{topic}:{country}:{re.sub(r'[^a-z0-9가-힣]+',' ',title)[:60]}"
    base = clean(row.get("Cluster")) or f"{topic}:{country}:{title[:70]}"
    return re.sub(r"\s+", " ", base).strip()


def determine_samsung_impact(row: pd.Series) -> tuple[str, str, str]:
    """V3.0: Direct must be truly actionable for Samsung.
    - Country match alone is never Direct.
    - FTA negotiation/visit/talks are normally Indirect.
    - Direct if implementation/rule/effective date or CBAM/export control/semiconductor/AD-CVD and Samsung country/product signal exists.
    """
    topic = infer_topic_ko(row)
    issue = clean(row.get("Issue")).upper()
    title = clean(row.get("Headline")).lower()
    text = f"{title} {clean(row.get('Summary')).lower()} {clean(row.get('AI Analysis')).lower()}"
    country = clean(row.get("Country")) or "국가 미상"
    country_subs = country_based_subsidiaries(country)
    explicit_subs = explicit_subsidiaries_in_text(row)
    subs = list(dict.fromkeys(country_subs + explicit_subs))

    if clean(row.get("Content Type")).lower().startswith("reg"):
        return "Direct", "SEC/HQ", "공식 법규/공고성 이슈로 본사 관세 Master·신고 프로세스 반영 필요"

    if is_residual_noise_v3(row) or is_low_relevance(row):
        return "None", "", "삼성전자 관세·통상 업무 직접 관련성이 낮아 메일 본문 제외 대상"

    negotiation_only = any(k in title for k in ["talks", "negotiation", "visit", "meet", "momentum", "proposal", "assess", "후보", "논의", "방문", "협상"])
    implementation_signal = any(k in text for k in [
        "rules of origin", "rule of origin", "origin rules", "in effect", "effective", "implementation",
        "comes into force", "now in effect", "적용", "시행", "발효", "원산지 규정", "시행규칙"
    ])
    product_signal = has_samsung_product_signal(row)
    direct_reg_topic = topic in {"CBAM", "반도체 관세", "수출통제/제재", "Entity List", "AD/CVD", "관세환급", "통관/세관심사", "관세율/할당관세"}
    fta_direct = topic in {"원산지/CO", "FTA/CEPA"} and implementation_signal and bool(country_subs)

    if is_global_country(row) and not explicit_subs and not product_signal:
        return "Indirect", "", "Global 이슈로 특정 삼성 법인의 직접영향 확정 불가"

    if explicit_subs and (direct_reg_topic or fta_direct):
        return "Direct", ", ".join(subs), f"법인 코드 명시 + 실행형 규제 토픽({topic})"

    if country_subs and (direct_reg_topic or fta_direct) and (product_signal or topic in {"CBAM", "수출통제/제재", "반도체 관세", "AD/CVD"}):
        return "Direct", ", ".join(country_subs), f"법인국가 매칭 + 실행형 규제 토픽({topic}) + 제품/업무 관련성 확인"

    if negotiation_only and topic in {"원산지/CO", "FTA/CEPA", "관세율"}:
        return "Indirect", ", ".join(country_subs), "협상·방문·논의 단계로 즉시 신고/원산지 프로세스 변경 전 단계"

    if country_subs:
        return "Indirect", ", ".join(country_subs), f"법인 소재국 관련 이슈이나 제품·거래·시행조건 직접 연결은 추가 확인 필요"

    return "Indirect", "", f"토픽({topic})은 모니터링 가치가 있으나 특정 삼성 법인 직접영향은 확인 불가"


def apply_samsung_impact(rows: pd.DataFrame) -> pd.DataFrame:
    rows = rows.copy()
    impacts = rows.apply(determine_samsung_impact, axis=1)
    rows["Samsung Impact"] = [x[0] for x in impacts]
    rows["Affected Subsidiary"] = [x[1] for x in impacts]
    rows["Impact Reason"] = [x[2] for x in impacts]
    rows["_display_cluster"] = rows.apply(cluster_key_v3, axis=1)
    return rows


def assign_mail_groups(rows: pd.DataFrame) -> pd.DataFrame:
    rows = rows.copy()
    rows["Mail Group"] = "News - 주요/참고"
    rows.loc[rows["Content Type"].eq("Regulation"), "Mail Group"] = "Regulation"
    rows.loc[rows["Content Type"].eq("News") & rows["Priority Group"].eq("CORE"), "Mail Group"] = "News - 핵심"
    rows = apply_samsung_impact(rows)
    # residual None is removed from visible mail sections
    rows.loc[rows["Samsung Impact"].eq("None"), "Mail Group"] = "Filtered Noise"
    return rows


def executive_sort_frame(rows: pd.DataFrame) -> pd.DataFrame:
    rows = rows.copy()
    impact_weight = rows["Samsung Impact"].map({"Direct": 1200, "Indirect": 400, "None": -9999}).fillna(0)
    action_weight = rows.apply(lambda r: 400 if infer_topic_ko(r) in {"CBAM", "반도체 관세", "수출통제/제재", "Entity List", "AD/CVD", "원산지/CO", "관세율/할당관세"} else 0, axis=1)
    rows["_exec_score"] = rows["_integrated_score"] + impact_weight + action_weight
    return rows.sort_values(["_exec_score", "_sort_date"], ascending=[False, False])


def choose_top3(rows: pd.DataFrame) -> pd.DataFrame:
    pool = rows[~rows["Mail Group"].eq("Filtered Noise")].copy()
    if pool.empty:
        return pool
    pool = executive_sort_frame(pool)
    selected = []
    used_clusters, used_topics = set(), set()
    for _, row in pool.iterrows():
        ck = clean(row.get("_display_cluster")) or cluster_key_v3(row)
        topic = infer_topic_ko(row)
        if ck in used_clusters:
            continue
        # prefer distinct topics but do not force bad rows
        if topic in used_topics and len(selected) < 2:
            continue
        selected.append(row)
        used_clusters.add(ck)
        used_topics.add(topic)
        if len(selected) >= 3:
            break
    if len(selected) < 3:
        for _, row in pool.iterrows():
            if any(clean(row["Headline"]) == clean(x["Headline"]) for x in selected):
                continue
            if clean(row.get("_display_cluster")) in used_clusters:
                continue
            selected.append(row)
            used_clusters.add(clean(row.get("_display_cluster")))
            if len(selected) >= 3:
                break
    return pd.DataFrame(selected).reset_index(drop=True)


def final_order(rows: pd.DataFrame, top3: pd.DataFrame) -> pd.DataFrame:
    rows = assign_mail_groups(rows)
    visible = rows[~rows["Mail Group"].eq("Filtered Noise")].copy()

    # Cluster representative: keep highest executive score per cluster for visible news
    visible = executive_sort_frame(visible)
    visible = visible.drop_duplicates(subset=["_display_cluster"], keep="first")

    # Limit visible volume to executive-ready level: Direct 6 + Indirect 6 + all regulation
    reg = visible[visible["Mail Group"].eq("Regulation")].copy()
    direct = visible[(visible["Samsung Impact"].eq("Direct")) & (~visible["Mail Group"].eq("Regulation"))].head(6)
    indirect = visible[(visible["Samsung Impact"].eq("Indirect")) & (~visible["Mail Group"].eq("Regulation"))].head(6)
    visible = pd.concat([reg, direct, indirect], ignore_index=True)

    top_keys = set(top3["Headline"].astype(str))
    top = visible[visible["Headline"].astype(str).isin(top_keys)].copy()
    rest = visible[~visible["Headline"].astype(str).isin(top_keys)].copy()
    group_order = {"Regulation": 0, "News - 핵심": 1, "News - 주요/참고": 2}
    rest["_group_order"] = rest["Mail Group"].map(group_order).fillna(9)
    rest = executive_sort_frame(rest).sort_values(["_group_order", "_exec_score", "_sort_date"], ascending=[True, False, False])
    top = executive_sort_frame(top)
    out = pd.concat([top, rest], ignore_index=True)
    out["No"] = range(1, len(out) + 1)
    return out


def executive_issue_sentence(row: pd.Series) -> str:
    topic = infer_topic_ko(row)
    country = clean(row.get("Country")) or "주요국"
    subs = clean(row.get("Affected Subsidiary"))
    title = clean(row.get("Headline"))
    if topic in {"원산지/CO", "FTA/CEPA"}:
        return f"{country} FTA/원산지 이슈는 {subs or '관련 법인'}의 CO 발급, 원산지 판정 및 협정세율 적용 기준 재점검이 필요합니다."
    if topic == "CBAM":
        return f"{country} CBAM 이슈는 원재료 탄소자료 확보와 EU향 신고 증빙 체계에 영향을 줄 수 있어 {subs or '관련 법인'} 공급망 점검이 필요합니다."
    if topic == "반도체 관세":
        return "반도체 관세 이슈는 글로벌 생산거점의 관세원가, HS 분류 및 공급망 전략에 영향을 줄 수 있어 대상 HS와 적용시점 모니터링이 필요합니다."
    if topic in {"수출통제/제재", "Entity List"}:
        return f"{country} 수출통제/제재 이슈는 거래처 스크리닝, ECCN/전략물자 분류 및 우회거래 점검이 필요합니다."
    if topic in {"관세율", "관세율/할당관세"}:
        return f"{country} 관세율 이슈는 대상 HS·세율·적용시점 확인과 법인별 관세원가 영향 산출이 필요합니다."
    return f"{title[:60]} 이슈는 삼성전자 관세·통상 관점에서 대상 국가·HS·법인 영향 확인이 필요합니다."


def build_required_actions(rows: pd.DataFrame, top3: pd.DataFrame) -> list[dict]:
    pool = pd.concat([
        top3,
        rows[rows["Samsung Impact"].eq("Direct")],
        rows[rows["Mail Group"].eq("Regulation")]
    ], ignore_index=True).drop_duplicates(subset=["Headline"], keep="first")
    pool = executive_sort_frame(pool).head(5)
    actions = []
    seen = set()
    for _, r in pool.iterrows():
        topic = infer_topic_ko(r)
        owner = EXECUTIVE_OWNER_MAP.get(topic, "관세기획/지역법인")
        subs = clean(r.get("Affected Subsidiary")) or clean(r.get("Country")) or "관련 법인"
        if topic in {"원산지/CO", "FTA/CEPA"}:
            action = f"{subs}: 협정문 원산지 기준, BOM 충족 여부, CO 발급·보관 증빙 점검"
        elif topic == "CBAM":
            action = f"{subs}: 알루미늄·철강 등 대상 원재료와 공급사 탄소자료 확보 현황 점검"
        elif topic == "반도체 관세":
            action = "DS/통관기획: 반도체 관련 대상 HS, 세율, 적용시점 및 원가 영향 산출"
        elif topic in {"수출통제/제재", "Entity List"}:
            action = f"{subs}: 거래처 스크리닝, ECCN/전략물자 분류, 우회거래 여부 확인"
        elif topic in {"관세율", "관세율/할당관세"}:
            action = f"{subs}: 대상 HS·관세율 변경 여부와 최근 수입신고 적용세율 점검"
        else:
            action = f"{subs}: 원문 기준 대상 국가·HS·제품군 영향 확인"
        key = (owner, action)
        if key in seen:
            continue
        seen.add(key)
        actions.append({"owner": owner, "action": action, "topic": topic})
        if len(actions) >= 4:
            break
    return actions


def build_required_actions_html(rows: pd.DataFrame, top3: pd.DataFrame) -> str:
    actions = build_required_actions(rows, top3)
    if not actions:
        return ""
    items = []
    for idx, a in enumerate(actions, 1):
        topic_value = clean(a.get('topic') or a.get('Issue') or a.get('issue') or a.get('Topic') or '-')
        action_value = clean(a.get('action') or a.get('Required Action') or a.get('required_action') or a.get('Action') or '-')
        owner_value = clean(a.get('owner') or a.get('Owner') or '-')
        items.append(
            f"<tr>"
            f"<td style='padding:8px;border:1px solid #d9d9d9;text-align:center;font-weight:bold;'>{idx}</td>"
            f"<td style='padding:8px;border:1px solid #d9d9d9;'>{html.escape(topic_value)}</td>"
            f"<td style='padding:8px;border:1px solid #d9d9d9;'>{html.escape(action_value)}</td>"
            f"<td style='padding:8px;border:1px solid #d9d9d9;text-align:center;'>{html.escape(owner_value)}</td>"
            f"</tr>"
        )
    return (
        "<h3 style='margin-top:22px;color:#C00000;'>2. Today's Required Actions</h3>"
        "<table style='border-collapse:collapse;width:100%;font-size:13px;margin-bottom:18px;'>"
        "<thead><tr style='background:#C00000;color:white;'>"
        "<th style='padding:8px;border:1px solid #d9d9d9;width:45px;'>No</th>"
        "<th style='padding:8px;border:1px solid #d9d9d9;width:120px;'>Issue</th>"
        "<th style='padding:8px;border:1px solid #d9d9d9;'>Required Action</th>"
        "<th style='padding:8px;border:1px solid #d9d9d9;width:150px;'>Owner</th>"
        "</tr></thead><tbody>" + "".join(items) + "</tbody></table>"
    )


def build_overall_review_html(rows: pd.DataFrame, top3: pd.DataFrame) -> str:
    visible = rows[~rows["Mail Group"].eq("Filtered Noise")].copy()
    regulation = visible[visible["Mail Group"].eq("Regulation")]
    news_core = visible[visible["Mail Group"].eq("News - 핵심")]
    news_usable = visible[visible["Mail Group"].eq("News - 주요/참고")]
    direct_rows = visible[visible["Samsung Impact"].eq("Direct")]
    indirect_rows = visible[visible["Samsung Impact"].eq("Indirect")]

    issues = []
    for _, r in top3.iterrows():
        issue = infer_topic_ko(r)
        if issue not in issues:
            issues.append(issue)
    issue_text = "·".join(issues[:3]) if issues else "관세·통상"

    narrative = [
        f"금일 GTI Radar는 {issue_text} 이슈를 중심으로 분석되었습니다."
    ]
    for _, r in top3.head(3).iterrows():
        narrative.append(executive_issue_sentence(r))
    narrative.append(
        f"오늘 식별된 이슈는 직접영향 {len(direct_rows)}건, 간접영향 {len(indirect_rows)}건이며, 원산지·FTA·관세율·수출통제 관련 실행과제를 우선 추진해야 합니다."
    )

    def mini_list(title: str, frame: pd.DataFrame, color: str) -> str:
        if frame.empty:
            return f"<div style='margin-top:8px;color:#777;'>{html.escape(title)}: 해당 없음</div>"
        frame = executive_sort_frame(frame).drop_duplicates(subset=["_display_cluster"], keep="first").head(5)
        items = []
        for _, r in frame.iterrows():
            subs = clean(r.get("Affected Subsidiary")) or "-"
            headline = clean(r.get("Headline"))
            if len(headline) > 58:
                headline = headline[:58].rstrip() + "…"
            items.append(
                f"<li style='margin:3px 0;'><b>{html.escape(infer_topic_ko(r))}</b> "
                f"[{html.escape(clean(r.get('Country')))} / {html.escape(subs)}] "
                f"{html.escape(headline)}</li>"
            )
        return (
            f"<div style='margin-top:12px;'><b style='color:{color};'>{html.escape(title)}</b>"
            f"<ol style='margin-top:6px;margin-bottom:4px;padding-left:22px;'>{''.join(items)}</ol></div>"
        )

    counts = (
        f"법규 {len(regulation)}건, 핵심 뉴스 {len(news_core)}건, 주요/참고 뉴스 {len(news_usable)}건 "
        f"| 직접영향 {len(direct_rows)}건, 간접영향 {len(indirect_rows)}건"
    )
    paragraphs = "".join(
        f"<div style='font-size:15px;font-weight:bold;line-height:1.85;margin-bottom:8px;'>{html.escape(x)}</div>"
        for x in narrative
    )
    return f"""
    <div style="margin-top:10px;margin-bottom:20px;padding:16px;background:#F4F6F8;border-left:6px solid #1F4E78;color:#222;">
      <div style="font-size:14px;color:#555;margin-bottom:10px;">금일 선별 결과: {html.escape(counts)}</div>
      {paragraphs}
      {mini_list('직접영향 Top 3', direct_rows, '#C00000')}
      {mini_list('간접영향 Top 5', indirect_rows, '#7F7F7F')}
    </div>
    """


def build_html(rows: pd.DataFrame, top3: pd.DataFrame) -> str:
    subject = f"[GTI Radar] Global Trade Intelligence | {RUN_DATE}"
    top_blocks = []
    for idx, row in top3.iterrows():
        top_blocks.append(f"""
        <div style="margin:14px 0 16px 0;padding:14px;border-left:5px solid #C00000;background:#FFF7F7;">
          <div style="font-size:15px;font-weight:bold;margin-bottom:6px;">Top {idx + 1}. {html_link(row['Headline'], row['URL'])}</div>
          <div style="font-size:12px;color:#555;margin-bottom:8px;">Type: {html.escape(clean(row['Content Type']))} | Group: {html.escape(clean(row['Mail Group']))} | Topic: {html.escape(display_topic(row))} | Samsung Impact: <b>{html.escape(clean(row.get('Samsung Impact')))}</b> | Subsidiary: {html.escape(clean(row.get('Affected Subsidiary')) or '-')} | Agency: {html.escape(clean(row['Agency']))} | Publish Date: {html.escape(clean(row['Date']))} | Country: {html.escape(clean(row['Country']))} | Risk: <span style="color:{risk_color(row['Risk'])};font-weight:bold;">{html.escape(clean(row['Risk']))}</span> | Score: {safe_num(row['Importance Score'])}</div>
          <div style="margin-top:7px;"><b>요약</b><br>{html.escape(force_korean_text(row, 'Summary', '요약 정보 확인 필요'))}</div>
          <div style="margin-top:7px;"><b>영향</b><br>{html.escape(force_korean_text(row, 'AI Analysis', '영향 검토 필요'))}</div>
          <div style="margin-top:7px;"><b>대응조치</b><br>{html.escape(force_korean_text(row, 'Action Plan', '담당 부서 확인 필요'))}</div>
        </div>
        """)

    visible = rows[~rows["Mail Group"].eq("Filtered Noise")].copy()
    direct_impact = visible[visible["Samsung Impact"].eq("Direct")].copy()
    regulation = visible[visible["Mail Group"].eq("Regulation")]
    news_core = visible[visible["Mail Group"].eq("News - 핵심")]
    news_usable = visible[visible["Mail Group"].eq("News - 주요/참고")]
    return f"""<!DOCTYPE html>
<html lang="ko">
<head><meta charset="utf-8"><title>{html.escape(subject)}</title></head>
<body style="font-family:Arial,'Malgun Gothic',sans-serif;font-size:13px;color:#222;line-height:1.5;">
  <div style="max-width:1320px;margin:0 auto;">
    <h2 style="margin-bottom:4px;color:#1F4E78;">[GTI Radar] Global Trade Intelligence</h2>
    <div style="font-size:14px;margin-bottom:4px;"><b>Date:</b> {RUN_DATE}</div>
    <div style="font-size:12px;color:#555;margin-bottom:16px;">Focus: Samsung Electronics Customs & Trade Intelligence</div>

    <h3 style="margin-top:18px;margin-bottom:6px;">1. Executive Summary</h3>
    {build_overall_review_html(visible, top3)}

    {build_required_actions_html(visible, top3)}

    <h3 style="color:#C00000;margin-top:22px;">3. Top 3 Deep Analysis</h3>
    {''.join(top_blocks)}

    {build_table('4. Samsung Direct Impact', direct_impact, '#7030A0')}
    {build_table('5. Regulation', regulation, '#1F4E78')}
    {build_table('6. News CORE', news_core, '#548235')}
    {build_table('6. News USABLE / Reference', news_usable, '#7F7F7F')}
    <p style="margin-top:18px;color:#666;font-size:12px;">첨부 Excel에는 STEP4 선별 결과와 메일 그룹이 함께 포함되어 있습니다.</p>
  </div>
</body>
</html>"""


def save_excel(rows: pd.DataFrame, top3: pd.DataFrame, paths: dict[str, Path]) -> None:
    wb = Workbook()
    visible = rows[~rows["Mail Group"].eq("Filtered Noise")].copy()
    filtered_noise = rows[rows["Mail Group"].eq("Filtered Noise")].copy()
    sheets = [
        ("GTI Radar", visible),
        ("Top3", top3),
        ("Required Actions", pd.DataFrame(build_required_actions(visible, top3))),
        ("Regulation", visible[visible["Mail Group"].eq("Regulation")]),
        ("News CORE", visible[visible["Mail Group"].eq("News - 핵심")]),
        ("News USABLE", visible[visible["Mail Group"].eq("News - 주요/참고")]),
        ("Filtered Noise", filtered_noise),
    ]
    first = True
    for name, frame in sheets:
        ws = wb.active if first else wb.create_sheet(name)
        first = False
        ws.title = name[:31]
        if name == "Required Actions":
            cols = ["topic", "action", "owner"]
            ws.append(["Issue", "Required Action", "Owner"])
            for _, rr in frame.iterrows():
                ws.append([rr.get("topic", ""), rr.get("action", ""), rr.get("owner", "")])
        else:
            ws.append(OUTPUT_COLUMNS)
            for _, row in frame.iterrows():
                append_row(ws, row)
        widths = {"A": 10, "B": 18, "C": 26, "D": 16, "E": 20, "F": 42, "G": 14, "H": 52, "I": 56, "J": 56, "K": 70, "L": 18, "M": 22, "N": 8, "O": 14, "P": 16, "Q": 20, "R": 32, "S": 38, "T": 30, "U": 38}
        for col, width in widths.items():
            ws.column_dimensions[col].width = width
        style_sheet(ws)

    runlog = wb.create_sheet("Run Log")
    runlog.append(["item", "value"])
    runlog.append(["regulation_input", str(REGULATION_INPUT_FILE)])
    runlog.append(["news_input", str(NEWS_INPUT_FILE)])
    runlog.append(["run_date", RUN_DATE])
    runlog.append(["visible_total", len(visible)])
    runlog.append(["filtered_noise", len(filtered_noise)])
    runlog.append(["regulation_rows", int(visible["Mail Group"].eq("Regulation").sum())])
    runlog.append(["news_core_rows", int(visible["Mail Group"].eq("News - 핵심").sum())])
    runlog.append(["news_usable_rows", int(visible["Mail Group"].eq("News - 주요/참고").sum())])
    runlog.append(["direct_rows", int(visible["Samsung Impact"].eq("Direct").sum())])
    runlog.append(["indirect_rows", int(visible["Samsung Impact"].eq("Indirect").sum())])
    style_sheet(runlog)

    wb.save(paths["mail_xlsx"])
    wb.save(paths["analysis"])
    visible[OUTPUT_COLUMNS].to_excel(paths["cumulative"], index=False)


def main() -> None:
    paths = output_paths()
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    rows = read_step4_results()
    rows = assign_mail_groups(rows)
    top3 = choose_top3(rows)
    rows = final_order(rows, top3)
    # Rebuild top3 after final ordering to ensure it is visible and deduplicated
    top3 = choose_top3(rows)
    html_body = build_html(rows, top3)
    save_excel(rows, top3, paths)
    paths["mail_html"].write_text(html_body, encoding="utf-8")
    send_email(html_body, paths["mail_xlsx"])
    visible = rows[~rows["Mail Group"].eq("Filtered Noise")]
    print(f"[DONE] HTML: {paths['mail_html']}")
    print(f"[DONE] XLSX: {paths['mail_xlsx']}")
    print(f"[ROWS] total={len(visible)}, regulation={(visible['Mail Group'] == 'Regulation').sum()}, news_core={(visible['Mail Group'] == 'News - 핵심').sum()}, news_usable={(visible['Mail Group'] == 'News - 주요/참고').sum()}, direct={(visible['Samsung Impact'] == 'Direct').sum()}, indirect={(visible['Samsung Impact'] == 'Indirect').sum()}")


# =========================================================
# V3.0.2 QUALITY OVERRIDES
# - Final forced-noise removal in STEP5
# - India-UK FTA/CBAM cluster = 1 representative
# - Direct impact cap and stricter Direct rules
# - Executive Summary mini-list count corrected
# =========================================================

FORCED_NOISE_V302 = [
    # politics / local / election
    "시장 후보", "후보", "당진시장", "시민사회", "정책협약", "지방선거", "선거", "공약", "정당", "의원",
    # food / supplement / consumer retail
    "영양제", "supplement", "vitamin", "프리미엄 영양제", "화장품", "마스크팩", "k-beauty", "k beauty",
    "우유", "유제품", "낙농", "치즈", "버터", "돼지고기", "축산", "농정", "농민", "농업", "농가",
    # weak lifestyle / unrelated
    "블라우스", "의류", "패션", "관광", "여행", "공항", "개인 세금", "출입국", "수상", "관세대상",
]

DIRECT_MAX_V302 = int(os.getenv("GTI_DIRECT_MAX", "5"))
INDIRECT_MAX_V302 = int(os.getenv("GTI_INDIRECT_MAX", "5"))
VISIBLE_MAX_V302 = int(os.getenv("GTI_VISIBLE_MAX", "10"))


def _text_all_v302(row: pd.Series) -> str:
    return " ".join(clean(row.get(f)) for f in ["Headline", "Summary", "AI Analysis", "Action Plan", "Country", "Agency", "Issue"]).lower()


def is_forced_noise_v302(row: pd.Series) -> bool:
    text = _text_all_v302(row)
    # These terms should be removed even if STEP4 accidentally tagged FTA/Origin.
    if any(p.lower() in text for p in FORCED_NOISE_V302):
        # Keep only if there is unmistakable Samsung/regulatory direct signal.
        hard_keep = ["samsung electronics", "삼성전자", "entity list", "export control", "수출통제", "uflpa", "cbam", "반도체"]
        return not any(k in text for k in hard_keep)
    return False


def cluster_key_v3(row: pd.Series) -> str:
    topic = infer_topic_ko(row)
    country = clean(row.get("Country")).upper()
    title = clean(row.get("Headline")).lower()
    text = f"{country} {title} {clean(row.get('Summary')).lower()}"

    # Strong cluster normalization: all India-UK FTA/CBAM/steel articles are one issue.
    if ("INDIA" in country and ("UNITED KINGDOM" in country or "UK" in country)) or ("india" in text and "uk" in text):
        if any(k in text for k in ["fta", "trade pact", "trade secretary", "piyush", "steel", "cbam", "carbon tax", "levy", "concessions", "rollout"]):
            return "INDIA_UK_FTA_CBAM_STEEL"
    if ("VIETNAM" in country or "베트남" in country) and any(k in text for k in ["acfta", "asean", "china", "trung quốc", "free trade"]):
        return "VIETNAM_ACFTA_ORIGIN"
    if ("INDIA" in country or "india" in text) and "oman" in text and any(k in text for k in ["fta", "cepa", "zero-duty", "zero duty"]):
        return "INDIA_OMAN_CEPA"
    if ("KOREA" in country or "한국" in country) and any(k in text for k in ["uflpa", "forced labor", "강제노동"]):
        return "KOREA_UFLPA_SUPPLY_CHAIN"
    if topic == "CBAM" and any(k in text for k in ["home appliance", "appliance", "stainless", "steelnews"]):
        return "GLOBAL_CBAM_APPLIANCE_STEEL"
    if topic == "반도체 관세":
        return "SEMICONDUCTOR_TARIFF_GENERAL"
    base = clean(row.get("Cluster")) or f"{topic}:{country}:{title[:90]}"
    return re.sub(r"\s+", " ", base).strip()


def determine_samsung_impact(row: pd.Series) -> tuple[str, str, str]:
    topic = infer_topic_ko(row)
    issue = clean(row.get("Issue")).upper()
    title = clean(row.get("Headline")).lower()
    text = _text_all_v302(row)
    country = clean(row.get("Country")) or "국가 미상"
    country_subs = country_based_subsidiaries(country)
    explicit_subs = explicit_subsidiaries_in_text(row)
    subs = list(dict.fromkeys(country_subs + explicit_subs))

    if clean(row.get("Content Type")).lower().startswith("reg"):
        return "Direct", "SEC/HQ", "공식 법규/공고성 이슈로 본사 관세 Master·신고 프로세스 반영 필요"

    if is_forced_noise_v302(row) or is_residual_noise_v3(row) or is_low_relevance(row):
        return "None", "", "삼성전자 관세·통상 업무 관련성이 낮은 Noise로 메일 본문 제외"

    negotiation_only = any(k in title for k in ["talks", "negotiation", "visit", "meet", "momentum", "proposal", "rollout", "to clarify", "secretary", "논의", "방문", "협상"])
    implementation_signal = any(k in text for k in [
        "rules of origin", "rule of origin", "origin rules", "in effect", "effective", "implementation", "comes into force",
        "now in effect", "적용", "시행", "발효", "원산지 규정", "시행규칙", "zero-duty access", "zero duty access"
    ])
    product_signal = has_samsung_product_signal(row)
    direct_hard_topic = topic in {"CBAM", "반도체 관세", "수출통제/제재", "Entity List", "AD/CVD", "관세환급", "통관/세관심사", "관세율/할당관세"}
    origin_direct = topic in {"원산지/CO", "FTA/CEPA"} and implementation_signal and not negotiation_only

    # FTA talks/visits should not be Direct unless implementation/rules are explicit.
    if negotiation_only and topic in {"원산지/CO", "FTA/CEPA", "관세율"}:
        return "Indirect", ", ".join(country_subs), "FTA 협상·방문·논의 단계로 즉시 원산지/신고 프로세스 변경 전 단계"

    if is_global_country(row) and not explicit_subs and not (direct_hard_topic and product_signal):
        return "Indirect", "", "Global 이슈로 특정 삼성 법인의 직접영향 확정 불가"

    if explicit_subs and (direct_hard_topic or origin_direct):
        return "Direct", ", ".join(subs), f"법인 코드 명시 + 실행형 규제 토픽({topic})"

    if country_subs and (direct_hard_topic or origin_direct) and (product_signal or topic in {"CBAM", "수출통제/제재", "반도체 관세", "AD/CVD"}):
        return "Direct", ", ".join(country_subs), f"법인국가 매칭 + 실행형 규제 토픽({topic}) + 제품/업무 관련성 확인"

    if country_subs:
        return "Indirect", ", ".join(country_subs), f"법인 소재국 관련 이슈이나 제품·거래·시행조건 직접 연결은 추가 확인 필요"
    return "Indirect", "", f"토픽({topic})은 모니터링 가치가 있으나 특정 삼성 법인 직접영향은 확인 불가"


def apply_samsung_impact(rows: pd.DataFrame) -> pd.DataFrame:
    rows = rows.copy()
    impacts = rows.apply(determine_samsung_impact, axis=1)
    rows["Samsung Impact"] = [x[0] for x in impacts]
    rows["Affected Subsidiary"] = [x[1] for x in impacts]
    rows["Impact Reason"] = [x[2] for x in impacts]
    rows["_display_cluster"] = rows.apply(cluster_key_v3, axis=1)
    return rows


def assign_mail_groups(rows: pd.DataFrame) -> pd.DataFrame:
    rows = rows.copy()
    rows["Mail Group"] = "News - 주요/참고"
    rows.loc[rows["Content Type"].eq("Regulation"), "Mail Group"] = "Regulation"
    rows.loc[rows["Content Type"].eq("News") & rows["Priority Group"].eq("CORE"), "Mail Group"] = "News - 핵심"
    rows = apply_samsung_impact(rows)
    rows.loc[rows["Samsung Impact"].eq("None"), "Mail Group"] = "Filtered Noise"
    return rows


def executive_sort_frame(rows: pd.DataFrame) -> pd.DataFrame:
    rows = rows.copy()
    impact_weight = rows["Samsung Impact"].map({"Direct": 1200, "Indirect": 250, "None": -9999}).fillna(0)
    topic_weight = rows.apply(lambda r: 450 if infer_topic_ko(r) in {"CBAM", "반도체 관세", "수출통제/제재", "Entity List", "AD/CVD", "원산지/CO", "관세율/할당관세"} else 0, axis=1)
    rows["_exec_score"] = rows["_integrated_score"] + impact_weight + topic_weight
    return rows.sort_values(["_exec_score", "_sort_date"], ascending=[False, False])


def choose_top3(rows: pd.DataFrame) -> pd.DataFrame:
    pool = rows[~rows["Mail Group"].eq("Filtered Noise")].copy()
    if pool.empty:
        return pool
    pool = executive_sort_frame(pool)
    selected, used_clusters, used_topics = [], set(), set()
    for _, row in pool.iterrows():
        ck = clean(row.get("_display_cluster")) or cluster_key_v3(row)
        topic = infer_topic_ko(row)
        if ck in used_clusters:
            continue
        if topic in used_topics and len(selected) < 2:
            continue
        selected.append(row)
        used_clusters.add(ck)
        used_topics.add(topic)
        if len(selected) >= 3:
            break
    if len(selected) < 3:
        for _, row in pool.iterrows():
            ck = clean(row.get("_display_cluster")) or cluster_key_v3(row)
            if ck in used_clusters:
                continue
            if any(clean(row["Headline"]) == clean(x["Headline"]) for x in selected):
                continue
            selected.append(row)
            used_clusters.add(ck)
            if len(selected) >= 3:
                break
    return pd.DataFrame(selected).reset_index(drop=True)


def final_order(rows: pd.DataFrame, top3: pd.DataFrame) -> pd.DataFrame:
    rows = assign_mail_groups(rows)
    visible = rows[~rows["Mail Group"].eq("Filtered Noise")].copy()
    visible = executive_sort_frame(visible)

    # Keep one representative per issue cluster. Regulation is kept separately.
    reg = visible[visible["Mail Group"].eq("Regulation")].copy()
    nonreg = visible[~visible["Mail Group"].eq("Regulation")].copy()
    nonreg = nonreg.drop_duplicates(subset=["_display_cluster"], keep="first")

    direct = nonreg[nonreg["Samsung Impact"].eq("Direct")].head(DIRECT_MAX_V302)
    indirect = nonreg[nonreg["Samsung Impact"].eq("Indirect")].head(INDIRECT_MAX_V302)
    visible = pd.concat([reg, direct, indirect], ignore_index=True)
    visible = executive_sort_frame(visible).head(max(VISIBLE_MAX_V302, len(reg) + len(direct)))

    top_keys = set(top3["Headline"].astype(str)) if top3 is not None and not top3.empty else set()
    top = visible[visible["Headline"].astype(str).isin(top_keys)].copy()
    rest = visible[~visible["Headline"].astype(str).isin(top_keys)].copy()
    group_order = {"Regulation": 0, "News - 핵심": 1, "News - 주요/참고": 2}
    rest["_group_order"] = rest["Mail Group"].map(group_order).fillna(9)
    rest = executive_sort_frame(rest).sort_values(["_group_order", "_exec_score", "_sort_date"], ascending=[True, False, False])
    top = executive_sort_frame(top)
    out = pd.concat([top, rest], ignore_index=True)
    out["No"] = range(1, len(out) + 1)
    return out


def build_required_actions(rows: pd.DataFrame, top3: pd.DataFrame) -> list[dict]:
    pool = pd.concat([
        top3,
        rows[rows["Samsung Impact"].eq("Direct")],
        rows[rows["Mail Group"].eq("Regulation")]
    ], ignore_index=True).drop_duplicates(subset=["_display_cluster"], keep="first")
    pool = executive_sort_frame(pool).head(4)
    actions, seen = [], set()
    for _, r in pool.iterrows():
        topic = infer_topic_ko(r)
        owner = EXECUTIVE_OWNER_MAP.get(topic, "관세기획/지역법인")
        subs = clean(r.get("Affected Subsidiary")) or clean(r.get("Country")) or "관련 법인"
        if topic in {"원산지/CO", "FTA/CEPA"}:
            action = f"{subs}: 협정문 원산지 기준, BOM 충족 여부, CO 발급·보관 증빙 점검"
        elif topic == "CBAM":
            action = f"{subs}: 알루미늄·철강 등 대상 원재료와 공급사 탄소자료 확보 현황 점검"
        elif topic in {"수출통제/제재", "Entity List"}:
            action = f"{subs}: 거래처 스크리닝, ECCN/전략물자 분류, 우회거래 여부 확인"
        elif topic in {"관세율", "관세율/할당관세"}:
            action = f"{subs}: 대상 HS·관세율 변경 여부와 최근 수입신고 적용세율 점검"
        elif topic == "반도체 관세":
            action = "DS/통관기획: 반도체 관련 대상 HS, 세율, 적용시점 및 원가 영향 산출"
        else:
            action = f"{subs}: 원문 기준 대상 국가·HS·제품군 영향 확인"
        key = (topic, owner, action)
        if key in seen:
            continue
        seen.add(key)
        actions.append({"owner": owner, "action": action, "topic": topic})
        if len(actions) >= 3:
            break
    return actions


def build_overall_review_html(rows: pd.DataFrame, top3: pd.DataFrame) -> str:
    visible = rows[~rows["Mail Group"].eq("Filtered Noise")].copy()
    regulation = visible[visible["Mail Group"].eq("Regulation")]
    news_core = visible[visible["Mail Group"].eq("News - 핵심")]
    news_usable = visible[visible["Mail Group"].eq("News - 주요/참고")]
    direct_rows = visible[visible["Samsung Impact"].eq("Direct")]
    indirect_rows = visible[visible["Samsung Impact"].eq("Indirect")]

    issues = []
    for _, r in top3.iterrows():
        issue = infer_topic_ko(r)
        if issue not in issues:
            issues.append(issue)
    issue_text = "·".join(issues[:3]) if issues else "관세·통상"
    narrative = [f"금일 GTI Radar는 {issue_text} 이슈를 중심으로 분석되었습니다."]
    for _, r in top3.head(3).iterrows():
        narrative.append(executive_issue_sentence(r))
    narrative.append(f"오늘 식별된 이슈는 직접영향 {len(direct_rows)}건, 간접영향 {len(indirect_rows)}건이며, 원산지·FTA·관세율·수출통제 관련 실행과제를 우선 추진해야 합니다.")

    def mini_list(title: str, frame: pd.DataFrame, color: str, limit: int) -> str:
        if frame.empty:
            return f"<div style='margin-top:8px;color:#777;'>{html.escape(title)}: 해당 없음</div>"
        frame = executive_sort_frame(frame).drop_duplicates(subset=["_display_cluster"], keep="first").head(limit)
        items = []
        for _, r in frame.iterrows():
            subs = clean(r.get("Affected Subsidiary")) or "-"
            headline = clean(r.get("Headline"))
            if len(headline) > 58:
                headline = headline[:58].rstrip() + "…"
            items.append(f"<li style='margin:3px 0;'><b>{html.escape(infer_topic_ko(r))}</b> [{html.escape(clean(r.get('Country')))} / {html.escape(subs)}] {html.escape(headline)}</li>")
        return f"<div style='margin-top:12px;'><b style='color:{color};'>{html.escape(title)}</b><ol style='margin-top:6px;margin-bottom:4px;padding-left:22px;'>{''.join(items)}</ol></div>"

    counts = f"법규 {len(regulation)}건, 핵심 뉴스 {len(news_core)}건, 주요/참고 뉴스 {len(news_usable)}건 | 직접영향 {len(direct_rows)}건, 간접영향 {len(indirect_rows)}건"
    paragraphs = "".join(f"<div style='font-size:15px;font-weight:bold;line-height:1.85;margin-bottom:8px;'>{html.escape(x)}</div>" for x in narrative)
    return f"""
    <div style="margin-top:10px;margin-bottom:20px;padding:16px;background:#F4F6F8;border-left:6px solid #1F4E78;color:#222;">
      <div style="font-size:14px;color:#555;margin-bottom:10px;">금일 선별 결과: {html.escape(counts)}</div>
      {paragraphs}
      {mini_list('직접영향 Top 3', direct_rows, '#C00000', 3)}
      {mini_list('간접영향 Top 3', indirect_rows, '#7F7F7F', 3)}
    </div>
    """


# =========================================================
# V3.0.3 EXECUTIVE QUALITY OVERRIDES
# - Executive Summary rewritten in Samsung-impact/action style
# - Balanced Direct/Indirect portfolio
# - Final visible rows capped to executive-readable level
# - Strong cluster representatives for FTA/AD-CVD/CBAM duplicates
# - HTML counts include Filtered Noise correctly
# =========================================================

VISIBLE_MAX_V303 = int(os.getenv("GTI_VISIBLE_MAX", "12"))
DIRECT_MAX_V303 = int(os.getenv("GTI_DIRECT_MAX", "5"))
INDIRECT_MIN_V303 = int(os.getenv("GTI_INDIRECT_MIN", "3"))
INDIRECT_MAX_V303 = int(os.getenv("GTI_INDIRECT_MAX", "6"))
NEWS_CORE_MAX_V303 = int(os.getenv("GTI_NEWS_CORE_MAX", "8"))
NEWS_USABLE_MAX_V303 = int(os.getenv("GTI_NEWS_USABLE_MAX", "4"))

FORCED_NOISE_V303 = list(dict.fromkeys(FORCED_NOISE_V302 + [
    "틴랩", "뉴질랜드 프리미엄 영양제", "김기재", "당진시장", "시민사회", "기후위기 극복 에너지 전환",
    "carmakers", "automotive", "자동차 공급망", "ev shift", "israeli extremists", "extremists",
    "tourism", "여행", "소매", "retail", "정치", "선거", "공약", "후보",
]))


def is_forced_noise_v303(row: pd.Series) -> bool:
    text = _text_all_v302(row)
    if any(p.lower() in text for p in FORCED_NOISE_V303):
        hard_keep = ["samsung electronics", "삼성전자", "semiconductor", "반도체", "entity list", "export control", "수출통제", "uflpa", "section 301", "section 232"]
        return not any(k in text for k in hard_keep)
    return False


def cluster_key_v303(row: pd.Series) -> str:
    topic = infer_topic_ko(row)
    title = clean(row.get("Headline")).lower()
    country = clean(row.get("Country")).lower()
    text = _text_all_v302(row)
    if clean(row.get("Content Type")).lower().startswith("reg"):
        return "REGULATION:" + (clean(row.get("Cluster")) or clean(row.get("Headline"))[:80])
    if ("india" in text and "uk" in text) or ("india" in text and "united kingdom" in text):
        if any(k in text for k in ["fta", "cepa", "carbon tax", "steel levy", "scotch", "cbam"]):
            return "INDIA_UK_FTA_CBAM_STEEL"
    if any(k in text for k in ["usmca", "north america", "free-trade agreement", "북미자유무역", "mexico free-trade", "u.s. and mexico"]):
        return "USMCA_EXTENSION"
    if "brazil" in text and any(k in text for k in ["25%", "25 percent", "tariff", "trade practices"]):
        return "US_BRAZIL_25_TARIFF"
    if any(k in text for k in ["japan", "일본", "nhk"]) and any(k in text for k in ["anti-dumping", "antidumping", "반덤핑", "steel", "철강"]):
        return "JAPAN_STEEL_AD_KR_CN"
    if any(k in text for k in ["section 301", "section 232", "ustr"]):
        return "US_USTR_301_232_TARIFF"
    if "india" in text and "oman" in text and any(k in text for k in ["cepa", "zero-duty", "electronics", "rules of origin"]):
        return "INDIA_OMAN_CEPA_ELECTRONICS"
    if topic == "CBAM" and "eu" in text and "carbon" in text:
        return "EU_CBAM_GENERAL"
    base = clean(row.get("Cluster")) or f"{topic}:{country}:{title[:90]}"
    return re.sub(r"\s+", " ", base).strip()


def determine_samsung_impact(row: pd.Series) -> tuple[str, str, str]:
    topic = infer_topic_ko(row)
    issue = clean(row.get("Issue")).upper()
    title = clean(row.get("Headline")).lower()
    text = _text_all_v302(row)
    country = clean(row.get("Country")) or "국가 미상"
    country_subs = country_based_subsidiaries(country)
    explicit_subs = explicit_subsidiaries_in_text(row)
    subs = list(dict.fromkeys(country_subs + explicit_subs))

    if clean(row.get("Content Type")).lower().startswith("reg"):
        return "Direct", "SEC/HQ", "공식 법규/공고성 이슈로 본사 관세 Master·신고 프로세스 반영 필요"

    if is_forced_noise_v303(row) or is_residual_noise_v3(row) or is_low_relevance(row):
        return "None", "", "삼성전자 관세·통상 업무 관련성이 낮은 Noise로 메일 본문 제외"

    product_signal = has_samsung_product_signal(row)
    implementation_signal = any(k in text for k in [
        "rules of origin", "rule of origin", "origin rules", "in effect", "effective", "implementation", "comes into force",
        "now in effect", "적용", "시행", "발효", "원산지 규정", "zero-duty", "electronics", "전자", "semiconductor", "반도체"
    ])
    negotiation_only = any(k in title for k in ["talks", "negotiation", "visit", "meet", "proposal", "rollout", "to clarify", "secretary", "협상", "방문", "논의", "제안"])

    # Strict Direct: legal/regulatory implementation or Samsung-product exposure.
    if topic in {"수출통제/제재", "Entity List", "반도체 관세"} and (country_subs or explicit_subs or product_signal):
        return "Direct", ", ".join(subs or country_subs), f"{topic}은 삼성 제품·거래 심사에 직접 연결되는 고위험 토픽"
    if topic == "CBAM" and (product_signal or any(k in text for k in ["steel", "aluminum", "aluminium", "철강", "알루미늄"])) and (country_subs or "eu" in text):
        return "Direct", ", ".join(subs or country_subs), "CBAM 대상 원재료·EU향 공급망 증빙 영향 가능"
    if topic in {"원산지/CO", "FTA/CEPA"} and implementation_signal and (product_signal or "electronics" in text or "전자" in text) and country_subs:
        return "Direct", ", ".join(country_subs), "FTA/원산지 적용 또는 발효 단계이며 삼성 제품·법인 연결 가능성 확인"
    if topic in {"AD/CVD", "미국 301/232 관세", "관세율/할당관세"} and product_signal and country_subs:
        return "Direct", ", ".join(country_subs), "대상 품목이 삼성 제품·원재료와 연결될 가능성이 있어 직접 점검 필요"

    # Everything else remains monitorable but not direct.
    if country_subs:
        return "Indirect", ", ".join(country_subs), "법인 소재국 관련 정책이나 제품·거래·시행조건 직접 연결은 추가 확인 필요"
    return "Indirect", "", f"토픽({topic})은 모니터링 가치가 있으나 특정 삼성 법인 직접영향은 확인 불가"


def apply_samsung_impact(rows: pd.DataFrame) -> pd.DataFrame:
    rows = rows.copy()
    impacts = rows.apply(determine_samsung_impact, axis=1)
    rows["Samsung Impact"] = [x[0] for x in impacts]
    rows["Affected Subsidiary"] = [x[1] for x in impacts]
    rows["Impact Reason"] = [x[2] for x in impacts]
    rows["_display_cluster"] = rows.apply(cluster_key_v303, axis=1)
    return rows


def assign_mail_groups(rows: pd.DataFrame) -> pd.DataFrame:
    rows = rows.copy()
    rows["Mail Group"] = "News - 주요/참고"
    rows.loc[rows["Content Type"].eq("Regulation"), "Mail Group"] = "Regulation"
    rows.loc[rows["Content Type"].eq("News") & rows["Priority Group"].eq("CORE"), "Mail Group"] = "News - 핵심"
    rows = apply_samsung_impact(rows)
    rows.loc[rows["Samsung Impact"].eq("None"), "Mail Group"] = "Filtered Noise"
    return rows


def executive_sort_frame(rows: pd.DataFrame) -> pd.DataFrame:
    rows = rows.copy()
    impact_weight = rows["Samsung Impact"].map({"Direct": 900, "Indirect": 450, "None": -9999}).fillna(0)
    topic_weight = rows.apply(lambda r: 450 if infer_topic_ko(r) in {"수출통제/제재", "Entity List", "반도체 관세", "CBAM", "원산지/CO", "AD/CVD", "미국 301/232 관세"} else 0, axis=1)
    source_weight = rows["Agency"].astype(str).str.lower().map(lambda x: 120 if any(s in x for s in ["reuters", "bloomberg", "yna", "연합", "mlex", "nhk", "ustr", "law.go.kr"]) else 0)
    rows["_exec_score"] = rows["_integrated_score"] + impact_weight + topic_weight + source_weight
    return rows.sort_values(["_exec_score", "_sort_date"], ascending=[False, False])


def choose_top3(rows: pd.DataFrame) -> pd.DataFrame:
    pool = rows[~rows["Mail Group"].eq("Filtered Noise")].copy()
    if pool.empty:
        return pool
    pool = executive_sort_frame(pool).drop_duplicates(subset=["_display_cluster"], keep="first")
    selected, used_topics = [], set()
    # Pick 1 direct/actionable, then important indirect if needed, with topic diversity.
    for _, row in pool.iterrows():
        topic = infer_topic_ko(row)
        if topic in used_topics and len(selected) < 2:
            continue
        selected.append(row)
        used_topics.add(topic)
        if len(selected) >= 3:
            break
    return pd.DataFrame(selected).reset_index(drop=True)


def _rebalance_direct_indirect(nonreg: pd.DataFrame) -> pd.DataFrame:
    nonreg = executive_sort_frame(nonreg).drop_duplicates(subset=["_display_cluster"], keep="first")
    direct = nonreg[nonreg["Samsung Impact"].eq("Direct")].head(DIRECT_MAX_V303)
    indirect = nonreg[nonreg["Samsung Impact"].eq("Indirect")].head(INDIRECT_MAX_V303)
    # If not enough indirect, demote lower-ranked Direct candidates except the highest-value ones.
    if len(indirect) < INDIRECT_MIN_V303:
        extra = nonreg[~nonreg.index.isin(direct.index) & ~nonreg.index.isin(indirect.index)].head(INDIRECT_MIN_V303 - len(indirect)).copy()
        if not extra.empty:
            extra["Samsung Impact"] = "Indirect"
            extra["Impact Reason"] = "임원 보고 균형상 정책 모니터링 대상으로 표시"
            indirect = pd.concat([indirect, extra], ignore_index=False)
    return pd.concat([direct, indirect], ignore_index=False)


def final_order(rows: pd.DataFrame, top3: pd.DataFrame) -> pd.DataFrame:
    rows = assign_mail_groups(rows)
    visible = rows[~rows["Mail Group"].eq("Filtered Noise")].copy()
    reg = visible[visible["Mail Group"].eq("Regulation")].copy()
    nonreg = visible[~visible["Mail Group"].eq("Regulation")].copy()
    selected_nonreg = _rebalance_direct_indirect(nonreg)
    selected = pd.concat([reg, selected_nonreg], ignore_index=False)
    selected = executive_sort_frame(selected).drop_duplicates(subset=["_display_cluster"], keep="first")

    # Cap final visible rows. Keep regulation, then best news.
    reg2 = selected[selected["Mail Group"].eq("Regulation")]
    news2 = selected[~selected["Mail Group"].eq("Regulation")].head(max(0, VISIBLE_MAX_V303 - len(reg2)))
    visible_final = pd.concat([reg2, news2], ignore_index=True)

    # Section caps: core max 8, usable max 4; convert overflow core to usable instead of showing too many core rows.
    news_idx = visible_final[visible_final["Mail Group"].eq("News - 핵심")].index.tolist()
    if len(news_idx) > NEWS_CORE_MAX_V303:
        overflow = news_idx[NEWS_CORE_MAX_V303:]
        visible_final.loc[overflow, "Mail Group"] = "News - 주요/참고"

    noise = rows[rows["Mail Group"].eq("Filtered Noise")].copy()
    out = pd.concat([visible_final, noise], ignore_index=True)
    visible_mask = ~out["Mail Group"].eq("Filtered Noise")
    ordered_visible = executive_sort_frame(out[visible_mask]).copy()
    ordered_noise = out[~visible_mask].copy()
    ordered_visible["No"] = range(1, len(ordered_visible) + 1)
    if not ordered_noise.empty:
        ordered_noise["No"] = range(len(ordered_visible) + 1, len(ordered_visible) + len(ordered_noise) + 1)
    return pd.concat([ordered_visible, ordered_noise], ignore_index=True)


def executive_issue_sentence(row: pd.Series) -> str:
    topic = infer_topic_ko(row)
    country = clean(row.get("Country")) or "관련 국가"
    subs = clean(row.get("Affected Subsidiary")) or "관련 법인"
    headline = clean(row.get("Headline"))
    text = _text_all_v302(row)
    if topic in {"원산지/CO", "FTA/CEPA"}:
        if "oman" in text:
            return "Oman CEPA 원산지 규정 및 관세양허 확대는 SIEL의 중동 수출품 원산지 판정, CO 발급 및 협정세율 적용 기준 재점검이 필요한 사안입니다."
        if any(k in text for k in ["usmca", "북미자유무역", "mexico", "canada"]):
            return "북미 FTA 연장·개정 논의는 멕시코·미국·캐나다 법인의 원산지 기준 및 역내조달 전략에 영향을 줄 수 있어 정책 모니터링이 필요합니다."
        return f"{country} FTA/원산지 이슈는 {subs}의 원산지 판정, CO 발급 및 특혜세율 적용 기준 재점검이 필요합니다."
    if topic == "CBAM":
        return f"{country} CBAM 이슈는 원재료 탄소자료 확보와 EU향 신고 증빙 체계에 영향을 줄 수 있어 {subs} 공급망 점검이 필요합니다."
    if topic in {"AD/CVD", "미국 301/232 관세", "관세율/할당관세", "관세율"}:
        return f"{country} 관세·AD/CVD 이슈는 대상 HS, 공급국, 가격자료 및 원산지 방어자료를 중심으로 영향 분석이 필요합니다."
    if topic in {"반도체 관세", "수출통제/제재", "Entity List"}:
        return "반도체 관세·수출통제 정책 변화는 글로벌 생산거점의 관세원가, 거래처 스크리닝 및 공급망 전략에 영향을 줄 가능성이 있어 지속 모니터링이 필요합니다."
    if clean(row.get("Content Type")).lower().startswith("reg"):
        return "공식 법규 변경사항은 시행일·대상 HS·세율·신고기준을 확인하여 HQ 관세 Master와 신고 체크리스트에 반영해야 합니다."
    return f"{headline} 이슈는 대상 국가·HS·제품군 기준으로 삼성전자 관세·통상 영향을 확인해야 합니다."


def _priority_theme_sentence(top3: pd.DataFrame, rows: pd.DataFrame) -> str:
    text = " ".join(clean(x) for x in rows.get("Headline", pd.Series(dtype=str)).tolist()).lower()
    if "oman" in text and "india" in text:
        return "금일 GTI Radar는 인도·중동 FTA 원산지 규정 개정과 글로벌 관세정책 변화를 핵심 이슈로 식별하였습니다."
    topics = []
    for _, r in top3.iterrows():
        t = infer_topic_ko(r)
        if t not in topics:
            topics.append(t)
    return f"금일 GTI Radar는 {'·'.join(topics[:3]) if topics else '관세·통상'} 이슈를 핵심으로 식별하였습니다."


def build_overall_review_html(rows: pd.DataFrame, top3: pd.DataFrame) -> str:
    visible = rows[~rows["Mail Group"].eq("Filtered Noise")].copy()
    filtered = rows[rows["Mail Group"].eq("Filtered Noise")].copy()
    regulation = visible[visible["Mail Group"].eq("Regulation")]
    news_core = visible[visible["Mail Group"].eq("News - 핵심")]
    news_usable = visible[visible["Mail Group"].eq("News - 주요/참고")]
    direct_rows = visible[visible["Samsung Impact"].eq("Direct")]
    indirect_rows = visible[visible["Samsung Impact"].eq("Indirect")]

    narrative = [_priority_theme_sentence(top3, visible)]
    seen_sent = set()
    for _, r in top3.head(3).iterrows():
        sent = executive_issue_sentence(r)
        key = re.sub(r"[^가-힣A-Za-z0-9]+", "", sent[:80]).lower()
        if key not in seen_sent:
            narrative.append(sent)
            seen_sent.add(key)
    narrative.append(f"직접영향 {len(direct_rows)}건, 간접영향 {len(indirect_rows)}건이 식별되었으며 원산지·FTA·관세 대응을 우선 추진해야 합니다.")

    def mini_list(title: str, frame: pd.DataFrame, color: str, limit: int) -> str:
        if frame.empty:
            return f"<div style='margin-top:8px;color:#777;'>{html.escape(title)}: 해당 없음</div>"
        frame = executive_sort_frame(frame).drop_duplicates(subset=["_display_cluster"], keep="first").head(limit)
        items = []
        for _, r in frame.iterrows():
            subs = clean(r.get("Affected Subsidiary")) or "-"
            headline = clean(r.get("Headline"))
            if len(headline) > 58:
                headline = headline[:58].rstrip() + "…"
            items.append(f"<li style='margin:3px 0;'><b>{html.escape(infer_topic_ko(r))}</b> [{html.escape(clean(r.get('Country')))} / {html.escape(subs)}] {html.escape(headline)}</li>")
        return f"<div style='margin-top:12px;'><b style='color:{color};'>{html.escape(title)}</b><ol style='margin-top:6px;margin-bottom:4px;padding-left:22px;'>{''.join(items)}</ol></div>"

    counts = f"법규 {len(regulation)}건, 핵심 뉴스 {len(news_core)}건, 주요/참고 뉴스 {len(news_usable)}건 | 직접영향 {len(direct_rows)}건, 간접영향 {len(indirect_rows)}건, 제외 {len(filtered)}건"
    paragraphs = "".join(f"<div style='font-size:15px;font-weight:bold;line-height:1.85;margin-bottom:8px;'>{html.escape(x)}</div>" for x in narrative)
    return f"""
    <div style="margin-top:10px;margin-bottom:20px;padding:16px;background:#F4F6F8;border-left:6px solid #1F4E78;color:#222;">
      <div style="font-size:14px;color:#555;margin-bottom:10px;">금일 선별 결과: {html.escape(counts)}</div>
      {paragraphs}
      {mini_list('직접영향 Top 3', direct_rows, '#C00000', 3)}
      {mini_list('간접영향 Top 3', indirect_rows, '#7F7F7F', 3)}
    </div>
    """


def build_required_actions(rows: pd.DataFrame, top3: pd.DataFrame) -> list[dict]:
    visible = rows[~rows["Mail Group"].eq("Filtered Noise")].copy()
    pool = pd.concat([top3, visible[visible["Samsung Impact"].eq("Direct")]], ignore_index=True).drop_duplicates(subset=["_display_cluster"], keep="first")
    pool = executive_sort_frame(pool).head(6)
    actions, seen = [], set()
    for _, r in pool.iterrows():
        topic = infer_topic_ko(r)
        owner = EXECUTIVE_OWNER_MAP.get(topic, "관세기획/지역법인")
        subs = clean(r.get("Affected Subsidiary")) or clean(r.get("Country")) or "관련 법인"
        if topic in {"원산지/CO", "FTA/CEPA"}:
            action = f"{subs}: 원산지 기준, BOM 충족 여부, CO 발급·보관 증빙 점검"
        elif topic == "CBAM":
            action = f"{subs}: 대상 원재료와 공급사 탄소자료 확보 현황 점검"
        elif topic in {"AD/CVD", "미국 301/232 관세", "관세율/할당관세", "관세율"}:
            action = f"{subs}: 대상 HS·공급국·가격자료 및 관세율 적용 영향 확인"
        elif topic in {"수출통제/제재", "Entity List"}:
            action = f"{subs}: 거래처 스크리닝, ECCN/전략물자 분류, 우회거래 여부 확인"
        else:
            action = f"{subs}: 원문 기준 대상 국가·HS·제품군 영향 확인"
        key = (topic, owner, re.sub(r"[A-Z]{2,5}[,;/ ]*", "", action)[:60])
        if key in seen:
            continue
        seen.add(key)
        actions.append({"owner": owner, "action": action, "topic": topic})
        if len(actions) >= 3:
            break
    return actions


def build_html(rows: pd.DataFrame, top3: pd.DataFrame) -> str:
    subject = f"[GTI Radar] Global Trade Intelligence | {RUN_DATE}"
    visible = rows[~rows["Mail Group"].eq("Filtered Noise")].copy()
    top_blocks = []
    for idx, row in top3.iterrows():
        top_blocks.append(f"""
        <div style="margin:14px 0 16px 0;padding:14px;border-left:5px solid #C00000;background:#FFF7F7;">
          <div style="font-size:15px;font-weight:bold;margin-bottom:6px;">Top {idx + 1}. {html_link(row['Headline'], row['URL'])}</div>
          <div style="font-size:12px;color:#555;margin-bottom:8px;">Type: {html.escape(clean(row['Content Type']))} | Group: {html.escape(clean(row['Mail Group']))} | Topic: {html.escape(display_topic(row))} | Samsung Impact: <b>{html.escape(clean(row.get('Samsung Impact')))}</b> | Subsidiary: {html.escape(clean(row.get('Affected Subsidiary')) or '-')} | Agency: {html.escape(clean(row['Agency']))} | Publish Date: {html.escape(clean(row['Date']))} | Country: {html.escape(clean(row['Country']))} | Risk: <span style="color:{risk_color(row['Risk'])};font-weight:bold;">{html.escape(clean(row['Risk']))}</span> | Score: {safe_num(row['Importance Score'])}</div>
          <div style="margin-top:7px;"><b>요약</b><br>{html.escape(force_korean_text(row, 'Summary', '요약 정보 확인 필요'))}</div>
          <div style="margin-top:7px;"><b>영향</b><br>{html.escape(force_korean_text(row, 'AI Analysis', '영향 검토 필요'))}</div>
          <div style="margin-top:7px;"><b>대응조치</b><br>{html.escape(force_korean_text(row, 'Action Plan', '담당 부서 확인 필요'))}</div>
        </div>
        """)
    direct_impact = visible[visible["Samsung Impact"].eq("Direct")].copy()
    regulation = visible[visible["Mail Group"].eq("Regulation")]
    news_core = visible[visible["Mail Group"].eq("News - 핵심")]
    news_usable = visible[visible["Mail Group"].eq("News - 주요/참고")]
    return f"""<!DOCTYPE html>
<html lang="ko">
<head><meta charset="utf-8"><title>{html.escape(subject)}</title></head>
<body style="font-family:Arial,'Malgun Gothic',sans-serif;font-size:13px;color:#222;line-height:1.5;">
  <div style="max-width:1320px;margin:0 auto;">
    <h2 style="margin-bottom:4px;color:#1F4E78;">[GTI Radar] Global Trade Intelligence</h2>
    <div style="font-size:14px;margin-bottom:4px;"><b>Date:</b> {RUN_DATE}</div>
    <div style="font-size:12px;color:#555;margin-bottom:16px;">Focus: Samsung Electronics Customs & Trade Intelligence</div>

    <h3 style="margin-top:18px;margin-bottom:6px;">1. Executive Summary</h3>
    {build_overall_review_html(rows, top3)}

    {build_required_actions_html(rows, top3)}

    <h3 style="color:#C00000;margin-top:22px;">3. Top 3 Deep Analysis</h3>
    {''.join(top_blocks)}

    {build_table('4. Samsung Direct Impact', direct_impact, '#7030A0')}
    {build_table('5. Regulation', regulation, '#1F4E78')}
    {build_table('6. News CORE', news_core, '#548235')}
    {build_table('6. News USABLE / Reference', news_usable, '#7F7F7F')}
    <p style="margin-top:18px;color:#666;font-size:12px;">첨부 Excel에는 STEP4 선별 결과와 메일 그룹이 함께 포함되어 있습니다.</p>
  </div>
</body>
</html>"""


def main() -> None:
    paths = output_paths()
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    rows_raw = read_step4_results()
    rows_grouped = assign_mail_groups(rows_raw)
    top3_pre = choose_top3(rows_grouped)
    rows = final_order(rows_grouped, top3_pre)
    top3 = choose_top3(rows)
    html_body = build_html(rows, top3)
    save_excel(rows, top3, paths)
    paths["mail_html"].write_text(html_body, encoding="utf-8")
    send_email(html_body, paths["mail_xlsx"])
    visible = rows[~rows["Mail Group"].eq("Filtered Noise")]
    filtered = rows[rows["Mail Group"].eq("Filtered Noise")]
    print(f"[DONE] HTML: {paths['mail_html']}")
    print(f"[DONE] XLSX: {paths['mail_xlsx']}")
    print(f"[ROWS] total={len(visible)}, regulation={(visible['Mail Group'] == 'Regulation').sum()}, news_core={(visible['Mail Group'] == 'News - 핵심').sum()}, news_usable={(visible['Mail Group'] == 'News - 주요/참고').sum()}, filtered_noise={len(filtered)}, direct={(visible['Samsung Impact'] == 'Direct').sum()}, indirect={(visible['Samsung Impact'] == 'Indirect').sum()}")



# =========================================================
# V3.0.4 EXECUTIVE QUALITY OVERRIDES
# - Remove military/security/submarine noise that STEP4 may mis-topic as AD/CVD
# - Cap Direct to 4 and keep Indirect portfolio for executive balance
# - Prevent duplicate Top3 clusters
# - Replace long scraped body snippets with concise Samsung-impact sentences
# =========================================================

VISIBLE_MAX_V303 = int(os.getenv("GTI_VISIBLE_MAX", "10"))
DIRECT_MAX_V303 = int(os.getenv("GTI_DIRECT_MAX", "4"))
INDIRECT_MIN_V303 = int(os.getenv("GTI_INDIRECT_MIN", "3"))
INDIRECT_MAX_V303 = int(os.getenv("GTI_INDIRECT_MAX", "5"))
NEWS_CORE_MAX_V303 = int(os.getenv("GTI_NEWS_CORE_MAX", "7"))
NEWS_USABLE_MAX_V303 = int(os.getenv("GTI_NEWS_USABLE_MAX", "3"))

SECURITY_NOISE_V304 = [
    "nuclear-powered submarine", "nuclear powered submarine", "submarine", "security talks",
    "defense talks", "naval", "nuclear submarine", "핵추진", "핵 추진", "잠수함", "안보 회담", "안보협의", "국방", "방산",
]

FORCED_NOISE_V304 = list(dict.fromkeys(FORCED_NOISE_V303 + SECURITY_NOISE_V304 + [
    "solar firms", "solar market", "태양광 업체", "태양광 시장",  # 삼성전자 관세 GTI 관점에서는 직접영향 제외
]))


def _contains_security_noise_v304(row: pd.Series) -> bool:
    text = _text_all_v302(row)
    return any(k in text for k in SECURITY_NOISE_V304)


def is_forced_noise_v303(row: pd.Series) -> bool:
    text = _text_all_v302(row)
    # 군사/안보/잠수함은 삼성 관세·통상 GTI 메일에서 무조건 제외
    if _contains_security_noise_v304(row):
        return True
    if any(p.lower() in text for p in FORCED_NOISE_V304):
        hard_keep = [
            "semiconductor", "반도체", "entity list", "export control", "수출통제",
            "uflpa", "section 301", "section 232", "rules of origin", "origin rules", "electronics exporters",
        ]
        return not any(k in text for k in hard_keep)
    return False


def _compact_subs_v304(subs) -> str:
    if isinstance(subs, str):
        parts = [x.strip() for x in re.split(r"[,;/]+", subs) if x.strip()]
    else:
        parts = list(subs or [])
    parts = list(dict.fromkeys(parts))
    if not parts:
        return ""
    # North America/EU 같은 다수 법인은 임원용으로 과도하게 늘리지 않음
    if len(parts) > 4:
        return ", ".join(parts[:4]) + " 외"
    return ", ".join(parts)


def determine_samsung_impact(row: pd.Series) -> tuple[str, str, str]:
    topic = infer_topic_ko(row)
    issue = clean(row.get("Issue")).upper()
    text = _text_all_v302(row)
    country = clean(row.get("Country")) or "국가 미상"
    country_subs = country_based_subsidiaries(country)
    explicit_subs = explicit_subsidiaries_in_text(row)
    subs = list(dict.fromkeys(country_subs + explicit_subs))

    if clean(row.get("Content Type")).lower().startswith("reg"):
        return "Direct", "SEC/HQ", "공식 법규/공고성 이슈로 본사 관세 Master·신고 프로세스 반영 필요"

    if is_forced_noise_v303(row) or is_residual_noise_v3(row) or is_low_relevance(row):
        return "None", "", "삼성전자 관세·통상 업무 관련성이 낮은 Noise로 메일 본문 제외"

    product_signal = has_samsung_product_signal(row)
    implementation_signal = any(k in text for k in [
        "rules of origin", "rule of origin", "origin rules", "in effect", "effective", "implementation", "comes into force",
        "now in effect", "적용", "시행", "발효", "원산지 규정", "zero-duty", "electronics", "전자", "semiconductor", "반도체",
        "usmca", "북미자유무역", "fta implementation",
    ])
    trade_remedy_signal = any(k in text for k in [
        "tariff", "duty", "anti-dumping", "antidumping", "countervailing", "section 301", "section 232", "ustr",
        "관세", "반덤핑", "상계관세", "철강", "steel", "aluminum", "aluminium", "copper", "electronics", "전자", "semiconductor", "반도체",
    ])

    # Strict Direct: 실제 관세/원산지/수출통제 실행성 또는 제품 연결성이 있어야 함
    if topic in {"수출통제/제재", "Entity List", "반도체 관세"} and (country_subs or explicit_subs or product_signal):
        return "Direct", _compact_subs_v304(subs or country_subs), f"{topic}은 삼성 제품·거래 심사에 직접 연결되는 고위험 토픽"
    if topic == "CBAM" and any(k in text for k in ["steel", "aluminum", "aluminium", "철강", "알루미늄", "carbon border", "cbam"]):
        # 중국 자동차 우회수출 등 삼성 제품 직접성이 약한 경우는 Indirect
        if any(k in text for k in ["carmaker", "automotive", "자동차", "morocco", "모로코"]):
            return "Indirect", _compact_subs_v304(subs or country_subs), "CBAM·우회수출 정책 모니터링 대상이나 삼성 제품 직접영향은 추가 확인 필요"
        return "Direct", _compact_subs_v304(subs or country_subs), "CBAM 대상 원재료·EU향 공급망 증빙 영향 가능"
    if topic in {"원산지/CO", "FTA/CEPA"} and implementation_signal and country_subs:
        # 협상/연장 논의는 Direct보다 Indirect. 다만 USMCA처럼 북미 법인 운영 영향은 Direct 가능.
        if any(k in text for k in ["usmca", "북미자유무역", "north america", "u.s. and mexico"]):
            return "Direct", _compact_subs_v304(country_subs), "북미 역내 원산지·FTA 운영 기준에 영향을 줄 수 있어 법인 점검 필요"
        if product_signal or "electronics" in text or "전자" in text:
            return "Direct", _compact_subs_v304(country_subs), "FTA/원산지 적용 또는 발효 단계이며 삼성 제품·법인 연결 가능성 확인"
        return "Indirect", _compact_subs_v304(country_subs), "FTA/원산지 정책 모니터링 대상이나 대상 제품 직접성 확인 필요"
    if topic in {"AD/CVD", "미국 301/232 관세", "관세율/할당관세", "관세율"} and trade_remedy_signal and country_subs:
        # 품목이 철강/알루미늄/구리/전자/반도체 등 공급망과 연결될 때만 Direct
        if any(k in text for k in ["section 301", "section 232", "steel", "aluminum", "aluminium", "copper", "electronics", "전자", "semiconductor", "반도체", "25% tariff", "25 percent tariff"]):
            return "Direct", _compact_subs_v304(country_subs), "대상 품목이 삼성 제품·원재료와 연결될 가능성이 있어 직접 점검 필요"
        return "Indirect", _compact_subs_v304(country_subs), "관세·AD/CVD 정책 모니터링 대상이나 삼성 제품 직접성 확인 필요"

    if country_subs:
        return "Indirect", _compact_subs_v304(country_subs), "법인 소재국 관련 정책이나 제품·거래·시행조건 직접 연결은 추가 확인 필요"
    return "Indirect", "", f"토픽({topic})은 모니터링 가치가 있으나 특정 삼성 법인 직접영향은 확인 불가"


def cluster_key_v303(row: pd.Series) -> str:
    topic = infer_topic_ko(row)
    title = clean(row.get("Headline")).lower()
    country = clean(row.get("Country")).lower()
    text = _text_all_v302(row)
    if clean(row.get("Content Type")).lower().startswith("reg"):
        return "REGULATION:" + (clean(row.get("Cluster")) or clean(row.get("Headline"))[:80])
    if _contains_security_noise_v304(row):
        return "SECURITY_SUBMARINE_NOISE"
    if any(k in text for k in ["usmca", "north america", "free-trade agreement", "북미자유무역", "mexico free-trade", "u.s. and mexico"]):
        return "USMCA_EXTENSION"
    if ("india" in text and ("uk" in text or "united kingdom" in text)) and any(k in text for k in ["fta", "cepa", "carbon tax", "steel levy", "scotch", "cbam"]):
        return "INDIA_UK_FTA_CBAM_STEEL"
    if "brazil" in text and any(k in text for k in ["25%", "25 percent", "tariff", "trade practices"]):
        return "US_BRAZIL_25_TARIFF"
    if any(k in text for k in ["section 301", "section 232", "ustr"]):
        return "US_USTR_301_232_TARIFF"
    if any(k in text for k in ["japan", "일본", "nhk"]) and any(k in text for k in ["anti-dumping", "antidumping", "반덤핑", "steel", "철강"]):
        return "JAPAN_STEEL_AD_KR_CN"
    if "india" in text and "oman" in text and any(k in text for k in ["cepa", "zero-duty", "electronics", "rules of origin"]):
        return "INDIA_OMAN_CEPA_ELECTRONICS"
    if topic == "CBAM" and "eu" in text and "carbon" in text:
        return "EU_CBAM_GENERAL"
    base = clean(row.get("Cluster")) or f"{topic}:{country}:{title[:90]}"
    return re.sub(r"\s+", " ", base).strip()


def _exec_country_label_v304(row: pd.Series) -> str:
    text = _text_all_v302(row)
    country = clean(row.get("Country"))
    if any(k in text for k in ["usmca", "북미자유무역", "north america", "mexico", "canada"]):
        return "북미(미국·멕시코·캐나다)"
    if "india" in text and "oman" in text:
        return "인도·오만"
    if "india" in text and ("uk" in text or "united kingdom" in text):
        return "인도·영국"
    if "brazil" in text and ("united states" in text or "usa" in text or "u.s." in text):
        return "미국·브라질"
    if "eu" in text or "europe" in text:
        return "EU"
    if ";" in country or "," in country:
        return country.split(";")[0].split(",")[0].strip()
    return country or "관련 국가"


def executive_issue_sentence(row: pd.Series) -> str:
    topic = infer_topic_ko(row)
    text = _text_all_v302(row)
    country = _exec_country_label_v304(row)
    subs = clean(row.get("Affected Subsidiary")) or "관련 법인"
    if topic in {"원산지/CO", "FTA/CEPA"}:
        if "oman" in text:
            return "Oman CEPA 원산지 규정 및 관세양허 확대는 SIEL의 중동 수출품 원산지 판정, CO 발급 및 협정세율 적용 기준 재점검이 필요한 사안입니다."
        if any(k in text for k in ["usmca", "북미자유무역", "mexico", "canada"]):
            return "북미 FTA 연장·개정 논의는 멕시코·미국·캐나다 법인의 원산지 기준 및 역내조달 전략에 영향을 줄 수 있어 정책 모니터링과 법인별 영향 점검이 필요합니다."
        return f"{country} FTA/원산지 이슈는 {subs}의 원산지 판정, CO 발급 및 특혜세율 적용 기준 재점검이 필요합니다."
    if topic == "CBAM":
        return f"{country} CBAM 이슈는 원재료 탄소자료 확보와 EU향 신고 증빙 체계에 영향을 줄 수 있어 공급망 점검이 필요합니다."
    if topic in {"AD/CVD", "미국 301/232 관세", "관세율/할당관세", "관세율"}:
        return f"{country} 관세·AD/CVD 이슈는 대상 HS, 공급국, 가격자료 및 원산지 방어자료를 중심으로 영향 분석이 필요합니다."
    if topic in {"반도체 관세", "수출통제/제재", "Entity List"}:
        return "반도체 관세·수출통제 정책 변화는 글로벌 생산거점의 관세원가, 거래처 스크리닝 및 공급망 전략에 영향을 줄 가능성이 있어 지속 모니터링이 필요합니다."
    if clean(row.get("Content Type")).lower().startswith("reg"):
        return "공식 법규 변경사항은 시행일·대상 HS·세율·신고기준을 확인하여 HQ 관세 Master와 신고 체크리스트에 반영해야 합니다."
    return "대상 국가·HS·제품군 기준으로 삼성전자 관세·통상 영향을 확인해야 합니다."


def force_korean_text(row: pd.Series, field: str, fallback: str) -> str:
    """V3.0.4: 메일 본문에는 긴 원문 스크랩을 노출하지 않고 임원용 요약문만 표시."""
    topic = infer_topic_ko(row)
    country = _exec_country_label_v304(row)
    subs = clean(row.get("Affected Subsidiary")) or "관련 법인"
    headline = clean(row.get("Headline"))
    if field == "Summary":
        return f"{headline} 관련 {topic} 이슈입니다. 대상 국가, 적용 품목, 시행일 및 관세·통관 영향 여부 확인이 필요합니다."
    if field == "AI Analysis":
        if topic in {"원산지/CO", "FTA/CEPA"}:
            return f"{country} FTA/원산지 이슈로 {subs}의 CO 발급, 원산지 판정 및 협정세율 적용 기준 재점검이 필요합니다."
        if topic in {"AD/CVD", "미국 301/232 관세", "관세율/할당관세", "관세율"}:
            return f"{country} 관세·AD/CVD 이슈로 대상 HS, 공급국, 가격자료 및 원산지 증빙 영향 분석이 필요합니다. 삼성전자 관련 원재료·부품·완제품과의 직접 매칭 여부를 우선 확인해야 합니다."
        if topic == "CBAM":
            return f"{country} CBAM 이슈로 EU향 제품·원재료 탄소자료, 공급사 증빙 및 신고 대응 체계 점검이 필요합니다."
        if topic in {"수출통제/제재", "Entity List"}:
            return f"{country} 수출통제/제재 이슈로 거래처 스크리닝, ECCN/전략물자 분류 및 우회거래 점검이 필요합니다."
        if clean(row.get("Content Type")).lower().startswith("reg"):
            return "공식 법규/고시 변경사항으로 시행일, 대상 HS, 세율 및 신고기준을 HQ Master와 체크리스트에 반영해야 합니다."
        return fallback
    if field == "Action Plan":
        if topic in {"원산지/CO", "FTA/CEPA"}:
            return "① 대상 거래·HS 추출 ② 협정 원산지 기준 확인 ③ BOM/공정/공급자확인서 점검 ④ CO 발급·보관 증빙 업데이트"
        if topic in {"AD/CVD", "미국 301/232 관세", "관세율/할당관세", "관세율"}:
            return "① 대상 HS·공급국 확인 ② 수입·수출 신고 이력 매칭 ③ 가격자료·원산지 증빙 점검 ④ 관세사·법무와 대응 시나리오 수립"
        if topic == "CBAM":
            return "① 대상 HS·원재료 확인 ② EU향 거래 매핑 ③ 공급사 탄소자료 확보 ④ CBAM 신고·증빙 체계 업데이트"
        if topic in {"수출통제/제재", "Entity List"}:
            return "① 거래처 스크리닝 ② ECCN/전략물자 분류 확인 ③ 우회거래 여부 점검 ④ Hold/Release 기록화"
        if clean(row.get("Content Type")).lower().startswith("reg"):
            return "① 원문 확인 ② 대상 국가·HS·법인 매핑 ③ 신고·증빙 영향 점검 ④ 체크리스트 업데이트"
        return fallback
    return clean(row.get(field)) or fallback


def choose_top3(rows: pd.DataFrame) -> pd.DataFrame:
    pool = rows[~rows["Mail Group"].eq("Filtered Noise")].copy()
    if pool.empty:
        return pool
    pool = executive_sort_frame(pool).drop_duplicates(subset=["_display_cluster"], keep="first")
    selected, used_topics, used_clusters = [], set(), set()
    # Prioritize: regulation or high-value direct, then distinct indirect themes
    for _, row in pool.iterrows():
        cluster = clean(row.get("_display_cluster"))
        topic = infer_topic_ko(row)
        if cluster in used_clusters:
            continue
        if topic in used_topics and len(selected) < 3:
            continue
        selected.append(row)
        used_clusters.add(cluster)
        used_topics.add(topic)
        if len(selected) >= 3:
            break
    if len(selected) < 3:
        for _, row in pool.iterrows():
            cluster = clean(row.get("_display_cluster"))
            if cluster in used_clusters:
                continue
            selected.append(row)
            used_clusters.add(cluster)
            if len(selected) >= 3:
                break
    return pd.DataFrame(selected).reset_index(drop=True)


def _priority_theme_sentence(top3: pd.DataFrame, rows: pd.DataFrame) -> str:
    text = " ".join(clean(x) for x in rows.get("Headline", pd.Series(dtype=str)).tolist()).lower()
    if "oman" in text and "india" in text:
        return "금일 GTI Radar는 인도·중동 FTA 원산지 규정 개정과 글로벌 관세정책 변화를 핵심 이슈로 식별하였습니다."
    if any(k in text for k in ["usmca", "북미자유무역", "mexico", "canada"]):
        return "금일 GTI Radar는 북미 FTA 원산지 체계 변화와 글로벌 관세·AD/CVD 정책을 핵심 이슈로 식별하였습니다."
    topics = []
    for _, r in top3.iterrows():
        t = infer_topic_ko(r)
        if t not in topics:
            topics.append(t)
    return f"금일 GTI Radar는 {'·'.join(topics[:3]) if topics else '관세·통상'} 이슈를 핵심으로 식별하였습니다."



# =========================================================
# V3.0.5 EXECUTIVE FINAL OVERRIDES
# - Direct cap includes Regulation: Direct <= 4 total
# - Remove Suez/industrial-zone diplomatic investment noise
# - De-duplicate Required Actions by issue family
# - Shorten country/subsidiary labels in Executive Summary
# - Keep final report compact for executive reading
# =========================================================

VISIBLE_MAX_V305 = int(os.getenv("GTI_VISIBLE_MAX", "8"))
DIRECT_TOTAL_MAX_V305 = int(os.getenv("GTI_DIRECT_TOTAL_MAX", "4"))
INDIRECT_MIN_V305 = int(os.getenv("GTI_INDIRECT_MIN", "3"))
INDIRECT_MAX_V305 = int(os.getenv("GTI_INDIRECT_MAX", "3"))
NEWS_CORE_MAX_V305 = int(os.getenv("GTI_NEWS_CORE_MAX", "5"))
NEWS_USABLE_MAX_V305 = int(os.getenv("GTI_NEWS_USABLE_MAX", "3"))

DIPLOMATIC_INVESTMENT_NOISE_V305 = [
    "suez", "수에즈", "suez canal", "수에즈운하", "industrial zone", "산업구역", "외교장관",
    "investment zone", "economic zone", "경제구역", "special zone", "특구",
]


def _is_diplomatic_investment_noise_v305(row: pd.Series) -> bool:
    text = _text_all_v302(row)
    if any(k in text for k in DIPLOMATIC_INVESTMENT_NOISE_V305):
        keep = ["samsung", "삼성전자", "semiconductor", "반도체", "electronics exporter", "원산지 규정", "rules of origin"]
        return not any(k in text for k in keep)
    return False


def is_forced_noise_v303(row: pd.Series) -> bool:
    text = _text_all_v302(row)
    if _contains_security_noise_v304(row):
        return True
    if _is_diplomatic_investment_noise_v305(row):
        return True
    if any(p.lower() in text for p in FORCED_NOISE_V304):
        hard_keep = [
            "semiconductor", "반도체", "entity list", "export control", "수출통제",
            "uflpa", "section 301", "section 232", "rules of origin", "origin rules", "electronics exporters",
        ]
        return not any(k in text for k in hard_keep)
    return False


def _exec_country_short_v305(row: pd.Series) -> str:
    text = _text_all_v302(row)
    country = clean(row.get("Country"))
    if any(k in text for k in ["usmca", "북미자유무역", "mexico", "canada"]):
        return "북미"
    if "india" in text and "oman" in text:
        return "인도·오만"
    if "india" in text and ("uk" in text or "united kingdom" in text):
        return "인도·영국"
    if "eu" in text or "europe" in text or "cbam" in text:
        return "EU"
    if "japan" in text or "일본" in text:
        return "일본"
    if "brazil" in text:
        return "브라질"
    if "korea" in text or "한국" in text:
        return "한국"
    if ";" in country or "," in country:
        return country.split(";")[0].split(",")[0].strip()
    return country or "관련국"


def _compact_subs_v305(subs) -> str:
    s = _compact_subs_v304(subs)
    if not s:
        return "-"
    # executive short labels by region/group
    parts = [x.strip() for x in re.split(r"[,;/]+", s.replace("외", "")) if x.strip()]
    if any(p in parts for p in ["SEA", "SAS", "SSI", "SRA"]) and any(p in parts for p in ["SEM", "SAMEX", "SECA"]):
        return "북미 법인"
    if any(p in parts for p in ["SEUK", "SEG", "SEF", "SEI", "SEPOL"]):
        return "EU 법인"
    if len(parts) > 3:
        return ", ".join(parts[:3]) + " 외"
    return ", ".join(parts)


def determine_samsung_impact(row: pd.Series) -> tuple[str, str, str]:
    topic = infer_topic_ko(row)
    text = _text_all_v302(row)
    country = clean(row.get("Country")) or "국가 미상"
    country_subs = country_based_subsidiaries(country)
    explicit_subs = explicit_subsidiaries_in_text(row)
    subs = list(dict.fromkeys(country_subs + explicit_subs))

    if clean(row.get("Content Type")).lower().startswith("reg"):
        return "Direct", "SEC/HQ", "공식 법규/공고성 이슈로 HQ 관세 Master·신고 프로세스 반영 필요"

    if is_forced_noise_v303(row) or is_residual_noise_v3(row) or is_low_relevance(row):
        return "None", "", "삼성전자 관세·통상 업무 관련성이 낮아 메일 본문 제외"

    # High-value Direct only. Negotiation/monitoring-only items stay Indirect.
    if topic in {"수출통제/제재", "Entity List", "반도체 관세"} and (country_subs or explicit_subs or has_samsung_product_signal(row)):
        return "Direct", _compact_subs_v305(subs or country_subs), f"{topic}은 거래 심사·품목 분류·공급망 리스크에 직접 연결"

    if topic in {"원산지/CO", "FTA/CEPA"}:
        if "india" in text and "oman" in text and any(k in text for k in ["electronics", "전자", "zero-duty", "cepa", "rules of origin"]):
            return "Direct", _compact_subs_v305(country_subs or ["SIEL", "SGE"]), "인도·오만 CEPA 활용 가능성이 있어 제품별 CO·협정세율 검토 필요"
        if "india" in text and ("uk" in text or "united kingdom" in text) and any(k in text for k in ["implementation", "carbon tax", "steel", "fta"]):
            return "Direct", _compact_subs_v305(country_subs or ["SIEL", "SEUK"]), "인도·영국 FTA 이행·탄소세 쟁점으로 원산지·협정세율 점검 필요"
        if any(k in text for k in ["usmca", "북미자유무역", "u.s. and mexico", "free-trade agreement for another 16 years"]):
            return "Indirect", _compact_subs_v305(country_subs), "북미 FTA 연장·개정 모니터링 대상이며 실제 원산지 규정 변경 여부 확인 필요"
        if country_subs:
            return "Indirect", _compact_subs_v305(country_subs), "FTA/원산지 정책 모니터링 대상이나 제품 직접성 확인 필요"

    if topic == "CBAM":
        if any(k in text for k in ["automotive", "carmaker", "자동차", "morocco", "모로코", "우회수출"]):
            return "Indirect", _compact_subs_v305(subs or country_subs), "CBAM·우회수출 정책 모니터링 대상이며 삼성 제품 직접성은 추가 확인 필요"
        if any(k in text for k in ["steel", "aluminum", "aluminium", "철강", "알루미늄", "carbon border", "cbam"]):
            return "Direct", _compact_subs_v305(subs or country_subs), "CBAM 대상 원재료·EU향 증빙 체계 영향 가능"

    if topic in {"AD/CVD", "미국 301/232 관세", "관세율/할당관세", "관세율"}:
        if any(k in text for k in ["section 301", "section 232", "ustr", "semiconductor", "반도체"]):
            return "Direct", _compact_subs_v305(country_subs), "고위험 관세정책으로 HS·원산지·관세율 영향 점검 필요"
        if any(k in text for k in ["steel", "aluminum", "aluminium", "copper", "철강", "구리", "25% tariff", "25 percent tariff"]):
            return "Indirect", _compact_subs_v305(country_subs), "원재료·AD/CVD 모니터링 대상이며 삼성 거래 직접 매칭 확인 필요"
        if country_subs:
            return "Indirect", _compact_subs_v305(country_subs), "관세·AD/CVD 정책 모니터링 대상"

    if country_subs:
        return "Indirect", _compact_subs_v305(country_subs), "법인 소재국 관련 정책이나 제품·거래 직접성은 추가 확인 필요"
    return "Indirect", "", f"토픽({topic})은 모니터링 가치가 있으나 특정 삼성 법인 직접영향은 확인 불가"


def _set_group_v305(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df["Mail Group"] = "News - 주요/참고"
    df.loc[df["Content Type"].eq("Regulation"), "Mail Group"] = "Regulation"
    df.loc[df["Content Type"].eq("News") & df["Samsung Impact"].eq("Direct"), "Mail Group"] = "News - 핵심"
    return df


def final_order(rows: pd.DataFrame, top3: pd.DataFrame) -> pd.DataFrame:
    rows = apply_samsung_impact(rows.copy())
    rows["_display_cluster"] = rows.apply(cluster_key_v303, axis=1)
    rows = executive_sort_frame(rows)

    noise = rows[rows["Samsung Impact"].eq("None")].copy()
    noise["Mail Group"] = "Filtered Noise"

    visible_pool = rows[~rows["Samsung Impact"].eq("None")].copy()
    visible_pool = visible_pool.drop_duplicates(subset=["_display_cluster"], keep="first")

    reg = visible_pool[visible_pool["Content Type"].eq("Regulation")].head(1)
    nonreg = visible_pool[~visible_pool.index.isin(reg.index)].copy()

    direct_slots = max(0, DIRECT_TOTAL_MAX_V305 - int(reg["Samsung Impact"].eq("Direct").sum()))
    direct = nonreg[nonreg["Samsung Impact"].eq("Direct")].head(direct_slots)
    used = set(reg.index).union(set(direct.index))

    indirect = nonreg[(~nonreg.index.isin(used)) & nonreg["Samsung Impact"].eq("Indirect")].head(INDIRECT_MAX_V305)
    used |= set(indirect.index)

    # fill remaining slots with best remaining indirect first, then direct if report is too thin
    remaining_slots = max(0, VISIBLE_MAX_V305 - (len(reg) + len(direct) + len(indirect)))
    filler = nonreg[(~nonreg.index.isin(used)) & nonreg["Samsung Impact"].eq("Indirect")].head(remaining_slots)
    if len(filler) < remaining_slots:
        more = nonreg[(~nonreg.index.isin(used.union(set(filler.index))))].head(remaining_slots - len(filler))
        filler = pd.concat([filler, more], ignore_index=False)

    visible = pd.concat([direct, reg, indirect, filler], ignore_index=False)
    visible = _set_group_v305(visible)
    visible = executive_sort_frame(visible)

    # limit section sizes for executive readability
    core = visible[visible["Mail Group"].eq("News - 핵심")].head(NEWS_CORE_MAX_V305)
    regulation = visible[visible["Mail Group"].eq("Regulation")]
    usable = visible[(visible["Mail Group"].eq("News - 주요/참고")) & (~visible.index.isin(core.index))].head(NEWS_USABLE_MAX_V305)
    visible = pd.concat([core, regulation, usable], ignore_index=False)
    visible = executive_sort_frame(visible).head(VISIBLE_MAX_V305)

    out = pd.concat([visible, noise], ignore_index=True)
    out["Affected Subsidiary"] = out["Affected Subsidiary"].apply(_compact_subs_v305)
    out["No"] = range(1, len(out) + 1)
    return out.reset_index(drop=True)


def choose_top3(rows: pd.DataFrame) -> pd.DataFrame:
    pool = rows[~rows["Mail Group"].eq("Filtered Noise")].copy()
    if pool.empty:
        return pool
    pool = executive_sort_frame(pool).drop_duplicates(subset=["_display_cluster"], keep="first")
    selected, used_topics, used_clusters = [], set(), set()

    # Prefer one core FTA/origin direct, one regulation, one distinct indirect macro risk.
    priority_clusters = ["INDIA_OMAN_CEPA_ELECTRONICS", "INDIA_UK_FTA_CBAM_STEEL", "REGULATION", "EU_CBAM_GENERAL", "USMCA_EXTENSION"]
    for wanted in priority_clusters:
        for _, row in pool.iterrows():
            cluster = clean(row.get("_display_cluster"))
            if wanted == "REGULATION":
                ok = clean(row.get("Content Type")).lower().startswith("reg")
            else:
                ok = cluster == wanted
            if not ok or cluster in used_clusters:
                continue
            selected.append(row)
            used_clusters.add(cluster)
            used_topics.add(infer_topic_ko(row))
            break
        if len(selected) >= 3:
            break

    if len(selected) < 3:
        for _, row in pool.iterrows():
            cluster = clean(row.get("_display_cluster"))
            topic = infer_topic_ko(row)
            if cluster in used_clusters:
                continue
            if topic in used_topics and len(selected) < 3:
                continue
            selected.append(row)
            used_clusters.add(cluster)
            used_topics.add(topic)
            if len(selected) >= 3:
                break
    return pd.DataFrame(selected).reset_index(drop=True)


def executive_issue_sentence(row: pd.Series) -> str:
    topic = infer_topic_ko(row)
    text = _text_all_v302(row)
    country = _exec_country_short_v305(row)
    subs = clean(row.get("Affected Subsidiary")) or "관련 법인"
    if topic in {"원산지/CO", "FTA/CEPA"}:
        if "oman" in text:
            return "Oman CEPA 원산지 규정 및 관세양허 확대는 SIEL의 중동 수출품 원산지 판정, CO 발급 및 협정세율 적용 기준 재점검이 필요한 사안입니다."
        if "india" in text and ("uk" in text or "united kingdom" in text):
            return "인도·영국 FTA 이행 쟁점은 SIEL·SEUK 간 거래의 원산지 판정, CO 발급 및 협정세율 적용 기준 점검이 필요한 사안입니다."
        if any(k in text for k in ["usmca", "북미자유무역", "mexico", "canada"]):
            return "북미 FTA 연장·개정 논의는 멕시코·미국·캐나다 법인의 원산지 기준 및 역내조달 전략에 영향을 줄 수 있어 정책 모니터링이 필요합니다."
        return f"{country} FTA/원산지 이슈는 {subs}의 원산지 판정, CO 발급 및 특혜세율 적용 기준 검토가 필요합니다."
    if topic == "CBAM":
        return "EU CBAM 이슈는 원재료 탄소자료 확보와 EU향 신고 증빙 체계에 영향을 줄 수 있어 공급망 점검이 필요합니다."
    if topic in {"AD/CVD", "미국 301/232 관세", "관세율/할당관세", "관세율"}:
        return f"{country} 관세·AD/CVD 이슈는 대상 HS, 공급국, 가격자료 및 원산지 방어자료를 중심으로 영향 분석이 필요합니다."
    if clean(row.get("Content Type")).lower().startswith("reg"):
        return "공식 법규 변경사항은 시행일·대상 HS·세율·신고기준을 확인하여 HQ 관세 Master와 신고 체크리스트에 반영해야 합니다."
    return "대상 국가·HS·제품군 기준으로 삼성전자 관세·통상 영향을 확인해야 합니다."


def build_overall_review_html(rows: pd.DataFrame, top3: pd.DataFrame) -> str:
    visible = rows[~rows["Mail Group"].eq("Filtered Noise")].copy()
    filtered = rows[rows["Mail Group"].eq("Filtered Noise")]
    direct_rows = visible[visible["Samsung Impact"].eq("Direct")]
    indirect_rows = visible[visible["Samsung Impact"].eq("Indirect")]
    regulation = visible[visible["Mail Group"].eq("Regulation")]
    news_core = visible[visible["Mail Group"].eq("News - 핵심")]
    news_usable = visible[visible["Mail Group"].eq("News - 주요/참고")]
    counts = f"법규 {len(regulation)}건, 핵심 뉴스 {len(news_core)}건, 주요/참고 뉴스 {len(news_usable)}건 | 직접영향 {len(direct_rows)}건, 간접영향 {len(indirect_rows)}건, 제외 {len(filtered)}건"

    sentences = []
    sentences.append(_priority_theme_sentence(top3, visible))
    for _, r in top3.iterrows():
        s = executive_issue_sentence(r)
        if s not in sentences:
            sentences.append(s)
        if len(sentences) >= 4:
            break
    sentences.append(f"직접영향 {len(direct_rows)}건, 간접영향 {len(indirect_rows)}건이 식별되었으며 원산지·FTA·관세 대응을 우선 추진해야 합니다.")

    def mini_list(title: str, frame: pd.DataFrame, color: str, maxn: int = 3) -> str:
        frame = executive_sort_frame(frame).drop_duplicates(subset=["_display_cluster"], keep="first").head(maxn)
        if frame.empty:
            return f"<div style='margin-top:8px;color:#777;'>{html.escape(title)}: 해당 없음</div>"
        items = []
        for _, r in frame.iterrows():
            headline = clean(r.get("Headline"))
            if len(headline) > 72:
                headline = headline[:72].rstrip() + "…"
            items.append(f"<li style='margin:3px 0;'><b>{html.escape(infer_topic_ko(r))}</b> [{html.escape(_exec_country_short_v305(r))} / {html.escape(clean(r.get('Affected Subsidiary')) or '-')}] {html.escape(headline)}</li>")
        return f"<div style='margin-top:12px;'><b style='color:{color};'>{html.escape(title)}</b><ol style='margin-top:6px;margin-bottom:4px;padding-left:22px;'>{''.join(items)}</ol></div>"

    summary_html = "".join(f"<div style='font-size:15px;font-weight:bold;line-height:1.85;margin-bottom:8px;'>{html.escape(s)}</div>" for s in sentences)
    return f"""
    <div style="margin-top:10px;margin-bottom:20px;padding:16px;background:#F4F6F8;border-left:6px solid #1F4E78;color:#222;">
      <div style="font-size:14px;color:#555;margin-bottom:10px;">금일 선별 결과: {html.escape(counts)}</div>
      {summary_html}
      {mini_list('직접영향 Top 3', direct_rows, '#C00000', 3)}
      {mini_list('간접영향 Top 3', indirect_rows, '#7F7F7F', 3)}
    </div>
    """


def build_required_actions(rows: pd.DataFrame, top3: pd.DataFrame) -> list[dict]:
    visible = rows[~rows["Mail Group"].eq("Filtered Noise")].copy()
    candidates = pd.concat([top3, visible], ignore_index=True).drop_duplicates(subset=["_display_cluster"], keep="first")
    actions, used_family = [], set()
    family_map = {
        "원산지/CO": "ORIGIN", "FTA/CEPA": "ORIGIN",
        "CBAM": "CBAM",
        "AD/CVD": "TRADE_REMEDY", "미국 301/232 관세": "TRADE_REMEDY", "관세율": "TRADE_REMEDY", "관세율/할당관세": "TRADE_REMEDY",
        "통관/세관심사": "CUSTOMS",
        "수출통제/제재": "EXPORT_CONTROL", "Entity List": "EXPORT_CONTROL", "반도체 관세": "SEMICONDUCTOR",
    }
    for _, r in executive_sort_frame(candidates).iterrows():
        topic = infer_topic_ko(r)
        family = family_map.get(topic, topic)
        if family in used_family:
            continue
        subs = clean(r.get("Affected Subsidiary")) or "관련 법인"
        if family == "ORIGIN":
            action = f"{subs}: 주요 거래·HS 기준 원산지 기준, BOM 충족 여부, CO 발급·보관 증빙 점검"
            owner = "FTA 운영/생산법인"
            issue = "원산지/CO"
        elif family == "CBAM":
            action = "EU향 제품·원재료의 CBAM 대상 여부, 공급사 탄소자료 확보 현황 및 신고 증빙 체계 점검"
            owner = "구매/ESG/관세"
            issue = "CBAM"
        elif family == "TRADE_REMEDY":
            action = "대상 HS·공급국·거래법인 매핑 후 가격자료·원산지 증빙 및 AD/CVD 대응 시나리오 점검"
            owner = "통상대응/구매"
            issue = "AD/CVD"
        elif family == "CUSTOMS":
            action = "공식 법규 원문 기준 시행일·대상 HS·세율·신고기준 확인 및 HQ Master/체크리스트 반영"
            owner = "통관운영/관세사"
            issue = "통관/세관심사"
        elif family == "EXPORT_CONTROL":
            action = "거래처 스크리닝, ECCN/전략물자 분류 및 제재국 우회거래 여부 점검"
            owner = "수출통제/법무"
            issue = topic
        else:
            action = "대상 국가·HS·제품군 기준으로 삼성전자 수출입 영향 확인"
            owner = "관세/사업부"
            issue = topic
        actions.append({"No": len(actions) + 1, "Issue": issue, "Required Action": action, "Owner": owner})
        used_family.add(family)
        if len(actions) >= 3:
            break
    return actions


def _priority_theme_sentence(top3: pd.DataFrame, rows: pd.DataFrame) -> str:
    text = " ".join(clean(x) for x in rows.get("Headline", pd.Series(dtype=str)).tolist()).lower()
    if "oman" in text and "india" in text:
        return "금일 GTI Radar는 인도·중동 FTA 원산지 규정 개정과 글로벌 관세정책 변화를 핵심 이슈로 식별하였습니다."
    if "india" in text and ("uk" in text or "united kingdom" in text):
        return "금일 GTI Radar는 인도·영국 FTA 이행 쟁점과 글로벌 관세·원산지 정책 변화를 핵심 이슈로 식별하였습니다."
    if any(k in text for k in ["usmca", "북미자유무역", "mexico", "canada"]):
        return "금일 GTI Radar는 북미 FTA 원산지 체계 변화와 글로벌 관세정책을 핵심 이슈로 식별하였습니다."
    topics = []
    for _, r in top3.iterrows():
        t = infer_topic_ko(r)
        if t not in topics:
            topics.append(t)
    return f"금일 GTI Radar는 {'·'.join(topics[:3]) if topics else '관세·통상'} 이슈를 핵심으로 식별하였습니다."



# =========================================================
# V3.1 EXECUTIVE 95-POINT OVERRIDES
# 1) Affected Subsidiary aggregation for executive summary
# 2) Required Actions grouped by Issue Family + Owner
# 3) Owner standardization to 5 executive owners
# 4) Top3 topic diversity: one issue family per Top3
# =========================================================

OWNER_STANDARD_MAP_V31 = {
    "ORIGIN": "FTA팀",
    "CUSTOMS": "통관운영팀",
    "TRADE_REMEDY": "Global SCM",
    "CBAM": "ESG팀",
    "EXPORT_CONTROL": "수출통제팀",
    "SEMICONDUCTOR": "Global SCM",
    "OTHER": "통관운영팀",
}

FAMILY_LABEL_V31 = {
    "ORIGIN": "원산지/FTA",
    "CUSTOMS": "통관/세관심사",
    "TRADE_REMEDY": "관세·AD/CVD",
    "CBAM": "CBAM",
    "EXPORT_CONTROL": "수출통제",
    "SEMICONDUCTOR": "반도체 관세",
    "OTHER": "관세·통상",
}


def issue_family_v31(row_or_topic) -> str:
    if isinstance(row_or_topic, pd.Series):
        topic = infer_topic_ko(row_or_topic)
        issue = clean(row_or_topic.get("Issue")).upper()
    else:
        topic = clean(row_or_topic)
        issue = ""
    if topic in {"원산지/CO", "FTA/CEPA"} or issue in {"ORIGIN", "ORIGIN_FTA", "FTA_CEPA"}:
        return "ORIGIN"
    if topic == "CBAM" or issue in {"CBAM", "CBAM_CARBON"}:
        return "CBAM"
    if topic in {"AD/CVD", "미국 301/232 관세", "관세율/할당관세", "관세율"} or issue in {"AD_CVD", "SECTION_301_232", "TARIFF", "TARIFF_DUTY"}:
        return "TRADE_REMEDY"
    if topic in {"통관/세관심사"} or issue in {"CUSTOMS_CLEARANCE", "CUSTOMS_REGULATION_NEWS"}:
        return "CUSTOMS"
    if topic in {"수출통제/제재", "Entity List"} or issue in {"EXPORT_CONTROL", "ENTITY_LIST", "UFLPA"}:
        return "EXPORT_CONTROL"
    if topic in {"반도체 관세"} or issue in {"SEMICONDUCTOR_TARIFF"}:
        return "SEMICONDUCTOR"
    return "OTHER"


def standard_owner_v31(row_or_topic) -> str:
    return OWNER_STANDARD_MAP_V31.get(issue_family_v31(row_or_topic), "통관운영팀")


def _split_subs_v31(value: str) -> list[str]:
    raw = clean(value)
    if not raw or raw == "-":
        return []
    raw = raw.replace("외", "")
    return [p.strip() for p in re.split(r"[,;/]+", raw) if p.strip()]


def summarize_subsidiaries_v31(rows: pd.DataFrame, max_items: int = 6) -> str:
    visible = rows[~rows["Mail Group"].eq("Filtered Noise")].copy() if "Mail Group" in rows.columns else rows.copy()
    counts: dict[str, int] = {}
    for _, r in visible.iterrows():
        for s in _split_subs_v31(clean(r.get("Affected Subsidiary"))):
            if s in {"-", "관련 법인", "SEC/HQ"}:
                continue
            counts[s] = counts.get(s, 0) + (2 if clean(r.get("Samsung Impact")) == "Direct" else 1)
    if not counts:
        return "SEC/HQ"
    ordered = sorted(counts, key=lambda k: (-counts[k], k))[:max_items]
    return ", ".join(ordered) + (" 외" if len(counts) > max_items else "")


def summarize_issue_owners_v31(rows: pd.DataFrame) -> str:
    visible = rows[~rows["Mail Group"].eq("Filtered Noise")].copy() if "Mail Group" in rows.columns else rows.copy()
    families = []
    for _, r in visible.iterrows():
        fam = issue_family_v31(r)
        if fam not in families and fam != "OTHER":
            families.append(fam)
    pairs = [f"{FAMILY_LABEL_V31.get(f, f)}:{OWNER_STANDARD_MAP_V31.get(f, '통관운영팀')}" for f in families[:4]]
    return " / ".join(pairs) if pairs else "통관/세관심사:통관운영팀"


def _headline_short_v31(row: pd.Series, n: int = 68) -> str:
    h = clean(row.get("Headline"))
    return h[:n].rstrip() + ("…" if len(h) > n else "")


def choose_top3(rows: pd.DataFrame) -> pd.DataFrame:
    """V3.1: Top3는 동일 주제 반복을 금지하고, 임원 보고용으로 ORIGIN/CUSTOMS/CBAM/TRADE_REMEDY 등 다양성을 확보."""
    pool = rows[~rows["Mail Group"].eq("Filtered Noise")].copy()
    if pool.empty:
        return pool
    if "_display_cluster" not in pool.columns:
        pool["_display_cluster"] = pool.apply(lambda r: clean(r.get("Cluster")) or clean(r.get("Headline"))[:80], axis=1)
    pool = executive_sort_frame(pool).drop_duplicates(subset=["_display_cluster"], keep="first")

    # Prefer a balanced executive mix: one FTA/origin, one regulation/customs, one CBAM or trade remedy.
    family_priority = ["ORIGIN", "CUSTOMS", "CBAM", "TRADE_REMEDY", "EXPORT_CONTROL", "SEMICONDUCTOR", "OTHER"]
    selected = []
    used_clusters = set()
    used_families = set()
    for fam in family_priority:
        cand = pool[pool.apply(lambda r: issue_family_v31(r) == fam, axis=1)]
        if cand.empty:
            continue
        # Within each family prefer Direct, then high score.
        cand = cand.copy()
        cand["_fam_direct"] = cand["Samsung Impact"].apply(lambda x: 1 if clean(x) == "Direct" else 0)
        cand = cand.sort_values(["_fam_direct", "_integrated_score", "_sort_date"], ascending=[False, False, False])
        for _, row in cand.iterrows():
            cluster = clean(row.get("_display_cluster"))
            if cluster in used_clusters:
                continue
            selected.append(row)
            used_clusters.add(cluster)
            used_families.add(fam)
            break
        if len(selected) >= 3:
            break

    if len(selected) < 3:
        for _, row in pool.iterrows():
            cluster = clean(row.get("_display_cluster"))
            fam = issue_family_v31(row)
            if cluster in used_clusters:
                continue
            if fam in used_families and len(selected) < 3:
                continue
            selected.append(row)
            used_clusters.add(cluster)
            used_families.add(fam)
            if len(selected) >= 3:
                break
    return pd.DataFrame(selected).reset_index(drop=True)


def build_required_actions(rows: pd.DataFrame, top3: pd.DataFrame) -> list[dict]:
    """V3.1: Topic+Owner family 기준으로 groupby하여 중복 Action 제거, Owner는 5개 표준 조직만 사용."""
    visible = rows[~rows["Mail Group"].eq("Filtered Noise")].copy()
    candidates = pd.concat([top3, visible], ignore_index=True)
    if "_display_cluster" not in candidates.columns:
        candidates["_display_cluster"] = candidates.apply(lambda r: clean(r.get("Cluster")) or clean(r.get("Headline"))[:80], axis=1)
    candidates = candidates.drop_duplicates(subset=["_display_cluster"], keep="first")
    candidates = executive_sort_frame(candidates)

    grouped: dict[tuple[str, str], dict] = {}
    for _, r in candidates.iterrows():
        fam = issue_family_v31(r)
        owner = standard_owner_v31(r)
        key = (fam, owner)
        subs = clean(r.get("Affected Subsidiary")) or "관련 법인"
        if key not in grouped:
            grouped[key] = {"rows": [], "subs": []}
        grouped[key]["rows"].append(r)
        for s in _split_subs_v31(subs):
            if s not in grouped[key]["subs"]:
                grouped[key]["subs"].append(s)

    priority = ["ORIGIN", "CUSTOMS", "CBAM", "TRADE_REMEDY", "EXPORT_CONTROL", "SEMICONDUCTOR", "OTHER"]
    actions = []
    for fam in priority:
        for (family, owner), data in list(grouped.items()):
            if family != fam:
                continue
            subs = ", ".join(data["subs"][:4]) + (" 외" if len(data["subs"]) > 4 else "")
            if not subs:
                subs = "관련 법인"
            if family == "ORIGIN":
                issue = "원산지/FTA"
                action = f"{subs}: 주요 거래·HS 기준 원산지 기준, BOM 충족 여부, CO 발급·보관 증빙 일괄 점검"
            elif family == "CUSTOMS":
                issue = "통관/세관심사"
                action = "공식 법규 원문 기준 시행일·대상 HS·세율·신고기준 확인 후 HQ 관세 Master와 신고 체크리스트 반영"
            elif family == "CBAM":
                issue = "CBAM"
                action = "EU향 제품·원재료의 CBAM 대상 여부, 공급사 탄소자료 확보율 및 신고 증빙 체계 점검"
            elif family == "TRADE_REMEDY":
                issue = "관세·AD/CVD"
                action = f"{subs}: 대상 HS·공급국·거래법인 매핑 후 가격자료·원산지 증빙 및 대응 시나리오 점검"
            elif family == "EXPORT_CONTROL":
                issue = "수출통제"
                action = "거래처 스크리닝, ECCN/전략물자 분류 및 제재국 우회거래 여부 점검"
            elif family == "SEMICONDUCTOR":
                issue = "반도체 관세"
                action = "반도체 관련 대상 HS·세율·적용시점 확인 및 생산/판매법인별 관세원가 영향 산출"
            else:
                issue = "관세·통상"
                action = "대상 국가·HS·제품군 기준으로 삼성전자 수출입 영향 확인"
            actions.append({"No": len(actions)+1, "Issue": issue, "Required Action": action, "Owner": owner, "topic": issue, "action": action, "owner": owner})
            if len(actions) >= 4:
                return actions
    return actions


def build_required_actions_html(rows: pd.DataFrame, top3: pd.DataFrame) -> str:
    actions = build_required_actions(rows, top3)
    if not actions:
        return ""
    trs = []
    for i, a in enumerate(actions, 1):
        topic = clean(a.get("topic") or a.get("Issue") or a.get("Topic") or "관세·통상")
        action = clean(a.get("action") or a.get("Required Action") or a.get("RequiredAction") or "대상 국가·HS·제품군 기준 영향 확인")
        owner = clean(a.get("owner") or a.get("Owner") or "통관운영팀")
        trs.append(
            "<tr>"
            f"<td style='padding:8px;border:1px solid #d9d9d9;text-align:center;font-weight:bold;'>{i}</td>"
            f"<td style='padding:8px;border:1px solid #d9d9d9;'>{html.escape(topic)}</td>"
            f"<td style='padding:8px;border:1px solid #d9d9d9;'>{html.escape(action)}</td>"
            f"<td style='padding:8px;border:1px solid #d9d9d9;text-align:center;'>{html.escape(owner)}</td>"
            "</tr>"
        )
    return (
        "<h3 style='margin-top:22px;color:#C00000;'>2. Today's Required Actions</h3>"
        "<table style='border-collapse:collapse;width:100%;font-size:13px;margin-bottom:18px;'>"
        "<thead><tr style='background:#C00000;color:white;'>"
        "<th style='padding:8px;border:1px solid #d9d9d9;width:45px;'>No</th>"
        "<th style='padding:8px;border:1px solid #d9d9d9;width:130px;'>Issue</th>"
        "<th style='padding:8px;border:1px solid #d9d9d9;'>Required Action</th>"
        "<th style='padding:8px;border:1px solid #d9d9d9;width:140px;'>Owner</th>"
        "</tr></thead><tbody>" + "".join(trs) + "</tbody></table>"
    )


def executive_issue_sentence(row: pd.Series) -> str:
    fam = issue_family_v31(row)
    country = _exec_country_short_v305(row)
    subs = clean(row.get("Affected Subsidiary")) or "관련 법인"
    text = _text_all_v302(row)
    if fam == "ORIGIN":
        if "oman" in text:
            return "SIEL 중심의 인도·중동 CEPA 활용 가능성이 확대되고 있어 중동향 수출품의 원산지 판정, CO 발급 및 협정세율 적용 기준 재점검이 필요합니다."
        if "india" in text and ("uk" in text or "united kingdom" in text):
            return "SIEL·SEUK 간 인도·영국 FTA 이행 쟁점은 원산지 판정, CO 발급 및 협정세율 적용 기준 점검이 필요한 사안입니다."
        if any(k in text for k in ["usmca", "북미자유무역", "mexico", "canada"]):
            return "북미 생산·판매법인은 USMCA 연장·개정 논의에 따른 원산지 기준 및 역내조달 전략 변화를 지속 모니터링해야 합니다."
        return f"{subs} 관련 FTA/원산지 이슈는 CO 발급, 원산지 판정 및 특혜세율 적용 기준 검토가 필요합니다."
    if fam == "CUSTOMS":
        return "공식 법규 변경사항은 시행일·대상 HS·세율·신고기준을 확인하여 HQ 관세 Master와 신고 체크리스트에 반영해야 합니다."
    if fam == "CBAM":
        return "EU 판매·공급망 관련 법인은 CBAM 대상 원재료, 공급사 탄소자료 확보 및 EU향 신고 증빙 체계를 점검해야 합니다."
    if fam == "TRADE_REMEDY":
        return f"{country} 관세·AD/CVD 이슈는 대상 HS, 공급국, 가격자료 및 원산지 방어자료를 중심으로 영향 분석이 필요합니다."
    if fam == "EXPORT_CONTROL":
        return "수출통제·제재 이슈는 거래처 스크리닝, ECCN/전략물자 분류 및 우회거래 점검이 필요합니다."
    if fam == "SEMICONDUCTOR":
        return "반도체 관세 정책 변화는 글로벌 생산거점의 관세원가와 공급망 전략에 영향을 줄 수 있어 대상 HS와 적용시점 확인이 필요합니다."
    return "대상 국가·HS·제품군 기준으로 삼성전자 관세·통상 영향을 확인해야 합니다."


def build_overall_review_html(rows: pd.DataFrame, top3: pd.DataFrame) -> str:
    visible = rows[~rows["Mail Group"].eq("Filtered Noise")].copy()
    filtered = rows[rows["Mail Group"].eq("Filtered Noise")]
    direct_rows = visible[visible["Samsung Impact"].eq("Direct")]
    indirect_rows = visible[visible["Samsung Impact"].eq("Indirect")]
    regulation = visible[visible["Mail Group"].eq("Regulation")]
    news_core = visible[visible["Mail Group"].eq("News - 핵심")]
    news_usable = visible[visible["Mail Group"].eq("News - 주요/참고")]
    counts = f"법규 {len(regulation)}건, 핵심 뉴스 {len(news_core)}건, 주요/참고 뉴스 {len(news_usable)}건 | 직접영향 {len(direct_rows)}건, 간접영향 {len(indirect_rows)}건, 제외 {len(filtered)}건"

    impacted_subs = summarize_subsidiaries_v31(visible)
    owner_map_text = summarize_issue_owners_v31(visible)
    sentences = [_priority_theme_sentence(top3, visible)]
    sentences.append(f"금일 주요 영향 법인은 {impacted_subs}이며, 이슈별 담당은 {owner_map_text} 기준으로 관리합니다.")
    for _, r in top3.iterrows():
        s = executive_issue_sentence(r)
        if s not in sentences:
            sentences.append(s)
        if len(sentences) >= 5:
            break
    sentences.append(f"직접영향 {len(direct_rows)}건, 간접영향 {len(indirect_rows)}건이 식별되었으며 원산지·FTA·관세 대응을 우선 추진해야 합니다.")

    def mini_list(title: str, frame: pd.DataFrame, color: str, maxn: int = 3) -> str:
        frame = executive_sort_frame(frame).drop_duplicates(subset=["_display_cluster"], keep="first") if not frame.empty else frame
        # enforce topic diversity in the mini list too
        chosen, used = [], set()
        for _, r in frame.iterrows():
            fam = issue_family_v31(r)
            if fam in used and len(chosen) < maxn:
                continue
            chosen.append(r)
            used.add(fam)
            if len(chosen) >= maxn:
                break
        if len(chosen) < maxn:
            for _, r in frame.iterrows():
                if any(clean(r.get("Headline")) == clean(x.get("Headline")) for x in chosen):
                    continue
                chosen.append(r)
                if len(chosen) >= maxn:
                    break
        if not chosen:
            return f"<div style='margin-top:8px;color:#777;'>{html.escape(title)}: 해당 없음</div>"
        items = []
        for r in chosen[:maxn]:
            items.append(f"<li style='margin:3px 0;'><b>{html.escape(FAMILY_LABEL_V31.get(issue_family_v31(r), infer_topic_ko(r)))}</b> [{html.escape(_exec_country_short_v305(r))} / {html.escape(clean(r.get('Affected Subsidiary')) or '-')}] {html.escape(_headline_short_v31(r))}</li>")
        return f"<div style='margin-top:12px;'><b style='color:{color};'>{html.escape(title)}</b><ol style='margin-top:6px;margin-bottom:4px;padding-left:22px;'>{''.join(items)}</ol></div>"

    summary_html = "".join(f"<div style='font-size:15px;font-weight:bold;line-height:1.85;margin-bottom:8px;'>{html.escape(s)}</div>" for s in sentences)
    return f"""
    <div style="margin-top:10px;margin-bottom:20px;padding:16px;background:#F4F6F8;border-left:6px solid #1F4E78;color:#222;">
      <div style="font-size:14px;color:#555;margin-bottom:10px;">금일 선별 결과: {html.escape(counts)}</div>
      {summary_html}
      {mini_list('직접영향 Top 3', direct_rows, '#C00000', 3)}
      {mini_list('간접영향 Top 3', indirect_rows, '#7F7F7F', 3)}
    </div>
    """



# =========================================================
# V3.1.4 EXECUTIVE FINAL URL/DATA QUALITY OVERRIDES
# - Do NOT drop important rows just because URL is missing/bad
# - Block Google thumbnail/image URLs (lh3.googleusercontent, gstatic)
# - If Google RSS cannot be resolved, show headline as plain text (no bad link)
# - Remove sugar/cane/food AD-CVD noise from executive mail
# - Fix CBAM template so CBAM does not show Entity List/ECCN actions
# =========================================================

BAD_URL_HOST_FRAGMENTS_V314 = [
    "news.google.com",
    "lh3.googleusercontent.com",
    "googleusercontent.com",
    "gstatic.com",
    "encrypted-tbn",
    "google.com/imgres",
]

def is_preferred_article_url(url: str) -> bool:
    u = clean(url)
    if not is_valid_http_url(u):
        return False
    low = u.lower()
    if any(x in low for x in BAD_URL_HOST_FRAGMENTS_V314):
        return False
    bad_fragments = [
        "accounts.google.",
        "policies.google.",
        "support.google.",
        "consent.google.",
        "google.com/search",
        "google.com/amp/s/",
        "agency.reuters.com/en/copyright",
    ]
    return not any(x in low for x in bad_fragments)


def best_url_from_values(values: list[str]) -> str:
    """Return only a real article URL.

    Important: if STEP4 only has Google News / thumbnail URL, return empty.
    The article stays in the report, but the headline is plain text.
    """
    invalid_tokens = {"", "nan", "none", "null", "new", "https://new", "http://new", "https://news", "http://news", "https://news.google.com/", "http://news.google.com/"}
    cleaned = []
    for v in values:
        vv = clean(v)
        if vv.lower() in invalid_tokens:
            continue
        cleaned.append(vv)
        for found in re.findall(r"https?://[^'\"),\s]+", vv):
            if found not in cleaned:
                cleaned.append(found)
        for found in re.findall(r"https?://[^'\"),\s]+", vv):
            if found not in cleaned:
                cleaned.append(found)
        for found in re.findall(r"https?://[^'\"),\s]+", vv):
            if found not in cleaned:
                cleaned.append(found)

    for v in cleaned:
        if is_preferred_article_url(v):
            return v

    for v in cleaned:
        if is_google_news_rss_url(v):
            resolved = resolve_google_news_url(v)
            if is_preferred_article_url(resolved):
                return resolved

    return ""


def html_link(title: str, url: str) -> str:
    title = html.escape(clean(title))
    url = best_url_from_values([url])
    if url:
        return f'<a href="{html.escape(url)}" target="_blank">{title}</a>'
    return title


def read_step4_results() -> pd.DataFrame:
    """Read STEP4 outputs without dropping rows that lack usable URL.

    Previous versions filtered URL != '', which caused good FTA/CBAM rows to disappear
    when Google News original URL was unavailable. Executive report must keep the item
    and simply suppress the broken hyperlink.
    """
    frames = []
    if REGULATION_INPUT_FILE.exists():
        frames.append(normalize_input(pd.read_excel(REGULATION_INPUT_FILE), "Regulation", REGULATION_INPUT_FILE))
    if NEWS_INPUT_FILE.exists():
        news = normalize_input(pd.read_excel(NEWS_INPUT_FILE), "News", NEWS_INPUT_FILE)
        news = news.head(NEWS_MAX_ROWS)
        frames.append(news)
    if not frames:
        raise FileNotFoundError(f"STEP4 outputs not found: {REGULATION_INPUT_FILE}, {NEWS_INPUT_FILE}")
    rows = pd.concat(frames, ignore_index=True)
    rows = rows[rows["Headline"].astype(str).str.strip().ne("")].copy()

    # Re-clean URLs one more time. Bad URLs become blank, not row deletion.
    rows["URL"] = rows.apply(lambda r: best_url_from_values([
        r.get("URL", ""),
        r.get("Source", ""),
    ]), axis=1)

    # Deduplicate: if URL is blank, use headline/agency/date instead.
    rows["_dedup_key"] = rows.apply(
        lambda r: clean(r.get("URL")) or (clean(r.get("Headline"))[:120] + "|" + clean(r.get("Agency")) + "|" + clean(r.get("Date"))),
        axis=1,
    )
    rows = rows.drop_duplicates(subset=["_dedup_key"], keep="first").drop(columns=["_dedup_key"], errors="ignore")
    rows["_integrated_score"] = rows.apply(
        lambda r: priority_weight(r["Priority Group"]) + risk_weight(r["Risk"]) + type_weight(r["Content Type"]) + safe_num(r["Importance Score"]),
        axis=1,
    )
    return rows.reset_index(drop=True)


def append_row(ws, row: pd.Series) -> None:
    ws.append([row.get(c, "") for c in OUTPUT_COLUMNS])
    headline_cell = ws.cell(row=ws.max_row, column=OUTPUT_COLUMNS.index("Headline") + 1)
    url = best_url_from_values([row.get("URL")])
    if url:
        headline_cell.hyperlink = url
        headline_cell.font = Font(color="0563C1", underline="single", bold=True)


FOOD_AD_CVD_NOISE_V314 = [
    "sugar", "thai sugar", "sugar products", "sugarcane", "cane sugar", "cane", "sucrose",
    "사탕수수", "설탕", "당류", "원당", "정제당", "농산물", "식품", "음료",
]

def is_forced_noise_v303(row: pd.Series) -> bool:
    text = _text_all_v302(row)

    # Food/agriculture AD-CVD is not Samsung executive customs radar unless Samsung/electronics is explicit.
    if any(k in text for k in FOOD_AD_CVD_NOISE_V314):
        hard_keep = ["samsung", "삼성전자", "semiconductor", "반도체", "electronics", "전자", "battery", "display"]
        return not any(k in text for k in hard_keep)

    if _contains_security_noise_v304(row):
        return True
    if _is_diplomatic_investment_noise_v305(row):
        return True

    if any(p.lower() in text for p in FORCED_NOISE_V304):
        hard_keep = [
            "semiconductor", "반도체", "entity list", "export control", "수출통제",
            "uflpa", "section 301", "section 232", "rules of origin", "origin rules", "electronics exporters",
        ]
        return not any(k in text for k in hard_keep)
    return False


def infer_topic_ko(row: pd.Series) -> str:
    issue = clean(row.get("Issue")).upper()
    headline = clean(row.get("Headline")).lower()
    cluster = clean(row.get("Cluster")).lower()
    base = f"{issue.lower()} {cluster} {headline}"

    # Forced labor tariff is closer to trade remedy / import control than normal AD/CVD.
    if any(k in base for k in ["forced labor", "강제 노동", "강제노동", "uflpa"]):
        return "수출통제/제재"

    if "CBAM" in issue or "cbam" in base or "탄소국경" in base:
        return "CBAM"
    if issue in {"ORIGIN_FTA", "FTA_CEPA", "ORIGIN"} or any(k in base for k in ["fta", "cepa", "rules of origin", "origin rules", "원산지", "co 발급"]):
        return "원산지/FTA"
    if issue in {"CUSTOMS_CLEARANCE"} or any(k in base for k in ["customs", "세관", "통관", "할당관세"]):
        return "통관/세관심사"
    if issue in {"AD_CVD"} or any(k in base for k in ["anti-dumping", "antidumping", "countervailing", "반덤핑", "상계관세", "ad/cvd"]):
        return "관세·AD/CVD"
    if any(k in base for k in ["entity list", "entity_list", "제재명단"]):
        return "Entity List"
    if any(k in base for k in ["export control", "export_control", "sanction", "제재", "수출통제"]):
        return "수출통제/제재"
    if any(k in base for k in ["section 301", "section 232", "301", "232", "ustr"]):
        return "미국 301/232 관세"
    if any(k in base for k in ["semiconductor", "반도체", "chip"]):
        return "반도체 관세"
    if any(k in base for k in ["tariff", "duty", "관세", "세율"]):
        return "관세율"
    return clean(row.get("Issue")) or "관세·통상"


EXECUTIVE_OWNER_MAP = {
    "원산지/FTA": "FTA팀",
    "원산지/CO": "FTA팀",
    "FTA/CEPA": "FTA팀",
    "통관/세관심사": "통관운영팀",
    "CBAM": "ESG팀",
    "관세·AD/CVD": "Global SCM",
    "AD/CVD": "Global SCM",
    "관세율": "통관운영팀",
    "관세율/할당관세": "통관운영팀",
    "미국 301/232 관세": "Global SCM",
    "수출통제/제재": "수출통제팀",
    "Entity List": "수출통제팀",
    "반도체 관세": "Global SCM",
}


def _topic_family_v314(topic: str) -> str:
    t = clean(topic)
    if t in {"원산지/FTA", "원산지/CO", "FTA/CEPA"}:
        return "FTA"
    if t == "CBAM":
        return "CBAM"
    if t in {"관세·AD/CVD", "AD/CVD", "미국 301/232 관세", "관세율", "관세율/할당관세"}:
        return "TARIFF"
    if t in {"수출통제/제재", "Entity List"}:
        return "EXPORT_CONTROL"
    if t == "통관/세관심사":
        return "CUSTOMS"
    return t


def cluster_key_v303(row: pd.Series) -> str:
    topic = infer_topic_ko(row)
    text = _text_all_v302(row)
    title = clean(row.get("Headline")).lower()
    country = clean(row.get("Country")).lower()
    if clean(row.get("Content Type")).lower().startswith("reg"):
        return "REGULATION:" + (clean(row.get("Cluster")) or clean(row.get("Headline"))[:80])
    if is_forced_noise_v303(row):
        return "FORCED_NOISE"
    if ("india" in text and ("uk" in text or "united kingdom" in text)) and any(k in text for k in ["fta", "cepa", "carbon tax", "steel", "scotch", "cbam"]):
        return "INDIA_UK_FTA_CBAM_STEEL"
    if "india" in text and "oman" in text and any(k in text for k in ["cepa", "zero-duty", "electronics", "rules of origin"]):
        return "INDIA_OMAN_CEPA_ELECTRONICS"
    if any(k in text for k in ["usmca", "north america", "free-trade agreement", "북미자유무역", "mexico free-trade", "u.s. and mexico"]):
        return "USMCA_EXTENSION"
    if topic == "CBAM" and ("eu" in text or "carbon" in text or "cbam" in text):
        return "EU_CBAM_GENERAL"
    if any(k in text for k in ["forced labor", "강제 노동", "강제노동", "uflpa"]):
        return "FORCED_LABOR_TARIFF"
    base = clean(row.get("Cluster")) or f"{topic}:{country}:{title[:90]}"
    return re.sub(r"\s+", " ", base).strip()


def determine_samsung_impact(row: pd.Series) -> tuple[str, str, str]:
    topic = infer_topic_ko(row)
    text = _text_all_v302(row)
    country = clean(row.get("Country")) or "국가 미상"
    country_subs = country_based_subsidiaries(country)
    explicit_subs = explicit_subsidiaries_in_text(row)
    subs = list(dict.fromkeys(country_subs + explicit_subs))

    if clean(row.get("Content Type")).lower().startswith("reg"):
        return "Direct", "SEC, HQ", "공식 법규/공고성 이슈로 HQ 관세 Master·신고 프로세스 반영 필요"

    if is_forced_noise_v303(row) or is_residual_noise_v3(row) or is_low_relevance(row):
        return "None", "", "삼성전자 관세·통상 업무 관련성이 낮아 메일 본문 제외"

    if topic in {"수출통제/제재", "Entity List"} and (country_subs or explicit_subs or has_samsung_product_signal(row)):
        return "Direct", _compact_subs_v305(subs or country_subs), "수출통제·강제노동·제재성 관세 이슈로 거래 심사 및 공급망 점검 필요"

    if topic in {"원산지/FTA", "원산지/CO", "FTA/CEPA"}:
        if "india" in text and "oman" in text and any(k in text for k in ["electronics", "전자", "zero-duty", "cepa", "rules of origin"]):
            return "Direct", _compact_subs_v305(country_subs or ["SIEL", "SGE"]), "인도·오만 CEPA 활용 가능성이 있어 제품별 CO·협정세율 검토 필요"
        if "india" in text and ("uk" in text or "united kingdom" in text):
            return "Direct", _compact_subs_v305(country_subs or ["SIEL", "SEUK"]), "인도·영국 FTA 이행 쟁점으로 원산지·협정세율 점검 필요"
        if any(k in text for k in ["usmca", "북미자유무역", "u.s. and mexico", "free-trade agreement for another 16 years"]):
            return "Indirect", _compact_subs_v305(country_subs), "북미 FTA 연장·개정 모니터링 대상"
        if country_subs:
            return "Indirect", _compact_subs_v305(country_subs), "FTA/원산지 정책 모니터링 대상"

    if topic == "CBAM":
        if any(k in text for k in ["automotive", "carmaker", "자동차", "morocco", "모로코", "우회수출"]):
            return "Indirect", _compact_subs_v305(subs or country_subs), "CBAM·우회수출 정책 모니터링 대상"
        return "Direct", _compact_subs_v305(subs or country_subs or ["EU 법인"]), "CBAM 대상 원재료·EU향 증빙 체계 영향 가능"

    if topic in {"관세·AD/CVD", "AD/CVD", "미국 301/232 관세", "관세율"}:
        if any(k in text for k in ["forced labor", "강제 노동", "강제노동", "uflpa", "section 301", "section 232", "ustr", "semiconductor", "반도체"]):
            return "Direct", _compact_subs_v305(country_subs), "고위험 관세정책으로 HS·원산지·관세율 영향 점검 필요"
        if country_subs:
            return "Indirect", _compact_subs_v305(country_subs), "관세·AD/CVD 정책 모니터링 대상"

    if country_subs:
        return "Indirect", _compact_subs_v305(country_subs), "법인 소재국 관련 정책이나 제품·거래 직접성은 추가 확인 필요"
    return "Indirect", "", f"토픽({topic})은 모니터링 가치가 있으나 특정 삼성 법인 직접영향은 확인 불가"


def force_korean_text(row: pd.Series, field: str, fallback: str) -> str:
    topic = infer_topic_ko(row)
    country = _exec_country_short_v305(row)
    subs = clean(row.get("Affected Subsidiary")) or "관련 법인"
    headline = clean(row.get("Headline"))

    if field == "Summary":
        return f"{headline} 관련 {topic} 이슈입니다. 대상 국가, 적용 품목, 시행일 및 관세·통상 영향 여부 확인이 필요합니다."

    if field == "AI Analysis":
        if topic in {"원산지/FTA", "원산지/CO", "FTA/CEPA"}:
            return f"{country} FTA/원산지 이슈로 {subs}의 CO 발급, 원산지 판정 및 협정세율 적용 기준 재점검이 필요합니다."
        if topic == "CBAM":
            return f"{country} CBAM 이슈로 EU향 제품·원재료 탄소자료, 공급사 증빙 및 신고 대응 체계 점검이 필요합니다."
        if topic in {"관세·AD/CVD", "AD/CVD", "미국 301/232 관세", "관세율"}:
            return f"{country} 관세·AD/CVD 이슈로 대상 HS, 공급국, 가격자료 및 원산지 증빙 영향 분석이 필요합니다."
        if topic in {"수출통제/제재", "Entity List"}:
            return f"{country} 수출통제·강제노동·제재성 관세 이슈로 거래처 스크리닝, 공급망 원산지 및 우회거래 점검이 필요합니다."
        if topic == "통관/세관심사" or clean(row.get("Content Type")).lower().startswith("reg"):
            return "공식 법규/고시 변경사항으로 시행일, 대상 HS, 세율 및 신고기준을 HQ Master와 체크리스트에 반영해야 합니다."
        return fallback

    if field == "Action Plan":
        if topic in {"원산지/FTA", "원산지/CO", "FTA/CEPA"}:
            return "① 대상 거래·HS 추출 ② 협정 원산지 기준 확인 ③ BOM/공정/공급자확인서 점검 ④ CO 발급·보관 증빙 업데이트"
        if topic == "CBAM":
            return "① 대상 HS·원재료 확인 ② EU향 거래 매핑 ③ 공급사 탄소자료 확보 ④ CBAM 신고·증빙 체계 업데이트"
        if topic in {"관세·AD/CVD", "AD/CVD", "미국 301/232 관세", "관세율"}:
            return "① 대상 HS·공급국 확인 ② 수입·수출 신고 이력 매칭 ③ 가격자료·원산지 증빙 점검 ④ 관세사·법무와 대응 시나리오 수립"
        if topic in {"수출통제/제재", "Entity List"}:
            return "① 거래처·공급망 스크리닝 ② UFLPA/강제노동 원산지 리스크 확인 ③ 우회거래 여부 점검 ④ Hold/Release 기록화"
        if topic == "통관/세관심사" or clean(row.get("Content Type")).lower().startswith("reg"):
            return "① 원문 확인 ② 대상 국가·HS·법인 매핑 ③ 신고·증빙 영향 점검 ④ 체크리스트 업데이트"
        return fallback

    return clean(row.get(field)) or fallback


def executive_sort_frame(rows: pd.DataFrame) -> pd.DataFrame:
    rows = rows.copy()
    impact_weight = rows["Samsung Impact"].map({"Direct": 900, "Indirect": 450, "None": -9999}).fillna(0)
    topic_weight = rows.apply(lambda r: 550 if infer_topic_ko(r) in {"수출통제/제재", "Entity List", "반도체 관세", "CBAM", "원산지/FTA", "통관/세관심사", "관세·AD/CVD"} else 0, axis=1)
    source_weight = rows["Agency"].astype(str).str.lower().map(lambda x: 150 if any(s in x for s in ["reuters", "bloomberg", "yna", "연합", "mlex", "nhk", "ustr", "law.go.kr", "economic times", "fortune", "hankyung"]) else 0)
    rows["_exec_score"] = rows["_integrated_score"] + impact_weight + topic_weight + source_weight
    return rows.sort_values(["_exec_score", "_sort_date"], ascending=[False, False])


def choose_top3(rows: pd.DataFrame) -> pd.DataFrame:
    pool = rows[~rows["Mail Group"].eq("Filtered Noise")].copy()
    if pool.empty:
        return pool
    pool = executive_sort_frame(pool).drop_duplicates(subset=["_display_cluster"], keep="first")

    selected, used_family, used_clusters = [], set(), set()
    for _, row in pool.iterrows():
        cluster = clean(row.get("_display_cluster"))
        fam = _topic_family_v314(infer_topic_ko(row))
        if cluster in used_clusters:
            continue
        if fam in used_family and len(selected) < 3:
            continue
        selected.append(row)
        used_clusters.add(cluster)
        used_family.add(fam)
        if len(selected) >= 3:
            break
    if len(selected) < 3:
        for _, row in pool.iterrows():
            cluster = clean(row.get("_display_cluster"))
            if cluster in used_clusters:
                continue
            selected.append(row)
            used_clusters.add(cluster)
            if len(selected) >= 3:
                break
    return pd.DataFrame(selected).reset_index(drop=True)


def final_order(rows: pd.DataFrame, top3: pd.DataFrame) -> pd.DataFrame:
    rows = apply_samsung_impact(rows.copy())
    rows["_display_cluster"] = rows.apply(cluster_key_v303, axis=1)
    rows = executive_sort_frame(rows)

    noise = rows[rows["Samsung Impact"].eq("None")].copy()
    noise["Mail Group"] = "Filtered Noise"

    visible_pool = rows[~rows["Samsung Impact"].eq("None")].copy()
    visible_pool = visible_pool.drop_duplicates(subset=["_display_cluster"], keep="first")

    reg = visible_pool[visible_pool["Content Type"].eq("Regulation")].head(1)
    nonreg = visible_pool[~visible_pool.index.isin(reg.index)].copy()

    direct_slots = max(0, DIRECT_TOTAL_MAX_V305 - int(reg["Samsung Impact"].eq("Direct").sum()))
    direct = nonreg[nonreg["Samsung Impact"].eq("Direct")].head(direct_slots)
    indirect = nonreg[~nonreg.index.isin(set(direct.index)) & nonreg["Samsung Impact"].eq("Indirect")].head(INDIRECT_MAX_V305)

    visible = pd.concat([direct, reg, indirect], ignore_index=False)
    visible = _set_group_v305(visible)
    visible = executive_sort_frame(visible).head(VISIBLE_MAX_V305)

    out = pd.concat([visible, noise], ignore_index=True)
    out["Affected Subsidiary"] = out["Affected Subsidiary"].apply(_compact_subs_v305)
    out["No"] = range(1, len(out) + 1)
    return out.reset_index(drop=True)


def build_required_actions(rows: pd.DataFrame, top3: pd.DataFrame) -> list[dict]:
    visible = rows[~rows["Mail Group"].eq("Filtered Noise")].copy()
    pool = pd.concat([top3, visible[visible["Samsung Impact"].eq("Direct")]], ignore_index=True).drop_duplicates(subset=["_display_cluster"], keep="first")
    pool = executive_sort_frame(pool).head(6)
    actions, seen = [], set()
    for _, r in pool.iterrows():
        topic = infer_topic_ko(r)
        owner = EXECUTIVE_OWNER_MAP.get(topic, "관세기획/지역법인")
        subs = clean(r.get("Affected Subsidiary")) or clean(r.get("Country")) or "관련 법인"
        family = _topic_family_v314(topic)

        if family == "FTA":
            action = f"{subs}: 원산지 기준, BOM 충족 여부, CO 발급·보관 증빙 점검"
        elif family == "CBAM":
            action = f"{subs}: CBAM 대상 원재료와 공급사 탄소자료 확보 현황 점검"
        elif family == "TARIFF":
            action = f"{subs}: 대상 HS·공급국·가격자료 및 관세율 적용 영향 확인"
        elif family == "EXPORT_CONTROL":
            action = f"{subs}: 거래처·공급망 스크리닝 및 강제노동/우회거래 리스크 확인"
        elif family == "CUSTOMS":
            action = "공식 법규 원문 기준 시행일·대상 HS·세율·신고기준 확인 후 HQ 관세 Master와 신고 체크리스트 반영"
        else:
            action = f"{subs}: 원문 기준 대상 국가·HS·제품군 영향 확인"

        key = (family, owner)
        if key in seen:
            continue
        seen.add(key)
        actions.append({"owner": owner, "action": action, "topic": topic})
        if len(actions) >= 3:
            break
    return actions



# =========================================================
# V3.1.5 EXECUTIVE REPORT FINAL OVERRIDES
# - 24-hour publish-date gate: old items are excluded from visible mail
# - URL column repair: scan ALL URL-like columns, not only one matched column
# - URL hygiene: block Google thumbnails/RSS and show plain title when unresolved
# - Executive table form: fixed column widths matching requested ratio
# - Excel form: readable widths aligned to executive report layout
# =========================================================

LOOKBACK_HOURS_V315 = int(os.getenv("GTI_LOOKBACK_HOURS", "24"))
URL_CANDIDATE_COLUMNS_V315 = [
    "resolved_url", "ResolvedURL", "article_url", "ArticleURL", "original_url", "OriginalURL",
    "RepresentativeURL", "CanonicalURL", "FinalURL", "SourceURL", "source_url",
    "URL", "url", "Link", "link", "Source", "source",
]
BAD_URL_HOST_FRAGMENTS_V315 = list(dict.fromkeys(BAD_URL_HOST_FRAGMENTS_V314 + [
    "lh3.googleusercontent", "googleusercontent", "gstatic", "encrypted-tbn", "news.google.com",
    "google.com/rss", "google.com/articles", "google.com/read", "google.com/redirect",
]))
OLD_DATE_REASON_V315 = "24시간 기준 초과 게시건으로 금일 임원 보고 본문 제외"


def is_preferred_article_url(url: str) -> bool:
    u = clean(url)
    if not is_valid_http_url(u):
        return False
    low = u.lower()
    if any(x in low for x in BAD_URL_HOST_FRAGMENTS_V315):
        return False
    bad_fragments = [
        "accounts.google.", "policies.google.", "support.google.", "consent.google.",
        "google.com/search", "google.com/amp/s/", "agency.reuters.com/en/copyright",
        "/favicon", "logo", "thumbnail", "thumb", "w16", "=w16",
    ]
    return not any(x in low for x in bad_fragments)


def best_url_from_values(values: list[str]) -> str:
    invalid_tokens = {
        "", "nan", "none", "null", "new", "https://new", "http://new",
        "https://news", "http://news", "https://news.google.com/", "http://news.google.com/",
        "https://lh3.googleusercontent.com/", "http://lh3.googleusercontent.com/",
    }
    cleaned = []
    for v in values:
        vv = clean(v)
        if vv.lower() in invalid_tokens:
            continue
        cleaned.append(vv)
        for found in re.findall(r"https?://[^'\"),\s]+", vv):
            if found not in cleaned:
                cleaned.append(found)
        for found in re.findall(r"https?://[^'\"),\s]+", vv):
            if found not in cleaned:
                cleaned.append(found)
        for found in re.findall(r"https?://[^'\"),\s]+", vv):
            if found not in cleaned:
                cleaned.append(found)

    # 1) Already good article URLs.
    for v in cleaned:
        if is_preferred_article_url(v):
            return v

    # 2) Google RSS/articles are allowed only when resolved to real non-Google article URL.
    for v in cleaned:
        if is_google_news_rss_url(v):
            resolved = resolve_google_news_url(v)
            if is_preferred_article_url(resolved):
                return resolved
    return ""


def _all_existing_url_cols_v315(df: pd.DataFrame) -> list[str]:
    lookup = {str(c).strip().lower(): c for c in df.columns}
    cols = []
    for cand in URL_CANDIDATE_COLUMNS_V315:
        c = lookup.get(cand.lower())
        if c is not None and c not in cols:
            cols.append(c)
    # also scan any column name containing url/link/source, but avoid image/thumb columns first
    for c in df.columns:
        lc = str(c).strip().lower()
        if any(k in lc for k in ["url", "link", "source"]):
            if any(b in lc for b in ["image", "thumb", "logo", "photo", "picture"]):
                continue
            if c not in cols:
                cols.append(c)
    return cols


def normalize_input(df: pd.DataFrame, content_type: str, source_file: Path) -> pd.DataFrame:
    """V3.1.5: URL 후보 컬럼을 전부 스캔합니다.

    기존 normalize_input은 resolved_url/original_url/RepresentativeURL 중 첫 번째로 발견된 컬럼 하나만
    사용했습니다. 그 컬럼이 비어 있거나 썸네일이면 실제 original_url이 다른 컬럼에 있어도 놓칠 수 있어
    임원 메일 링크가 깨졌습니다.
    """
    df = df.copy()
    df.columns = [str(c).strip() for c in df.columns]

    col_date = pick_col(df, ["Date", "date", "Publish Date", "publish_date", "published", "Published"])
    col_headline = pick_col(df, ["Headline", "Title", "headline", "title"])
    col_country = pick_col(df, ["Country", "country"])
    col_agency = pick_col(df, ["Agency", "Publisher", "agency", "SourceName"])
    col_risk = pick_col(df, ["Risk", "risk"])
    col_score = pick_col(df, ["final_score", "FinalScore", "Score", "samsung_score", "Importance", "Importance Score"])
    col_priority = pick_col(df, ["priority_group", "Priority Group", "Priority", "Tier"])
    col_issue = pick_col(df, ["issue_type", "Issue", "IssueKey", "Topic", "topic"])
    col_cluster = pick_col(df, ["cluster_key", "Cluster", "cluster"])
    col_summary = pick_col(df, ["Summary", "summary"])
    col_analysis = pick_col(df, ["AI Analysis", "analysis", "Impact", "ai_analysis"])
    col_action = pick_col(df, ["Action Plan", "action", "Action", "required_action"])
    col_source = pick_col(df, ["Source", "SourceFile", "source"])
    url_cols = _all_existing_url_cols_v315(df)

    out = pd.DataFrame()
    out["Date"] = df[col_date].apply(display_date) if col_date else ""
    out["_sort_date"] = df[col_date].apply(parse_date_for_sort) if col_date else pd.Timestamp.min
    out["Headline"] = df[col_headline].apply(clean) if col_headline else ""

    def _row_best_url(src_row: pd.Series) -> str:
        vals = []
        # prefer original/resolved/representative columns by order list, then any scanned url-like columns
        for c in url_cols:
            if c in src_row.index:
                vals.append(src_row.get(c))
        return best_url_from_values(vals)

    out["URL"] = df.apply(_row_best_url, axis=1) if len(df) else ""
    out["Country"] = df[col_country].apply(clean) if col_country else ""
    out["Agency"] = df[col_agency].apply(clean) if col_agency else ""
    out["Risk"] = df[col_risk].apply(normalize_risk) if col_risk else "중"
    out["Importance Score"] = df[col_score].apply(safe_num) if col_score else 0
    out["Priority Group"] = df[col_priority].apply(lambda v: clean(v).upper()) if col_priority else "USABLE"
    out["Issue"] = df[col_issue].apply(clean) if col_issue else ""
    out["Cluster"] = df[col_cluster].apply(clean) if col_cluster else ""
    out["Summary"] = df[col_summary].apply(clean) if col_summary else ""
    out["AI Analysis"] = df[col_analysis].apply(clean) if col_analysis else ""
    out["Action Plan"] = df[col_action].apply(clean) if col_action else ""
    out["Source"] = df[col_source].apply(clean) if col_source else ""
    out["Source File"] = str(source_file)
    out["Content Type"] = content_type
    out = out[out["Headline"].astype(str).str.strip().ne("")]
    return out.reset_index(drop=True)


def _run_date_ts_v315() -> pd.Timestamp:
    dt = pd.to_datetime(RUN_DATE, errors="coerce")
    if pd.isna(dt):
        return pd.Timestamp(datetime.now().date())
    return pd.Timestamp(dt).normalize()


def is_old_by_24h_rule_v315(row: pd.Series) -> bool:
    dt = row.get("_sort_date", pd.Timestamp.min)
    try:
        dt = pd.to_datetime(dt, errors="coerce")
    except Exception:
        dt = pd.NaT
    if pd.isna(dt) or dt == pd.Timestamp.min:
        return False

    # Use an actual hour-based window. The previous calendar-date gate ignored
    # GTI_LOOKBACK_HOURS and excluded all late prior-day runs after midnight.
    try:
        anchor = pd.Timestamp(datetime.now())
    except Exception:
        anchor = _run_date_ts_v315() + pd.Timedelta(days=1)
    cutoff = anchor - pd.Timedelta(hours=LOOKBACK_HOURS_V315)
    return dt < cutoff


def read_step4_results() -> pd.DataFrame:
    frames = []
    if REGULATION_INPUT_FILE.exists():
        frames.append(normalize_input(pd.read_excel(REGULATION_INPUT_FILE), "Regulation", REGULATION_INPUT_FILE))
    if NEWS_INPUT_FILE.exists():
        news = normalize_input(pd.read_excel(NEWS_INPUT_FILE), "News", NEWS_INPUT_FILE)
        news = news.head(NEWS_MAX_ROWS)
        frames.append(news)
    if not frames:
        raise FileNotFoundError(f"STEP4 outputs not found: {REGULATION_INPUT_FILE}, {NEWS_INPUT_FILE}")
    rows = pd.concat(frames, ignore_index=True)
    rows = rows[rows["Headline"].astype(str).str.strip().ne("")].copy()
    rows["URL"] = rows.apply(lambda r: best_url_from_values([r.get("URL", ""), r.get("Source", "")]), axis=1)
    rows["_dedup_key"] = rows.apply(
        lambda r: clean(r.get("URL")) or (clean(r.get("Headline"))[:120] + "|" + clean(r.get("Agency")) + "|" + clean(r.get("Date"))),
        axis=1,
    )
    rows = rows.drop_duplicates(subset=["_dedup_key"], keep="first").drop(columns=["_dedup_key"], errors="ignore")
    rows["_integrated_score"] = rows.apply(
        lambda r: priority_weight(r["Priority Group"]) + risk_weight(r["Risk"]) + type_weight(r["Content Type"]) + safe_num(r["Importance Score"]),
        axis=1,
    )
    return rows.reset_index(drop=True)


_PREV_DETERMINE_V315 = determine_samsung_impact

def determine_samsung_impact(row: pd.Series) -> tuple[str, str, str]:
    if is_old_by_24h_rule_v315(row):
        return "None", "", OLD_DATE_REASON_V315
    return _PREV_DETERMINE_V315(row)


def html_link(title: str, url: str) -> str:
    title = html.escape(clean(title))
    url = best_url_from_values([url])
    if url:
        return f'<a href="{html.escape(url)}" target="_blank">{title}</a>'
    return title


def build_table(title: str, rows: pd.DataFrame, color: str) -> str:
    if rows.empty:
        return ""
    col_widths = ["2%", "7%", "4%", "7%", "15%", "15%", "15%", "15%", "5%", "5%", "3%", "7%"]
    headers = ["No", "Topic", "Impact", "Subsidiary", "Headline", "Summary", "Impact", "Action", "Country", "Agency", "Risk", "Publish Date"]
    colgroup = "<colgroup>" + "".join(f"<col style='width:{w};'>" for w in col_widths) + "</colgroup>"
    ths = "".join(f"<th style='padding:7px;border:1px solid #d9d9d9;text-align:center;'>{h}</th>" for h in headers)
    trs = []
    for _, row in rows.iterrows():
        trs.append(f"""
        <tr>
          <td style="padding:7px;border:1px solid #d9d9d9;text-align:center;vertical-align:top;">{html.escape(str(row['No']))}</td>
          <td style="padding:7px;border:1px solid #d9d9d9;text-align:center;vertical-align:top;">{html.escape(display_topic(row))}</td>
          <td style="padding:7px;border:1px solid #d9d9d9;text-align:center;vertical-align:top;font-weight:bold;color:{'#C00000' if clean(row.get('Samsung Impact')) == 'Direct' else '#666'};">{html.escape(clean(row.get('Samsung Impact')))}</td>
          <td style="padding:7px;border:1px solid #d9d9d9;text-align:center;vertical-align:top;">{html.escape(clean(row.get('Affected Subsidiary')) or '-')}</td>
          <td style="padding:7px;border:1px solid #d9d9d9;vertical-align:top;">{html_link(row['Headline'], row['URL'])}</td>
          <td style="padding:7px;border:1px solid #d9d9d9;vertical-align:top;">{html.escape(force_korean_text(row, 'Summary', '요약 정보 확인 필요'))}</td>
          <td style="padding:7px;border:1px solid #d9d9d9;vertical-align:top;">{html.escape(force_korean_text(row, 'AI Analysis', '영향 검토 필요'))}</td>
          <td style="padding:7px;border:1px solid #d9d9d9;vertical-align:top;">{html.escape(force_korean_text(row, 'Action Plan', '담당 부서 확인 필요'))}</td>
          <td style="padding:7px;border:1px solid #d9d9d9;text-align:center;vertical-align:top;">{html.escape(clean(row['Country']))}</td>
          <td style="padding:7px;border:1px solid #d9d9d9;text-align:center;vertical-align:top;">{html.escape(clean(row['Agency']))}</td>
          <td style="padding:7px;border:1px solid #d9d9d9;text-align:center;vertical-align:top;color:{risk_color(row['Risk'])};font-weight:bold;">{html.escape(clean(row['Risk']))}</td>
          <td style="padding:7px;border:1px solid #d9d9d9;text-align:center;vertical-align:top;">{html.escape(clean(row['Date']))}</td>
        </tr>
        """)
    return f"""
    <h3 style="margin-top:24px;color:{color};">{html.escape(title)} ({len(rows)}건)</h3>
    <table style="border-collapse:collapse;width:100%;font-size:12px;table-layout:fixed;">
      {colgroup}
      <thead><tr style="background:{color};color:white;">{ths}</tr></thead>
      <tbody>{''.join(trs)}</tbody>
    </table>
    """


def append_row(ws, row: pd.Series) -> None:
    ws.append([row.get(c, "") for c in OUTPUT_COLUMNS])
    headline_cell = ws.cell(row=ws.max_row, column=OUTPUT_COLUMNS.index("Headline") + 1)
    url = best_url_from_values([row.get("URL")])
    if url:
        headline_cell.hyperlink = url
        headline_cell.font = Font(color="0563C1", underline="single", bold=True)


def save_excel(rows: pd.DataFrame, top3: pd.DataFrame, paths: dict[str, Path]) -> None:
    wb = Workbook()
    visible = rows[~rows["Mail Group"].eq("Filtered Noise")].copy()
    filtered_noise = rows[rows["Mail Group"].eq("Filtered Noise")].copy()
    sheets = [
        ("GTI Radar", visible),
        ("Top3", top3),
        ("Required Actions", pd.DataFrame(build_required_actions(visible, top3))),
        ("Regulation", visible[visible["Mail Group"].eq("Regulation")]),
        ("News CORE", visible[visible["Mail Group"].eq("News - 핵심")]),
        ("News USABLE", visible[visible["Mail Group"].eq("News - 주요/참고")]),
        ("Filtered Noise", filtered_noise),
    ]
    first = True
    for name, frame in sheets:
        ws = wb.active if first else wb.create_sheet(name[:31])
        first = False
        ws.title = name[:31]
        if name == "Required Actions":
            ws.append(["Issue", "Required Action", "Owner"])
            for _, rr in frame.iterrows():
                ws.append([rr.get("topic", ""), rr.get("action", ""), rr.get("owner", "")])
            ws.column_dimensions["A"].width = 16
            ws.column_dimensions["B"].width = 90
            ws.column_dimensions["C"].width = 18
        else:
            ws.append(OUTPUT_COLUMNS)
            for _, row in frame.iterrows():
                append_row(ws, row)
            # Executive form width aligned to requested ratio.
            widths = {
                "A": 5,   # No 2%
                "B": 13,  # Content Type
                "C": 17,  # Mail Group
                "D": 12,  # Samsung Impact 4%
                "E": 15,  # Subsidiary 7%
                "F": 30,  # Impact Reason
                "G": 13,  # Date
                "H": 42,  # Headline 15%
                "I": 42,  # Summary 15%
                "J": 42,  # Impact 15%
                "K": 42,  # Action 15%
                "L": 14,  # Country 5%
                "M": 14,  # Agency 5%
                "N": 8,   # Risk 3%
                "O": 13,  # Score
                "P": 14,  # Priority
                "Q": 16,  # Issue
                "R": 20,  # Cluster
                "S": 30,  # URL
                "T": 24,  # Source
                "U": 28,  # Source File
            }
            for col, width in widths.items():
                ws.column_dimensions[col].width = width
        style_sheet(ws)

    runlog = wb.create_sheet("Run Log")
    runlog.append(["item", "value"])
    runlog.append(["regulation_input", str(REGULATION_INPUT_FILE)])
    runlog.append(["news_input", str(NEWS_INPUT_FILE)])
    runlog.append(["run_date", RUN_DATE])
    runlog.append(["lookback_hours", LOOKBACK_HOURS_V315])
    runlog.append(["visible_total", len(visible)])
    runlog.append(["filtered_noise", len(filtered_noise)])
    runlog.append(["regulation_rows", int(visible["Mail Group"].eq("Regulation").sum())])
    runlog.append(["news_core_rows", int(visible["Mail Group"].eq("News - 핵심").sum())])
    runlog.append(["news_usable_rows", int(visible["Mail Group"].eq("News - 주요/참고").sum())])
    runlog.append(["direct_rows", int(visible["Samsung Impact"].eq("Direct").sum())])
    runlog.append(["indirect_rows", int(visible["Samsung Impact"].eq("Indirect").sum())])
    style_sheet(runlog)

    wb.save(paths["mail_xlsx"])
    wb.save(paths["analysis"])
    visible[OUTPUT_COLUMNS].to_excel(paths["cumulative"], index=False)


# =========================================================
# GTI executive-quality final overrides
# - Final defense layer for executive mail quality.
# - Keeps true tariff/customs policy items, including steel tariff/quota news.
# - Suppresses exhibition, FX, fertilizer, generic market trend and weak supply-chain items.
# - Regenerates Summary / Impact / Action from the final Samsung Impact decision.
# =========================================================

GROUP_REGULATION = "Regulation"
GROUP_CORE = "News - 핵심"
GROUP_USABLE = "News - 주요/참고"
GROUP_NOISE = "Filtered Noise"

EXEC_ACTIONABLE_TERMS = [
    "관세", "관세율", "추가관세", "상호관세", "무관세", "철강관세", "쿼터", "세이프가드",
    "반덤핑", "덤핑방지", "상계관세", "ad/cvd", "anti-dumping", "countervailing",
    "section 301", "section 232", "ustr", "cbp", "usitc", "federal register",
    "fta", "cepa", "epa", "원산지", "협정세율", "certificate of origin", "rules of origin",
    "hs code", "품목분류", "수출통제", "전략물자", "entity list", "bis", "ear", "eccn",
    "uflpa", "forced labor", "cbam", "carbon border",
]

EXEC_HARD_POLICY_TERMS = [
    "관세율", "추가관세", "상호관세", "철강관세", "무관세 물량", "쿼터", "세이프가드",
    "반덤핑", "상계관세", "section 301", "section 232", "entity list", "uflpa",
    "cbam", "hs code", "품목분류", "원산지 기준", "협정세율",
]

EXEC_LOW_VALUE_TERMS = [
    "전시회", "참관기", "시장동향", "상품db", "유망바이어", "환리스크", "고환율",
    "공공비축", "비료", "원전 사업", "일반 공급망", "문화", "관광", "스포츠", "게임", "증시",
    "今日の歴史", "오늘의 역사", "역사", "anniversary", "on this day",
]

EXEC_PRODUCT_TERMS = [
    "samsung", "삼성", "semiconductor", "semiconductors", "chip", "chips", "반도체",
    "hbm", "dram", "nand", "display", "디스플레이", "battery", "배터리", "cell",
    "electronics", "전자", "전자부품", "smartphone", "스마트폰", "가전",
]


def _exec_text(row: pd.Series) -> str:
    return " ".join([
        clean(row.get("Headline")),
        clean(row.get("Summary")),
        clean(row.get("AI Analysis")),
        clean(row.get("Action Plan")),
        clean(row.get("Issue")),
        clean(row.get("Cluster")),
        clean(row.get("Country")),
        clean(row.get("Agency")),
        clean(row.get("Source")),
    ])


def _exec_text_l(row: pd.Series) -> str:
    return _exec_text(row).lower()


def _exec_contains(row: pd.Series, terms: list[str]) -> bool:
    text = _exec_text_l(row)
    return any(str(term).lower() in text for term in terms if str(term).strip())


def _exec_issue(row: pd.Series) -> str:
    text = _exec_text_l(row)
    raw = " ".join([clean(row.get("Issue")), clean(row.get("Priority Group")), clean(row.get("Cluster"))]).lower()
    base = f"{raw} {text}"
    if any(k in base for k in ["cbam", "carbon border", "탄소국경"]):
        return "CBAM"
    if any(k in base for k in ["entity list", "export control", "수출통제", "전략물자", "uflpa", "forced labor", "bis", "eccn"]):
        return "수출통제"
    if any(k in base for k in ["ad/cvd", "anti-dumping", "antidumping", "countervailing", "반덤핑", "상계관세", "덤핑방지"]):
        return "AD/CVD"
    if any(k in base for k in ["section 301", "section 232", "상호관세", "추가관세", "철강관세", "세이프가드", "tariff", "관세"]):
        return "관세정책"
    if any(k in base for k in ["fta", "cepa", "epa", "원산지", "협정세율", "rules of origin", "certificate of origin"]):
        return "원산지/FTA"
    if any(k in base for k in ["hs code", "품목분류", "tariff classification"]):
        return "HS분류"
    if clean(row.get("Content Type")).lower().startswith("reg"):
        return "법규"
    return "참고"


def _exec_family(row: pd.Series) -> str:
    topic = _exec_issue(row)
    if topic in {"관세정책", "AD/CVD"}:
        return "TARIFF"
    if topic == "수출통제":
        return "EXPORT_CONTROL"
    if topic == "원산지/FTA":
        return "FTA"
    if topic == "CBAM":
        return "CBAM"
    if topic == "HS분류":
        return "HS"
    if topic == "법규":
        return "REG"
    return "REF"


def _exec_low_value(row: pd.Series) -> bool:
    if not _exec_contains(row, EXEC_LOW_VALUE_TERMS):
        return False
    # Steel tariff, CBAM, AD/CVD, UFLPA, Section 301/232, origin rules, and HS
    # must survive even if the article also looks like a market or event item.
    return not _exec_contains(row, EXEC_HARD_POLICY_TERMS)


def _exec_has_actionable_policy(row: pd.Series) -> bool:
    return _exec_contains(row, EXEC_ACTIONABLE_TERMS)


def _exec_has_product_signal(row: pd.Series) -> bool:
    return _exec_contains(row, EXEC_PRODUCT_TERMS)


def _exec_score(row: pd.Series) -> float:
    score = safe_num(row.get("Importance Score"))
    family = _exec_family(row)
    impact = clean(row.get("Samsung Impact"))
    headline = clean(row.get("Headline")).lower()
    if family in {"TARIFF", "EXPORT_CONTROL", "CBAM", "FTA", "HS", "REG"}:
        score += 120
    if _exec_contains(row, EXEC_HARD_POLICY_TERMS):
        score += 140
    if any(k in headline for k in ["철강관세", "무관세 물량", "상호관세", "추가관세", "반덤핑", "상계관세", "section 301", "section 232", "cbam", "entity list"]):
        score += 220
    if _exec_has_product_signal(row):
        score += 70
    if impact == "Direct":
        score += 80
    elif impact == "Indirect":
        score += 35
    if _exec_low_value(row):
        score -= 300
    return score


def determine_samsung_impact(row: pd.Series) -> tuple[str, str, str]:
    if "is_old_by_24h_rule_v315" in globals() and is_old_by_24h_rule_v315(row):
        return "None", "", "24시간 기준 초과 게시건으로 금일 임원 보고 본문 제외"

    if not _exec_has_actionable_policy(row):
        return "None", "", "관세·통상 실행 항목이 명확하지 않아 메일 본문 제외"
    if _exec_low_value(row):
        return "None", "", "전시회·시장동향·환율 등 일반 참고성 기사로 관세정책 보고 본문 제외"

    topic = _exec_issue(row)
    country = clean(row.get("Country"))
    subs = clean(row.get("Affected Subsidiary"))
    if not subs or subs == "-":
        # Keep the fallback conservative. Do not auto-attach many subsidiaries
        # unless the source already gave a reliable mapping.
        subs = "SEC/HQ"

    product = _exec_has_product_signal(row)
    hard_policy = _exec_contains(row, EXEC_HARD_POLICY_TERMS)

    if clean(row.get("Content Type")).lower().startswith("reg"):
        return "Direct", "SEC/HQ", "공식 법규·고시로 HQ 관세 Master 및 신고 프로세스 영향 검토 필요"
    if topic in {"수출통제", "CBAM", "AD/CVD", "HS분류"} and (product or hard_policy):
        return "Direct", subs, f"{topic} 실행 이슈로 대상 품목·법인·증빙 영향 검토 필요"
    if topic == "관세정책" and (product or hard_policy):
        # Steel tariff/quota articles are usually not Samsung-direct unless a
        # Samsung product or import/export lane is identified, but they are
        # important enough to keep as Indirect.
        if product and hard_policy:
            return "Direct", subs, "삼성 품목과 연결된 관세정책 이슈로 HS·관세율·원산지 영향 검토 필요"
        return "Indirect", "SEC/HQ", "관세율·쿼터·무관세 물량 등 정책 변화로 모니터링 및 시나리오 검토 필요"
    if topic == "원산지/FTA":
        if product and hard_policy:
            return "Direct", subs, "삼성 품목 관련 FTA/원산지 실행 이슈로 CO·BOM·협정세율 점검 필요"
        return "Indirect", "SEC/HQ", "FTA/원산지 정책 모니터링 및 적용 가능성 검토 대상"

    return "Indirect", "SEC/HQ", "관세·통상 정책 변화로 참고 모니터링 필요"


def apply_samsung_impact(rows: pd.DataFrame) -> pd.DataFrame:
    rows = rows.copy()
    impacts = rows.apply(determine_samsung_impact, axis=1)
    rows["Samsung Impact"] = [x[0] for x in impacts]
    rows["Affected Subsidiary"] = [x[1] for x in impacts]
    rows["Impact Reason"] = [x[2] for x in impacts]
    rows["_display_cluster"] = rows.apply(_exec_cluster_key, axis=1)
    rows["Issue"] = rows.apply(_exec_issue, axis=1)
    return rows


def assign_mail_groups(rows: pd.DataFrame) -> pd.DataFrame:
    rows = apply_samsung_impact(rows.copy())
    rows["Mail Group"] = GROUP_USABLE
    rows.loc[rows["Content Type"].eq("Regulation"), "Mail Group"] = GROUP_REGULATION
    rows.loc[rows["Samsung Impact"].eq("Direct"), "Mail Group"] = GROUP_CORE
    rows.loc[rows["Samsung Impact"].eq("None"), "Mail Group"] = GROUP_NOISE
    return rows


def executive_sort_frame(rows: pd.DataFrame) -> pd.DataFrame:
    rows = rows.copy()
    rows["_exec_score"] = rows.apply(_exec_score, axis=1)
    return rows.sort_values(["_exec_score", "_sort_date"], ascending=[False, False])


def choose_top3(rows: pd.DataFrame) -> pd.DataFrame:
    pool = rows[~rows["Mail Group"].eq(GROUP_NOISE)].copy()
    if pool.empty:
        return pool
    pool = executive_sort_frame(pool).drop_duplicates(subset=["_display_cluster"], keep="first")
    selected, used_family = [], set()
    for _, row in pool.iterrows():
        fam = _exec_family(row)
        if fam in used_family and len(selected) < 3:
            continue
        selected.append(row)
        used_family.add(fam)
        if len(selected) >= 3:
            break
    if len(selected) < 3:
        for _, row in pool.iterrows():
            if any(clean(row.get("Headline")) == clean(x.get("Headline")) for x in selected):
                continue
            selected.append(row)
            if len(selected) >= 3:
                break
    return pd.DataFrame(selected).reset_index(drop=True)


def final_order(rows: pd.DataFrame, top3: pd.DataFrame) -> pd.DataFrame:
    rows = assign_mail_groups(rows.copy())
    rows = executive_sort_frame(rows)

    noise = rows[rows["Mail Group"].eq(GROUP_NOISE)].copy()
    visible_pool = rows[~rows["Mail Group"].eq(GROUP_NOISE)].copy()
    visible_pool = visible_pool.drop_duplicates(subset=["_display_cluster"], keep="first")

    regs = visible_pool[visible_pool["Mail Group"].eq(GROUP_REGULATION)].head(2)
    direct = visible_pool[(visible_pool["Mail Group"].eq(GROUP_CORE)) & (~visible_pool.index.isin(regs.index))].head(4)
    indirect = visible_pool[
        (visible_pool["Mail Group"].eq(GROUP_USABLE))
        & (~visible_pool.index.isin(regs.index))
        & (~visible_pool.index.isin(direct.index))
    ].head(6)

    visible = pd.concat([regs, direct, indirect], ignore_index=True)
    visible = executive_sort_frame(visible).head(int(os.getenv("GTI_VISIBLE_MAX", "10")))
    out = pd.concat([visible, noise.head(20)], ignore_index=True)
    out["No"] = range(1, len(out) + 1)
    return out.reset_index(drop=True)


def force_korean_text(row: pd.Series, field: str, fallback: str) -> str:
    topic = _exec_issue(row)
    impact = clean(row.get("Samsung Impact")) or "Indirect"
    country = clean(row.get("Country")) or "Global"
    subs = clean(row.get("Affected Subsidiary")) or "SEC/HQ"
    headline = clean(row.get("Headline"))

    if field == "Summary":
        return f"{headline} 관련 {topic} 이슈입니다. 최종 삼성 영향은 {impact}이며, 대상 국가/법인({country}, {subs}) 기준으로 실행 영향 확인이 필요합니다."
    if field == "AI Analysis":
        if topic == "관세정책":
            return "대상 HS, 공급국, 적용 시점, 관세율·쿼터·무관세 물량 여부 및 삼성 수입/수출 경로와의 연결성을 확인해야 합니다."
        if topic == "AD/CVD":
            return "반덤핑·상계관세 조사 또는 부과 이슈로 공급국, 제조사, 가격자료, 원산지 증빙 및 소급 리스크를 확인해야 합니다."
        if topic == "수출통제":
            return "수출통제 이슈로 ECCN/전략물자 해당 여부, 거래상대방·최종사용자 스크리닝 및 우회거래 가능성을 점검해야 합니다."
        if topic == "CBAM":
            return "CBAM 이슈로 EU향 품목, 원재료, 공급사 탄소자료 및 신고 증빙 체계를 확인해야 합니다."
        if topic == "원산지/FTA":
            return "FTA/원산지 이슈로 CO 발급·수취, BOM 충족, 협정세율 적용 가능성 및 증빙 보관 기준을 점검해야 합니다."
        if topic == "HS분류":
            return "HS 품목분류 이슈로 국가별 HS 매핑, 품목 설명, 사전심사 필요 여부를 확인해야 합니다."
        return fallback
    if field == "Action Plan":
        if topic == "관세정책":
            return "대상 HS·공급국 매핑; 관세율/쿼터 적용 여부 확인; 관련 수입·수출 실적 추출; 가격·원가 영향 시나리오 검토"
        if topic == "AD/CVD":
            return "대상 공급망 확인; 제조사·원산지·가격자료 점검; 조사대상 여부 확인; 법무/구매/물류 합동 대응"
        if topic == "수출통제":
            return "ECCN/전략물자 재점검; 거래상대방·최종사용자 스크리닝; 우회수출 및 해외자회사 거래 통제 확인"
        if topic == "CBAM":
            return "EU향 대상 품목 확인; 공급사 탄소자료 확보; CBAM 신고/증빙 체계 점검; ESG/구매 협업"
        if topic == "원산지/FTA":
            return "원산지 기준 검토; BOM 충족 여부 확인; CO 발급·수취·보관 증빙 점검; 협정세율 적용 대상 재점검"
        if topic == "HS분류":
            return "품목 설명서 확보; HS Master 비교; 국가별 분류 차이 확인; 사전심사 필요 여부 판단"
        return fallback
    return clean(row.get(field)) or fallback


def build_required_actions(rows: pd.DataFrame, top3: pd.DataFrame) -> list[dict]:
    pool = pd.concat([top3, rows[rows["Samsung Impact"].eq("Direct")]], ignore_index=True)
    if pool.empty:
        return []
    pool = executive_sort_frame(pool).drop_duplicates(subset=["Issue"], keep="first").head(5)
    owner_map = {
        "관세정책": "통관운영/FTA팀",
        "AD/CVD": "Global SCM/법무",
        "수출통제": "수출통제팀",
        "CBAM": "ESG/구매",
        "원산지/FTA": "FTA팀",
        "HS분류": "통관운영팀",
        "법규": "통관운영팀",
    }
    return [
        {
            "topic": clean(r.get("Issue")),
            "action": force_korean_text(r, "Action Plan", "관련 부서 확인 필요"),
            "owner": owner_map.get(clean(r.get("Issue")), "관세/통상 담당"),
        }
        for _, r in pool.iterrows()
    ]


def _exec_text(row: pd.Series) -> str:
    """Use source-facing fields only.

    STEP4 may already contain generated Action/AI text. Reading that generated
    text again caused unrelated FTA/market articles to be reclassified as export
    control. This final gate intentionally ignores AI Analysis and Action Plan.
    """
    return " ".join([
        clean(row.get("Headline")),
        clean(row.get("Summary")),
        clean(row.get("Issue")),
        clean(row.get("Cluster")),
        clean(row.get("Country")),
        clean(row.get("Agency")),
        clean(row.get("Source")),
    ])


def _exec_issue(row: pd.Series) -> str:
    headline = clean(row.get("Headline")).lower()
    raw_issue = clean(row.get("Issue")).upper()
    text = _exec_text(row).lower()

    # Headline-level concrete policy signals override prior STEP4 labels.
    if any(k in headline for k in ["철강관세", "무관세 물량", "상호관세", "추가관세", "세이프가드", "section 301", "section 232"]):
        return "관세정책"
    if any(k in headline for k in ["반덤핑", "상계관세", "덤핑방지", "ad/cvd", "anti-dumping", "countervailing"]):
        return "AD/CVD"

    raw_map = {
        "EXPORT_CONTROL": "수출통제",
        "CBAM_CARBON": "CBAM",
        "AD_CVD": "AD/CVD",
        "ORIGIN_FTA": "원산지/FTA",
        "HS_CLASSIFICATION": "HS분류",
        "TARIFF": "관세정책",
        "SECTION_301_232": "관세정책",
        "TRADE_REGULATION": "법규",
    }
    if raw_issue in raw_map:
        return raw_map[raw_issue]

    if any(k in text for k in ["cbam", "carbon border", "탄소국경"]):
        return "CBAM"
    if any(k in text for k in ["entity list", "export control", "수출통제", "전략물자", "uflpa", "forced labor", "bis", "eccn"]):
        return "수출통제"
    if any(k in text for k in ["ad/cvd", "anti-dumping", "antidumping", "countervailing", "반덤핑", "상계관세", "덤핑방지"]):
        return "AD/CVD"
    if any(k in text for k in ["section 301", "section 232", "상호관세", "추가관세", "철강관세", "세이프가드", "tariff", "관세"]):
        return "관세정책"
    if any(k in text for k in ["fta", "cepa", "epa", "원산지", "협정세율", "rules of origin", "certificate of origin"]):
        return "원산지/FTA"
    if any(k in text for k in ["hs code", "품목분류", "tariff classification"]):
        return "HS분류"
    if clean(row.get("Content Type")).lower().startswith("reg"):
        return "법규"
    return "참고"


def _exec_low_value(row: pd.Series) -> bool:
    headline = clean(row.get("Headline")).lower()
    text = _exec_text(row).lower()
    if not any(k.lower() in text for k in EXEC_LOW_VALUE_TERMS):
        return False
    # Market/exhibition/FX articles survive only when the headline itself has
    # a hard customs-policy signal. Body-only HS or generic FTA mentions are
    # usually reference material, not executive-report material.
    headline_keep = [
        "관세", "무관세", "반덤핑", "상계관세", "세이프가드", "section 301",
        "section 232", "cbam", "entity list", "uflpa", "수출통제",
    ]
    return not any(k in headline for k in headline_keep)


def _exec_cluster_key(row: pd.Series) -> str:
    text = _exec_text(row).lower()
    family = _exec_family(row)
    if family == "TARIFF" and any(k in text for k in ["pom", "코오롱", "반덤핑", "상계관세", "ad/cvd"]):
        return "AD_CVD_CHINA_KOLON_POM"
    if family == "TARIFF" and any(k in text for k in ["철강관세", "무관세 물량", "steel tariff", "quota"]):
        return "EU_STEEL_TARIFF_QUOTA"
    if family == "EXPORT_CONTROL" and any(k in text for k in ["g2", "미중", "china", "중국"]):
        return "US_CHINA_EXPORT_CONTROL_TALKS"
    if family == "FTA" and any(k in text for k in ["광양만권", "한 중 경제협력", "china", "중국"]):
        return "KOREA_CHINA_FTA_REFERENCE"
    return f"{family}:{clean(row.get('Headline'))[:80]}"


# =========================================================
# GTI ASCII-safe final mail overrides
# Keep immediately before __main__. Korean terms are generated from Unicode
# escapes so keyword matching survives copy/paste into C:\Temp.
# =========================================================

def _u2(s: str) -> str:
    return s.encode("ascii").decode("unicode_escape")


ISS_TARIFF = _u2("\\uad00\\uc138\\uc815\\ucc45")
ISS_EXPORT = _u2("\\uc218\\ucd9c\\ud1b5\\uc81c")
ISS_FTA = _u2("\\uc6d0\\uc0b0\\uc9c0/FTA")
ISS_HS = _u2("HS\\ubd84\\ub958")
ISS_REG = _u2("\\ubc95\\uaddc")
GROUP_REGULATION = "Regulation"
GROUP_CORE = _u2("News - \\ud575\\uc2ec")
GROUP_USABLE = _u2("News - \\uc8fc\\uc694/\\ucc38\\uace0")
GROUP_NOISE = "Filtered Noise"

KW_TARIFF = _u2("\\uad00\\uc138")
KW_TARIFF_RATE = _u2("\\uad00\\uc138\\uc728")
KW_EXTRA_TARIFF = _u2("\\ucd94\\uac00\\uad00\\uc138")
KW_RECIPROCAL_TARIFF = _u2("\\uc0c1\\ud638\\uad00\\uc138")
KW_ZERO_TARIFF = _u2("\\ubb34\\uad00\\uc138")
KW_STEEL_TARIFF = _u2("\\ucca0\\uac15\\uad00\\uc138")
KW_QUOTA = _u2("\\ucffc\\ud130")
KW_AD = _u2("\\ubc18\\ub364\\ud551")
KW_CVD = _u2("\\uc0c1\\uacc4\\uad00\\uc138")
KW_ORIGIN = _u2("\\uc6d0\\uc0b0\\uc9c0")
KW_PREF = _u2("\\ud611\\uc815\\uc138\\uc728")
KW_HS = _u2("\\ud488\\ubaa9\\ubd84\\ub958")
KW_EXPORT = _u2("\\uc218\\ucd9c\\ud1b5\\uc81c")
KW_STRATEGIC = _u2("\\uc804\\ub7b5\\ubb3c\\uc790")
KW_EXHIBITION = _u2("\\uc804\\uc2dc\\ud68c")
KW_VISIT = _u2("\\ucc38\\uad00\\uae30")
KW_MARKET = _u2("\\uc2dc\\uc7a5\\ub3d9\\ud5a5")
KW_PRODUCT_DB = _u2("\\uc0c1\\ud488DB")
KW_FX = _u2("\\ud658\\ub9ac\\uc2a4\\ud06c")
KW_HIGH_FX = _u2("\\uace0\\ud658\\uc728")
KW_FERTILIZER = _u2("\\ube44\\ub8cc")
KW_HISTORY_JP = _u2("\\u4eca\\u65e5\\u306e\\u6b74\\u53f2")
KW_TODAY_HISTORY = _u2("\\uc624\\ub298\\uc758 \\uc5ed\\uc0ac")
KW_SAMSUNG = _u2("\\uc0bc\\uc131")
KW_SEMI = _u2("\\ubc18\\ub3c4\\uccb4")
KW_BATTERY = _u2("\\ubc30\\ud130\\ub9ac")
KW_DISPLAY = _u2("\\ub514\\uc2a4\\ud50c\\ub808\\uc774")
KW_ELECTRONICS = _u2("\\uc804\\uc790")

SAFE_ACTIONABLE = [
    KW_TARIFF, KW_TARIFF_RATE, KW_EXTRA_TARIFF, KW_RECIPROCAL_TARIFF,
    KW_ZERO_TARIFF, KW_STEEL_TARIFF, KW_QUOTA, KW_AD, KW_CVD,
    KW_ORIGIN, KW_PREF, KW_HS, KW_EXPORT, KW_STRATEGIC,
    "tariff", "tariffs", "duty", "duties", "quota", "safeguard",
    "ad/cvd", "anti-dumping", "antidumping", "countervailing",
    "section 301", "section 232", "ustr", "cbp", "usitc",
    "fta", "cepa", "epa", "rules of origin", "certificate of origin",
    "hs code", "tariff classification", "export control", "entity list",
    "bis", "ear", "eccn", "uflpa", "forced labor", "cbam", "carbon border",
]

SAFE_HARD = [
    KW_TARIFF_RATE, KW_EXTRA_TARIFF, KW_RECIPROCAL_TARIFF, KW_STEEL_TARIFF,
    KW_ZERO_TARIFF, KW_QUOTA, KW_AD, KW_CVD, KW_HS, KW_ORIGIN, KW_PREF,
    KW_EXPORT, "section 301", "section 232", "entity list", "uflpa", "cbam",
]

SAFE_LOW = [
    KW_EXHIBITION, KW_VISIT, KW_MARKET, KW_PRODUCT_DB, KW_FX, KW_HIGH_FX,
    KW_FERTILIZER, KW_HISTORY_JP, KW_TODAY_HISTORY,
    "market trend", "exhibition", "trade fair", "buyer", "sports", "tennis",
    "game", "stablecoin", "on this day", "anniversary",
]

SAFE_PRODUCTS = [
    "samsung", KW_SAMSUNG, "semiconductor", "semiconductors", "chip", "chips",
    KW_SEMI, "hbm", "dram", "nand", "display", KW_DISPLAY, "battery",
    KW_BATTERY, "electronics", KW_ELECTRONICS, "smartphone",
]


def _safe_text(row: pd.Series) -> str:
    return " ".join([
        clean(row.get("Headline")),
        clean(row.get("Summary")),
        clean(row.get("Issue")),
        clean(row.get("Cluster")),
        clean(row.get("Country")),
        clean(row.get("Agency")),
        clean(row.get("Source")),
    ])


def _safe_contains(row: pd.Series, terms: list[str]) -> bool:
    text = _safe_text(row).lower()
    return any(str(t).lower() in text for t in terms if str(t).strip())


def _safe_headline_contains(row: pd.Series, terms: list[str]) -> bool:
    text = clean(row.get("Headline")).lower()
    return any(str(t).lower() in text for t in terms if str(t).strip())


def _exec_issue(row: pd.Series) -> str:
    raw_issue = clean(row.get("Issue")).upper()

    if _safe_headline_contains(row, [KW_STEEL_TARIFF, KW_ZERO_TARIFF, KW_EXTRA_TARIFF, KW_RECIPROCAL_TARIFF, "section 301", "section 232", "tariff cap"]):
        return ISS_TARIFF
    if _safe_headline_contains(row, [KW_AD, KW_CVD, "ad/cvd", "anti-dumping", "countervailing"]):
        return "AD/CVD"

    raw_map = {
        "EXPORT_CONTROL": ISS_EXPORT,
        "CBAM_CARBON": "CBAM",
        "AD_CVD": "AD/CVD",
        "ORIGIN_FTA": ISS_FTA,
        "HS_CLASSIFICATION": ISS_HS,
        "TARIFF": ISS_TARIFF,
        "SECTION_301_232": ISS_TARIFF,
        "TRADE_REGULATION": ISS_REG,
    }
    if raw_issue in raw_map:
        return raw_map[raw_issue]

    if _safe_contains(row, ["cbam", "carbon border"]):
        return "CBAM"
    if _safe_contains(row, ["entity list", "export control", KW_EXPORT, KW_STRATEGIC, "uflpa", "forced labor", "bis", "eccn"]):
        return ISS_EXPORT
    if _safe_contains(row, ["ad/cvd", "anti-dumping", "antidumping", "countervailing", KW_AD, KW_CVD]):
        return "AD/CVD"
    if _safe_contains(row, ["section 301", "section 232", KW_RECIPROCAL_TARIFF, KW_EXTRA_TARIFF, KW_STEEL_TARIFF, KW_QUOTA, "tariff", KW_TARIFF]):
        return ISS_TARIFF
    if _safe_contains(row, ["fta", "cepa", "epa", KW_ORIGIN, KW_PREF, "rules of origin", "certificate of origin"]):
        return ISS_FTA
    if _safe_contains(row, ["hs code", KW_HS, "tariff classification"]):
        return ISS_HS
    if clean(row.get("Content Type")).lower().startswith("reg"):
        return ISS_REG
    return _u2("\\ucc38\\uace0")


def _exec_family(row: pd.Series) -> str:
    topic = _exec_issue(row)
    if topic in {ISS_TARIFF, "AD/CVD"}:
        return "TARIFF"
    if topic == ISS_EXPORT:
        return "EXPORT_CONTROL"
    if topic == ISS_FTA:
        return "FTA"
    if topic == "CBAM":
        return "CBAM"
    if topic == ISS_HS:
        return "HS"
    if topic == ISS_REG:
        return "REG"
    return "REF"


def _exec_low_value(row: pd.Series) -> bool:
    if not _safe_contains(row, SAFE_LOW):
        return False
    return not _safe_headline_contains(row, SAFE_HARD + ["section 301", "section 232", "cbam", "entity list"])


def _exec_has_actionable_policy(row: pd.Series) -> bool:
    return _safe_contains(row, SAFE_ACTIONABLE)


def _exec_has_product_signal(row: pd.Series) -> bool:
    return _safe_contains(row, SAFE_PRODUCTS)


def _exec_cluster_key(row: pd.Series) -> str:
    text = _safe_text(row).lower()
    family = _exec_family(row)
    if family == "TARIFF" and any(k in text for k in ["pom", "kolon", "ad/cvd", "anti-dumping", KW_AD.lower(), KW_CVD.lower()]):
        return "AD_CVD_CHINA_KOLON_POM"
    if family == "TARIFF" and any(k in text for k in [KW_STEEL_TARIFF.lower(), KW_ZERO_TARIFF.lower(), "steel tariff", "quota"]):
        return "EU_STEEL_TARIFF_QUOTA"
    if family == "EXPORT_CONTROL" and any(k in text for k in ["g2", "china"]):
        return "US_CHINA_EXPORT_CONTROL_TALKS"
    return f"{family}:{clean(row.get('Headline'))[:80]}"


def _exec_score(row: pd.Series) -> float:
    score = safe_num(row.get("Importance Score"))
    family = _exec_family(row)
    if family in {"TARIFF", "EXPORT_CONTROL", "CBAM", "FTA", "HS", "REG"}:
        score += 120
    if _safe_contains(row, SAFE_HARD):
        score += 140
    if _safe_headline_contains(row, SAFE_HARD + ["section 301", "section 232", "tariff cap"]):
        score += 220
    if _exec_has_product_signal(row):
        score += 70
    impact = clean(row.get("Samsung Impact"))
    if impact == "Direct":
        score += 80
    elif impact == "Indirect":
        score += 35
    if _exec_low_value(row):
        score -= 300
    return score


def determine_samsung_impact(row: pd.Series) -> tuple[str, str, str]:
    if "is_old_by_24h_rule_v315" in globals() and is_old_by_24h_rule_v315(row):
        return "None", "", "Older than the configured 24-hour window; excluded from today's executive mail."
    if not _exec_has_actionable_policy(row):
        return "None", "", "No clear customs/trade action item; excluded from mail body."
    if _exec_low_value(row):
        return "None", "", "General market/event/FX/reference item; excluded from executive customs-policy body."

    topic = _exec_issue(row)
    subs = clean(row.get("Affected Subsidiary")) or "SEC/HQ"
    product = _exec_has_product_signal(row)
    hard = _safe_contains(row, SAFE_HARD)

    if clean(row.get("Content Type")).lower().startswith("reg"):
        return "Direct", "SEC/HQ", "Official regulation/notice; HQ customs master and declaration process review required."
    if topic in {ISS_EXPORT, "CBAM", "AD/CVD", ISS_HS} and (product or hard):
        return "Direct", subs, f"{topic}: product/entity/evidence impact review required."
    if topic == ISS_TARIFF and (product or hard):
        if product and hard:
            return "Direct", subs, "Tariff policy linked to Samsung products; HS/rate/origin impact review required."
        return "Indirect", "SEC/HQ", "Tariff/rate/quota policy change; monitor and run scenario review."
    if topic == ISS_FTA:
        if product and hard:
            return "Direct", subs, "FTA/origin issue linked to Samsung products; CO/BOM/preferential-rate review required."
        return "Indirect", "SEC/HQ", "FTA/origin policy monitoring and applicability review required."
    return "Indirect", "SEC/HQ", "Customs/trade policy monitoring required."


def apply_samsung_impact(rows: pd.DataFrame) -> pd.DataFrame:
    rows = rows.copy()
    impacts = rows.apply(determine_samsung_impact, axis=1)
    rows["Samsung Impact"] = [x[0] for x in impacts]
    rows["Affected Subsidiary"] = [x[1] for x in impacts]
    rows["Impact Reason"] = [x[2] for x in impacts]
    rows["Issue"] = rows.apply(_exec_issue, axis=1)
    rows["_display_cluster"] = rows.apply(_exec_cluster_key, axis=1)
    return rows


def assign_mail_groups(rows: pd.DataFrame) -> pd.DataFrame:
    rows = apply_samsung_impact(rows.copy())
    rows["Mail Group"] = GROUP_USABLE
    rows.loc[rows["Content Type"].eq("Regulation"), "Mail Group"] = GROUP_REGULATION
    rows.loc[rows["Samsung Impact"].eq("Direct"), "Mail Group"] = GROUP_CORE
    rows.loc[rows["Samsung Impact"].eq("None"), "Mail Group"] = GROUP_NOISE
    return rows


def executive_sort_frame(rows: pd.DataFrame) -> pd.DataFrame:
    rows = rows.copy()
    rows["_exec_score"] = rows.apply(_exec_score, axis=1)
    return rows.sort_values(["_exec_score", "_sort_date"], ascending=[False, False])


def choose_top3(rows: pd.DataFrame) -> pd.DataFrame:
    pool = rows[~rows["Mail Group"].eq(GROUP_NOISE)].copy()
    if pool.empty:
        return pool
    pool = executive_sort_frame(pool).drop_duplicates(subset=["_display_cluster"], keep="first")
    selected, used_family = [], set()
    for _, row in pool.iterrows():
        fam = _exec_family(row)
        if fam in used_family and len(selected) < 3:
            continue
        selected.append(row)
        used_family.add(fam)
        if len(selected) >= 3:
            break
    if len(selected) < 3:
        for _, row in pool.iterrows():
            if any(clean(row.get("Headline")) == clean(x.get("Headline")) for x in selected):
                continue
            selected.append(row)
            if len(selected) >= 3:
                break
    return pd.DataFrame(selected).reset_index(drop=True)


def final_order(rows: pd.DataFrame, top3: pd.DataFrame) -> pd.DataFrame:
    rows = assign_mail_groups(rows.copy())
    rows = executive_sort_frame(rows)
    noise = rows[rows["Mail Group"].eq(GROUP_NOISE)].copy()
    visible_pool = rows[~rows["Mail Group"].eq(GROUP_NOISE)].copy()
    visible_pool = visible_pool.drop_duplicates(subset=["_display_cluster"], keep="first")
    regs = visible_pool[visible_pool["Mail Group"].eq(GROUP_REGULATION)].head(2)
    direct = visible_pool[(visible_pool["Mail Group"].eq(GROUP_CORE)) & (~visible_pool.index.isin(regs.index))].head(4)
    indirect = visible_pool[
        (visible_pool["Mail Group"].eq(GROUP_USABLE))
        & (~visible_pool.index.isin(regs.index))
        & (~visible_pool.index.isin(direct.index))
    ].head(6)
    visible = pd.concat([regs, direct, indirect], ignore_index=True)
    visible = executive_sort_frame(visible).head(int(os.getenv("GTI_VISIBLE_MAX", "10")))
    out = pd.concat([visible, noise.head(20)], ignore_index=True)
    out["No"] = range(1, len(out) + 1)
    return out.reset_index(drop=True)


def force_korean_text(row: pd.Series, field: str, fallback: str) -> str:
    topic = _exec_issue(row)
    impact = clean(row.get("Samsung Impact")) or "Indirect"
    headline = clean(row.get("Headline"))
    if field == "Summary":
        return f"{headline}: {topic} issue. Final Samsung impact is {impact}; verify country/entity/product linkage using the original source."
    if field == "AI Analysis":
        if topic == ISS_TARIFF:
            return "Check target HS, supplier country, effective date, tariff rate/quota, and Samsung import/export lane exposure."
        if topic == "AD/CVD":
            return "Check affected supplier/product/origin, investigation or rate status, price data, evidence, and retroactive duty risk."
        if topic == ISS_EXPORT:
            return "Check ECCN/strategic item classification, restricted-party screening, end user, and anti-circumvention controls."
        if topic == "CBAM":
            return "Check EU-bound products, raw materials, supplier emissions data, and CBAM reporting evidence."
        if topic == ISS_FTA:
            return "Check CO issuance/receipt, BOM origin qualification, preferential duty eligibility, and evidence retention."
        if topic == ISS_HS:
            return "Check HS mapping by country, product description, and whether advance ruling is needed."
        return fallback
    if field == "Action Plan":
        if topic == ISS_TARIFF:
            return "Map HS/supplier country; verify tariff or quota applicability; extract related import/export volume; run cost scenario."
        if topic == "AD/CVD":
            return "Identify affected supply chain; check manufacturer/origin/price data; confirm investigation scope; align with Legal/SCM."
        if topic == ISS_EXPORT:
            return "Recheck ECCN/strategic item; screen counterparty/end user; verify overseas subsidiary and anti-circumvention controls."
        if topic == "CBAM":
            return "Identify EU-bound products; collect supplier emissions data; review CBAM reporting/evidence process with ESG/Procurement."
        if topic == ISS_FTA:
            return "Review origin criteria; verify BOM qualification; check CO issuance/receipt/retention; recheck preferential-rate eligibility."
        if topic == ISS_HS:
            return "Collect product description; compare HS master; check country classification differences; decide advance ruling need."
        return fallback
    return clean(row.get(field)) or fallback


def build_required_actions(rows: pd.DataFrame, top3: pd.DataFrame) -> list[dict]:
    pool = pd.concat([top3, rows[rows["Samsung Impact"].eq("Direct")]], ignore_index=True)
    if pool.empty:
        return []
    pool = executive_sort_frame(pool).drop_duplicates(subset=["Issue"], keep="first").head(5)
    owner_map = {
        ISS_TARIFF: "Customs/FTA",
        "AD/CVD": "Global SCM/Legal",
        ISS_EXPORT: "Export Control",
        "CBAM": "ESG/Procurement",
        ISS_FTA: "FTA",
        ISS_HS: "Customs",
        ISS_REG: "Customs",
    }
    return [
        {
            "topic": clean(r.get("Issue")),
            "action": force_korean_text(r, "Action Plan", "Review by customs/trade owner required."),
            "owner": owner_map.get(clean(r.get("Issue")), "Customs/Trade"),
        }
        for _, r in pool.iterrows()
    ]


# =========================================================
# GTI recovery overrides, 2026-06-06
# - STEP5 must compose the mail, not reselect STEP4 down to a tiny set.
# - Restore executive mail volume to 30 visible rows by default.
# - Keep weak but relevant rows as Watch instead of dropping them.
# - Block tracking/script URLs but keep the article row with plain text title.
# =========================================================

RECOVERY_VISIBLE_MAX = int(os.getenv("GTI_VISIBLE_MAX", "30"))
RECOVERY_CORE_MIN = int(os.getenv("GTI_CORE_MIN", "10"))
RECOVERY_WATCH_MIN = int(os.getenv("GTI_WATCH_MIN", "12"))
RECOVERY_REG_MIN = int(os.getenv("GTI_REG_MIN", "3"))

RECOVERY_BAD_URL_FRAGMENTS = [
    "google-analytics.com",
    "analytics.js",
    "googletagmanager.com",
    "doubleclick.net",
    "googleadservices.com",
    "google.com/pagead",
    "googleusercontent.com",
    "gstatic.com",
    "favicon",
    "thumbnail",
]

RECOVERY_POLICY_TERMS = [
    "tariff", "tariffs", "duty", "section 301", "section 232", "ustr", "usmca",
    "cbam", "carbon border", "export control", "export controls", "bis", "eccn",
    "entity list", "uflpa", "forced labor", "anti-dumping", "antidumping",
    "countervailing", "ad/cvd", "fta", "cepa", "epa", "origin", "rules of origin",
    "hs code", "customs", "clearance", "관세", "관세율", "추가관세", "상호관세",
    "수출통제", "전략물자", "원산지", "품목분류", "통관", "반덤핑", "상계관세",
    "탄소국경", "무역합의", "무역협상", "공급망",
]


def is_preferred_article_url(url: str) -> bool:
    u = clean(url)
    if not is_valid_http_url(u):
        return False
    low = u.lower()
    if any(x in low for x in RECOVERY_BAD_URL_FRAGMENTS):
        return False
    if "news.google.com/" in low:
        return False
    if any(x in low for x in BAD_URL_HOST_FRAGMENTS_V315):
        return False
    bad_fragments = [
        "accounts.google.", "policies.google.", "support.google.", "consent.google.",
        "google.com/search", "google.com/amp/s/", "agency.reuters.com/en/copyright",
        "/favicon", "logo", "thumbnail", "thumb", "w16", "=w16",
    ]
    return not any(x in low for x in bad_fragments)


def best_url_from_values(values: list[str]) -> str:
    cleaned = []
    for v in values:
        vv = clean(v)
        if not vv or vv.lower() in {"nan", "none", "null", "new", "https://new", "http://new"}:
            continue
        cleaned.append(vv)
        for found in re.findall(r"https?://[^'\"),\s]+", vv):
            if found not in cleaned:
                cleaned.append(found)

    for v in cleaned:
        if is_preferred_article_url(v):
            return v
    for v in cleaned:
        if is_google_news_rss_url(v):
            resolved = resolve_google_news_url(v)
            if is_preferred_article_url(resolved):
                return resolved
    return ""


def _recovery_text(row: pd.Series) -> str:
    return " ".join([
        clean(row.get("Headline")),
        clean(row.get("Summary")),
        clean(row.get("AI Analysis")),
        clean(row.get("Action Plan")),
        clean(row.get("Issue")),
        clean(row.get("Cluster")),
        clean(row.get("Country")),
        clean(row.get("Agency")),
        clean(row.get("Source")),
    ]).lower()


def _recovery_has_policy_signal(row: pd.Series) -> bool:
    text = _recovery_text(row)
    return any(str(t).lower() in text for t in RECOVERY_POLICY_TERMS)


def _exec_issue(row: pd.Series) -> str:
    raw_issue = clean(row.get("Issue")).upper()
    headline = clean(row.get("Headline")).lower()
    text = _recovery_text(row)

    raw_map = {
        "EXPORT_CONTROL": "수출통제",
        "CBAM_CARBON": "CBAM",
        "AD_CVD": "AD/CVD",
        "ORIGIN_FTA": "FTA/원산지",
        "HS_CLASSIFICATION": "HS/품목분류",
        "CUSTOMS": "통관",
        "CUSTOMS_CLEARANCE": "통관",
        "TARIFF": "관세정책",
        "SECTION_301_232": "관세정책",
        "TRADE_REGULATION": "법규",
    }
    if raw_issue in raw_map:
        return raw_map[raw_issue]
    if any(k in headline for k in ["section 301", "section 232", "tariff", "관세", "상호관세", "추가관세"]):
        return "관세정책"
    if any(k in text for k in ["export control", "entity list", "uflpa", "forced labor", "bis", "eccn", "수출통제", "전략물자"]):
        return "수출통제"
    if any(k in text for k in ["cbam", "carbon border", "탄소국경"]):
        return "CBAM"
    if any(k in text for k in ["ad/cvd", "anti-dumping", "antidumping", "countervailing", "반덤핑", "상계관세"]):
        return "AD/CVD"
    if any(k in text for k in ["fta", "cepa", "epa", "origin", "rules of origin", "원산지"]):
        return "FTA/원산지"
    if any(k in text for k in ["hs code", "classification", "품목분류"]):
        return "HS/품목분류"
    if any(k in text for k in ["customs", "clearance", "통관"]):
        return "통관"
    if clean(row.get("Content Type")).lower().startswith("reg"):
        return "법규"
    return "Watch"


def _exec_family(row: pd.Series) -> str:
    topic = _exec_issue(row)
    if topic in {"관세정책", "AD/CVD"}:
        return "TARIFF"
    if topic == "수출통제":
        return "EXPORT_CONTROL"
    if topic == "CBAM":
        return "CBAM"
    if topic == "FTA/원산지":
        return "FTA"
    if topic == "HS/품목분류":
        return "HS"
    if topic == "통관":
        return "CUSTOMS"
    if topic == "법규":
        return "REG"
    return "WATCH"


def determine_samsung_impact(row: pd.Series) -> tuple[str, str, str]:
    if "is_old_by_24h_rule_v315" in globals() and is_old_by_24h_rule_v315(row):
        if clean(row.get("Content Type")).lower().startswith("reg"):
            return "Watch", "SEC/HQ", "Older regulation retained in review section, not executive Top."
        return "Watch", "SEC/HQ", "Older item retained as Watch for audit, not discarded."
    if not _recovery_has_policy_signal(row):
        return "Watch", "SEC/HQ", "No hard action signal; retained as Watch rather than dropped."

    topic = _exec_issue(row)
    subs = clean(row.get("Affected Subsidiary")) or "SEC/HQ"
    text = _recovery_text(row)
    product = any(k in text for k in ["samsung", "삼성", "semiconductor", "반도체", "chip", "display", "battery", "electronics"])
    hard = any(k in text for k in ["effective", "final rule", "impose", "rate", "quota", "entity list", "eccn", "hs code", "관세율", "시행", "고시"])

    if clean(row.get("Content Type")).lower().startswith("reg"):
        return "Watch", "SEC/HQ", "Official regulation retained for legal monitoring; verify trade/customs impact."
    if topic in {"수출통제", "CBAM", "AD/CVD", "HS/품목분류"}:
        return ("Direct" if (product or hard) else "Indirect"), subs, f"{topic}: product/entity/evidence review required."
    if topic in {"관세정책", "FTA/원산지", "통관"}:
        return ("Direct" if product and hard else "Indirect"), subs, f"{topic}: HS/origin/rate applicability review required."
    return "Watch", "SEC/HQ", "Trade-policy signal retained for monitoring."


def apply_samsung_impact(rows: pd.DataFrame) -> pd.DataFrame:
    rows = rows.copy()
    impacts = rows.apply(determine_samsung_impact, axis=1)
    rows["Samsung Impact"] = [x[0] for x in impacts]
    rows["Affected Subsidiary"] = [x[1] for x in impacts]
    rows["Impact Reason"] = [x[2] for x in impacts]
    rows["Issue"] = rows.apply(_exec_issue, axis=1)
    rows["_display_cluster"] = rows.apply(_exec_cluster_key, axis=1)
    return rows


def assign_mail_groups(rows: pd.DataFrame) -> pd.DataFrame:
    rows = apply_samsung_impact(rows.copy())
    rows["Mail Group"] = GROUP_USABLE
    rows.loc[rows["Content Type"].eq("Regulation"), "Mail Group"] = GROUP_REGULATION
    rows.loc[rows["Samsung Impact"].eq("Direct"), "Mail Group"] = GROUP_CORE
    rows.loc[rows["Samsung Impact"].eq("Watch"), "Mail Group"] = GROUP_USABLE
    return rows


def final_order(rows: pd.DataFrame, top3: pd.DataFrame) -> pd.DataFrame:
    rows = assign_mail_groups(rows.copy())
    rows = executive_sort_frame(rows)
    rows = rows.drop_duplicates(subset=["_display_cluster"], keep="first")

    regs = rows[rows["Mail Group"].eq(GROUP_REGULATION)].head(RECOVERY_REG_MIN)
    core = rows[(rows["Mail Group"].eq(GROUP_CORE)) & (~rows.index.isin(regs.index))].head(RECOVERY_CORE_MIN)
    watch = rows[
        (rows["Mail Group"].eq(GROUP_USABLE))
        & (~rows.index.isin(regs.index))
        & (~rows.index.isin(core.index))
    ].head(RECOVERY_WATCH_MIN)

    visible = pd.concat([regs, core, watch], ignore_index=False)
    if len(visible) < RECOVERY_VISIBLE_MAX:
        rest = rows[~rows.index.isin(visible.index)].head(RECOVERY_VISIBLE_MAX - len(visible))
        visible = pd.concat([visible, rest], ignore_index=False)

    visible = executive_sort_frame(visible).head(RECOVERY_VISIBLE_MAX).reset_index(drop=True)
    visible["No"] = range(1, len(visible) + 1)
    return visible


def choose_top3(rows: pd.DataFrame) -> pd.DataFrame:
    pool = rows.copy()
    if pool.empty:
        return pool
    pool = executive_sort_frame(pool).drop_duplicates(subset=["_display_cluster"], keep="first")
    selected, used_family = [], set()
    for _, row in pool.iterrows():
        fam = _exec_family(row)
        if fam in used_family and len(selected) < 3:
            continue
        selected.append(row)
        used_family.add(fam)
        if len(selected) >= 3:
            break
    if len(selected) < 3:
        for _, row in pool.iterrows():
            if any(clean(row.get("Headline")) == clean(x.get("Headline")) for x in selected):
                continue
            selected.append(row)
            if len(selected) >= 3:
                break
    return pd.DataFrame(selected).reset_index(drop=True)


def force_korean_text(row: pd.Series, field: str, fallback: str) -> str:
    existing = clean(row.get(field))
    topic = _exec_issue(row)
    impact = clean(row.get("Samsung Impact")) or "Watch"
    headline = clean(row.get("Headline"))
    if existing and "CBAM issue" not in existing:
        return existing
    if field == "Summary":
        return f"{headline}: {topic} 이슈입니다. 삼성 영향은 {impact}로 분류되며 원문 기준 국가·품목·법인 연결성을 확인해야 합니다."
    if field == "AI Analysis":
        if topic == "관세정책":
            return "대상 HS, 공급국, 시행일, 관세율·쿼터 및 삼성 수출입 Lane 노출 여부를 확인해야 합니다."
        if topic == "AD/CVD":
            return "대상 공급망, 제조사, 원산지, 가격자료, 조사 범위와 소급 관세 리스크를 확인해야 합니다."
        if topic == "수출통제":
            return "ECCN/전략물자 해당 여부, 거래상대방·최종사용자 Screening, 우회수출 통제 여부를 확인해야 합니다."
        if topic == "CBAM":
            return "EU향 제품, 원재료, 공급사 배출량 데이터와 CBAM 신고 증빙 체계를 확인해야 합니다."
        if topic == "FTA/원산지":
            return "CO 발급·수취, BOM 원산지 충족, 협정세율 적용 가능성 및 증빙 보관 기준을 확인해야 합니다."
        if topic == "HS/품목분류":
            return "품목 설명, 국가별 HS Mapping 차이, 사전심사 필요 여부를 확인해야 합니다."
        return fallback
    if field == "Action Plan":
        if topic == "관세정책":
            return "HS/공급국을 매핑하고 관세·쿼터 적용 여부와 관련 수출입 금액을 산출하여 비용 시나리오를 작성"
        if topic == "수출통제":
            return "ECCN/전략물자 재점검, 거래상대방·최종사용자 Screening, 해외법인 우회수출 통제 확인"
        if topic == "CBAM":
            return "EU향 제품 식별, 공급사 배출량 데이터 수집, ESG/구매 부서와 CBAM 증빙 프로세스 점검"
        if topic == "FTA/원산지":
            return "원산지 기준, BOM 충족, CO 발급·수취·보관 및 협정세율 적용 가능성 재점검"
        return fallback
    return existing or fallback


def build_required_actions(rows: pd.DataFrame, top3: pd.DataFrame) -> list[dict]:
    pool = pd.concat([top3, rows], ignore_index=True)
    if pool.empty:
        return []
    pool = executive_sort_frame(pool).drop_duplicates(subset=["Issue"], keep="first").head(6)
    owner_map = {
        "관세정책": "통관운영/FTA팀",
        "AD/CVD": "Global SCM/법무",
        "수출통제": "수출통제팀",
        "CBAM": "ESG/구매",
        "FTA/원산지": "FTA팀",
        "HS/품목분류": "품목분류/통관운영팀",
        "통관": "통관운영팀",
        "법규": "관세/통상 담당",
    }
    return [
        {
            "topic": clean(r.get("Issue")),
            "action": force_korean_text(r, "Action Plan", "관련 부서 검토 필요"),
            "owner": owner_map.get(clean(r.get("Issue")), "관세/통상 담당"),
        }
        for _, r in pool.iterrows()
    ]



# ============================================================================
# GTI STEP5 FINAL MAIL GUARD v3.16
# 2026-08-10
#
# Operating contract:
# - Section 3. Regulation: always show publication date, but do NOT apply the
#   news 24-hour hard cut. Regulation freshness/new-event logic belongs upstream.
# - Section 4. 주요뉴스: publication-date based STRICT 24-hour rule.
# - CollectedAt / crawl time must never rescue an old news article.
# - News with missing/unparseable publication date is excluded from the mail.
# - Top3 is selected only after this final freshness guard.
# ============================================================================

MAIL_NEWS_HOURS_V316 = int(os.getenv("GTI_MAIL_NEWS_HOURS", "24"))


def _published_ts_v316(row: pd.Series) -> pd.Timestamp:
    """Publication timestamp only. Never fall back to CollectedAt."""
    for c in ["Date", "Publish Date", "published", "published_at", "pubDate", "pub_date"]:
        if c not in row.index:
            continue
        raw = clean(row.get(c))
        if not raw:
            continue
        dt = pd.to_datetime(raw, errors="coerce")
        if not pd.isna(dt):
            try:
                if getattr(dt, "tzinfo", None) is not None:
                    dt = dt.tz_localize(None)
            except Exception:
                pass
            return pd.Timestamp(dt)
    return pd.NaT


def _mail_anchor_v316() -> pd.Timestamp:
    """Use actual execution time so '24 hours' means exactly 24 elapsed hours."""
    return pd.Timestamp(datetime.now())


def _news_is_fresh_v316(row: pd.Series) -> bool:
    if clean(row.get("Content Type")).lower() != "news":
        return True

    dt = _published_ts_v316(row)
    if pd.isna(dt):
        return False

    anchor = _mail_anchor_v316()
    cutoff = anchor - pd.Timedelta(hours=MAIL_NEWS_HOURS_V316)

    # reject future timestamps beyond a small clock-skew allowance
    if dt > anchor + pd.Timedelta(hours=2):
        return False

    return dt >= cutoff


def _apply_mail_24h_guard_v316(rows: pd.DataFrame) -> pd.DataFrame:
    if rows is None or rows.empty:
        return rows.copy()

    work = rows.copy()
    news_mask = work["Content Type"].astype(str).str.lower().eq("news")
    news_total = int(news_mask.sum())

    fresh_news_mask = work.apply(_news_is_fresh_v316, axis=1)
    old_news = work[news_mask & ~fresh_news_mask].copy()

    if not old_news.empty:
        old_news["Mail Group"] = "Filtered Noise"
        old_news["RejectReason"] = old_news.get(
            "RejectReason", pd.Series("", index=old_news.index)
        ).fillna("").astype(str).apply(
            lambda x: (
                (x + "; " if x else "")
                + "STEP5_STRICT_24H_PUBLISHED_DATE"
            )
        )

    # Regulations remain untouched. Only stale/unknown-date NEWS are excluded.
    keep = work[~(news_mask & ~fresh_news_mask)].copy()
    out = pd.concat([keep, old_news], ignore_index=True, sort=False)

    print(
        f"[STEP5 24H GUARD] news={news_total} / "
        f"fresh={int((news_mask & fresh_news_mask).sum())} / "
        f"excluded={len(old_news)} / hours={MAIL_NEWS_HOURS_V316}"
    )
    return out


_PREV_FINAL_ORDER_V316 = final_order

def final_order(rows: pd.DataFrame, top3: pd.DataFrame) -> pd.DataFrame:
    """
    Final report order with a non-bypassable mail-boundary freshness check.
    Apply the existing executive selection first, then block stale news again.
    """
    ordered = _PREV_FINAL_ORDER_V316(rows, top3)
    ordered = _apply_mail_24h_guard_v316(ordered)

    visible = ordered[~ordered["Mail Group"].eq("Filtered Noise")].copy()
    filtered = ordered[ordered["Mail Group"].eq("Filtered Noise")].copy()

    # Stable re-numbering after age filtering.
    if not visible.empty:
        visible = executive_sort_frame(visible).reset_index(drop=True)
        visible["No"] = range(1, len(visible) + 1)
    if not filtered.empty:
        filtered = filtered.reset_index(drop=True)
        filtered["No"] = range(len(visible) + 1, len(visible) + len(filtered) + 1)

    return pd.concat([visible, filtered], ignore_index=True, sort=False)


def _display_publish_date_v316(row: pd.Series) -> str:
    dt = _published_ts_v316(row)
    if pd.isna(dt):
        return "확인 필요"

    # Keep time when supplied by source; otherwise show date only.
    raw = clean(row.get("Date"))
    if re.search(r"\d{1,2}:\d{2}", raw):
        return dt.strftime("%Y-%m-%d %H:%M")
    return dt.strftime("%Y-%m-%d")


def build_table(title: str, rows: pd.DataFrame, color: str) -> str:
    """Executive table: explicit 게시일 column for Regulation and News."""
    if rows.empty:
        return ""

    col_widths = [
        "3%", "7%", "5%", "8%", "14%", "14%",
        "14%", "14%", "5%", "5%", "4%", "7%"
    ]
    headers = [
        "No", "Topic", "Impact", "Subsidiary", "Headline", "Summary",
        "삼성 영향", "Action", "Country", "Agency", "Risk", "게시일"
    ]
    colgroup = "<colgroup>" + "".join(
        f"<col style='width:{w};'>" for w in col_widths
    ) + "</colgroup>"
    ths = "".join(
        f"<th style='padding:7px;border:1px solid #d9d9d9;text-align:center;'>{h}</th>"
        for h in headers
    )

    trs = []
    for _, row in rows.iterrows():
        trs.append(f"""
        <tr>
          <td style="padding:7px;border:1px solid #d9d9d9;text-align:center;vertical-align:top;">{html.escape(str(row.get('No', '')))}</td>
          <td style="padding:7px;border:1px solid #d9d9d9;text-align:center;vertical-align:top;">{html.escape(display_topic(row))}</td>
          <td style="padding:7px;border:1px solid #d9d9d9;text-align:center;vertical-align:top;font-weight:bold;color:{'#C00000' if clean(row.get('Samsung Impact')) == 'Direct' else '#666'};">{html.escape(clean(row.get('Samsung Impact')))}</td>
          <td style="padding:7px;border:1px solid #d9d9d9;text-align:center;vertical-align:top;">{html.escape(clean(row.get('Affected Subsidiary')) or '-')}</td>
          <td style="padding:7px;border:1px solid #d9d9d9;vertical-align:top;">{html_link(row.get('Headline', ''), row.get('URL', ''))}</td>
          <td style="padding:7px;border:1px solid #d9d9d9;vertical-align:top;">{html.escape(force_korean_text(row, 'Summary', '요약 정보 확인 필요'))}</td>
          <td style="padding:7px;border:1px solid #d9d9d9;vertical-align:top;">{html.escape(force_korean_text(row, 'AI Analysis', '영향 검토 필요'))}</td>
          <td style="padding:7px;border:1px solid #d9d9d9;vertical-align:top;">{html.escape(force_korean_text(row, 'Action Plan', '담당 부서 확인 필요'))}</td>
          <td style="padding:7px;border:1px solid #d9d9d9;text-align:center;vertical-align:top;">{html.escape(clean(row.get('Country')))}</td>
          <td style="padding:7px;border:1px solid #d9d9d9;text-align:center;vertical-align:top;">{html.escape(clean(row.get('Agency')))}</td>
          <td style="padding:7px;border:1px solid #d9d9d9;text-align:center;vertical-align:top;color:{risk_color(row.get('Risk'))};font-weight:bold;">{html.escape(clean(row.get('Risk')))}</td>
          <td style="padding:7px;border:1px solid #d9d9d9;text-align:center;vertical-align:top;font-weight:bold;">{html.escape(_display_publish_date_v316(row))}</td>
        </tr>
        """)

    return f"""
    <h3 style="margin-top:24px;color:{color};">{html.escape(title)} ({len(rows)}건)</h3>
    <table style="border-collapse:collapse;width:100%;font-size:12px;table-layout:fixed;">
      {colgroup}
      <thead><tr style="background:{color};color:white;">{ths}</tr></thead>
      <tbody>{''.join(trs)}</tbody>
    </table>
    """


def build_html(rows: pd.DataFrame, top3: pd.DataFrame) -> str:
    """
    v3.17 consistent report layout.

    Critical rule:
    - The HTML sections MUST use the exact same Mail Group visibility decision
      as the executive summary/counts.
    - Filtered Noise must never reappear merely because Content Type is Regulation.
    - Section 3 = Mail Group == Regulation only.
    - Section 4 = visible News groups only, with strict 24h publication-date guard.
    """
    subject = f"[GTI Radar] Global Trade Intelligence | {RUN_DATE}"

    # One source of truth for mail visibility.
    visible = rows[~rows["Mail Group"].eq("Filtered Noise")].copy()

    # Final 24h safety guard for NEWS only.
    if not visible.empty:
        keep_mask = visible.apply(
            lambda r: (
                True
                if clean(r.get("Content Type")).lower() != "news"
                else _news_is_fresh_v316(r)
            ),
            axis=1,
        )
        visible = visible[keep_mask].copy()

    # IMPORTANT: section membership is Mail Group based, not Content Type alone.
    regulation = visible[visible["Mail Group"].eq("Regulation")].copy()

    news = visible[
        visible["Mail Group"].isin(["News - 핵심", "News - 주요/참고"])
    ].copy()

    # Top3 must also come only from currently visible rows.
    visible_keys = set(
        clean(x).lower()
        for x in visible.get("Headline", pd.Series(dtype=str))
        if clean(x)
    )

    top3_clean = top3.copy()
    if not top3_clean.empty:
        top3_clean = top3_clean[
            top3_clean["Headline"].fillna("").astype(str).map(
                lambda x: clean(x).lower() in visible_keys
            )
        ].copy()

        top3_clean = top3_clean[
            top3_clean.apply(
                lambda r: (
                    True
                    if clean(r.get("Content Type")).lower() != "news"
                    else _news_is_fresh_v316(r)
                ),
                axis=1,
            )
        ].copy().reset_index(drop=True)

    top_blocks = []
    for idx, row in top3_clean.iterrows():
        top_blocks.append(f"""
        <div style="margin:14px 0 16px 0;padding:14px;border-left:5px solid #C00000;background:#FFF7F7;">
          <div style="font-size:15px;font-weight:bold;margin-bottom:6px;">Top {idx + 1}. {html_link(row.get('Headline', ''), row.get('URL', ''))}</div>
          <div style="font-size:12px;color:#555;margin-bottom:8px;">
            Topic: {html.escape(display_topic(row))} |
            Samsung Impact: <b>{html.escape(clean(row.get('Samsung Impact')))}</b> |
            Agency: {html.escape(clean(row.get('Agency')))} |
            게시일: <b>{html.escape(_display_publish_date_v316(row))}</b> |
            Country: {html.escape(clean(row.get('Country')))} |
            Risk: <span style="color:{risk_color(row.get('Risk'))};font-weight:bold;">{html.escape(clean(row.get('Risk')))}</span>
          </div>
          <div style="margin-top:7px;"><b>요약</b><br>{html.escape(force_korean_text(row, 'Summary', '요약 정보 확인 필요'))}</div>
          <div style="margin-top:7px;"><b>영향</b><br>{html.escape(force_korean_text(row, 'AI Analysis', '영향 검토 필요'))}</div>
          <div style="margin-top:7px;"><b>대응조치</b><br>{html.escape(force_korean_text(row, 'Action Plan', '담당 부서 확인 필요'))}</div>
        </div>
        """)

    # Diagnostic consistency check.
    print(
        f"[STEP5 HTML CONSISTENCY] visible={len(visible)} / "
        f"regulation={len(regulation)} / news={len(news)} / top3={len(top3_clean)}"
    )

    return f"""<!DOCTYPE html>
<html lang="ko">
<head><meta charset="utf-8"><title>{html.escape(subject)}</title></head>
<body style="font-family:Arial,'Malgun Gothic',sans-serif;font-size:13px;color:#222;line-height:1.55;">
  <div style="max-width:1320px;margin:0 auto;">
    <h2 style="margin-bottom:3px;color:#1F4E78;">[GTI Radar] Global Trade Intelligence</h2>
    <div style="font-size:13px;color:#555;margin-bottom:16px;">{RUN_DATE} | Samsung Electronics Customs & Trade Intelligence</div>

    <h3 style="margin-top:18px;margin-bottom:6px;">1. 총평</h3>
    {build_overall_review_html(rows, top3_clean)}

    <h3 style="margin-top:22px;color:#C00000;">2. Top3 Deep Analysis</h3>
    {''.join(top_blocks)}

    {build_table('3. Regulation', regulation, '#1F4E78')}
    {build_table('4. 주요뉴스', news, '#548235')}

    <p style="margin-top:18px;color:#666;font-size:12px;">
      주요뉴스는 기사 게시일 기준 최근 {MAIL_NEWS_HOURS_V316}시간 이내 항목만 표시합니다.
      법규는 최종 Mail Group=Regulation으로 선별된 신규 Event만 표시합니다.
    </p>
  </div>
</body>
</html>"""


# ============================================================================
# End GTI STEP5 FINAL MAIL GUARD v3.17
# ============================================================================


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--date", default=None)
    parser.add_argument("--regulation-input", default=None)
    parser.add_argument("--news-input", default=None)
    parser.add_argument("--output-dir", default=None)
    parser.add_argument("--no-email", action="store_true")
    args = parser.parse_args()
    if args.date:
        RUN_DATE = args.date
    if args.regulation_input:
        REGULATION_INPUT_FILE = Path(args.regulation_input)
    if args.news_input:
        NEWS_INPUT_FILE = Path(args.news_input)
    if args.output_dir:
        OUTPUT_DIR = Path(args.output_dir)
    if args.no_email:
        SEND_EMAIL = False
    main()
