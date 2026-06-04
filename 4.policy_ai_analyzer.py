# -*- coding: utf-8 -*-
"""
GTI STEP4-2 : News AI Analysis / 50-row Ranking Engine - GTI95 RELAXED FINAL

Input
-----
- C:/Temp/3-2.news_article_summary.xlsx  (preferred)
- C:/Temp/3-2.news_summary.xlsx          (fallback)

Reference masters
-----------------
- C:/Temp/GTI_TOPIC_MASTER.xlsx
- C:/Temp/SUBSIDIARY_MASTER.xlsx

Output
------
- C:/Temp/4-2.news_ai_summary.xlsx
- C:/Temp/4-2.news_ai_cumulative.xlsx
- C:/Temp/4-2.news_ai_audit_candidates.xlsx

GTI95 scoring rule
------------------
Final Score = TopicScore * 40% + SamsungImpactScore * 30% + ActionScore * 20% + UrgencyScore * 10%

Purpose
-------
STEP4-2 is a ranking engine. It creates a 50-row candidate pool for STEP5 Top30:
- Direct / Indirect / None impact classification
- affected subsidiary/product mapping
- required action and owner
- executive message for Samsung Electronics customs/trade management
"""
from __future__ import annotations

import os
import re
import json
import time
import tempfile
from datetime import datetime
from pathlib import Path
from urllib.parse import urlparse, quote_plus, unquote
from difflib import SequenceMatcher

import requests

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


BASE_DIR = Path(os.getenv("GTI_BASE_DIR", r"C:\Temp"))
INPUT_CANDIDATES = [
    Path(os.getenv("GTI_NEWS_ARTICLE_INPUT", BASE_DIR / "3-2.news_article_summary.xlsx")),
    Path(os.getenv("GTI_NEWS_SUMMARY_INPUT", BASE_DIR / "3-2.news_summary.xlsx")),
]
TOPIC_MASTER_INPUT = Path(os.getenv("GTI_TOPIC_MASTER_INPUT", BASE_DIR / "GTI_TOPIC_MASTER.xlsx"))
SUBSIDIARY_MASTER_INPUT = Path(os.getenv("GTI_SUBSIDIARY_MASTER_INPUT", BASE_DIR / "SUBSIDIARY_MASTER.xlsx"))

OUTPUT_DAILY = Path(os.getenv("GTI_NEWS_OUTPUT", BASE_DIR / "4-2.news_ai_summary.xlsx"))
OUTPUT_CUMUL = Path(os.getenv("GTI_NEWS_CUMULATIVE", BASE_DIR / "4-2.news_ai_cumulative.xlsx"))
OUTPUT_AUDIT = Path(os.getenv("GTI_NEWS_AUDIT", BASE_DIR / "4-2.news_ai_audit_candidates.xlsx"))

TOP_N = int(os.getenv("GTI_NEWS_TOP_N", "30"))  # Executive mail standard Top30; override with GTI_NEWS_TOP_N
MAX_REFERENCE = int(os.getenv("GTI_MAX_REFERENCE", "15"))
MAX_PER_CLUSTER = int(os.getenv("GTI_MAX_PER_CLUSTER", "1"))
MAX_PER_SOURCE = int(os.getenv("GTI_MAX_PER_SOURCE", "12"))

# URL repair / validation settings. Keep keys in Windows environment variables, not in source code.
ENABLE_URL_REPAIR = os.getenv("GTI_ENABLE_URL_REPAIR", "Y").strip().upper() != "N"
ENABLE_PDF_TEXT_CHECK = os.getenv("GTI_ENABLE_PDF_TEXT_CHECK", "Y").strip().upper() != "N"
URL_REPAIR_TIMEOUT = int(os.getenv("GTI_URL_REPAIR_TIMEOUT", "8"))
SERPAPI_KEY = os.getenv("SERPAPI_KEY", "").strip()
NAVER_CLIENT_ID = os.getenv("NAVER_CLIENT_ID", "").strip()
NAVER_CLIENT_SECRET = os.getenv("NAVER_CLIENT_SECRET", "").strip()
NEWS_API_ENDPOINT = os.getenv("NEWS_API", "https://newsapi.org/v2/everything").strip()
NEWS_API_KEY = os.getenv("NEWS_API_KEY", os.getenv("NEWS_KEY", "")).strip()

MISSING_TEXT = "본문에서 확인 불가"

# ---------------------------------------------------------
# Keyword dictionaries
# ---------------------------------------------------------

FALLBACK_TOPICS = [
    {
        "topic": "Entity List",
        "issue_type": "EXPORT_CONTROL",
        "score": 100,
        "keywords": ["entity list", "bis", "restricted party", "blacklist", "제재", "수출통제", "거래제한"],
        "standard_action": "ECCN 재점검; 거래상대방 Screening; 최종사용자/최종용도 확인",
        "action_owner": "수출통제팀",
    },
    {
        "topic": "Export Control",
        "issue_type": "EXPORT_CONTROL",
        "score": 100,
        "keywords": ["export control", "export controls", "ear", "eccn", "dual-use", "수출통제", "전략물자"],
        "standard_action": "ECCN/전략물자 해당 여부 점검; 수출허가 필요성 검토; Hold & Review 적용",
        "action_owner": "수출통제팀",
    },
    {
        "topic": "Semiconductor Tariff",
        "issue_type": "SEMICONDUCTOR_TARIFF",
        "score": 95,
        "keywords": ["semiconductor tariff", "chip tariff", "semiconductor", "chip", "반도체 관세", "반도체"],
        "standard_action": "반도체 관련 HS Mapping 확인; 생산/판매법인별 관세비용 영향 분석; 공급망 대체 시나리오 점검",
        "action_owner": "Global SCM",
    },
    {
        "topic": "Section 301/232",
        "issue_type": "SECTION_301_232",
        "score": 95,
        "keywords": ["section 301", "section 232", "301조", "232조", "ustr", "additional tariff"],
        "standard_action": "미국향/미국발 거래 HS별 추가관세 적용 여부 확인; 관세비용 시뮬레이션 수행",
        "action_owner": "통관운영팀",
    },
    {
        "topic": "AD/CVD",
        "issue_type": "AD_CVD",
        "score": 92,
        "keywords": ["anti-dumping", "antidumping", "countervailing", "dumping", "반덤핑", "덤핑방지관세", "상계관세"],
        "standard_action": "대상국·공급자·HS 기준 AD/CVD 적용 여부 확인; 잠정/확정세율 및 소급 가능성 점검",
        "action_owner": "통관운영팀",
    },
    {
        "topic": "CBAM",
        "issue_type": "CBAM_CARBON",
        "score": 90,
        "keywords": ["cbam", "carbon border", "탄소국경", "carbon tariff"],
        "standard_action": "EU 수출품 CBAM 대상 여부 확인; 배출량 데이터·공급자 증빙 수집 체계 점검",
        "action_owner": "ESG/통상지원팀",
    },
    {
        "topic": "FTA / Origin",
        "issue_type": "ORIGIN_FTA",
        "score": 88,
        "keywords": ["fta", "cepa", "epa", "rules of origin", "origin", "원산지", "협정세율", "co 발급", "certificate of origin"],
        "standard_action": "CO 발급/수취 기준 검토; 원산지 판정근거·BOM·공급자확인서 재점검",
        "action_owner": "FTA팀",
    },
    {
        "topic": "HS Classification",
        "issue_type": "HS_CLASSIFICATION",
        "score": 86,
        "keywords": ["hs code", "hs classification", "tariff classification", "품목분류", "hs코드", "세번"],
        "standard_action": "HS Mapping Master 점검; 최근 신고 HS와 ERP 품목마스터 대조; 사전심사 필요 여부 검토",
        "action_owner": "품목분류/통관운영팀",
    },
    {
        "topic": "Tariff",
        "issue_type": "TARIFF",
        "score": 84,
        "keywords": ["tariff", "tariffs", "import duty", "customs duty", "관세", "관세율", "할당관세"],
        "standard_action": "HS별 관세율 변경 여부 확인; 수입원가 영향 분석; 법인/관세사 신고로직 업데이트",
        "action_owner": "통관운영팀",
    },
    {
        "topic": "Customs Regulation",
        "issue_type": "CUSTOMS",
        "score": 78,
        "keywords": ["customs", "clearance", "declaration", "통관", "수입신고", "수출신고", "세관", "관세청"],
        "standard_action": "신고필드·증빙·관세사 업무지침 변경 여부 확인; 법인별 통관 체크리스트 업데이트",
        "action_owner": "통관운영팀",
    },
]

PRODUCT_TERMS = [
    "samsung", "삼성", "semiconductor", "semiconductors", "chip", "chips", "반도체", "hbm", "dram", "nand",
    "memory", "wafer", "smartphone", "mobile", "galaxy", "스마트폰", "갤럭시", "display", "oled", "디스플레이",
    "battery", "배터리", "electronics", "전자", "server", "network", "tv", "television",
]

DIRECT_POLICY_ISSUES = {
    "EXPORT_CONTROL", "SEMICONDUCTOR_TARIFF", "SECTION_301_232", "AD_CVD", "ORIGIN_FTA", "HS_CLASSIFICATION", "TARIFF", "CBAM_CARBON"
}

NONE_NOISE_TERMS = [
    "award", "awards", "수상", "학회", "세미나", "포럼", "workshop", "conference", "교육", "채용", "인턴",
    "history", "통계", "마약", "밀수", "drug", "smuggling", "travel", "flight", "airport", "opinion", "column",
    "fashion", "restaurant", "recipe", "gold", "jewelry", "crypto", "bitcoin", "sports", "weather",
]

# Security/defense stories often contain the word "tariff" in a long article body
# but are not GTI customs/trade-compliance issues.  They must not be promoted
# to AD/CVD or Tariff merely by generic body text.
SECURITY_NON_TRADE_TERMS = [
    "nuclear-powered submarine", "nuclear submarine", "submarine", "잠수함",
    "uranium enrichment", "우라늄 농축", "security talks", "defense cooperation",
    "방산 협력", "안보 협의", "military", "defense", "국방", "aerospace explosion",
]

# Terms that prove the article is about an actionable trade/customs matter.
# Generic words such as "tariff" alone are intentionally not sufficient when
# the headline is dominated by security, politics, event, or industry fluff.
TRADE_ACTION_SIGNALS = [
    "entity list", "bis", "export control", "export controls", "export restriction", "export restrictions",
    "eccn", "ear", "restricted party", "section 301", "section 232", "ustr",
    "anti-dumping", "antidumping", "countervailing", "ad/cvd",
    "customs duty", "import duty", "additional tariff", "tariff rate",
    "rules of origin", "certificate of origin", "free trade agreement", "fta", "cepa", "epa", "hs code",
    "customs clearance", "customs declaration", "cbam",
    "ai chip", "ai chips", "nvidia", "h200", "huawei", "chip export", "semiconductor export",
    "수출통제", "전략물자", "수출 제한", "수출규제", "제재", "엔비디아", "화웨이", "ai칩", "반도체 수출", "우회로", "해외 자회사",
    "반덤핑", "덤핑방지관세", "상계관세", "추가관세", "관세율", "원산지",
    "협정세율", "품목분류", "수입신고", "통관", "탄소국경",
]

FTA_CONTEXT_TERMS = [
    "free trade agreement", "trade agreement", "economic partnership", "comprehensive economic partnership",
    "fta", "cepa", "epa", "rules of origin", "certificate of origin", "origin rule",
    "tariff concession", "preferential tariff", "협정", "자유무역", "경제동반자", "원산지", "협정세율", "관세양허",
]

CEPA_FALSE_POSITIVE_TERMS = [
    "center for european policy analysis", "cepa.org", "center for european", "policy analysis (cepa)"
]

URGENCY_HIGH = ["effective", "effective date", "in force", "implementation", "final rule", "impose", "imposed", "levy", "시행", "발효", "공포", "확정", "부과", "적용"]
URGENCY_MED = ["proposal", "proposed", "investigation", "notice", "consultation", "review", "개정", "입법예고", "행정예고", "조사개시", "검토", "발표"]
URGENCY_LOW = ["plan", "may", "could", "discuss", "talks", "전망", "논의", "계획", "가능성"]

OUTPUT_COLUMNS = [
    "rank", "Date", "Headline", "URL", "URL_Quality", "URL_Repair_Source", "PDF_Text_Status", "Country", "Agency", "Publisher",
    "priority_group", "mail_section", "selected", "Risk", "final_score",
    "topic", "topic_score", "samsung_impact", "samsung_impact_score", "subsidiary_score", "action_score", "urgency_score",
    "topic_keyword", "topic_reason", "issue_type", "cluster_key",
    "RegulationRelated", "RegulationTransferType",
    "affected_subsidiary", "affected_subsidiaries", "affected_products", "subsidiary_products", "subsidiary_reason",
    "impact_production_subsidiaries", "impact_sales_subsidiaries", "impact_products",
    "fta_impact", "export_control_impact", "hs_impact", "tariff_impact",
    "RequiredAction", "ActionOwner", "ExecutiveMessage",
    "samsung_score", "samsung_reason", "Summary", "AI Analysis", "Action Plan",
    "KeywordMatches", "SelectReason", "RejectReason", "Source", "SourceFile",
    "original_url", "article_body", "ClusterHeadlines", "article_extract_status", "article_source_type",
    "effective_date_hint", "change_detail_hint", "hs_hint", "tariff_rate_hint",
    "last_checked",
]

AUDIT_COLUMNS = OUTPUT_COLUMNS + ["audit_decision", "audit_reason", "base_score", "source_cap_reason"]


def log(msg: str) -> None:
    print(f"[{datetime.now():%Y-%m-%d %H:%M:%S}] {msg}")


def safe_str(value) -> str:
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except Exception:
        pass
    return str(value).strip()


def norm(value) -> str:
    text = safe_str(value).lower()
    text = re.sub(r"https?://\S+", " ", text)
    text = re.sub(r"[^0-9a-z가-힣/%\.\-\s]", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def contains_any(text, keywords) -> bool:
    n = norm(text)
    return any(norm(k) in n for k in keywords if safe_str(k))


def keyword_in_text(keyword: str, text: str) -> bool:
    """Safer keyword match than plain substring.

    - Prevents short tokens such as AD from matching ordinary words.
    - Uses word boundaries for latin/alphanumeric terms.
    - Keeps normal substring behavior for Korean terms.
    """
    k = norm(keyword)
    t = norm(text)
    if not k or not t:
        return False
    # Short latin tokens create many false positives: AD, IN, US, etc.
    if re.fullmatch(r"[a-z]{1,2}", k):
        return False
    if re.fullmatch(r"[a-z0-9][a-z0-9\-/\. ]*[a-z0-9]", k):
        return re.search(rf"(?<![a-z0-9]){re.escape(k)}(?![a-z0-9])", t) is not None
    return k in t


def is_security_non_trade(row: pd.Series) -> bool:
    headline = safe_str(row.get("headline"))
    body = full_text(row)
    # If the headline is clearly security/defense and does not itself mention
    # actionable trade terms, treat it as non-GTI even if the long body mentions tariff.
    if contains_any(headline, SECURITY_NON_TRADE_TERMS) and not contains_any(headline, TRADE_ACTION_SIGNALS):
        return True
    # Body-only defense noise also excluded unless a real trade action signal exists.
    if contains_any(body, SECURITY_NON_TRADE_TERMS) and not contains_any(body, TRADE_ACTION_SIGNALS):
        return True
    return False


def has_article_product_signal(row: pd.Series) -> bool:
    text = full_text(row)
    if contains_any(text, PRODUCT_TERMS):
        return True
    if safe_str(row.get("hs_hint")):
        return True
    # change_detail_hint may hold actual product names from STEP3.
    change = safe_str(row.get("change_detail_hint"))
    if change and not contains_any(change, ["tariff", "관세", "rate", "policy", "agreement", "talks", "협상", "정책"]):
        return True
    return False


def pick_col(df: pd.DataFrame, names: list[str]) -> str | None:
    lookup = {str(c).strip().lower(): c for c in df.columns}
    for name in names:
        if name.lower() in lookup:
            return lookup[name.lower()]
    return None


def first_existing(paths: list[Path]) -> Path | None:
    for p in paths:
        if p.exists():
            return p
    return None


def truncate(text, limit=700):
    return re.sub(r"\s+", " ", safe_str(text)).strip()[:limit]


def full_text(row: pd.Series) -> str:
    """Step4 판단용 전체 텍스트.

    제목만으로 판단하지 않고 STEP3가 넘겨준 article_body,
    ClusterHeadlines, change_detail_hint, hs/tariff hints까지 함께 본다.
    특히 Google News/RSS 계열은 제목이 짧고 본문/클러스터 제목에
    실제 FTA·CBAM·수출통제 단서가 들어있는 경우가 많다.
    """
    cols = [
        "headline", "Headline", "title", "summary", "description",
        "article_body", "ArticleBody", "body", "content",
        "ClusterHeadlines", "cluster_headlines", "cluster_titles", "RelatedHeadlines",
        "country", "Country", "agency", "Agency", "publisher", "Publisher",
        "keyword_matches", "KeywordMatches", "select_reason", "SelectReason",
        "reject_reason", "RejectReason", "source", "Source",
        "effective_date_hint", "EffectiveDateHint", "change_detail_hint", "ChangeDetailHint",
        "hs_hint", "HsHint", "tariff_rate_hint", "TariffRateHint",
    ]
    return " ".join(safe_str(row.get(c)) for c in cols)


def canonical_title(value: str) -> str:
    text = safe_str(value)
    text = re.sub(r"\s+-\s+[^-]{2,80}$", "", text)
    text = re.sub(r"\s+\|\s+.*$", "", text)
    return norm(text)


BAD_URL_PATTERNS = [
    "lh3.googleusercontent.com",
    "googleusercontent.com",
    "gstatic.com",
    "ggpht.com",
    "news.google.com/rss/articles",
    "news.google.com/articles",
    "news.google.com/",
]

ARTICLE_URL_CANDIDATE_COLUMNS = [
    "original_url", "canonical_url", "article_url", "source_url", "resolved_url",
    "originallink", "originallink_url", "real_url", "final_url", "url", "link",
]


def is_bad_article_url(value: str) -> bool:
    """Return True when URL is not a real article URL.

    Google News sometimes stores thumbnail image URLs in URL fields
    (lh3.googleusercontent.com) or RSS redirect URLs.  STEP5 should not
    receive those URLs, so STEP4 normalizes them here.
    """
    url = safe_str(value)
    if not url or url.lower() in {"nan", "none", "null", "-"}:
        return True
    u = url.lower().strip()
    if not (u.startswith("http://") or u.startswith("https://")):
        return True
    if any(bad in u for bad in BAD_URL_PATTERNS):
        return True
    # image-like URLs are not article sources
    if re.search(r"\.(png|jpg|jpeg|gif|webp|svg)(\?|$)", u):
        return True
    return False


def is_pdf_url(value: str) -> bool:
    u = safe_str(value).lower()
    return u.startswith(("http://", "https://")) and (".pdf" in u or "filedownload" in u or "download" in u)


def normalize_url(value: str) -> str:
    url = safe_str(value)
    if not url:
        return ""
    url = url.replace("&amp;", "&").strip()
    # Some feeds wrap URLs inside query parameters such as ?url= or ?u=
    for key in ["url=", "u=", "link="]:
        if key in url and "http" in url.split(key, 1)[-1]:
            tail = url.split(key, 1)[-1].split("&", 1)[0]
            decoded = unquote(tail)
            if decoded.startswith(("http://", "https://")):
                url = decoded
                break
    return url


def choose_clean_url(row: pd.Series) -> tuple[str, str, str]:
    """Choose the best article URL from available candidate columns.

    Returns: (url, URL_Quality, URL_Repair_Source)
    """
    for col in ARTICLE_URL_CANDIDATE_COLUMNS:
        if col in row.index:
            url = normalize_url(row.get(col))
            if url and not is_bad_article_url(url):
                quality = "GAZETTE_PDF" if is_pdf_url(url) else "ORIGINAL_ARTICLE"
                return url, quality, f"column:{col}"
    return "", "URL_MISSING_OR_BAD", "none"



GOOGLE_NEWS_RESOLVE_CACHE = {}


def is_google_news_rss_url(value: str) -> bool:
    u = safe_str(value).lower()
    return "news.google.com/rss/articles/" in u or "news.google.com/articles/" in u


def _google_news_article_id(value: str) -> str:
    try:
        parsed = urlparse(safe_str(value))
        parts = [p for p in parsed.path.split("/") if p]
        return parts[-1] if parts else ""
    except Exception:
        return ""


def _extract_article_url_from_google_text(text: str) -> str:
    if not text:
        return ""
    variants = [text]
    try:
        variants.append(text.encode("utf-8", errors="ignore").decode("unicode_escape", errors="ignore"))
    except Exception:
        pass
    for t in variants:
        for pat in [
            r"data-n-au=[\"'](https?://[^\"']+)[\"']",
            r"data-url=[\"'](https?://[^\"']+)[\"']",
            r"href=[\"'](https?://[^\"']+)[\"']",
            r"url=(https?%3A%2F%2F[^&\"'<>]+)",
            r"(https?:\\/\\/[^\"'<>\\]+)",
            r"(https?://[^\"'<>\s]+)",
        ]:
            for m in re.finditer(pat, t, flags=re.I):
                cand = normalize_url(unquote(m.group(1).replace("\\/", "/"))).rstrip(".,;?곥?")
                if cand and not is_bad_article_url(cand):
                    return cand
    return ""


def _decode_google_news_batchexecute(article_id: str, page_text: str) -> str:
    if not article_id or not page_text:
        return ""
    sg = ""
    ts = ""
    m = re.search(r'data-n-a-sg=["\']([^"\']+)["\']', page_text)
    if m:
        sg = m.group(1)
    m = re.search(r'data-n-a-ts=["\']([^"\']+)["\']', page_text)
    if m:
        ts = m.group(1)
    if not sg or not ts:
        return ""
    try:
        req_obj = [
            "garturlreq",
            [["en-US", "US", ["FINANCE_TOP_INDICES", "WEB_TEST_1_0_0"], None, None, 1, 1, "US:en", None, 180, None, None, None, None, None, 0, None, None, [int(ts), 0]],
             "en-US", "US", 1, [2, 3, 4, 8], 1, 0, "655000234", 0, 0, None, 0],
            article_id,
            int(ts),
            sg,
        ]
        f_req = [[["Fbv4je", json.dumps(req_obj, ensure_ascii=False, separators=(",", ":")), None, "generic"]]]
        resp = requests.post(
            "https://news.google.com/_/DotsSplashUi/data/batchexecute?rpcids=Fbv4je",
            data={"f.req": json.dumps(f_req, ensure_ascii=False, separators=(",", ":"))},
            headers={
                "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/120 Safari/537.36",
                "Accept-Language": "ko-KR,ko;q=0.9,en-US;q=0.8,en;q=0.7",
                "Content-Type": "application/x-www-form-urlencoded;charset=UTF-8",
                "Referer": "https://news.google.com/",
            },
            timeout=URL_REPAIR_TIMEOUT,
        )
        if resp.status_code == 200:
            return _extract_article_url_from_google_text(resp.text)
    except Exception:
        return ""
    return ""


def resolve_google_news_url(value: str) -> str:
    u = safe_str(value)
    if not is_google_news_rss_url(u):
        return u if u and not is_bad_article_url(u) else ""
    if u in GOOGLE_NEWS_RESOLVE_CACHE:
        return GOOGLE_NEWS_RESOLVE_CACHE[u]

    resolved = ""
    page_text = ""
    article_id = _google_news_article_id(u)
    try:
        resp = requests.get(
            u,
            headers={
                "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/120 Safari/537.36",
                "Accept-Language": "ko-KR,ko;q=0.9,en-US;q=0.8,en;q=0.7",
            },
            allow_redirects=True,
            timeout=URL_REPAIR_TIMEOUT,
        )
        final_url = normalize_url(resp.url)
        if final_url and not is_bad_article_url(final_url):
            resolved = final_url
        page_text = resp.text[:500000]
        if not resolved:
            resolved = _extract_article_url_from_google_text(page_text)
        if not resolved:
            resolved = _decode_google_news_batchexecute(article_id, page_text)
    except Exception:
        resolved = ""

    if not resolved or is_bad_article_url(resolved):
        resolved = ""
    GOOGLE_NEWS_RESOLVE_CACHE[u] = resolved
    return resolved


def google_news_urls_from_row(row: pd.Series) -> list[str]:
    urls = []
    for col in ["url", "original_url", "google_url", "source_url", "resolved_url", "Source", "source", "RelatedURLs", "RepresentativeURL"]:
        if col in row.index:
            for found in re.findall(r"https?://[^\s\)\]\}'\"]+", safe_str(row.get(col)), flags=re.I):
                found = found.rstrip(".,;?곥?")
                if is_google_news_rss_url(found) and found not in urls:
                    urls.append(found)
    return urls[:5]

def urls_from_text(*texts: str) -> list[str]:
    joined = "\n".join(safe_str(t) for t in texts if safe_str(t))
    urls = re.findall(r"https?://[^\s\)\]\}\>'\"]+", joined, flags=re.I)
    cleaned = []
    for u in urls:
        u = normalize_url(u).rstrip(".,;、。)")
        if u and not is_bad_article_url(u) and u not in cleaned:
            cleaned.append(u)
    return cleaned[:10]


def search_url_by_serpapi(headline: str, agency: str = "", prefer_pdf: bool = False) -> tuple[str, str]:
    if not ENABLE_URL_REPAIR or not SERPAPI_KEY:
        return "", ""
    q = f"{headline} {agency}".strip()
    if prefer_pdf:
        q = f"{q} 관보 PDF OR gwanbo"
    try:
        params = {
            "engine": "google",
            "q": q,
            "api_key": SERPAPI_KEY,
            "num": 5,
            "hl": "en",
        }
        if not prefer_pdf:
            params["tbm"] = "nws"
        resp = requests.get("https://serpapi.com/search.json", params=params, timeout=URL_REPAIR_TIMEOUT)
        if resp.status_code != 200:
            return "", f"serpapi_http_{resp.status_code}"
        data = resp.json()
        results = data.get("news_results") or data.get("organic_results") or []
        best_url, best_score = "", 0.0
        for item in results[:8]:
            link = normalize_url(item.get("link") or item.get("url"))
            title = safe_str(item.get("title"))
            source = safe_str(item.get("source"))
            if not link or is_bad_article_url(link):
                continue
            score = similar_title(headline, title)
            if agency and agency.lower() in (source + " " + link).lower():
                score += 0.12
            if prefer_pdf and ("pdf" in link.lower() or "관보" in title):
                score += 0.10
            if score > best_score:
                best_url, best_score = link, score
        if best_url and (best_score >= 0.35 or prefer_pdf):
            return best_url, f"serpapi_score_{best_score:.2f}"
        return "", "serpapi_no_match"
    except Exception as e:
        return "", f"serpapi_error:{type(e).__name__}"


def search_url_by_naver(headline: str, agency: str = "") -> tuple[str, str]:
    if not ENABLE_URL_REPAIR or not (NAVER_CLIENT_ID and NAVER_CLIENT_SECRET):
        return "", ""
    try:
        headers = {
            "X-Naver-Client-Id": NAVER_CLIENT_ID,
            "X-Naver-Client-Secret": NAVER_CLIENT_SECRET,
        }
        q = f"{headline} {agency}".strip()
        resp = requests.get(
            "https://openapi.naver.com/v1/search/news.json",
            headers=headers,
            params={"query": q, "display": 5, "sort": "sim"},
            timeout=URL_REPAIR_TIMEOUT,
        )
        if resp.status_code != 200:
            return "", f"naver_http_{resp.status_code}"
        items = resp.json().get("items", [])
        best_url, best_score = "", 0.0
        for item in items:
            link = normalize_url(item.get("originallink") or item.get("link"))
            title = re.sub(r"<[^>]+>", "", safe_str(item.get("title")))
            if not link or is_bad_article_url(link):
                continue
            score = similar_title(headline, title)
            if score > best_score:
                best_url, best_score = link, score
        if best_url and best_score >= 0.35:
            return best_url, f"naver_score_{best_score:.2f}"
        return "", "naver_no_match"
    except Exception as e:
        return "", f"naver_error:{type(e).__name__}"


def search_url_by_newsapi(headline: str, agency: str = "") -> tuple[str, str]:
    if not ENABLE_URL_REPAIR or not NEWS_API_KEY:
        return "", ""
    try:
        q = f"\"{headline[:120]}\""
        resp = requests.get(
            NEWS_API_ENDPOINT,
            params={"q": q, "language": "en", "sortBy": "relevancy", "pageSize": 5, "apiKey": NEWS_API_KEY},
            timeout=URL_REPAIR_TIMEOUT,
        )
        if resp.status_code != 200:
            return "", f"newsapi_http_{resp.status_code}"
        articles = resp.json().get("articles", [])
        best_url, best_score = "", 0.0
        for item in articles:
            link = normalize_url(item.get("url"))
            title = safe_str(item.get("title"))
            source = safe_str((item.get("source") or {}).get("name"))
            if not link or is_bad_article_url(link):
                continue
            score = similar_title(headline, title)
            if agency and agency.lower() in source.lower():
                score += 0.10
            if score > best_score:
                best_url, best_score = link, score
        if best_url and best_score >= 0.35:
            return best_url, f"newsapi_score_{best_score:.2f}"
        return "", "newsapi_no_match"
    except Exception as e:
        return "", f"newsapi_error:{type(e).__name__}"


def search_gazette_pdf_url(row: pd.Series) -> tuple[str, str]:
    """Find Official Gazette PDF for Korean 법령/관보 style rows when URL/body is weak."""
    title = safe_str(row.get("headline") or row.get("title"))
    text = " ".join([title, safe_str(row.get("agency")), safe_str(row.get("article_body"))[:500]])
    if not contains_any(text, ["관보", "대통령령", "총리령", "부령", "고시", "공고", "시행령", "시행규칙"]):
        return "", "gazette_not_applicable"
    q_title = re.sub(r"\s+", " ", title).strip()
    return search_url_by_serpapi(q_title, "관보 PDF", prefer_pdf=True)


def repair_article_url(row: pd.Series) -> tuple[str, str, str]:
    """Final URL quality gate for STEP4.

    This function repairs thumbnail/RSS URLs using row candidates, article body,
    SerpAPI, Naver News API, and NewsAPI. If all fail, URL is left blank and
    URL_Quality is URL_MISSING_OR_BAD so Top30 selection can exclude it.
    """
    url, quality, source = choose_clean_url(row)
    if url:
        return url, quality, source

    for google_url in google_news_urls_from_row(row):
        resolved = resolve_google_news_url(google_url)
        if resolved and not is_bad_article_url(resolved):
            return resolved, "REPAIRED_GOOGLE_NEWS", "google_news_resolver"

    body_urls = urls_from_text(row.get("article_body"), row.get("summary"), row.get("description"))
    if body_urls:
        u = body_urls[0]
        return u, "REPAIRED_FROM_BODY", "body_url_regex"

    title = safe_str(row.get("headline") or row.get("title"))
    agency = safe_str(row.get("agency") or row.get("publisher") or row.get("source"))
    if not title:
        return "", "URL_MISSING_OR_BAD", "no_title"

    # Korean official notices/gazettes: try PDF search first.
    pdf_url, pdf_reason = search_gazette_pdf_url(row)
    if pdf_url and not is_bad_article_url(pdf_url):
        return pdf_url, "GAZETTE_PDF", f"serpapi_gazette:{pdf_reason}"

    for func, label in [
        (search_url_by_serpapi, "SERPAPI"),
        (search_url_by_naver, "NAVER"),
        (search_url_by_newsapi, "NEWSAPI"),
    ]:
        u, reason = func(title, agency)
        if u and not is_bad_article_url(u):
            return u, f"REPAIRED_{label}", reason
        time.sleep(0.05)

    return "", "URL_MISSING_OR_BAD", "repair_failed"


def fetch_pdf_text(url: str, max_chars: int = 5000) -> tuple[str, str]:
    if not ENABLE_PDF_TEXT_CHECK or not is_pdf_url(url):
        return "", "NOT_PDF"
    try:
        resp = requests.get(url, timeout=URL_REPAIR_TIMEOUT, headers={"User-Agent": "Mozilla/5.0"})
        if resp.status_code != 200 or not resp.content:
            return "", f"PDF_HTTP_{resp.status_code}"
        with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as tmp:
            tmp.write(resp.content)
            tmp_path = tmp.name
        text = ""
        try:
            try:
                from pypdf import PdfReader
            except Exception:
                from PyPDF2 import PdfReader
            reader = PdfReader(tmp_path)
            pages = []
            for page in reader.pages[:5]:
                pages.append(page.extract_text() or "")
            text = "\n".join(pages).strip()
        finally:
            try:
                os.remove(tmp_path)
            except Exception:
                pass
        if text:
            return text[:max_chars], "PDF_TEXT_OK"
        return "", "PDF_TEXT_EMPTY"
    except Exception as e:
        return "", f"PDF_ERROR:{type(e).__name__}"


def enrich_pdf_body(row: pd.Series) -> tuple[str, str]:
    """If row URL is PDF and body is weak, add extracted PDF text."""
    body = safe_str(row.get("article_body"))
    url = safe_str(row.get("url"))
    if not is_pdf_url(url):
        return body, "NOT_PDF"
    if len(body) >= 1000:
        return body, "PDF_BODY_ALREADY_PRESENT"
    text, status = fetch_pdf_text(url)
    if text:
        merged = (body + "\n\n[PDF_TEXT_EXTRACT]\n" + text).strip()
        return merged, status
    return body, status

def similar_title(a: str, b: str) -> float:
    return SequenceMatcher(None, canonical_title(a), canonical_title(b)).ratio()


# ---------------------------------------------------------
# Load and normalize inputs
# ---------------------------------------------------------

def load_input() -> tuple[pd.DataFrame, Path]:
    path = first_existing(INPUT_CANDIDATES)
    if not path:
        raise FileNotFoundError("No STEP4-2 input found: " + " | ".join(str(p) for p in INPUT_CANDIDATES))
    df = pd.read_excel(path)
    log(f"LOAD {path}: {len(df)} rows")
    return df, path


def normalize_columns(raw: pd.DataFrame) -> pd.DataFrame:
    df = pd.DataFrame()
    mapping = {
        "date": ["Date", "date", "published", "Published"],
        "headline": ["Headline", "headline", "title", "Title"],
        "url": ["URL", "url", "link", "Link"],
        "country": ["Country", "country"],
        "agency": ["Agency", "agency", "SourceAgency"],
        "publisher": ["Publisher", "publisher", "source", "Source"],
        "summary": ["Summary", "summary", "ArticleSummary"],
        "keyword_matches": ["KeywordMatches", "keyword_matches", "keywords"],
        "select_reason": ["SelectReason", "select_reason", "FilterReason"],
        "reject_reason": ["RejectReason", "reject_reason"],
        "source": ["Source", "source"],
        "source_file": ["SourceFile", "source_file"],
        "original_url": ["OriginalURLResolved", "OriginalURL", "original_url", "canonical_url"],
        "canonical_url": ["canonical_url", "CanonicalURL"],
        "article_url": ["article_url", "ArticleURL", "article_link", "ArticleLink"],
        "source_url": ["source_url", "SourceURL", "sourceLink", "source_link"],
        "resolved_url": ["resolved_url", "ResolvedURL", "final_url", "FinalURL"],
        "article_body": ["article_body", "ArticleBody", "body", "content"],
        "ClusterHeadlines": ["ClusterHeadlines", "cluster_headlines", "cluster_titles", "RelatedHeadlines"],
        "article_extract_status": ["article_extract_status", "ArticleExtractStatus"],
        "article_source_type": ["article_source_type", "ArticleSourceType"],
        "regulation_related": ["RegulationRelated", "regulation_related"],
        "regulation_transfer_type": ["RegulationTransferType", "regulation_transfer_type"],
        "effective_date_hint": ["effective_date_hint", "EffectiveDateHint"],
        "change_detail_hint": ["change_detail_hint", "ChangeDetailHint"],
        "hs_hint": ["hs_hint", "HsHint", "HS_Hint"],
        "tariff_rate_hint": ["tariff_rate_hint", "TariffRateHint"],
        "score": ["Score", "score"],
    }
    for out_col, candidates in mapping.items():
        c = pick_col(raw, candidates)
        df[out_col] = raw[c] if c in raw.columns else ""

    # Recover URL/title from Excel HYPERLINK formula if present.
    m = df["headline"].astype(str).str.extract(r'=HYPERLINK\("([^"]+)","([^"]+)"\)', expand=True)
    mask = m[0].notna()
    df.loc[mask & df["url"].astype(str).str.strip().eq(""), "url"] = m.loc[mask, 0]
    df.loc[mask, "headline"] = m.loc[mask, 1]

    df["headline"] = df["headline"].astype(str).map(lambda x: re.sub(r"\s+", " ", x).strip())

    url_pick = df.apply(choose_clean_url, axis=1, result_type="expand")
    df["url"] = url_pick[0]
    df["URL_Quality"] = url_pick[1]
    df["URL_Repair_Source"] = url_pick[2]

    # Keep original_url clean as a useful audit field as well.
    df["original_url"] = df["url"].where(
        df["url"].astype(str).str.strip().ne(""),
        df["original_url"].astype(str).str.strip()
    )

    df = df[df["headline"].astype(str).str.strip().ne("")].copy()
    return df.reset_index(drop=True)


# ---------------------------------------------------------
# Topic and subsidiary master
# ---------------------------------------------------------

def load_topic_master() -> list[dict]:
    if not TOPIC_MASTER_INPUT.exists():
        log(f"[TOPIC] Topic master not found: {TOPIC_MASTER_INPUT} -> fallback")
        return FALLBACK_TOPICS
    try:
        df = pd.read_excel(TOPIC_MASTER_INPUT)
    except Exception as e:
        log(f"[TOPIC] Failed to load topic master: {e} -> fallback")
        return FALLBACK_TOPICS

    topic_col = pick_col(df, ["Topic", "topic", "Issue", "issue", "Keyword", "keyword"])
    score_col = pick_col(df, ["Score", "score", "Priority", "priority", "Weight", "weight"])
    keywords_col = pick_col(df, ["Keywords", "keywords", "Keyword", "keyword", "SearchTerms", "search_terms"])
    issue_col = pick_col(df, ["IssueType", "issue_type", "Category", "category"])
    action_col = pick_col(df, ["Standard Action", "StandardAction", "standard_action", "Action", "action"])
    owner_col = pick_col(df, ["Action Owner", "ActionOwner", "Owner", "owner", "ActionOwnerTeam"])

    if not topic_col:
        log("[TOPIC] Topic master has no Topic column -> fallback")
        return FALLBACK_TOPICS

    records: list[dict] = []
    for _, r in df.iterrows():
        topic = safe_str(r.get(topic_col))
        if not topic:
            continue
        score = pd.to_numeric(pd.Series([r.get(score_col) if score_col else 70]), errors="coerce").fillna(70).iloc[0]
        keywords_text = safe_str(r.get(keywords_col)) if keywords_col else topic
        keywords = [x.strip() for x in re.split(r"[;,/|\n]+", keywords_text) if x.strip()]
        if topic not in keywords:
            keywords.append(topic)
        issue_type = safe_str(r.get(issue_col)) if issue_col else infer_issue_type_from_topic(topic)
        records.append({
            "topic": topic,
            "score": int(min(max(score, 0), 100)),
            "keywords": keywords,
            "issue_type": issue_type or infer_issue_type_from_topic(topic),
            "standard_action": safe_str(r.get(action_col)) if action_col else default_action_by_issue(issue_type),
            "action_owner": safe_str(r.get(owner_col)) if owner_col else default_owner_by_issue(issue_type),
        })
    log(f"[TOPIC] Loaded topic master: {TOPIC_MASTER_INPUT} rows={len(records)}")
    return records or FALLBACK_TOPICS


def infer_issue_type_from_topic(topic: str) -> str:
    t = norm(topic)
    if contains_any(t, ["entity", "bis", "export control", "수출통제", "제재"]):
        return "EXPORT_CONTROL"
    if contains_any(t, ["semiconductor", "chip", "반도체"]):
        return "SEMICONDUCTOR_TARIFF"
    if contains_any(t, ["301", "232", "ustr"]):
        return "SECTION_301_232"
    if contains_any(t, ["ad", "cvd", "dump", "반덤핑", "상계"]):
        return "AD_CVD"
    if contains_any(t, ["cbam", "carbon", "탄소"]):
        return "CBAM_CARBON"
    if contains_any(t, ["fta", "origin", "cepa", "원산지"]):
        return "ORIGIN_FTA"
    if contains_any(t, ["hs", "classification", "품목분류"]):
        return "HS_CLASSIFICATION"
    if contains_any(t, ["tariff", "duty", "관세"]):
        return "TARIFF"
    if contains_any(t, ["customs", "통관", "세관"]):
        return "CUSTOMS"
    return "TRADE_GENERAL"


def default_action_by_issue(issue_type: str) -> str:
    return {
        "EXPORT_CONTROL": "ECCN/전략물자 해당 여부와 거래상대방 Screening을 즉시 점검",
        "SEMICONDUCTOR_TARIFF": "반도체 관련 HS Mapping 및 생산/판매법인 관세비용 영향 분석",
        "SECTION_301_232": "미국 추가관세 대상 HS와 수입신고 적용세율 점검",
        "AD_CVD": "AD/CVD 대상국·공급자·HS 기준 적용 여부 점검",
        "CBAM_CARBON": "CBAM 대상품목 및 배출량 증빙 수집 체계 점검",
        "ORIGIN_FTA": "CO 발급/수취 및 원산지 판정근거 재점검",
        "HS_CLASSIFICATION": "HS Mapping Master와 신고 HS 정합성 점검",
        "TARIFF": "HS별 관세율 변경 및 수입원가 영향 분석",
        "CUSTOMS": "통관 신고필드·증빙·관세사 업무지침 변경 여부 확인",
    }.get(issue_type, "직접 영향 가능성이 있는 국가·품목·거래선 여부를 모니터링")


def default_owner_by_issue(issue_type: str) -> str:
    return {
        "EXPORT_CONTROL": "수출통제팀",
        "SEMICONDUCTOR_TARIFF": "Global SCM",
        "SECTION_301_232": "통관운영팀",
        "AD_CVD": "통관운영팀",
        "CBAM_CARBON": "ESG/통상지원팀",
        "ORIGIN_FTA": "FTA팀",
        "HS_CLASSIFICATION": "품목분류/통관운영팀",
        "TARIFF": "통관운영팀",
        "CUSTOMS": "통관운영팀",
    }.get(issue_type, "GTI 운영팀")


def load_subsidiary_master() -> pd.DataFrame:
    if not SUBSIDIARY_MASTER_INPUT.exists():
        log(f"[SUBSIDIARY] Subsidiary master not found: {SUBSIDIARY_MASTER_INPUT}")
        return pd.DataFrame(columns=["Country", "Subsidiary", "Product"])
    try:
        df = pd.read_excel(SUBSIDIARY_MASTER_INPUT)
    except Exception as e:
        log(f"[SUBSIDIARY] Failed to load subsidiary master: {e}")
        return pd.DataFrame(columns=["Country", "Subsidiary", "Product"])
    country_col = pick_col(df, ["Country", "country", "국가"])
    sub_col = pick_col(df, ["Subsidiary", "subsidiary", "법인", "Company", "company"])
    product_col = pick_col(df, ["Product", "Products", "product", "products", "품목", "제품"])
    out = pd.DataFrame()
    out["Country"] = df[country_col] if country_col else ""
    out["Subsidiary"] = df[sub_col] if sub_col else ""
    out["Product"] = df[product_col] if product_col else ""
    out = out[out["Country"].astype(str).str.strip().ne("")]
    log(f"[SUBSIDIARY] Loaded subsidiary master: {SUBSIDIARY_MASTER_INPUT} rows={len(out)}")
    return out.reset_index(drop=True)


TOPIC_MASTER = load_topic_master()
SUBSIDIARY_MASTER = load_subsidiary_master()


def issue_gate_pass(issue_type: str, matched_keywords: list[str], row: pd.Series) -> bool:
    """Issue-specific guardrails to prevent generic false positives."""
    text = full_text(row)
    headline = safe_str(row.get("headline"))
    issue = safe_str(issue_type)
    matched_norm = [norm(x) for x in matched_keywords if safe_str(x)]

    if is_security_non_trade(row):
        return False

    if issue == "AD_CVD":
        return contains_any(text, ["anti-dumping", "antidumping", "countervailing", "ad/cvd", "반덤핑", "덤핑방지관세", "상계관세"])

    if issue == "EXPORT_CONTROL":
        return contains_any(text, [
            "entity list", "bis", "export control", "export controls", "export restriction", "export restrictions",
            "eccn", "ear", "restricted party", "ai chip", "ai chips", "nvidia", "h200", "huawei",
            "chip export", "semiconductor export", "military university", "수출통제", "수출 제한", "수출규제",
            "전략물자", "제재", "ai칩", "엔비디아", "화웨이", "우회로", "해외 자회사", "군사 연계",
        ])

    if issue == "SEMICONDUCTOR_TARIFF":
        return contains_any(text, ["semiconductor", "semiconductors", "chip", "chips", "ai chip", "ai chips", "nvidia", "h200", "반도체", "ai칩", "엔비디아"])

    if issue == "SECTION_301_232":
        return contains_any(text, ["section 301", "section 232", "301조", "232조", "ustr", "trade expansion act", "reciprocity law", "무역법 301"])

    if issue == "ORIGIN_FTA":
        # CEPA can mean both a trade agreement and the think tank "Center for European Policy Analysis".
        # Treat CEPA as FTA only when trade-agreement context exists.
        if contains_any(text, CEPA_FALSE_POSITIVE_TERMS):
            return False
        if "cepa" in matched_norm and not contains_any(text, FTA_CONTEXT_TERMS):
            return False
        # Generic body-only FTA mentions are not enough unless headline or structured hints show agreement/origin context.
        if contains_any(headline, FTA_CONTEXT_TERMS):
            return True
        if safe_str(row.get("change_detail_hint")) and contains_any(row.get("change_detail_hint"), FTA_CONTEXT_TERMS):
            return True
        return contains_any(text, ["rules of origin", "certificate of origin", "origin rule", "tariff concession", "preferential tariff", "원산지", "협정세율", "관세양허"])

    if issue == "HS_CLASSIFICATION":
        return contains_any(text, ["hs code", "hs classification", "tariff classification", "품목분류", "hs코드", "세번"])

    if issue == "TARIFF":
        # A tariff article should have tariff language in the headline OR a more concrete tariff signal in the body.
        return contains_any(headline, ["tariff", "tariffs", "duty", "관세", "관세율", "추가관세", "할당관세"]) or contains_any(text, ["additional tariff", "tariff rate", "import duty", "customs duty", "관세율", "추가관세", "할당관세", "보복 관세"])

    if issue == "CUSTOMS":
        return contains_any(text, ["customs", "clearance", "declaration", "customs audit", "통관", "수입신고", "수출신고", "세관", "관세청"])

    if issue == "CBAM_CARBON":
        return contains_any(text, ["cbam", "carbon border", "carbon border levy", "탄소국경", "탄소국경세"])

    return True


def detect_topic(row: pd.Series) -> tuple[str, int, str, str, str]:
    text = full_text(row)
    headline = safe_str(row.get("headline"))
    ntext = norm(text)
    nhead = norm(headline)

    if is_security_non_trade(row):
        return "REFERENCE", 0, "", "Security/defense non-trade article excluded from GTI topic scoring", "TRADE_GENERAL"

    # High-confidence pre-classification. This prevents CBAM/FTA/AD-CVD
    # articles from being mapped to the wrong action template by generic
    # words inside long article bodies.
    if contains_any(text, ["cbam", "carbon border", "carbon border adjustment", "탄소국경", "탄소국경조정"]):
        return "CBAM", 95, "CBAM; carbon border", "High-confidence CBAM signal", "CBAM_CARBON"
    if contains_any(text, ["entity list", "bureau of industry and security", "bis entity", "export control", "eccn", "수출통제", "전략물자"]):
        return "Export Control", 95, "Entity List/BIS/Export Control", "High-confidence export control signal", "EXPORT_CONTROL"
    if contains_any(text, ["dumping margin", "anti-dumping", "antidumping", "countervailing", "ad/cvd", "덤핑방지관세", "반덤핑", "상계관세"]):
        return "AD/CVD", 92, "AD/CVD", "High-confidence AD/CVD signal", "AD_CVD"
    # CEPA ambiguity guard: Center for European Policy Analysis is not FTA/CEPA.
    if (contains_any(text, ["fta", "free trade agreement", "cepa", "rules of origin", "certificate of origin", "원산지", "협정세율"])
        and not contains_any(text, ["center for european policy analysis"])):
        return "원산지/CO", 90, "FTA/CEPA/Origin", "High-confidence FTA/Origin signal", "ORIGIN_FTA"

    best = None
    for rec in TOPIC_MASTER:
        raw_keywords = rec.get("keywords", [])
        matched = [kw for kw in raw_keywords if keyword_in_text(kw, text)]
        if not matched:
            continue
        issue_type = rec.get("issue_type", "TRADE_GENERAL")
        if not issue_gate_pass(issue_type, matched, row):
            continue

        # Weight headline matches much higher than body-only generic matches.
        headline_matches = [kw for kw in matched if keyword_in_text(kw, headline)]
        base = int(rec.get("score", 70))
        score = base + min(len(matched) * 2, 8) + min(len(headline_matches) * 6, 12)

        # Penalize generic body-only tariff/customs matches; they often appear in unrelated articles.
        if not headline_matches and issue_type in {"TARIFF", "CUSTOMS", "AD_CVD"}:
            score -= 25
        if issue_type in {"TARIFF", "CUSTOMS"} and not has_article_product_signal(row) and not safe_str(row.get("hs_hint")):
            score -= 10

        score = max(0, min(score, 100))
        candidate = (
            rec.get("topic", "REFERENCE"),
            score,
            "; ".join(matched[:5]),
            f"GTI topic matched: {'; '.join(matched[:3])}; headline_match={bool(headline_matches)}",
            issue_type,
        )
        if best is None or candidate[1] > best[1]:
            best = candidate

    if best and best[1] >= 35:
        return best
    return "REFERENCE", 0, "", "No actionable GTI topic after guardrails", "TRADE_GENERAL"


COUNTRY_ALIASES = {
    "USA": ["usa", "u.s.", "united states", "미국", "washington", "ustr", "bis"],
    "United States": ["usa", "u.s.", "united states", "미국", "washington", "ustr", "bis"],
    "China": ["china", "chinese", "중국"],
    "Vietnam": ["vietnam", "viet nam", "베트남"],
    "India": ["india", "indian", "인도"],
    "Mexico": ["mexico", "멕시코"],
    "Poland": ["poland", "폴란드"],
    "Brazil": ["brazil", "브라질"],
    "Korea": ["korea", "south korea", "한국", "대한민국"],
    "EU": ["eu", "european union", "european commission", "유럽연합"],
    "Oman": ["oman", "오만"],
}


def detect_country(row: pd.Series) -> str:
    current = safe_str(row.get("country"))
    text = norm(full_text(row))
    found = []
    if current:
        found.append(current)
    for country, aliases in COUNTRY_ALIASES.items():
        if any(norm(a) in text for a in aliases):
            found.append(country)
    # stable unique
    unique = []
    for x in found:
        if x and x not in unique:
            unique.append(x)
    return "; ".join(unique[:4]) if unique else "Global"


def country_matches(master_country: str, article_country: str, text: str) -> bool:
    mc = safe_str(master_country)
    ac = safe_str(article_country)
    if not mc:
        return False
    if norm(mc) in norm(ac) or norm(mc) in norm(text):
        return True
    aliases = COUNTRY_ALIASES.get(mc, [])
    return any(norm(a) in norm(ac) or norm(a) in norm(text) for a in aliases)


def detect_affected_subsidiaries(row: pd.Series) -> tuple[str, str, str, str]:
    if SUBSIDIARY_MASTER.empty:
        return MISSING_TEXT, derive_affected_products(row), MISSING_TEXT, "SUBSIDIARY_MASTER 없음"
    text = full_text(row)
    country = safe_str(row.get("country")) or detect_country(row)
    hits = []
    for _, r in SUBSIDIARY_MASTER.iterrows():
        if country_matches(r.get("Country"), country, text):
            hits.append(r)
    if not hits:
        return MISSING_TEXT, derive_affected_products(row), MISSING_TEXT, "영향 법인 매핑 없음"
    subs = []
    products = []
    for r in hits:
        sub = safe_str(r.get("Subsidiary"))
        prod = safe_str(r.get("Product"))
        if sub and sub not in subs:
            subs.append(sub)
        if prod and prod not in products:
            products.append(prod)
    affected_products = derive_affected_products(row) or "; ".join(products[:6]) or MISSING_TEXT
    return "; ".join(subs[:10]) or MISSING_TEXT, affected_products, "; ".join(products[:10]) or MISSING_TEXT, f"SUBSIDIARY_MASTER country match: {country}"

# Subsidiary impact axis evaluation for Step5 executive summary.
# The current SUBSIDIARY_MASTER has Country/Subsidiary/Product only, so production/sales
# classification is inferred by subsidiary code and product context.  This keeps Step4
# independent from ERP while still producing decision-grade GTI fields.
PRODUCTION_SUBSIDIARY_HINTS = {"SAS", "SEV", "SEVT", "SIEL", "SEIN", "SEHC", "SEHZ", "SEH", "SEPM", "SESS"}
SALES_SUBSIDIARY_HINTS = {"SEA", "SEUK", "SEG", "SEF", "SEI", "SECA", "SELA", "SEDA", "SEBN"}

def _split_subsidiaries(subs: str) -> list[str]:
    return [x.strip() for x in re.split(r"[;,/\n]+", safe_str(subs)) if x.strip() and x.strip() != MISSING_TEXT]

def classify_subsidiary_axis(subs: str) -> tuple[str, str]:
    prod, sales = [], []
    for s in _split_subsidiaries(subs):
        code = s.upper()
        if code in PRODUCTION_SUBSIDIARY_HINTS:
            prod.append(s)
        elif code in SALES_SUBSIDIARY_HINTS:
            sales.append(s)
        else:
            # Conservative fallback: keep unknown as sales/market-facing if not known production.
            sales.append(s)
    return "; ".join(dict.fromkeys(prod).keys()) or MISSING_TEXT, "; ".join(dict.fromkeys(sales).keys()) or MISSING_TEXT

def evaluate_impact_axes(row: pd.Series) -> tuple[str, str, str, str, str, str, str]:
    issue = safe_str(row.get("issue_type"))
    text = full_text(row)
    subs = safe_str(row.get("affected_subsidiaries"))
    products = safe_str(row.get("affected_products")) or safe_str(row.get("subsidiary_products")) or derive_affected_products(row)
    prod_subs, sales_subs = classify_subsidiary_axis(subs)

    fta_impact = "Y" if issue == "ORIGIN_FTA" or contains_any(text, ["fta", "cepa", "epa", "rules of origin", "certificate of origin", "원산지", "협정세율", "co 발급"]) else "N"
    export_control_impact = "Y" if issue in {"EXPORT_CONTROL", "SEMICONDUCTOR_TARIFF"} or contains_any(text, ["entity list", "bis", "export control", "eccn", "ear", "수출통제", "전략물자", "최종사용자", "우회수출"]) else "N"
    hs_impact = "Y" if issue in {"HS_CLASSIFICATION", "TARIFF", "AD_CVD", "SECTION_301_232"} or safe_str(row.get("hs_hint")) or contains_any(text, ["hs code", "품목분류", "세번"]) else "N"
    tariff_impact = "Y" if issue in {"TARIFF", "AD_CVD", "SECTION_301_232", "SEMICONDUCTOR_TARIFF"} or safe_str(row.get("tariff_rate_hint")) or contains_any(text, ["tariff", "duty", "관세", "관세율", "덤핑방지관세", "상계관세"]) else "N"

    impact_products = products if products and products != MISSING_TEXT else (safe_str(row.get("subsidiary_products")) or MISSING_TEXT)
    return prod_subs, sales_subs, impact_products, fta_impact, export_control_impact, hs_impact, tariff_impact

def samsung_impact_axis_score(row: pd.Series) -> int:
    score = 0
    if safe_str(row.get("impact_production_subsidiaries")) != MISSING_TEXT:
        score += 25
    if safe_str(row.get("impact_sales_subsidiaries")) != MISSING_TEXT:
        score += 20
    if safe_str(row.get("impact_products")) != MISSING_TEXT:
        score += 20
    if safe_str(row.get("fta_impact")) == "Y":
        score += 15
    if safe_str(row.get("export_control_impact")) == "Y":
        score += 20
    if safe_str(row.get("hs_impact")) == "Y":
        score += 10
    if safe_str(row.get("tariff_impact")) == "Y":
        score += 10
    return min(score, 100)


def derive_affected_products(row: pd.Series) -> str:
    hints = [safe_str(row.get("change_detail_hint")), safe_str(row.get("hs_hint")), safe_str(row.get("tariff_rate_hint"))]
    text = full_text(row)
    product_hits = []
    for p in PRODUCT_TERMS:
        if norm(p) in norm(text) and p not in product_hits:
            product_hits.append(p)
    result = "; ".join([x for x in hints if x][:2] + product_hits[:5])
    return truncate(result, 300)


# ---------------------------------------------------------
# Scoring
# ---------------------------------------------------------

def is_noise(row: pd.Series) -> str:
    text = full_text(row)
    headline = safe_str(row.get("headline"))
    issue = safe_str(row.get("issue_type"))
    topic_score = int(row.get("topic_score", 0) or 0)

    if is_security_non_trade(row):
        return "SECURITY_DEFENSE_NON_TRADE"

    # Explicitly exclude domestic politics, agriculture, ceremonies and education unless
    # a strong GTI policy topic is present in the headline.
    if contains_any(headline, ["농지", "농협", "민생", "국정", "대통령", "장관", "전수조사", "개혁", "교실", "세미나", "포럼", "교육", "수상", "행사"]):
        if not contains_any(headline, TRADE_ACTION_SIGNALS):
            return "LOW_VALUE_DOMESTIC_OR_EVENT_NOISE"

    if contains_any(text, NONE_NOISE_TERMS):
        # Keep if strongly matched to GTI topic and product/policy impact.
        if topic_score >= 82 and (contains_any(text, PRODUCT_TERMS) or issue in DIRECT_POLICY_ISSUES):
            return ""
        return "LOW_VALUE_NOISE"

    if topic_score == 0 and not contains_any(text, PRODUCT_TERMS):
        return "NO_TOPIC_NO_PRODUCT_SIGNAL"

    # FTA false positives from think tanks or general policy commentary.
    if issue == "ORIGIN_FTA" and contains_any(text, CEPA_FALSE_POSITIVE_TERMS):
        return "CEPA_THINKTANK_FALSE_POSITIVE"

    return ""


def samsung_impact(row: pd.Series) -> str:
    """Re-evaluate Samsung impact by subsidiary/product/FTA/export-control axes.

    Direct is not a simple country match.  It requires at least one mapped
    production/sales subsidiary and a concrete product, FTA, HS/tariff, or
    export-control impact axis.  Indirect is used for actionable global policy
    with no mapped subsidiary or weaker product signal.  Reference is used for
    watch-list items that are trade-related but not executable for Samsung.
    """
    if is_noise(row) or is_security_non_trade(row):
        return "Reference"
    topic_score = int(row.get("topic_score", 0) or 0)
    issue = safe_str(row.get("issue_type"))
    if topic_score == 0:
        return "Reference"

    prod_subs = safe_str(row.get("impact_production_subsidiaries"))
    sales_subs = safe_str(row.get("impact_sales_subsidiaries"))
    product = safe_str(row.get("impact_products"))
    has_sub_axis = prod_subs != MISSING_TEXT or sales_subs != MISSING_TEXT
    has_product_axis = product != MISSING_TEXT or bool(safe_str(row.get("hs_hint"))) or has_article_product_signal(row)
    has_policy_axis = any(safe_str(row.get(c)) == "Y" for c in ["fta_impact", "export_control_impact", "hs_impact", "tariff_impact"])
    actionable_policy = issue in DIRECT_POLICY_ISSUES and topic_score >= 70

    if has_sub_axis and actionable_policy and (has_product_axis or has_policy_axis):
        return "Direct"
    if actionable_policy and (has_policy_axis or has_article_product_signal(row)):
        return "Indirect"
    if topic_score >= 75 and has_policy_axis:
        return "Indirect"
    return "Reference"



def calculate_subsidiary_score(row: pd.Series) -> int:
    """Score Samsung subsidiary relevance from SUBSIDIARY_MASTER and impact axes.

    This is the 30% component of GTI95 FinalScore.  It is intentionally
    separate from samsung_impact_score so the Excel output shows exactly how
    the affected subsidiary axis contributed to ranking.
    """
    subs = safe_str(row.get("affected_subsidiaries"))
    prod_subs = safe_str(row.get("impact_production_subsidiaries"))
    sales_subs = safe_str(row.get("impact_sales_subsidiaries"))
    products = safe_str(row.get("impact_products"))
    issue = safe_str(row.get("issue_type"))
    text = full_text(row)

    has_subs = subs not in {"", MISSING_TEXT}
    has_prod_subs = prod_subs not in {"", MISSING_TEXT}
    has_sales_subs = sales_subs not in {"", MISSING_TEXT}
    has_product = products not in {"", MISSING_TEXT} or has_article_product_signal(row) or bool(safe_str(row.get("hs_hint")))
    has_policy_axis = any(safe_str(row.get(c)) == "Y" for c in ["fta_impact", "export_control_impact", "hs_impact", "tariff_impact"])

    if has_subs and (has_product or has_policy_axis):
        score = 90
    elif has_subs:
        score = 70
    elif issue in DIRECT_POLICY_ISSUES and (has_product or has_policy_axis):
        score = 45
    elif contains_any(text, PRODUCT_TERMS):
        score = 30
    else:
        score = 0

    if has_prod_subs and has_sales_subs:
        score += 10
    elif has_prod_subs or has_sales_subs:
        score += 5

    return int(max(0, min(score, 100)))

def samsung_impact_score(row: pd.Series) -> int:
    impact = safe_str(row.get("samsung_impact"))
    axis_score = samsung_impact_axis_score(row)
    issue = safe_str(row.get("issue_type"))
    if impact == "Direct":
        base = 70
    elif impact == "Indirect":
        base = 45
    else:
        base = 10
    if issue in {"EXPORT_CONTROL", "SEMICONDUCTOR_TARIFF", "SECTION_301_232"}:
        base += 10
    if issue in {"ORIGIN_FTA", "AD_CVD", "CBAM_CARBON", "HS_CLASSIFICATION", "TARIFF"}:
        base += 5
    return min(100, int(base + axis_score * 0.35))

def urgency_score(row: pd.Series) -> int:
    text = full_text(row)
    if safe_str(row.get("effective_date_hint")) or contains_any(text, URGENCY_HIGH):
        return 100
    if safe_str(row.get("tariff_rate_hint")) or contains_any(text, URGENCY_MED):
        return 70
    if contains_any(text, URGENCY_LOW):
        return 40
    return 20


def required_action(row: pd.Series) -> tuple[str, str]:
    """Issue-specific action template for Step5 Today's Required Actions.

    This intentionally overrides generic topic-master actions for high-risk
    topics so CBAM never receives Entity List/ECCN actions, FTA never receives
    CBAM actions, etc.
    """
    issue = safe_str(row.get("issue_type"))
    topic = safe_str(row.get("topic"))
    subs = safe_str(row.get("affected_subsidiaries"))
    hs = safe_str(row.get("hs_hint"))
    tariff = safe_str(row.get("tariff_rate_hint"))

    if issue == "CBAM_CARBON":
        return "EU CBAM 대상 품목 여부 확인; 탄소배출 데이터 확보; 공급사 증빙 수집; EU향 신고/보고 영향 점검", "ESG/통상지원팀"
    if issue == "EXPORT_CONTROL":
        return "ECCN/전략물자 해당 여부 재점검; 거래상대방·최종사용자 Screening; 우회수출 및 해외자회사 거래 통제 확인", "수출통제팀"
    if issue == "ORIGIN_FTA":
        return "원산지 기준 검토; BOM 충족 여부 확인; CO 발급·수취·보관 증빙 점검; 협정세율 적용 대상 재점검", "FTA팀"
    if issue == "AD_CVD":
        return "대상국·공급자·HS별 AD/CVD 적용 여부 확인; 신고세율 및 소급 리스크 점검; 관세사 신고 가이드 배포", "통관운영팀"
    if issue == "HS_CLASSIFICATION":
        return "HS Mapping Master 정비; 신고 HS와 품목분류 기준 정합성 점검; 국가별 HS 불일치 후보 추출", "품목분류/통관운영팀"
    if issue in {"TARIFF", "SECTION_301_232"}:
        detail = []
        if hs:
            detail.append(f"HS {hs} 영향 확인")
        if tariff:
            detail.append(f"관세율 {tariff} 반영 여부 점검")
        prefix = "; ".join(detail) + "; " if detail else ""
        return prefix + "대상 국가·HS·제품군별 관세율 변경 및 수입원가 영향 분석; 관세 Master 반영", "통관운영팀"
    if issue == "SEMICONDUCTOR_TARIFF":
        return "반도체 관련 HS Mapping 확인; 생산/판매법인 관세원가 영향 분석; 공급망·가격전가 시나리오 점검", "Global SCM"
    if issue == "CUSTOMS":
        return "수입·수출 신고필드와 증빙 요건 확인; 관세사 업무지침 업데이트; 신고 체크리스트 반영", "통관운영팀"

    # Fallback to topic master only for non-core trade general topics.
    for rec in TOPIC_MASTER:
        if safe_str(rec.get("topic")) == topic:
            action = safe_str(rec.get("standard_action")) or default_action_by_issue(issue)
            owner = safe_str(rec.get("action_owner")) or default_owner_by_issue(issue)
            return action, owner
    return default_action_by_issue(issue), default_owner_by_issue(issue)

def action_score(row: pd.Series) -> int:
    impact = safe_str(row.get("samsung_impact"))
    topic_score = int(row.get("topic_score", 0) or 0)
    urgency = int(row.get("urgency_score", 0) or 0)
    action = safe_str(row.get("RequiredAction"))
    if impact == "None" or topic_score == 0:
        return 0
    score = 60
    if action:
        score += 20
    if impact == "Direct":
        score += 10
    if urgency >= 70:
        score += 10
    return min(score, 100)


def calculate_final_score(row: pd.Series) -> float:
    topic = float(row.get("topic_score", 0) or 0)
    samsung = float(row.get("subsidiary_score", row.get("samsung_impact_score", 0)) or 0)
    action = float(row.get("action_score", 0) or 0)
    urgency = float(row.get("urgency_score", 0) or 0)
    return round(topic * 0.40 + samsung * 0.30 + action * 0.20 + urgency * 0.10, 1)


def risk_level(row: pd.Series) -> str:
    score = float(row.get("final_score", 0) or 0)
    impact = safe_str(row.get("samsung_impact"))
    urgency = int(row.get("urgency_score", 0) or 0)
    if score >= 80 or (impact == "Direct" and urgency >= 70):
        return "상"
    if score >= 60:
        return "중"
    return "하"


def priority_group(row: pd.Series) -> str:
    score = float(row.get("final_score", 0) or 0)
    impact = safe_str(row.get("samsung_impact"))
    if impact in {"None", "Reference"}:
        if score >= 65 and int(row.get("topic_score", 0) or 0) >= 75:
            return "REFERENCE"
        return "NOISE"
    if score < 35:
        return "NOISE"
    if score >= 80 or impact == "Direct":
        return "CORE"
    if score >= 60:
        return "USABLE"
    return "REFERENCE"


# ---------------------------------------------------------
# Message generation
# ---------------------------------------------------------

def make_executive_message(row: pd.Series) -> str:
    issue = safe_str(row.get("issue_type"))
    topic = safe_str(row.get("topic"))
    prod_subs = safe_str(row.get("impact_production_subsidiaries"))
    sales_subs = safe_str(row.get("impact_sales_subsidiaries"))
    products = safe_str(row.get("impact_products")) or safe_str(row.get("affected_products"))
    action = safe_str(row.get("RequiredAction"))
    owner = safe_str(row.get("ActionOwner"))
    impact = safe_str(row.get("samsung_impact"))
    country = safe_str(row.get("country"))

    axes = []
    if prod_subs and prod_subs != MISSING_TEXT:
        axes.append(f"생산법인 {prod_subs}")
    if sales_subs and sales_subs != MISSING_TEXT:
        axes.append(f"판매법인 {sales_subs}")
    if products and products != MISSING_TEXT:
        axes.append(f"품목 {products}")
    if safe_str(row.get("fta_impact")) == "Y":
        axes.append("FTA/원산지")
    if safe_str(row.get("export_control_impact")) == "Y":
        axes.append("수출통제")
    if safe_str(row.get("hs_impact")) == "Y":
        axes.append("HS")
    if safe_str(row.get("tariff_impact")) == "Y":
        axes.append("관세율")
    axis_text = "; ".join(axes) if axes else "삼성 직접 영향 제한"

    if impact == "Direct":
        return f"{country} {topic} 이슈는 {axis_text}에 직접 연결됩니다. {owner} 주관으로 {action} 조치가 필요합니다."
    if impact == "Indirect":
        return f"{country} {topic} 정책 변화는 {axis_text} 관점에서 간접 영향 가능성이 있습니다. {owner} 중심 모니터링 및 사전점검이 필요합니다."
    return f"{topic} 관련 참고 이슈이나 현재 생산법인·판매법인·품목·FTA·수출통제 축에서 직접 실행항목은 제한적입니다."

def make_summary(row: pd.Series) -> str:
    topic = safe_str(row.get("topic"))
    impact = safe_str(row.get("samsung_impact"))
    country = safe_str(row.get("country"))
    axes = []
    for label, col in [("생산법인", "impact_production_subsidiaries"), ("판매법인", "impact_sales_subsidiaries"), ("품목", "impact_products")]:
        val = safe_str(row.get(col))
        if val and val != MISSING_TEXT:
            axes.append(f"{label}: {val}")
    if safe_str(row.get("fta_impact")) == "Y": axes.append("FTA/원산지 영향")
    if safe_str(row.get("export_control_impact")) == "Y": axes.append("수출통제 영향")
    if safe_str(row.get("hs_impact")) == "Y": axes.append("HS 영향")
    if safe_str(row.get("tariff_impact")) == "Y": axes.append("관세율 영향")
    axis_text = "; ".join(axes) if axes else "직접 영향축 제한"
    return f"{country} {topic} 이슈입니다. Samsung Impact={impact}, FinalScore={safe_str(row.get('final_score'))}. 판단축: {axis_text}."

def make_ai_analysis(row: pd.Series) -> str:
    prod_subs = safe_str(row.get("impact_production_subsidiaries")) or MISSING_TEXT
    sales_subs = safe_str(row.get("impact_sales_subsidiaries")) or MISSING_TEXT
    products = safe_str(row.get("impact_products")) or safe_str(row.get("affected_products")) or MISSING_TEXT
    issue = safe_str(row.get("issue_type"))
    evidence = []
    if safe_str(row.get("change_detail_hint")):
        evidence.append(f"변경단서: {truncate(row.get('change_detail_hint'), 180)}")
    if safe_str(row.get("hs_hint")):
        evidence.append(f"HS: {safe_str(row.get('hs_hint'))}")
    if safe_str(row.get("tariff_rate_hint")):
        evidence.append(f"관세율: {safe_str(row.get('tariff_rate_hint'))}")
    if safe_str(row.get("ClusterHeadlines")):
        evidence.append(f"클러스터: {truncate(row.get('ClusterHeadlines'), 220)}")
    if safe_str(row.get("article_body")):
        evidence.append(f"본문근거: {truncate(row.get('article_body'), 240)}")
    evidence_text = " / ".join(evidence[:4])

    axis_text = f"생산법인: {prod_subs}; 판매법인: {sales_subs}; 영향품목: {products}; FTA={safe_str(row.get('fta_impact'))}; 수출통제={safe_str(row.get('export_control_impact'))}; HS={safe_str(row.get('hs_impact'))}; 관세={safe_str(row.get('tariff_impact'))}."
    if issue == "ORIGIN_FTA":
        core = "FTA CO 발급/수취, 원산지 판정기준, BOM 충족, 협정세율 적용 가능성을 우선 검토해야 합니다."
    elif issue in {"TARIFF", "SEMICONDUCTOR_TARIFF", "SECTION_301_232"}:
        core = "HS별 관세율 변경, 추가관세 적용, 수입원가 및 가격전가 영향을 분석해야 합니다."
    elif issue == "EXPORT_CONTROL":
        core = "ECCN/전략물자 해당 여부, 거래상대방 및 최종사용자 Screening, 우회수출·해외자회사 통제가 필요합니다."
    elif issue == "CBAM_CARBON":
        core = "EU CBAM 대상 품목 여부, 탄소배출 데이터, 공급사 증빙 및 신고/보고 영향 점검이 필요합니다."
    elif issue == "AD_CVD":
        core = "대상국·공급자·HS 기준으로 반덤핑/상계관세 적용 여부와 소급 리스크를 점검해야 합니다."
    elif issue == "HS_CLASSIFICATION":
        core = "HS Mapping Master와 신고 HS 정합성, 국가별 HS 불일치 후보를 점검해야 합니다."
    else:
        core = "관세·통관·원산지·수출통제 관점에서 삼성 영향 여부를 점검해야 합니다."
    return f"{axis_text} {core}" + (f" 근거: {evidence_text}" if evidence_text else "")


# ---------------------------------------------------------
# Candidate building and selection
# ---------------------------------------------------------

def cluster_key(row: pd.Series) -> str:
    text = norm(full_text(row))
    topic = safe_str(row.get("topic")) or "REFERENCE"
    if "semiconductor" in text or "chip tariff" in text or "반도체" in text:
        return "SEMICONDUCTOR_TARIFF"
    if ("oman" in text or "오만" in text) and ("cepa" in text or "origin" in text or "원산지" in text):
        return "OMAN_CEPA_ORIGIN"
    if "india" in text and ("cepa" in text or "fta" in text or "tariff" in text):
        return "INDIA_FTA_TARIFF"
    if "entity list" in text or "bis" in text:
        return "US_BIS_ENTITY_LIST"
    title_key = canonical_title(row.get("headline"))[:90]
    return f"{topic}:{title_key}"


def build_candidates(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    # STEP4 final URL quality gate: repair thumbnail/RSS URLs before scoring/selection.
    url_repair = out.apply(repair_article_url, axis=1, result_type="expand")
    out["url"] = url_repair[0]
    out["URL_Quality"] = url_repair[1]
    out["URL_Repair_Source"] = url_repair[2]
    pdf_enrich = out.apply(enrich_pdf_body, axis=1, result_type="expand")
    out["article_body"] = pdf_enrich[0]
    out["PDF_Text_Status"] = pdf_enrich[1]
    out["country"] = out.apply(detect_country, axis=1)
    topic = out.apply(detect_topic, axis=1, result_type="expand")
    out["topic"] = topic[0]
    out["topic_score"] = pd.to_numeric(topic[1], errors="coerce").fillna(0).astype(int)
    out["topic_keyword"] = topic[2]
    out["topic_reason"] = topic[3]
    out["issue_type"] = topic[4]

    subs = out.apply(detect_affected_subsidiaries, axis=1, result_type="expand")
    out["affected_subsidiaries"] = subs[0]
    out["affected_subsidiary"] = subs[0]
    out["affected_products"] = subs[1]
    out["subsidiary_products"] = subs[2]
    out["subsidiary_reason"] = subs[3]

    impact_axes = out.apply(evaluate_impact_axes, axis=1, result_type="expand")
    out["impact_production_subsidiaries"] = impact_axes[0]
    out["impact_sales_subsidiaries"] = impact_axes[1]
    out["impact_products"] = impact_axes[2]
    out["fta_impact"] = impact_axes[3]
    out["export_control_impact"] = impact_axes[4]
    out["hs_impact"] = impact_axes[5]
    out["tariff_impact"] = impact_axes[6]

    actions = out.apply(required_action, axis=1, result_type="expand")
    out["RequiredAction"] = actions[0]
    out["ActionOwner"] = actions[1]

    out["samsung_impact"] = out.apply(samsung_impact, axis=1)
    out["samsung_impact_score"] = out.apply(samsung_impact_score, axis=1)
    out["subsidiary_score"] = out.apply(calculate_subsidiary_score, axis=1)
    out["urgency_score"] = out.apply(urgency_score, axis=1)
    out["action_score"] = out.apply(action_score, axis=1)
    out["final_score"] = out.apply(calculate_final_score, axis=1)
    out["Risk"] = out.apply(risk_level, axis=1)
    out["priority_group"] = out.apply(priority_group, axis=1)
    out["cluster_key"] = out.apply(cluster_key, axis=1)
    out["base_score"] = out["topic_score"]
    out["samsung_score"] = out["samsung_impact_score"]
    out["samsung_reason"] = out.apply(lambda r: f"Impact={r.get('samsung_impact')}; Subsidiary={r.get('affected_subsidiaries')}; Product={r.get('affected_products')}", axis=1)
    out["ExecutiveMessage"] = out.apply(make_executive_message, axis=1)
    out["Summary"] = out.apply(make_summary, axis=1)
    out["AI Analysis"] = out.apply(make_ai_analysis, axis=1)
    out["Action Plan"] = out["RequiredAction"]
    out["mail_section"] = "주요뉴스"
    if "RegulationRelated" in out.columns:
        reg_mask = (
            out["RegulationRelated"]
            .fillna("")
            .astype(str)
            .str.upper()
            .eq("Y")
        )
        out.loc[reg_mask, "mail_section"] = "법규 관련 주요뉴스"
    out["selected"] = "N"
    out["audit_reason"] = out.apply(is_noise, axis=1)
    out.loc[out["priority_group"].ne("NOISE") & out["audit_reason"].eq(""), "audit_reason"] = "GTI95_candidate"
    out["audit_decision"] = out["priority_group"].map({"NOISE": "EXCLUDED"}).fillna("CANDIDATE")
    out["source_cap_reason"] = ""
    out["last_checked"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    return out


def issue_bucket(row_or_value) -> str:
    """Normalize various STEP3/STEP4 topic labels into portfolio buckets."""
    if isinstance(row_or_value, pd.Series):
        raw = " ".join([
            safe_str(row_or_value.get("issue_type")),
            safe_str(row_or_value.get("topic")),
            safe_str(row_or_value.get("topic_keyword")),
        ])
    else:
        raw = safe_str(row_or_value)
    t = norm(raw)
    if contains_any(t, ["export control", "entity list", "bis", "uflpa", "ear", "sanction", "수출통제", "전략물자"]):
        return "EXPORT_CONTROL"
    if contains_any(t, ["cbam", "carbon border", "탄소국경"]):
        return "CBAM"
    if contains_any(t, ["origin", "fta", "cepa", "epa", "rules of origin", "원산지", "co 발급", "협정세율"]):
        return "ORIGIN_FTA"
    if contains_any(t, ["hs classification", "hs_classification", "hs code", "품목분류", "tariff classification"]):
        return "HS_CLASSIFICATION"
    if contains_any(t, ["ad/cvd", "ad_cvd", "anti dumping", "anti-dumping", "antidumping", "countervailing", "덤핑방지", "반덤핑", "상계관세"]):
        return "AD_CVD"
    if contains_any(t, ["section 301", "section 232", "semiconductor tariff", "tariff", "duty", "관세", "관세율", "할당관세"]):
        return "TARIFF"
    if contains_any(t, ["customs", "clearance", "통관", "세관", "수입신고", "수출신고"]):
        return "CUSTOMS"
    return "REFERENCE" if contains_any(t, ["reference", "trade general", "general"]) else "OTHER" 


# Executive-mail portfolio policy.  STEP3 can have many tariff articles, but
# STEP4 Top30 must not be dominated by TARIFF.  These values are intentionally
# configurable for daily tuning.
TOPIC_MAX_CAP = {
    "TARIFF": int(os.getenv("GTI_MAX_TARIFF", "12")),
    "EXPORT_CONTROL": int(os.getenv("GTI_MAX_EXPORT_CONTROL", "8")),
    "ORIGIN_FTA": int(os.getenv("GTI_MAX_ORIGIN_FTA", "6")),
    "HS_CLASSIFICATION": int(os.getenv("GTI_MAX_HS", "5")),
    "AD_CVD": int(os.getenv("GTI_MAX_AD_CVD", "5")),
    "CBAM": int(os.getenv("GTI_MAX_CBAM", "4")),
    "CUSTOMS": int(os.getenv("GTI_MAX_CUSTOMS", "5")),
    "REFERENCE": int(os.getenv("GTI_MAX_REFERENCE_TOPIC", "4")),
    "OTHER": int(os.getenv("GTI_MAX_OTHER", "3")),
}

TOPIC_MIN_TARGET = {
    "EXPORT_CONTROL": int(os.getenv("GTI_MIN_EXPORT_CONTROL", "5")),
    "ORIGIN_FTA": int(os.getenv("GTI_MIN_ORIGIN_FTA", "4")),
    "HS_CLASSIFICATION": int(os.getenv("GTI_MIN_HS", "3")),
    "AD_CVD": int(os.getenv("GTI_MIN_AD_CVD", "2")),
    "CBAM": int(os.getenv("GTI_MIN_CBAM", "2")),
    "CUSTOMS": int(os.getenv("GTI_MIN_CUSTOMS", "2")),
    "TARIFF": int(os.getenv("GTI_MIN_TARIFF", "6")),
}


def evidence_score(row: pd.Series) -> int:
    """Prefer rows with real article body, cluster evidence and STEP3 hints."""
    body_len = len(safe_str(row.get("article_body")))
    cluster_len = len(safe_str(row.get("ClusterHeadlines")))
    change_len = len(safe_str(row.get("change_detail_hint")))
    hint_len = len(safe_str(row.get("hs_hint"))) + len(safe_str(row.get("tariff_rate_hint")))
    score = 0
    if body_len >= 800:
        score += 18
    elif body_len >= 300:
        score += 10
    elif body_len >= 120:
        score += 5
    if cluster_len >= 120:
        score += 10
    elif cluster_len >= 40:
        score += 5
    if change_len >= 80:
        score += 10
    elif change_len >= 20:
        score += 5
    if hint_len > 0:
        score += 5
    return min(score, 35)


def cluster_size_value(row: pd.Series) -> int:
    for col in ["ClusterSize", "cluster_size", "RelatedCount", "related_count", "IssueMergeCount", "issue_merge_count"]:
        if col in row.index:
            try:
                return int(float(row.get(col) or 0))
            except Exception:
                pass
    # Fallback: count separators in ClusterHeadlines.
    ch = safe_str(row.get("ClusterHeadlines"))
    if ch:
        return min(10, max(1, ch.count(";") + ch.count("|") + ch.count("\n") + 1))
    return 1


def rank_candidates(df: pd.DataFrame) -> pd.DataFrame:
    tier_order = {"CORE": 0, "USABLE": 1, "REFERENCE": 2, "NOISE": 9}
    impact_order = {"Direct": 0, "Indirect": 1, "Reference": 2, "None": 9}
    out = df.copy()
    out["topic_bucket"] = out.apply(issue_bucket, axis=1)
    out["evidence_score"] = out.apply(evidence_score, axis=1)
    out["cluster_size"] = out.apply(cluster_size_value, axis=1)
    out["cluster_score"] = out["cluster_size"].clip(lower=1, upper=10) * 2
    out["portfolio_score"] = (
        pd.to_numeric(out.get("final_score", 0), errors="coerce").fillna(0)
        + pd.to_numeric(out["evidence_score"], errors="coerce").fillna(0)
        + pd.to_numeric(out["cluster_score"], errors="coerce").fillna(0)
    ).round(1)
    out["_tier_order"] = out["priority_group"].map(tier_order).fillna(9)
    out["_impact_order"] = out["samsung_impact"].map(impact_order).fillna(9)
    return out.sort_values(
        ["_tier_order", "_impact_order", "portfolio_score", "final_score", "topic_score", "urgency_score"],
        ascending=[True, True, False, False, False, False],
        kind="stable",
    )


def select_news(candidates: pd.DataFrame) -> pd.DataFrame:
    """Select executive Top30 with portfolio balance.

    STEP3 may produce 200+ TARIFF candidates.  STEP4 must rank all candidates
    but produce a balanced Top30 for executive mail:
    - TARIFF capped at 10~12 by default
    - Export Control, Origin/FTA, HS, AD/CVD, CBAM receive minimum backfill
    - Samsung Direct, larger ClusterSize, article_body and change_detail_hint are preferred
    - URL issues are quality flags, not hard deletion when evidence exists
    """
    ranked = rank_candidates(candidates)
    candidates["topic_bucket"] = ranked["topic_bucket"]
    candidates["evidence_score"] = ranked["evidence_score"]
    candidates["cluster_size"] = ranked["cluster_size"]
    candidates["portfolio_score"] = ranked["portfolio_score"]

    selected_rows = []
    seen_titles = []
    cluster_counts = {}
    source_counts = {}
    bucket_counts = {}
    reference_count = 0

    def has_reportable_evidence(row) -> bool:
        body = safe_str(row.get("article_body"))
        cluster_titles = safe_str(row.get("ClusterHeadlines"))
        change_hint = safe_str(row.get("change_detail_hint"))
        summary = safe_str(row.get("article_summary")) + " " + safe_str(row.get("Summary"))
        evidence_len = len(body) + len(cluster_titles) + len(change_hint) + len(summary)
        agency = safe_str(row.get("publisher") or row.get("agency") or row.get("source"))
        trusted_agency = contains_any(agency, [
            "Reuters", "Bloomberg", "AP", "Yonhap", "연합뉴스", "Korea Customs",
            "관세청", "USTR", "CBP", "USITC", "Federal Register", "WTO",
            "European Commission", "TAXUD", "EU", "MOIT", "Ministry",
            "Vietnam.vn", "TradingView", "Economic Times", "Hankyung", "한국경제",
            "CNBC", "MSN", "Yahoo", "Fortune", "SteelOrbis", "Politico",
        ])
        return evidence_len >= 180 or trusted_agency

    def bucket_cap(bucket: str, phase: str) -> int:
        cap = TOPIC_MAX_CAP.get(bucket, TOPIC_MAX_CAP.get("OTHER", 3))
        # During loose final fill, allow slight overflow except for TARIFF.
        if phase == "loose" and bucket != "TARIFF":
            cap += 2
        return cap

    def try_add(idx, row, phase: str = "ranking", enforce_bucket_cap: bool = True) -> bool:
        nonlocal reference_count
        row = row.copy()
        bucket = issue_bucket(row)
        row["topic_bucket"] = bucket

        if safe_str(row.get("priority_group")) == "NOISE":
            candidates.loc[idx, "audit_decision"] = "EXCLUDED"
            candidates.loc[idx, "audit_reason"] = safe_str(candidates.loc[idx].get("audit_reason")) or "NOISE"
            return False

        # URL is not a hard exclusion when STEP3 evidence exists.  It is carried
        # into STEP5 as NEED_REVIEW so users can still read the analysis.
        if safe_str(row.get("URL_Quality")) == "URL_MISSING_OR_BAD":
            if has_reportable_evidence(row):
                candidates.loc[idx, "URL_Quality"] = "NEED_REVIEW"
                candidates.loc[idx, "URL_Repair_Source"] = safe_str(row.get("URL_Repair_Source")) or "step4_portfolio_need_review"
                row["URL_Quality"] = "NEED_REVIEW"
                row["URL_Repair_Source"] = candidates.loc[idx, "URL_Repair_Source"]
            else:
                candidates.loc[idx, "audit_decision"] = "EXCLUDED"
                candidates.loc[idx, "audit_reason"] = "NO_URL_AND_NO_EVIDENCE"
                return False

        title = safe_str(row.get("headline"))
        if any(similar_title(title, old) >= 0.90 for old in seen_titles):
            candidates.loc[idx, "audit_decision"] = "EXCLUDED"
            candidates.loc[idx, "audit_reason"] = "DUPLICATE_SIMILAR_TITLE"
            return False

        ck = safe_str(row.get("cluster_key"))
        if cluster_counts.get(ck, 0) >= MAX_PER_CLUSTER:
            candidates.loc[idx, "audit_decision"] = "EXCLUDED"
            candidates.loc[idx, "audit_reason"] = "DUPLICATE_CLUSTER_CAP"
            return False

        source = safe_str(row.get("publisher") or row.get("agency") or row.get("source"))
        if source and source_counts.get(source, 0) >= MAX_PER_SOURCE and phase != "loose":
            candidates.loc[idx, "audit_decision"] = "EXCLUDED"
            candidates.loc[idx, "audit_reason"] = "SOURCE_CAP"
            candidates.loc[idx, "source_cap_reason"] = source
            return False

        if enforce_bucket_cap and bucket_counts.get(bucket, 0) >= bucket_cap(bucket, phase):
            candidates.loc[idx, "audit_decision"] = "EXCLUDED"
            candidates.loc[idx, "audit_reason"] = f"TOPIC_BUCKET_CAP_{bucket}"
            return False

        if bucket == "REFERENCE" and reference_count >= MAX_REFERENCE and phase != "loose":
            candidates.loc[idx, "audit_decision"] = "EXCLUDED"
            candidates.loc[idx, "audit_reason"] = "REFERENCE_CAP"
            return False

        selected_rows.append(row)
        seen_titles.append(title)
        cluster_counts[ck] = cluster_counts.get(ck, 0) + 1
        source_counts[source] = source_counts.get(source, 0) + 1
        bucket_counts[bucket] = bucket_counts.get(bucket, 0) + 1
        if bucket == "REFERENCE" or safe_str(row.get("priority_group")) == "REFERENCE":
            reference_count += 1
        candidates.loc[idx, "selected"] = "Y"
        candidates.loc[idx, "audit_decision"] = "SELECTED"
        candidates.loc[idx, "audit_reason"] = f"selected_by_STEP4_portfolio_{phase}_{bucket}"
        return True

    # Pass 0: Direct impact first, but still respect TARIFF cap.
    direct_pool = ranked[ranked["samsung_impact"].astype(str).eq("Direct") & ranked["priority_group"].ne("NOISE")]
    for idx, row in direct_pool.iterrows():
        if len(selected_rows) >= TOP_N:
            break
        try_add(idx, row, phase="direct")

    # Pass 1: minimum targets for underrepresented high-value topics.
    min_order = ["EXPORT_CONTROL", "ORIGIN_FTA", "HS_CLASSIFICATION", "AD_CVD", "CBAM", "CUSTOMS", "TARIFF"]
    for bucket in min_order:
        if len(selected_rows) >= TOP_N:
            break
        target = TOPIC_MIN_TARGET.get(bucket, 0)
        if target <= 0:
            continue
        pool = ranked[(ranked["topic_bucket"].eq(bucket)) & ranked["priority_group"].ne("NOISE")]
        for idx, row in pool.iterrows():
            if len(selected_rows) >= TOP_N or bucket_counts.get(bucket, 0) >= target:
                break
            if candidates.loc[idx, "selected"] == "Y":
                continue
            try_add(idx, row, phase="min_target")

    # Pass 2: score fill under bucket caps.
    for idx, row in ranked.iterrows():
        if len(selected_rows) >= TOP_N:
            break
        if candidates.loc[idx, "selected"] == "Y":
            continue
        try_add(idx, row, phase="score")

    # Pass 3: if still short, allow non-tariff buckets to overflow slightly.
    if len(selected_rows) < TOP_N:
        for idx, row in ranked.iterrows():
            if len(selected_rows) >= TOP_N:
                break
            if candidates.loc[idx, "selected"] == "Y":
                continue
            if safe_str(row.get("priority_group")) == "NOISE":
                continue
            bucket = issue_bucket(row)
            enforce = bucket == "TARIFF"  # Never let tariff dominate final Top30.
            try_add(idx, row, phase="loose", enforce_bucket_cap=enforce)

    if not selected_rows:
        return pd.DataFrame(columns=candidates.columns)
    selected = pd.DataFrame(selected_rows).reset_index(drop=True)
    selected = selected.sort_values(
        ["portfolio_score", "final_score", "topic_score", "urgency_score"],
        ascending=[False, False, False, False],
        kind="stable",
    ).reset_index(drop=True)
    selected = selected.head(TOP_N).copy()
    selected["rank"] = range(1, len(selected) + 1)
    selected["selected"] = "Y"
    selected["last_checked"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    return selected

def prepare_output(df: pd.DataFrame, columns: list[str]) -> pd.DataFrame:
    if df.empty:
        return pd.DataFrame(columns=columns)
    out = df.rename(columns={
        "date": "Date",
        "headline": "Headline",
        "url": "URL",
        "country": "Country",
        "agency": "Agency",
        "publisher": "Publisher",
        "keyword_matches": "KeywordMatches",
        "select_reason": "SelectReason",
        "reject_reason": "RejectReason",
        "source": "Source",
        "source_file": "SourceFile",
        "regulation_related": "RegulationRelated",
        "regulation_transfer_type": "RegulationTransferType",
    }).copy()
    for col in columns:
        if col not in out.columns:
            out[col] = ""
    out = out[out["Headline"].astype(str).str.strip().ne("")].copy()
    return out[columns]


def clean_cumulative(existing: pd.DataFrame, daily: pd.DataFrame) -> pd.DataFrame:
    frames = []
    if existing is not None and not existing.empty:
        frames.append(existing)
    if daily is not None and not daily.empty:
        frames.append(daily)
    if not frames:
        return pd.DataFrame(columns=OUTPUT_COLUMNS)
    combined = pd.concat(frames, ignore_index=True, sort=False)
    for col in OUTPUT_COLUMNS:
        if col not in combined.columns:
            combined[col] = ""
    combined = combined[combined["Headline"].astype(str).str.strip().ne("")]
    combined["_dedup_key"] = combined.apply(lambda r: safe_str(r.get("URL")) or canonical_title(r.get("Headline")), axis=1)
    combined = combined.drop_duplicates(subset=["_dedup_key"], keep="first")
    tier_order = {"CORE": 0, "USABLE": 1, "REFERENCE": 2, "NOISE": 9}
    impact_order = {"Direct": 0, "Indirect": 1, "Reference": 2, "None": 9}
    combined["_tier_order"] = combined["priority_group"].map(tier_order).fillna(9)
    combined["_impact_order"] = combined["samsung_impact"].map(impact_order).fillna(9)
    combined["final_score"] = pd.to_numeric(combined["final_score"], errors="coerce").fillna(0)
    combined = combined.sort_values(["_tier_order", "_impact_order", "final_score"], ascending=[True, True, False], kind="stable")
    combined = combined.drop(columns=["_dedup_key", "_tier_order", "_impact_order"], errors="ignore").head(TOP_N).reset_index(drop=True)
    combined["rank"] = range(1, len(combined) + 1)
    return combined[OUTPUT_COLUMNS]


def write_excel(path: Path, df: pd.DataFrame, sheet_name: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    try:
        with pd.ExcelWriter(path, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)
        format_excel(path)
    except PermissionError:
        backup = path.with_name(f"{path.stem}_{datetime.now():%Y%m%d_%H%M%S}{path.suffix}")
        with pd.ExcelWriter(backup, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)
        format_excel(backup)
        log(f"LOCKED -> SAVE BACKUP: {backup}")


def format_excel(path: Path) -> None:
    if not path.exists():
        return
    wb = load_workbook(path)
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        header_fill = PatternFill("solid", fgColor="1F4E78")
        header_font = Font(color="FFFFFF", bold=True)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        widths = {
            "Headline": 60, "ExecutiveMessage": 70, "RequiredAction": 55, "ActionOwner": 22,
            "Summary": 55, "AI Analysis": 60, "Action Plan": 55, "URL": 45,
            "affected_subsidiaries": 30, "affected_products": 35, "topic_reason": 35,
        }
        for idx in range(1, ws.max_column + 1):
            h = str(ws.cell(1, idx).value or "")
            ws.column_dimensions[get_column_letter(idx)].width = widths.get(h, 16)
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
    wb.save(path)


def main() -> None:
    print("[STEP4-2] News analysis start - GTI95")
    raw, input_path = load_input()
    df = normalize_columns(raw)
    candidates = build_candidates(df)
    selected = select_news(candidates)

    daily = prepare_output(selected, OUTPUT_COLUMNS)
    audit = prepare_output(candidates, AUDIT_COLUMNS)

    if OUTPUT_CUMUL.exists():
        existing = pd.read_excel(OUTPUT_CUMUL)
        log(f"cumulative existing load: {len(existing)} rows")
    else:
        existing = pd.DataFrame(columns=OUTPUT_COLUMNS)
        log("cumulative file missing -> new create")
    cumulative = clean_cumulative(existing, daily)

    write_excel(OUTPUT_DAILY, daily, "news_ai_summary")
    write_excel(OUTPUT_CUMUL, cumulative, "news_ai_cumulative")
    write_excel(OUTPUT_AUDIT, audit, "news_ai_audit")

    log(f"STEP4-2 COMPLETE: selected={len(daily)} / audit={len(audit)} / cumulative={len(cumulative)}")
    log(f"SAVE: {OUTPUT_DAILY}")
    log(f"SAVE: {OUTPUT_CUMUL}")
    log(f"SAVE: {OUTPUT_AUDIT}")


# =========================================================
# GTI executive-quality overrides
# - Reduce false positives from market/exhibition/FX/general supply articles.
# - Keep only actionable customs/trade-policy topics in STEP4 selection.
# - Make Direct impact stricter so STEP5 does not receive inflated Direct rows.
# =========================================================

ACTIONABLE_POLICY_TERMS = [
    "관세", "관세율", "추가관세", "상호관세", "무관세", "철강관세", "세이프가드",
    "반덤핑", "상계관세", "덤핑방지", "ad/cvd", "anti-dumping", "countervailing",
    "section 301", "section 232", "ustr", "cbp", "usitc", "federal register",
    "fta", "cepa", "epa", "원산지", "협정세율", "certificate of origin", "rules of origin",
    "hs code", "품목분류", "수출통제", "전략물자", "entity list", "bis", "ear", "eccn",
    "uflpa", "forced labor", "cbam", "carbon border",
]

LOW_VALUE_REPORT_TERMS = [
    "전시회", "참관기", "시장동향", "상품db", "유망바이어", "환리스크", "고환율",
    "공공비축", "비료", "원전 사업", "관광", "문화", "스포츠", "게임", "증시",
]

STRONG_SAMSUNG_PRODUCT_TERMS = [
    "samsung", "삼성", "semiconductor", "semiconductors", "chip", "chips", "반도체",
    "hbm", "dram", "nand", "display", "디스플레이", "battery", "배터리", "cell",
    "smartphone", "스마트폰", "전자부품", "electronics", "가전",
]

TARIFF_POLICY_TERMS = [
    "관세", "관세율", "추가관세", "상호관세", "무관세", "철강관세", "세이프가드",
    "tariff", "tariffs", "duty", "duties", "section 301", "section 232", "ustr",
]


def _gti_text(row: pd.Series) -> str:
    return " ".join([
        safe_str(row.get("headline")),
        safe_str(row.get("summary")),
        safe_str(row.get("description")),
        safe_str(row.get("article_body")),
        safe_str(row.get("ClusterHeadlines")),
        safe_str(row.get("change_detail_hint")),
        safe_str(row.get("hs_hint")),
        safe_str(row.get("tariff_rate_hint")),
        safe_str(row.get("keyword_matches")),
        safe_str(row.get("keyword")),
    ])


def _has_actionable_policy(row: pd.Series) -> bool:
    text = _gti_text(row)
    return contains_any(text, ACTIONABLE_POLICY_TERMS)


def _has_strong_product(row: pd.Series) -> bool:
    return contains_any(_gti_text(row), STRONG_SAMSUNG_PRODUCT_TERMS)


def _low_value_for_executive(row: pd.Series) -> bool:
    text = _gti_text(row)
    if not contains_any(text, LOW_VALUE_REPORT_TERMS):
        return False
    # Do not suppress a low-value looking item when it also has a concrete
    # policy hook that customs teams can act on.
    return not contains_any(text, [
        "관세율", "추가관세", "상호관세", "반덤핑", "상계관세", "세이프가드",
        "section 301", "section 232", "cbam", "entity list", "uflpa",
        "원산지 기준", "협정세율", "hs code", "품목분류",
    ])


def detect_topic(row: pd.Series) -> tuple[str, int, str, str, str]:
    text = _gti_text(row)
    headline = safe_str(row.get("headline"))

    if is_security_non_trade(row) or _low_value_for_executive(row):
        return "REFERENCE", 0, "", "Low-value or non-trade item for executive customs report", "TRADE_GENERAL"

    if contains_any(text, ["cbam", "carbon border", "탄소국경"]):
        return "CBAM", 95, "CBAM", "Concrete CBAM signal", "CBAM_CARBON"
    if contains_any(text, ["entity list", "bis", "export control", "export controls", "eccn", "ear", "uflpa", "forced labor", "수출통제", "전략물자"]):
        return "Export Control", 95, "Export Control", "Concrete export-control signal", "EXPORT_CONTROL"
    if contains_any(text, ["anti-dumping", "antidumping", "countervailing", "ad/cvd", "반덤핑", "상계관세", "덤핑방지"]):
        return "AD/CVD", 92, "AD/CVD", "Concrete AD/CVD signal", "AD_CVD"
    if contains_any(text, ["section 301", "section 232", "ustr", "상호관세", "추가관세", "철강관세", "세이프가드"]):
        return "Tariff / Section 301/232", 92, "Tariff", "Concrete tariff policy signal", "TARIFF"
    if contains_any(text, ["rules of origin", "certificate of origin", "원산지", "협정세율"]) or (
        contains_any(text, ["fta", "cepa", "epa"]) and contains_any(text, ["관세", "무관세", "tariff", "origin", "원산지", "협정"])
    ):
        return "FTA / Origin", 88, "FTA/Origin", "Concrete FTA/origin signal", "ORIGIN_FTA"
    if contains_any(text, ["hs code", "품목분류", "tariff classification"]):
        return "HS Classification", 84, "HS", "HS classification signal", "HS_CLASSIFICATION"
    if contains_any(headline, TARIFF_POLICY_TERMS):
        return "Tariff", 82, "Tariff headline", "Tariff signal in headline", "TARIFF"

    return "REFERENCE", 0, "", "No actionable customs/trade-policy topic after executive guardrail", "TRADE_GENERAL"


def is_noise(row: pd.Series) -> str:
    issue = safe_str(row.get("issue_type"))
    if issue == "TRADE_GENERAL":
        return "NO_TOPIC_NO_ACTIONABLE_POLICY_SIGNAL"
    if _low_value_for_executive(row):
        return "LOW_VALUE_EXECUTIVE_REPORT_NOISE"
    if not _has_actionable_policy(row):
        return "NO_ACTIONABLE_CUSTOMS_TRADE_SIGNAL"
    return ""


def samsung_impact(row: pd.Series) -> str:
    issue = safe_str(row.get("issue_type"))
    text = _gti_text(row)
    has_product = _has_strong_product(row)
    has_subsidiary = safe_str(row.get("affected_subsidiaries")) not in {"", MISSING_TEXT}
    has_action_hint = contains_any(text, [
        "시행", "발효", "부과", "적용", "final rule", "effective", "impose", "levy",
        "관세율", "hs code", "품목분류", "원산지", "협정세율", "entity list", "uflpa",
        "수출통제", "cbam", "반덤핑", "상계관세",
    ])

    if issue == "TRADE_GENERAL" or _low_value_for_executive(row):
        return "Reference"
    if issue in {"EXPORT_CONTROL", "CBAM_CARBON", "AD_CVD", "TARIFF", "SECTION_301_232", "HS_CLASSIFICATION"}:
        return "Direct" if (has_product or has_subsidiary or has_action_hint) else "Indirect"
    if issue == "ORIGIN_FTA":
        return "Direct" if ((has_product or has_subsidiary) and has_action_hint) else "Indirect"
    return "Indirect" if _has_actionable_policy(row) else "Reference"


def samsung_impact_score(row: pd.Series) -> int:
    impact = safe_str(row.get("samsung_impact") or samsung_impact(row))
    if impact == "Direct":
        return 86
    if impact == "Indirect":
        return 58
    return 15


def priority_group(row: pd.Series) -> str:
    if safe_str(row.get("audit_reason")) or safe_str(row.get("issue_type")) == "TRADE_GENERAL":
        return "NOISE"
    impact = safe_str(row.get("samsung_impact") or samsung_impact(row))
    score = float(row.get("final_score", 0) or 0)
    if impact == "Direct" and score >= 78:
        return "CORE"
    if impact in {"Direct", "Indirect"} and score >= 58:
        return "USABLE"
    return "REFERENCE"


def make_summary(row: pd.Series) -> str:
    topic = safe_str(row.get("topic")) or safe_str(row.get("issue_type"))
    impact = safe_str(row.get("samsung_impact") or samsung_impact(row))
    country = safe_str(row.get("country")) or "Global"
    return f"{country} {topic} 관련 관세·통상 이슈입니다. 삼성 영향은 {impact}로 분류되며, 원문 기준으로 관세율·원산지·수출통제·CBAM 등 실행 항목 확인이 필요합니다."


def make_ai_analysis(row: pd.Series) -> str:
    issue = safe_str(row.get("issue_type"))
    if issue == "ORIGIN_FTA":
        return "FTA/원산지 이슈로 CO 발급·수취, BOM 충족, 협정세율 적용 가능성 및 증빙 보관 기준 점검이 필요합니다."
    if issue in {"TARIFF", "SECTION_301_232", "AD_CVD"}:
        return "관세정책 이슈로 대상 HS, 공급국, 적용 시점, 관세율 및 가격 전가 영향을 확인해야 합니다."
    if issue == "EXPORT_CONTROL":
        return "수출통제 이슈로 ECCN/전략물자 해당 여부, 거래상대방·최종사용자 스크리닝 및 우회거래 가능성 점검이 필요합니다."
    if issue == "CBAM_CARBON":
        return "CBAM 이슈로 EU향 품목, 원재료, 공급사 탄소자료 및 신고 증빙 체계를 확인해야 합니다."
    if issue == "HS_CLASSIFICATION":
        return "HS 품목분류 이슈로 국가별 HS 매핑, 품목 설명, 사전심사 필요 여부를 확인해야 합니다."
    return "관세·통상 업무 관련 실행 영향이 제한적이므로 참고 모니터링 대상으로 관리합니다."


# =========================================================
# GTI ASCII-safe executive-quality overrides
# Keep this block immediately before __main__.
# All Korean keywords use Unicode escapes so copying to C:\Temp cannot corrupt
# customs/tariff matching rules.
# =========================================================

def _u(s: str) -> str:
    return s.encode("ascii").decode("unicode_escape")


K_TARIFF = _u("\\uad00\\uc138")
K_TARIFF_RATE = _u("\\uad00\\uc138\\uc728")
K_EXTRA_TARIFF = _u("\\ucd94\\uac00\\uad00\\uc138")
K_RECIPROCAL_TARIFF = _u("\\uc0c1\\ud638\\uad00\\uc138")
K_ZERO_TARIFF = _u("\\ubb34\\uad00\\uc138")
K_STEEL_TARIFF = _u("\\ucca0\\uac15\\uad00\\uc138")
K_QUOTA = _u("\\ucffc\\ud130")
K_AD = _u("\\ubc18\\ub364\\ud551")
K_CVD = _u("\\uc0c1\\uacc4\\uad00\\uc138")
K_ORIGIN = _u("\\uc6d0\\uc0b0\\uc9c0")
K_PREF_RATE = _u("\\ud611\\uc815\\uc138\\uc728")
K_HS_CLASS = _u("\\ud488\\ubaa9\\ubd84\\ub958")
K_EXPORT_CONTROL = _u("\\uc218\\ucd9c\\ud1b5\\uc81c")
K_STRATEGIC = _u("\\uc804\\ub7b5\\ubb3c\\uc790")
K_EXHIBITION = _u("\\uc804\\uc2dc\\ud68c")
K_VISIT = _u("\\ucc38\\uad00\\uae30")
K_MARKET = _u("\\uc2dc\\uc7a5\\ub3d9\\ud5a5")
K_PRODUCT_DB = _u("\\uc0c1\\ud488DB")
K_FX = _u("\\ud658\\ub9ac\\uc2a4\\ud06c")
K_HIGH_FX = _u("\\uace0\\ud658\\uc728")
K_FERTILIZER = _u("\\ube44\\ub8cc")
K_TODAY_HISTORY = _u("\\uc624\\ub298\\uc758 \\uc5ed\\uc0ac")
K_HISTORY_JP = _u("\\u4eca\\u65e5\\u306e\\u6b74\\u53f2")
K_SAMSUNG = _u("\\uc0bc\\uc131")
K_SEMI = _u("\\ubc18\\ub3c4\\uccb4")
K_DISPLAY = _u("\\ub514\\uc2a4\\ud50c\\ub808\\uc774")
K_BATTERY = _u("\\ubc30\\ud130\\ub9ac")
K_ELECTRONICS = _u("\\uc804\\uc790")

GTI_ACTIONABLE_TERMS_SAFE = [
    K_TARIFF, K_TARIFF_RATE, K_EXTRA_TARIFF, K_RECIPROCAL_TARIFF, K_ZERO_TARIFF,
    K_STEEL_TARIFF, K_QUOTA, K_AD, K_CVD, K_ORIGIN, K_PREF_RATE, K_HS_CLASS,
    K_EXPORT_CONTROL, K_STRATEGIC,
    "tariff", "tariffs", "duty", "duties", "quota", "safeguard",
    "ad/cvd", "anti-dumping", "antidumping", "countervailing",
    "section 301", "section 232", "ustr", "cbp", "usitc", "federal register",
    "fta", "cepa", "epa", "rules of origin", "certificate of origin",
    "hs code", "tariff classification", "export control", "entity list",
    "bis", "ear", "eccn", "uflpa", "forced labor", "cbam", "carbon border",
]

GTI_LOW_VALUE_TERMS_SAFE = [
    K_EXHIBITION, K_VISIT, K_MARKET, K_PRODUCT_DB, K_FX, K_HIGH_FX, K_FERTILIZER,
    K_TODAY_HISTORY, K_HISTORY_JP,
    "market trend", "buyer", "exhibition", "trade fair", "on this day",
    "anniversary", "tennis", "sports", "game", "stablecoin",
]

GTI_PRODUCT_TERMS_SAFE = [
    "samsung", K_SAMSUNG, "semiconductor", "semiconductors", "chip", "chips",
    K_SEMI, "hbm", "dram", "nand", "display", K_DISPLAY, "battery", K_BATTERY,
    "electronics", K_ELECTRONICS, "smartphone",
]


def _gti_text_safe(row: pd.Series) -> str:
    return " ".join([
        safe_str(row.get("headline")),
        safe_str(row.get("summary")),
        safe_str(row.get("description")),
        safe_str(row.get("article_body")),
        safe_str(row.get("ClusterHeadlines")),
        safe_str(row.get("change_detail_hint")),
        safe_str(row.get("hs_hint")),
        safe_str(row.get("tariff_rate_hint")),
        safe_str(row.get("keyword_matches")),
        safe_str(row.get("keyword")),
    ])


def _contains_safe(row: pd.Series, terms: list[str]) -> bool:
    text = norm(_gti_text_safe(row))
    return any(norm(t) in text for t in terms if safe_str(t))


def _headline_contains_safe(row: pd.Series, terms: list[str]) -> bool:
    text = norm(row.get("headline"))
    return any(norm(t) in text for t in terms if safe_str(t))


def _low_value_safe(row: pd.Series) -> bool:
    if not _contains_safe(row, GTI_LOW_VALUE_TERMS_SAFE):
        return False
    hard = [
        K_STEEL_TARIFF, K_ZERO_TARIFF, K_EXTRA_TARIFF, K_RECIPROCAL_TARIFF,
        K_AD, K_CVD, "section 301", "section 232", "cbam", "entity list",
        "uflpa", K_EXPORT_CONTROL, K_HS_CLASS,
    ]
    return not _headline_contains_safe(row, hard)


def _has_product_safe(row: pd.Series) -> bool:
    return _contains_safe(row, GTI_PRODUCT_TERMS_SAFE)


def detect_topic(row: pd.Series) -> tuple[str, int, str, str, str]:
    if is_security_non_trade(row) or _low_value_safe(row):
        return "REFERENCE", 0, "", "Low-value or non-trade item for executive customs report", "TRADE_GENERAL"

    if _contains_safe(row, ["cbam", "carbon border"]):
        return "CBAM", 95, "CBAM", "Concrete CBAM signal", "CBAM_CARBON"
    if _contains_safe(row, ["entity list", "bis", "export control", "eccn", "ear", "uflpa", "forced labor", K_EXPORT_CONTROL, K_STRATEGIC]):
        return "Export Control", 95, "Export Control", "Concrete export-control signal", "EXPORT_CONTROL"
    if _contains_safe(row, ["anti-dumping", "antidumping", "countervailing", "ad/cvd", K_AD, K_CVD]):
        return "AD/CVD", 92, "AD/CVD", "Concrete AD/CVD signal", "AD_CVD"
    if _contains_safe(row, ["section 301", "section 232", "ustr", K_RECIPROCAL_TARIFF, K_EXTRA_TARIFF, K_STEEL_TARIFF, K_QUOTA]):
        return "Tariff / Section 301/232", 92, "Tariff", "Concrete tariff policy signal", "TARIFF"
    if _contains_safe(row, ["rules of origin", "certificate of origin", K_ORIGIN, K_PREF_RATE]) or (
        _contains_safe(row, ["fta", "cepa", "epa"]) and _contains_safe(row, [K_TARIFF, K_ZERO_TARIFF, "tariff", "origin", K_ORIGIN])
    ):
        return "FTA / Origin", 88, "FTA/Origin", "Concrete FTA/origin signal", "ORIGIN_FTA"
    if _contains_safe(row, ["hs code", "tariff classification", K_HS_CLASS]):
        return "HS Classification", 84, "HS", "HS classification signal", "HS_CLASSIFICATION"
    if _headline_contains_safe(row, [K_TARIFF, K_TARIFF_RATE, K_EXTRA_TARIFF, K_RECIPROCAL_TARIFF, K_ZERO_TARIFF, K_STEEL_TARIFF, "tariff", "duty"]):
        return "Tariff", 82, "Tariff headline", "Tariff signal in headline", "TARIFF"
    return "REFERENCE", 0, "", "No actionable customs/trade-policy topic after executive guardrail", "TRADE_GENERAL"


def is_noise(row: pd.Series) -> str:
    if safe_str(row.get("issue_type")) == "TRADE_GENERAL":
        return "NO_TOPIC_NO_ACTIONABLE_POLICY_SIGNAL"
    if _low_value_safe(row):
        return "LOW_VALUE_EXECUTIVE_REPORT_NOISE"
    if not _contains_safe(row, GTI_ACTIONABLE_TERMS_SAFE):
        return "NO_ACTIONABLE_CUSTOMS_TRADE_SIGNAL"
    return ""


def samsung_impact(row: pd.Series) -> str:
    issue = safe_str(row.get("issue_type"))
    if issue == "TRADE_GENERAL" or _low_value_safe(row):
        return "Reference"
    product = _has_product_safe(row)
    action_hint = _contains_safe(row, [
        K_TARIFF_RATE, K_STEEL_TARIFF, K_AD, K_CVD, K_HS_CLASS, K_ORIGIN,
        K_PREF_RATE, K_EXPORT_CONTROL, "entity list", "uflpa", "cbam",
        "effective", "impose", "levy", "final rule",
    ])
    if issue in {"EXPORT_CONTROL", "CBAM_CARBON", "AD_CVD", "TARIFF", "SECTION_301_232", "HS_CLASSIFICATION"}:
        return "Direct" if (product or action_hint) else "Indirect"
    if issue == "ORIGIN_FTA":
        return "Direct" if (product and action_hint) else "Indirect"
    return "Indirect" if _contains_safe(row, GTI_ACTIONABLE_TERMS_SAFE) else "Reference"


def samsung_impact_score(row: pd.Series) -> int:
    impact = safe_str(row.get("samsung_impact") or samsung_impact(row))
    return 86 if impact == "Direct" else 58 if impact == "Indirect" else 15


def priority_group(row: pd.Series) -> str:
    if safe_str(row.get("audit_reason")) or safe_str(row.get("issue_type")) == "TRADE_GENERAL":
        return "NOISE"
    impact = safe_str(row.get("samsung_impact") or samsung_impact(row))
    score = float(row.get("final_score", 0) or 0)
    if impact == "Direct" and score >= 78:
        return "CORE"
    if impact in {"Direct", "Indirect"} and score >= 58:
        return "USABLE"
    return "REFERENCE"


def make_summary(row: pd.Series) -> str:
    topic = safe_str(row.get("topic")) or safe_str(row.get("issue_type"))
    impact = safe_str(row.get("samsung_impact") or samsung_impact(row))
    country = safe_str(row.get("country")) or "Global"
    return f"{country} {topic} customs/trade-policy issue. Samsung impact is classified as {impact}; verify tariff, origin, export-control, CBAM, HS and evidence requirements against the original source."


def make_ai_analysis(row: pd.Series) -> str:
    issue = safe_str(row.get("issue_type"))
    if issue == "ORIGIN_FTA":
        return "Check CO issuance/receipt, BOM origin qualification, preferential duty eligibility, and evidence retention."
    if issue in {"TARIFF", "SECTION_301_232", "AD_CVD"}:
        return "Check target HS, supplier country, effective date, tariff or AD/CVD rate, and price/cost impact."
    if issue == "EXPORT_CONTROL":
        return "Check ECCN/strategic item classification, restricted-party screening, end user, and anti-circumvention controls."
    if issue == "CBAM_CARBON":
        return "Check EU-bound products, raw materials, supplier emissions data, and CBAM reporting evidence."
    if issue == "HS_CLASSIFICATION":
        return "Check HS mapping by country, product description, and whether advance ruling is needed."
    return "Keep as reference monitoring unless a concrete customs/trade action is identified."


if __name__ == "__main__":
    main()
