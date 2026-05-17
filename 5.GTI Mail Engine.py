# -*- coding: utf-8 -*-
r"""
5.GTI Mail Engine_FINAL_URL_AI.py
GTI Radar STEP5 Mail Engine - FINAL STABLE VERSION

목적
- C:\temp\3.news_ai_summary.xlsx 의 URL/제목을 기준으로 뉴스 본문을 직접 확인
- Google News / Google redirect URL 최대한 원문 URL로 복원
- Gemini API로 기사별 Summary / AI Analysis / Action Plan 생성
- Gemini 실패 시에도 제목/본문 기반 rule-based 분석으로 복붙 문구 방지
- Top30 GTI 메일 Excel + HTML 생성
- 선택적으로 SMTP 메일 발송

필수 입력
- C:\temp\3.news_ai_summary.xlsx

출력
- C:\temp\GTI_Radar_YYYY-MM-DD_Top30.xlsx
- C:\temp\GTI_Radar_YYYY-MM-DD_Top30_Email.html
- C:\temp\mail_cumulative.xlsx

환경변수 예시
PowerShell 현재 창 테스트:
  $env:GEMINI_API_KEY="본인 Gemini Key"
  $env:GEMINI_MODEL="gemini-1.5-flash"
  $env:GTI_SEND_EMAIL="Y"
  $env:GTI_SMTP_USER="kch8872@naver.com"
  $env:GTI_SMTP_PASS="네이버앱비밀번호"
  $env:GTI_MAIL_TO="수신자메일"

주의
- API Key / SMTP Password는 코드에 직접 넣지 마십시오.
"""

from __future__ import annotations

import os
import re
import ssl
import json
import html
import time
import smtplib
import traceback
from pathlib import Path
from datetime import datetime
from email.message import EmailMessage
from email.utils import formataddr
from urllib.parse import urlparse, parse_qs, unquote

import pandas as pd
import requests
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

try:
    from bs4 import BeautifulSoup
except Exception:
    BeautifulSoup = None


# ============================================================
# 0. CONFIG
# ============================================================
BASE_DIR = Path(os.getenv("GTI_BASE_DIR", r"C:\temp"))
TODAY = datetime.now().strftime("%Y-%m-%d")
NOW_STR = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

INPUT_CANDIDATES = [
    BASE_DIR / "3.news_ai_summary.xlsx",
    BASE_DIR / "4.news_ai_analysis.xlsx",
    BASE_DIR / "news_ai_summary.xlsx",
    BASE_DIR / "news_raw.xlsx",
]

OUTPUT_XLSX = BASE_DIR / f"GTI_Radar_{TODAY}_Top30.xlsx"
OUTPUT_HTML = BASE_DIR / f"GTI_Radar_{TODAY}_Top30_Email.html"
MAIL_CUMULATIVE = BASE_DIR / "mail_cumulative.xlsx"
SUBJECT = f"[GTI Radar] Global Trade Intelligence | {TODAY}"

GEMINI_API_KEY = (os.getenv("GEMINI_API_KEY") or "").strip()
GEMINI_MODEL = (os.getenv("GEMINI_MODEL") or "gemini-1.5-flash").strip()
USE_GEMINI = bool(GEMINI_API_KEY)

SMTP_HOST = os.getenv("GTI_SMTP_HOST", "smtp.naver.com")
SMTP_PORT = int(os.getenv("GTI_SMTP_PORT", "465"))
SMTP_USER = (os.getenv("GTI_SMTP_USER") or os.getenv("GTI_MAIL_ID") or "").strip()
SMTP_PASS = (os.getenv("GTI_SMTP_PASS") or os.getenv("GTI_MAIL_PW") or "").strip()
MAIL_FROM_NAME = os.getenv("GTI_MAIL_FROM_NAME", "GTI Radar")
SEND_EMAIL = str(os.getenv("GTI_SEND_EMAIL", "N")).strip().upper() == "Y"
FALLBACK_TO = os.getenv("GTI_MAIL_TO", "").strip()

TOP_N = int(os.getenv("GTI_TOP_N", "30"))
FETCH_LIMIT = int(os.getenv("GTI_FETCH_LIMIT", "45"))     # 본문/Gemini 분석 대상 후보 수
REQUEST_TIMEOUT = int(os.getenv("GTI_TIMEOUT", "10"))

FOCUS_COUNTRIES = ["KR", "CN", "VN", "IN", "US", "MX", "BR", "EU"]
PRODUCTS = ["Mobile", "Consumer Electronics", "Network Equipment", "Semiconductor", "Display", "Medical"]

RISK_ORDER = {"상": 1, "중": 2, "하": 3}
SECTION_ORDER = {
    "1.직접 영향": 1,
    "2.관세/통상 정책": 2,
    "3.수입규제/조사": 3,
    "4.기타 모니터링": 4,
}

TRADE_KEYWORDS = [
    "tariff", "customs", "duty", "fta", "trade", "origin", "hs code", "valuation",
    "anti-dumping", "countervailing", "safeguard", "export control", "import regulation",
    "section 301", "section 232", "cbam", "ustr", "cbp", "wto", "wco",
    "관세", "통관", "세관", "무역", "통상", "원산지", "품목분류", "과세가격",
    "반덤핑", "상계관세", "세이프가드", "수출통제", "수입규제", "무역협정", "환급",
]

HIGH_IMPACT_TERMS = [
    "tariff hike", "raise duty", "increased tariff", "anti-dumping", "countervailing",
    "section 301", "section 232", "export control", "import ban", "sanction",
    "관세 인상", "반덤핑", "상계관세", "수입금지", "수출통제", "제재", "조사 착수",
]

NOISE_TERMS = [
    "sports", "football", "baseball", "concert", "festival", "celebrity", "movie",
    "stock price", "crypto", "bitcoin", "gold price only", "weather", "crime",
    "cigarette", "students", "immigration", "visa", "opt fraud",
    "연예", "스포츠", "축구", "야구", "콘서트", "주가", "코인", "비트코인",
    "담배", "밀수", "학생", "이민", "비자", "범죄", "날씨",
]

SAMSUNG_COUNTRY_MAP = {
    "KR": ["korea", "south korea", "한국", "대한민국"],
    "CN": ["china", "중국"],
    "VN": ["vietnam", "베트남"],
    "IN": ["india", "인도"],
    "US": ["united states", "u.s.", "usa", "미국"],
    "MX": ["mexico", "멕시코"],
    "BR": ["brazil", "브라질"],
    "EU": ["european union", "european commission", "eu", "유럽연합"],
}

AGENCY_MAP = [
    ("USTR", ["ustr", "u.s. trade representative", "미 무역대표부"]),
    ("U.S. Customs and Border Protection (CBP)", ["cbp", "u.s. customs", "customs and border protection", "미 세관"]),
    ("U.S. Department of Commerce", ["department of commerce", "u.s. commerce", "상무부"]),
    ("European Commission", ["european commission", "eu commission", "유럽연합 집행위원회"]),
    ("WTO", ["wto", "world trade organization", "세계무역기구"]),
    ("WCO", ["wco", "world customs organization", "세계관세기구"]),
    ("MOFCOM", ["mofcom", "중국 상무부"]),
    ("GACC", ["gacc", "중국 해관", "해관총서"]),
    ("Vietnam Customs / Trade Remedies Authority", ["vietnam customs", "trade remedies authority", "vietnam.vn", "베트남"]),
    ("Ministry of Commerce & Industry, India", ["india", "인도", "piyush goyal"]),
    ("관세청", ["관세청"]),
    ("산업통상자원부", ["산업통상자원부", "산업부"]),
    ("기획재정부", ["기획재정부"]),
]


# ============================================================
# 1. LOG / BASIC UTILS
# ============================================================
def log(msg: str) -> None:
    print(f"[{datetime.now():%Y-%m-%d %H:%M:%S}] {msg}", flush=True)


def clean_text(v) -> str:
    if v is None:
        return ""
    try:
        if pd.isna(v):
            return ""
    except Exception:
        pass
    s = str(v)
    s = html.unescape(s)
    s = s.replace("\u200b", " ")
    s = re.sub(r"\s+", " ", s)
    return s.strip()


def compact(v) -> str:
    return clean_text(v)


def safe_join_values(values) -> str:
    out = []
    for v in values:
        s = clean_text(v)
        if s and s.lower() not in ["nan", "none", "nat"]:
            out.append(s)
    return " ".join(out)


def safe_date(v) -> str:
    s = clean_text(v)
    if not s:
        return ""

    # pandas Series가 문자열화된 "date 2026-05-13 ... Name: ..." 형태 보정
    m = re.search(r"(20\d{2}[-/.]\d{1,2}[-/.]\d{1,2})(?:[ T](\d{1,2}:\d{2}(?::\d{2})?))?", s)
    if m:
        date_part = m.group(1).replace("/", "-").replace(".", "-")
        time_part = m.group(2) or "00:00"
        try:
            return pd.to_datetime(f"{date_part} {time_part}", errors="coerce").strftime("%Y-%m-%d %H:%M")
        except Exception:
            return f"{date_part} {time_part[:5]}"

    try:
        dt = pd.to_datetime(v, errors="coerce")
        if not pd.isna(dt):
            return dt.strftime("%Y-%m-%d %H:%M")
    except Exception:
        pass
    return s[:16]


def normalize_title(t: str) -> str:
    s = clean_text(t).lower()
    s = re.sub(r"[-|–—].*$", "", s)  # 언론사명 제거
    s = re.sub(r"[^0-9a-z가-힣一-龥ぁ-んァ-ン]+", " ", s)
    s = re.sub(r"\s+", " ", s)
    return s.strip()


def domain_of(url: str) -> str:
    try:
        return urlparse(url).netloc.lower().replace("www.", "")
    except Exception:
        return ""


# ============================================================
# 2. HTTP SESSION / URL RESOLUTION / BODY EXTRACT
# ============================================================
def make_session() -> requests.Session:
    session = requests.Session()
    retry = Retry(
        total=1,
        connect=1,
        read=1,
        backoff_factor=1.0,
        status_forcelist=[429, 500, 502, 503, 504],
        allowed_methods=["GET", "HEAD"],
        raise_on_status=False,
    )
    adapter = HTTPAdapter(max_retries=retry, pool_connections=20, pool_maxsize=20)
    session.mount("http://", adapter)
    session.mount("https://", adapter)
    session.headers.update({
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/124.0 Safari/537.36"
        ),
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        "Accept-Language": "ko-KR,ko;q=0.9,en-US;q=0.8,en;q=0.7",
        "Connection": "close",
    })
    return session


SESSION = make_session()


def resolve_redirect_url(url: str) -> str:
    url = clean_text(url)
    if not url:
        return ""

    # Google redirect: https://www.google.com/url?...&url=https://...
    try:
        parsed = urlparse(url)
        qs = parse_qs(parsed.query)
        if "url" in qs and qs["url"]:
            return unquote(qs["url"][0])
        if "q" in qs and qs["q"]:
            q = qs["q"][0]
            if q.startswith("http"):
                return unquote(q)
    except Exception:
        pass

    # Google News RSS URL은 requests redirect를 통해 최종 URL 확보 시도
    if "news.google.com/rss/articles" in url or "news.google.com/articles" in url:
        try:
            r = SESSION.get(url, timeout=REQUEST_TIMEOUT, allow_redirects=True)
            if r.url and "news.google.com" not in r.url:
                return r.url
        except Exception:
            return url

    return url


def fetch_html(url: str) -> tuple[str, str]:
    """return (final_url, html_text). 실패 시 ("", "") 대신 final_url만 반환 가능."""
    url = resolve_redirect_url(url)
    if not url:
        return "", ""
    try:
        r = SESSION.get(url, timeout=REQUEST_TIMEOUT, allow_redirects=True)
        final_url = r.url or url
        if r.status_code >= 400:
            log(f"[FETCH SKIP] HTTP {r.status_code}: {url[:120]}")
            return final_url, ""
        enc = r.encoding or r.apparent_encoding or "utf-8"
        r.encoding = enc
        return final_url, r.text or ""
    except Exception as e:
        log(f"[FETCH SKIP] {type(e).__name__}: {url[:120]}")
        return url, ""


def extract_article_text(html_text: str) -> str:
    if not html_text:
        return ""

    if BeautifulSoup is None:
        txt = re.sub(r"(?is)<(script|style|noscript).*?>.*?</\1>", " ", html_text)
        txt = re.sub(r"(?s)<[^>]+>", " ", txt)
        txt = html.unescape(txt)
        txt = re.sub(r"\s+", " ", txt)
        return txt[:4000].strip()

    soup = BeautifulSoup(html_text, "html.parser")

    for tag in soup(["script", "style", "noscript", "iframe", "svg", "footer", "header", "nav"]):
        tag.decompose()

    candidates = []

    for selector in ["article", "main", "[role=main]", ".article", ".article-body", ".news_body", "#articleBody"]:
        try:
            for node in soup.select(selector):
                text = " ".join([p.get_text(" ", strip=True) for p in node.find_all(["p", "div", "li"])])
                text = re.sub(r"\s+", " ", text).strip()
                if len(text) > 300:
                    candidates.append(text)
        except Exception:
            pass

    if not candidates:
        ps = [p.get_text(" ", strip=True) for p in soup.find_all("p")]
        ps = [p for p in ps if len(p) >= 30]
        text = " ".join(ps)
        text = re.sub(r"\s+", " ", text).strip()
        if text:
            candidates.append(text)

    if not candidates:
        text = soup.get_text(" ", strip=True)
        text = re.sub(r"\s+", " ", text).strip()
        candidates.append(text)

    text = max(candidates, key=len) if candidates else ""
    text = re.sub(r"(구독|광고|저작권|Copyright|All rights reserved).{0,120}", " ", text, flags=re.I)
    text = re.sub(r"\s+", " ", text).strip()
    return text[:6000]


# ============================================================
# 3. COLUMN NORMALIZE
# ============================================================
def find_input_file() -> Path:
    for p in INPUT_CANDIDATES:
        if p.exists():
            return p
    raise FileNotFoundError("입력 파일 없음: " + " / ".join(str(p) for p in INPUT_CANDIDATES))


def normalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    """
    입력 Excel 컬럼을 표준 컬럼으로 정규화합니다.

    중요 수정:
    - date / collected_at 같이 여러 컬럼이 같은 표준명(Date)으로 매핑되면
      pandas가 중복 컬럼명을 만들고, 이후 pd.DataFrame(rows)에서
      InvalidIndexError가 발생합니다.
    - 따라서 같은 표준명으로 매핑되는 컬럼은 첫 번째 non-empty 값을 합쳐
      표준 컬럼 1개만 남깁니다.
    """
    raw = df.copy().fillna("")

    def canon(col) -> str:
        lc = str(col).strip().lower()
        if lc in ["date", "publish date", "publish_date", "published", "news date", "뉴스 원본 게시일시", "원문등록일"]:
            return "Date"
        # collected_at은 원본 게시일이 아니라 수집일이므로 Date와 충돌시키지 않음
        if lc in ["collected_at", "last_checked", "checked_at", "수집일", "수집일시"]:
            return "last_checked"
        if lc in ["headline", "title", "news title", "뉴스 제목", "제목", "headlines"]:
            return "Headline"
        if lc in ["summary", "뉴스 본문", "뉴스 본문요약", "주요내용", "description", "content", "body", "article"]:
            return "Summary"
        if lc in ["ai analysis", "analysis", "impact", "ai_analysis", "전문관세사 분석", "ai분석"]:
            return "AI Analysis"
        if lc in ["action plan", "action", "action_plan", "대응방안", "action_plan"]:
            return "Action Plan"
        if lc in ["country", "국가", "대상 국가", "countries"]:
            return "Country"
        if lc in ["agency", "관련 기관", "정책기관", "관련기관", "organization"]:
            return "agency"
        if lc in ["risk", "importance", "중요도", "위험도"]:
            return "Risk"
        if lc in ["url", "link", "source url", "출처url", "링크", "출처 url"]:
            return "URL"
        if lc in ["source", "출처", "date source", "data source", "수집툴"]:
            return "source"
        if lc in ["score", "importance_score", "risk_score"]:
            return "score"
        return str(col).strip()

    # 원본 컬럼명별 표준명 생성
    canon_names = [canon(c) for c in raw.columns]

    # 중복 표준 컬럼 병합: 행 단위로 먼저 값이 있는 컬럼 선택
    out = pd.DataFrame(index=raw.index)
    for std in dict.fromkeys(canon_names):
        idxs = [i for i, name in enumerate(canon_names) if name == std]
        if len(idxs) == 1:
            coldata = raw.iloc[:, idxs[0]]
        else:
            part = raw.iloc[:, idxs].copy()
            coldata = part.apply(lambda r: next((clean_text(v) for v in r.values if clean_text(v)), ""), axis=1)
        out[std] = coldata.map(clean_text)

    # 필수 컬럼 보장
    for col in ["Date", "Headline", "Summary", "AI Analysis", "Action Plan", "Country", "agency", "Risk", "URL", "source", "score", "last_checked"]:
        if col not in out.columns:
            out[col] = ""
        out[col] = out[col].map(clean_text)

    # URL이 source 또는 Summary에 들어간 케이스 보정
    mask = (out["URL"].eq("")) & (out["source"].str.startswith("http", na=False))
    out.loc[mask, "URL"] = out.loc[mask, "source"]

    # 표준 컬럼을 앞쪽에 정렬하고 나머지 컬럼 유지
    fixed = ["Date", "Headline", "Summary", "AI Analysis", "Action Plan", "Country", "agency", "Risk", "URL", "source", "score", "last_checked"]
    others = [c for c in out.columns if c not in fixed]
    out = out[fixed + others].copy()

    # 최종 안전장치: 컬럼명 중복 완전 제거
    out = out.loc[:, ~out.columns.duplicated()].copy()
    return out


def headline_fallback(row: pd.Series) -> str:
    h = clean_text(row.get("Headline", ""))
    if h and h.lower() not in ["nan", "none"]:
        return h
    s = clean_text(row.get("Summary", ""))
    if s:
        return s[:100]
    u = clean_text(row.get("URL", ""))
    return domain_of(u) or "뉴스 제목 확인 필요"


# ============================================================
# 4. INFERENCE / SCORING
# ============================================================
def infer_country(text: str, current: str = "") -> str:
    raw = f"{current} {text}"
    low = raw.lower()
    found = []
    for code, keys in SAMSUNG_COUNTRY_MAP.items():
        if any(k.lower() in low for k in keys):
            found.append(code)
    return " / ".join(found[:2]) if found else clean_text(current) or "Global"


def infer_agency(text: str, current: str = "") -> str:
    low = text.lower()
    cur = clean_text(current)
    if cur and cur.lower() not in ["nan", "none", "google news", "google", "rss"]:
        return cur
    for agency, keys in AGENCY_MAP:
        if any(k.lower() in low for k in keys):
            return agency
    return "관련 정부/국제기관"


def infer_products(text: str) -> str:
    low = text.lower()
    products = []
    checks = [
        ("Mobile", ["smartphone", "mobile", "phone", "스마트폰", "모바일"]),
        ("Consumer Electronics", ["appliance", "tv", "refrigerator", "washing machine", "가전", "냉장고", "세탁기", "tv"]),
        ("Network Equipment", ["network", "telecom", "5g", "base station", "네트워크", "통신장비"]),
        ("Semiconductor", ["semiconductor", "chip", "hbm", "memory", "반도체", "칩", "메모리"]),
        ("Display", ["display", "oled", "panel", "디스플레이", "패널"]),
        ("Medical", ["medical", "healthcare", "의료기기"]),
    ]
    for name, keys in checks:
        if any(k in low for k in keys):
            products.append(name)
    return " / ".join(products[:3]) if products else "공통 공급망"


def normalize_risk(v: str) -> str:
    s = clean_text(v).lower()
    if "상" in s or "high" in s or "직접" in s:
        return "상"
    if "하" in s or "low" in s or "기타" in s:
        return "하"
    if "중" in s or "medium" in s or "간접" in s:
        return "중"
    return ""


def calc_score(row: pd.Series) -> int:
    text = safe_join_values([row.get("Headline", ""), row.get("Summary", ""), row.get("Country", ""), row.get("agency", ""), row.get("source", "")]).lower()
    score = 0

    if any(k.lower() in text for k in TRADE_KEYWORDS):
        score += 40
    if any(k.lower() in text for k in HIGH_IMPACT_TERMS):
        score += 30

    country = infer_country(text)
    if any(c in country for c in ["US", "CN", "VN", "IN", "MX", "BR", "EU", "KR"]):
        score += 20

    if any(k in text for k in ["samsung", "semiconductor", "smartphone", "display", "appliance", "electronics", "삼성", "반도체", "스마트폰", "디스플레이", "가전"]):
        score += 20

    if any(n.lower() in text for n in NOISE_TERMS):
        score -= 80

    risk = normalize_risk(row.get("Risk", ""))
    if risk == "상":
        score += 15
    elif risk == "중":
        score += 7

    try:
        score += int(float(row.get("score", 0)))
    except Exception:
        pass

    return score


def infer_risk(headline: str, body: str, country: str, products: str) -> str:
    text = f"{headline} {body}".lower()

    if any(n.lower() in text for n in NOISE_TERMS):
        return "하"

    if any(k.lower() in text for k in HIGH_IMPACT_TERMS):
        if any(c in country for c in ["US", "CN", "VN", "IN", "MX", "BR", "EU", "KR"]):
            return "상"

    if any(k.lower() in text for k in TRADE_KEYWORDS):
        if any(c in country for c in ["US", "CN", "VN", "IN", "MX", "BR", "EU", "KR"]):
            return "중"

    return "하"


def infer_section(risk: str, text: str) -> str:
    low = text.lower()
    if risk == "상":
        return "1.직접 영향"
    if any(k in low for k in ["tariff", "fta", "trade", "관세", "통상", "무역협정", "환급"]):
        return "2.관세/통상 정책"
    if any(k in low for k in ["anti-dumping", "countervailing", "investigation", "customs", "반덤핑", "상계관세", "조사", "통관"]):
        return "3.수입규제/조사"
    return "4.기타 모니터링"


def is_relevant(row: pd.Series) -> bool:
    text = safe_join_values(row.values).lower()
    has_trade = any(k.lower() in text for k in TRADE_KEYWORDS)
    has_country = any(k.lower() in text for keys in SAMSUNG_COUNTRY_MAP.values() for k in keys)
    is_noise = any(k.lower() in text for k in NOISE_TERMS)
    return (has_trade and has_country) or (has_trade and not is_noise)


def dedup_similar(df: pd.DataFrame, max_rows: int = 120) -> pd.DataFrame:
    rows = []
    seen_titles = []
    seen_urls = set()

    for _, r in df.iterrows():
        url = clean_text(r.get("URL", "")).lower()
        title = normalize_title(r.get("Headline", ""))

        if url and url in seen_urls:
            continue
        if not title:
            continue

        duplicate = False
        for st in seen_titles:
            # 단순 token overlap 기반 중복 제거
            a = set(title.split())
            b = set(st.split())
            if a and b:
                overlap = len(a & b) / max(1, min(len(a), len(b)))
                if overlap >= 0.75:
                    duplicate = True
                    break

        if duplicate:
            continue

        seen_titles.append(title)
        if url:
            seen_urls.add(url)
        # Series 그대로 append하면 중복 컬럼/index가 있을 때 InvalidIndexError 발생 가능
        rows.append(r.to_dict())

        if len(rows) >= max_rows:
            break

    if rows:
        out = pd.DataFrame(rows)
        out = out.loc[:, ~out.columns.duplicated()].copy()
        return out
    return df.head(max_rows).copy().loc[:, ~df.columns.duplicated()]


# ============================================================
# 5. GEMINI ANALYSIS
# ============================================================
def call_gemini(headline: str, body: str, url: str, country: str, agency: str) -> dict:
    if not USE_GEMINI:
        return {}

    prompt = f"""
당신은 삼성전자 본사 관세/통상 리스크 분석 담당자입니다.
아래 뉴스/게시물 내용을 기준으로 GTI 메일용 분석을 작성하세요.

분석 기준:
- 삼성전자 생산거점: Korea, China, Vietnam, India, Indonesia, Turkey, Slovakia, Poland, Mexico, Brazil
- 제품군: Mobile, Consumer Electronics, Network Equipment, Semiconductor, Display, Medical
- 관세율 변동/반덤핑/상계관세/수출통제/수입규제/FTA 원산지 영향은 중요도 상 또는 중
- 금/은/담배/학생비자/범죄/일반정치 등 삼성전자 관세업무 직접 관련 낮은 뉴스는 중요도 하
- 과장하지 말고 원문 기반으로 작성
- 반드시 한국어로 작성
- 아래 JSON만 출력

출력 JSON 형식:
{{
  "Summary": "뉴스 핵심 내용 2~3문장",
  "AI Analysis": "삼성전자 관세/통상 업무 영향 2~3문장. 관련 없으면 영향 낮음이라고 명확히 작성",
  "Action Plan": "관세 전문가 대응방안 1~3개 항목",
  "Country": "대표 국가 2개 이하",
  "agency": "관련 정부기관/국제기구 2개 이하",
  "Risk": "상/중/하",
  "Products": "관련 제품군"
}}

Headline: {headline}
Current Country: {country}
Current Agency: {agency}
URL: {url}

Body:
{body[:4500]}
""".strip()

    endpoint = f"https://generativelanguage.googleapis.com/v1beta/models/{GEMINI_MODEL}:generateContent"
    params = {"key": GEMINI_API_KEY}
    payload = {
        "contents": [{"parts": [{"text": prompt}]}],
        "generationConfig": {
            "temperature": 0.2,
            "topP": 0.8,
            "maxOutputTokens": 1200,
        },
    }

    try:
        r = requests.post(endpoint, params=params, json=payload, timeout=60)
        if r.status_code >= 400:
            log(f"[GEMINI SKIP] HTTP {r.status_code}: {r.text[:120]}")
            return {}

        data = r.json()
        text = data["candidates"][0]["content"]["parts"][0]["text"]
        text = text.strip()
        text = re.sub(r"^```json\s*", "", text, flags=re.I)
        text = re.sub(r"^```\s*", "", text)
        text = re.sub(r"\s*```$", "", text)
        m = re.search(r"\{.*\}", text, flags=re.S)
        if m:
            text = m.group(0)
        obj = json.loads(text)
        return obj if isinstance(obj, dict) else {}
    except Exception as e:
        log(f"[GEMINI SKIP] {type(e).__name__}: {e}")
        return {}


def fallback_summary(headline: str, body: str) -> str:
    body = clean_text(body)
    headline = clean_text(headline)

    if body and len(body) >= 250:
        sentences = re.split(r"(?<=[.!?。다])\s+", body)
        sentences = [clean_text(s) for s in sentences if len(clean_text(s)) >= 25]
        if sentences:
            out = " ".join(sentences[:3])
            return out[:420]

    # 제목만 있어도 반복 문구 대신 제목 기반 요약 생성
    return f"해당 뉴스는 '{headline}' 이슈를 다루고 있습니다. 제목 기준으로 관세·통상 관련성, 대상 국가, 관련 품목 또는 제도 변경 가능성을 확인해야 합니다."


def fallback_analysis(headline: str, body: str, country: str, products: str, risk: str) -> str:
    text = f"{headline} {body}".lower()

    if risk == "하":
        return "삼성전자 주요 제품·생산거점의 관세비용 또는 통관 프로세스에 대한 직접 영향은 낮아 보입니다. 다만 동일 이슈가 수입규제·관세조치로 확대되는지 모니터링이 필요합니다."

    if "anti-dumping" in text or "반덤핑" in text:
        return f"{country} 관련 반덤핑 또는 수입규제 이슈로, {products} 공급망 내 유사 HS 품목의 조사 확대 가능성을 점검해야 합니다. 베트남·인도·중국 등 생산법인의 수출입 품목과 규제 대상 품목의 중복 여부 확인이 필요합니다."

    if "fta" in text or "무역협정" in text:
        return f"{country} FTA/무역협정 변화는 원산지 기준, 협정세율 적용 가능성, 공급망 전환 전략에 영향을 줄 수 있습니다. {products} 관련 거래에서 원산지 판정 및 증빙 체계를 사전 점검해야 합니다."

    if "tariff" in text or "duty" in text or "관세" in text:
        return f"{country} 관세·수입세율 변동 이슈로, {products} 관련 원재료·완제품의 수입원가와 통관 신고 기준에 영향을 줄 수 있습니다. HS, 과세가격, 원산지 기준 변경 여부를 확인해야 합니다."

    if "customs" in text or "통관" in text or "세관" in text:
        return f"{country} 통관·세관 집행 동향으로, 신고 정확성·증빙관리·사후심사 리스크가 높아질 수 있습니다. 법인별 신고자료와 관세사 제출자료의 정합성 점검이 필요합니다."

    return f"{country} 통상정책 변화 가능성이 있어 {products} 공급망 관점에서 HS, 원산지, 과세가격, 수입규제 영향 여부를 확인해야 합니다."


def fallback_action(headline: str, body: str, country: str, risk: str) -> str:
    text = f"{headline} {body}".lower()

    if risk == "하":
        return "① 일일 모니터링 유지 ② 동일 이슈의 관세·수입규제 확대 여부 확인 ③ 삼성 관련 HS/품목과 직접 연결될 경우 재분류"

    if "anti-dumping" in text or "반덤핑" in text:
        return "① 규제 대상 HS/품목 확인 ② 생산법인 수출입 품목과 매칭 ③ 조사대상국·공급업체 거래 여부 확인 ④ 필요 시 관세사/법무 검토 요청"

    if "fta" in text or "무역협정" in text:
        return "① 협정문·원산지 기준 변경 여부 확인 ② 적용 가능 HS 및 생산공정 기준 검토 ③ 법인별 원산지 증빙 준비상태 점검"

    if "tariff" in text or "duty" in text or "관세" in text:
        return "① 세율 변경 대상 HS 확인 ② 수입원가 영향 시뮬레이션 ③ 기존 신고가격·원산지·거래구조 점검 ④ 법인별 대응 가이드 배포"

    return "① 원문 공고/법령 확인 ② 대상 국가·HS·제품군 매핑 ③ 법인별 수입/수출 신고 영향 검토 ④ 필요 시 대응계획 수립"


# ============================================================
# 6. PREPARE TOP NEWS
# ============================================================
def prepare_top(raw: pd.DataFrame) -> pd.DataFrame:
    df = normalize_columns(raw)

    # 핵심 버그 방지: 모든 값을 문자열로 안전 결합
    df["_pre_text"] = df.apply(lambda r: safe_join_values(r.values), axis=1)
    df["Headline"] = df.apply(headline_fallback, axis=1)
    df["Date"] = df["Date"].map(safe_date)
    df["URL"] = df["URL"].map(clean_text)

    # URL 없는 행 제거
    df = df[df["Headline"].map(lambda x: len(clean_text(x)) > 0)].copy()

    # relevance/score
    df["_relevant"] = df.apply(is_relevant, axis=1)
    df["_score"] = df.apply(calc_score, axis=1)
    df = df.sort_values(["_relevant", "_score"], ascending=[False, False]).copy()

    # 1차 중복 제거
    df["_url_key"] = df["URL"].str.lower().str.strip()
    df["_title_key"] = df["Headline"].map(normalize_title)
    if "_url_key" in df.columns:
        df = df.drop_duplicates(subset=["_url_key"], keep="first")
    df = df.drop_duplicates(subset=["_title_key"], keep="first")

    # 유사 제목 중복 제거
    candidates = dedup_similar(df, max_rows=FETCH_LIMIT).copy()
    log(f"[CANDIDATES] {len(candidates)} rows selected for body/AI analysis")

    rows = []
    for i, (_, r) in enumerate(candidates.iterrows(), start=1):
        headline = clean_text(r.get("Headline", ""))
        url = clean_text(r.get("URL", ""))
        source_summary = clean_text(r.get("Summary", ""))
        source_ai = clean_text(r.get("AI Analysis", ""))
        source_action = clean_text(r.get("Action Plan", ""))

        log(f"[{i}/{len(candidates)}] {headline[:70]}")

        final_url = resolve_redirect_url(url)
        article_body = ""

        # 기존 Summary가 충분하지 않으면 URL 본문 추출
        if len(source_summary) < 120 or "본문 정보가 제한" in source_summary:
            fetched_url, html_text = fetch_html(final_url)
            if fetched_url:
                final_url = fetched_url
            article_body = extract_article_text(html_text)
            time.sleep(0.4)

        base_body = article_body if len(article_body) >= 200 else source_summary
        if not base_body:
            base_body = headline

        country = infer_country(safe_join_values([headline, base_body, r.get("Country", "")]), r.get("Country", ""))
        agency = infer_agency(safe_join_values([headline, base_body, final_url, r.get("agency", "")]), r.get("agency", ""))
        products = infer_products(safe_join_values([headline, base_body]))
        pre_risk = normalize_risk(r.get("Risk", "")) or infer_risk(headline, base_body, country, products)

        ai = {}
        # 기존 분석이 복붙/부실이면 Gemini 재분석
        bad_existing = (
            not source_ai
            or "HS, 원산지, 과세가격" in source_ai
            or "본문 정보가 제한" in source_summary
            or len(source_ai) < 40
        )
        if bad_existing:
            ai = call_gemini(headline, base_body, final_url, country, agency)

        summary = clean_text(ai.get("Summary", "")) or fallback_summary(headline, base_body)
        analysis = clean_text(ai.get("AI Analysis", "")) or (
            source_ai if source_ai and "HS, 원산지, 과세가격" not in source_ai else fallback_analysis(headline, base_body, country, products, pre_risk)
        )
        action = clean_text(ai.get("Action Plan", "")) or (
            source_action if source_action and "대상 국가·HS·제품군" not in source_action else fallback_action(headline, base_body, country, pre_risk)
        )
        country = clean_text(ai.get("Country", "")) or country
        agency = clean_text(ai.get("agency", "")) or agency
        products = clean_text(ai.get("Products", "")) or products
        risk = normalize_risk(ai.get("Risk", "")) or pre_risk

        # Gemini가 관련 없는 뉴스를 과대평가한 경우 하향
        if any(n.lower() in f"{headline} {base_body}".lower() for n in NOISE_TERMS):
            risk = "하"

        section = infer_section(risk, safe_join_values([headline, summary, analysis]))

        rows.append({
            "Date": safe_date(r.get("Date", "")),
            "Headline": headline,
            "Summary": summary,
            "AI Analysis": analysis,
            "Action Plan": action,
            "Country": country,
            "agency": agency,
            "Risk": risk,
            "URL": final_url or url,
            "source": clean_text(r.get("source", "")) or domain_of(final_url or url),
            "Products": products,
            "Section": section,
            "score": calc_score(r) + (30 if risk == "상" else 15 if risk == "중" else 0),
            "body_len": len(article_body),
        })

    out = pd.DataFrame(rows)

    if out.empty:
        raise RuntimeError("분석 대상 뉴스가 없습니다. STEP3 결과 파일을 확인하세요.")

    out["_section_order"] = out["Section"].map(SECTION_ORDER).fillna(9)
    out["_risk_order"] = out["Risk"].map(RISK_ORDER).fillna(4)
    out = out.sort_values(
        ["_risk_order", "_section_order", "score", "Country"],
        ascending=[True, True, False, True]
    ).head(TOP_N).reset_index(drop=True)
    out.insert(0, "No", range(1, len(out) + 1))

    return out[[
        "No", "Section", "Date", "Headline", "Summary", "AI Analysis", "Action Plan",
        "Country", "agency", "Risk", "Products", "URL", "source", "score", "body_len"
    ]]


# ============================================================
# 7. EXCEL / HTML
# ============================================================
def save_excel(df: pd.DataFrame, path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "GTI Radar Top30"

    headers = ["No", "Section", "Date", "Headline", "Summary", "AI Analysis", "Action Plan",
               "Country", "agency", "Risk", "Products", "URL", "source", "score", "body_len"]
    ws.append(headers)

    for _, r in df.iterrows():
        ws.append([r.get(h, "") for h in headers])
        row_idx = ws.max_row
        url = clean_text(r.get("URL", ""))
        if url:
            cell = ws.cell(row=row_idx, column=4)
            cell.hyperlink = url
            cell.font = Font(color="0563C1", underline="single", bold=True)

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    widths = {
        "A": 6, "B": 16, "C": 18, "D": 46, "E": 55, "F": 58, "G": 52,
        "H": 14, "I": 28, "J": 10, "K": 22, "L": 36, "M": 18, "N": 10, "O": 10
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border
        ws.row_dimensions[row[0].row].height = 95

    for row in range(2, ws.max_row + 1):
        risk = ws.cell(row=row, column=10).value
        if risk == "상":
            ws.cell(row=row, column=10).fill = PatternFill("solid", fgColor="F4B183")
        elif risk == "중":
            ws.cell(row=row, column=10).fill = PatternFill("solid", fgColor="FFE699")
        else:
            ws.cell(row=row, column=10).fill = PatternFill("solid", fgColor="C6E0B4")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    ws2 = wb.create_sheet("Run Summary")
    summary_rows = [
        ["Generated At", NOW_STR],
        ["Subject", SUBJECT],
        ["Input", str(INPUT_FILE)],
        ["Output HTML", str(OUTPUT_HTML)],
        ["Rows", len(df)],
        ["Gemini", "ON" if USE_GEMINI else "OFF"],
        ["Model", GEMINI_MODEL],
    ]
    for row in summary_rows:
        ws2.append(row)
    ws2.column_dimensions["A"].width = 24
    ws2.column_dimensions["B"].width = 90

    wb.save(path)


def esc(v) -> str:
    return html.escape(clean_text(v)).replace("\n", "<br>")


def link_html(title: str, url: str) -> str:
    title = esc(title)
    url = clean_text(url)
    if url:
        return f"<a href='{html.escape(url)}' style='color:#0563c1;text-decoration:underline;font-weight:bold;'>{title}</a>"
    return title


def build_overall_review(df: pd.DataFrame) -> str:
    top = df.head(5)
    countries = []
    risks = df["Risk"].value_counts().to_dict()
    for c in top["Country"].tolist():
        for x in str(c).split("/"):
            x = x.strip()
            if x and x not in countries:
                countries.append(x)
    country_txt = ", ".join(countries[:5]) if countries else "주요 생산국"
    return f"금일 GTI Radar는 {country_txt} 관련 관세·통상 이슈를 중심으로 선별했습니다. 중요도 상 {risks.get('상',0)}건, 중 {risks.get('중',0)}건, 하 {risks.get('하',0)}건이며, 고위험 건은 대상 HS·원산지·수입규제 여부를 우선 확인해야 합니다."


def build_html(df: pd.DataFrame) -> str:
    top3 = df.head(3).copy()
    rest = df.iloc[3:].copy()

    top_blocks = []
    for _, r in top3.iterrows():
        top_blocks.append(f"""
        <div style="margin:16px 0 20px 0;padding:14px;border-left:5px solid #C00000;background:#fff7f7;">
          <div style="font-size:15px;margin-bottom:6px;">{int(r['No'])}️⃣ {link_html(r['Headline'], r['URL'])}</div>
          <div style="font-size:12px;color:#555;margin-bottom:8px;">
            Date: {esc(r['Date'])} | Country: {esc(r['Country'])} | Agency: {esc(r['agency'])} | Risk: <b>{esc(r['Risk'])}</b>
          </div>
          <div><b>Summary</b><br>{esc(r['Summary'])}</div>
          <div style="margin-top:8px;"><b>AI Analysis</b><br>{esc(r['AI Analysis'])}</div>
          <div style="margin-top:8px;"><b>Action Plan</b><br>{esc(r['Action Plan'])}</div>
        </div>
        """)

    rows = []
    for _, r in rest.iterrows():
        risk_bg = "#F4B183" if r["Risk"] == "상" else "#FFE699" if r["Risk"] == "중" else "#C6E0B4"
        rows.append(f"""
        <tr>
          <td style="padding:7px;border:1px solid #d9d9d9;text-align:center;">{int(r['No'])}</td>
          <td style="padding:7px;border:1px solid #d9d9d9;">{esc(r['Section'])}</td>
          <td style="padding:7px;border:1px solid #d9d9d9;">{link_html(r['Headline'], r['URL'])}</td>
          <td style="padding:7px;border:1px solid #d9d9d9;">{esc(r['Summary'])}</td>
          <td style="padding:7px;border:1px solid #d9d9d9;">{esc(r['AI Analysis'])}</td>
          <td style="padding:7px;border:1px solid #d9d9d9;">{esc(r['Action Plan'])}</td>
          <td style="padding:7px;border:1px solid #d9d9d9;text-align:center;">{esc(r['Country'])}</td>
          <td style="padding:7px;border:1px solid #d9d9d9;">{esc(r['agency'])}</td>
          <td style="padding:7px;border:1px solid #d9d9d9;text-align:center;background:{risk_bg};font-weight:bold;">{esc(r['Risk'])}</td>
          <td style="padding:7px;border:1px solid #d9d9d9;text-align:center;">{esc(r['Date'])}</td>
        </tr>
        """)

    return f"""<!DOCTYPE html>
<html lang="ko">
<head>
<meta charset="utf-8">
<title>{html.escape(SUBJECT)}</title>
</head>
<body style="font-family:Arial,'Malgun Gothic',sans-serif;font-size:13px;color:#222;line-height:1.5;">
<div style="max-width:1280px;margin:0 auto;">
  <h2 style="margin-bottom:4px;">[GTI Radar] Global Trade Intelligence</h2>
  <div style="font-size:14px;margin-bottom:4px;"><b>Date:</b> {TODAY}</div>
  <div style="font-size:12px;color:#555;margin-bottom:16px;">
    Coverage: Last 24 Hours | Focus: Samsung Electronics Customs & Trade Risk
  </div>

  <h3 style="margin-top:18px;margin-bottom:6px;">총평</h3>
  <p style="margin-top:0;">{esc(build_overall_review(df))}</p>

  <h3 style="color:#C00000;margin-top:22px;">🔴 TOP POLICY EVENTS</h3>
  {''.join(top_blocks)}

  <h3 style="color:#1F4E78;margin-top:24px;">🟦 EVENT LIST ({len(rest)})</h3>
  <table style="border-collapse:collapse;width:100%;font-size:12px;">
    <tr style="background:#1F4E78;color:white;">
      <th style="padding:7px;border:1px solid #d9d9d9;">No</th>
      <th style="padding:7px;border:1px solid #d9d9d9;">Section</th>
      <th style="padding:7px;border:1px solid #d9d9d9;">Headline</th>
      <th style="padding:7px;border:1px solid #d9d9d9;">Summary</th>
      <th style="padding:7px;border:1px solid #d9d9d9;">AI Analysis</th>
      <th style="padding:7px;border:1px solid #d9d9d9;">Action Plan</th>
      <th style="padding:7px;border:1px solid #d9d9d9;">Country</th>
      <th style="padding:7px;border:1px solid #d9d9d9;">Agency</th>
      <th style="padding:7px;border:1px solid #d9d9d9;">Risk</th>
      <th style="padding:7px;border:1px solid #d9d9d9;">Date</th>
    </tr>
    {''.join(rows)}
  </table>

  <p style="margin-top:18px;color:#666;font-size:12px;">
    ※ 본 메일은 GTI 자동 수집/분석 결과이며, 중요도 상 이슈는 원문 공고·법령 확인 후 법인별 영향 검토가 필요합니다.
  </p>
</div>
</body>
</html>"""


# ============================================================
# 8. MAIL / CUMULATIVE
# ============================================================
def load_recipients() -> list[str]:
    recipients = []
    for fn in ["00.xlsx", "mail.xlsx"]:
        fp = BASE_DIR / fn
        if fp.exists():
            try:
                rdf = pd.read_excel(fp)
                text = "\n".join(rdf.fillna("").astype(str).values.ravel().tolist())
                recipients.extend(re.findall(r"[A-Za-z0-9._%+\-]+@[A-Za-z0-9.\-]+\.[A-Za-z]{2,}", text))
            except Exception as e:
                log(f"[RECIPIENT READ SKIP] {fp}: {e}")

    if not recipients and FALLBACK_TO:
        recipients.extend([x.strip() for x in FALLBACK_TO.split(",") if x.strip()])

    return list(dict.fromkeys(recipients))


def send_email(html_body: str, attachments: list[Path]) -> None:
    recipients = load_recipients()

    if not recipients:
        log("[MAIL SKIP] 수신자 없음: 00.xlsx / mail.xlsx / GTI_MAIL_TO 확인")
        return

    if not SMTP_USER or not SMTP_PASS:
        log("[MAIL SKIP] SMTP_USER/PASS 없음")
        return

    msg = EmailMessage()
    msg["Subject"] = SUBJECT
    msg["From"] = formataddr((MAIL_FROM_NAME, SMTP_USER))
    msg["To"] = ", ".join(recipients)
    msg.set_content("GTI Radar 메일입니다. HTML 메일을 지원하는 클라이언트에서 확인해 주세요.")
    msg.add_alternative(html_body, subtype="html")

    for fp in attachments:
        if not fp.exists():
            continue
        data = fp.read_bytes()
        msg.add_attachment(
            data,
            maintype="application",
            subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename=fp.name,
        )

    context = ssl.create_default_context()
    with smtplib.SMTP_SSL(SMTP_HOST, SMTP_PORT, context=context) as server:
        server.login(SMTP_USER, SMTP_PASS)
        server.send_message(msg)

    log(f"[MAIL SENT] {len(recipients)} recipients")


def update_cumulative(df: pd.DataFrame) -> None:
    data = df.copy()
    data.insert(0, "mail_date", TODAY)
    data.insert(1, "subject", SUBJECT)

    if MAIL_CUMULATIVE.exists():
        try:
            old = pd.read_excel(MAIL_CUMULATIVE)
            data = pd.concat([old, data], ignore_index=True)
            data = data.drop_duplicates(subset=["mail_date", "Headline", "URL"], keep="last")
        except Exception as e:
            log(f"[CUMULATIVE READ SKIP] {e}")

    data.to_excel(MAIL_CUMULATIVE, index=False)


# ============================================================
# 9. MAIN
# ============================================================
def main() -> None:
    global INPUT_FILE
    BASE_DIR.mkdir(parents=True, exist_ok=True)
    INPUT_FILE = find_input_file()

    log("[START] GTI Mail Engine FINAL URL BODY + AI")
    log(f"BASE_DIR={BASE_DIR}")
    log(f"INPUT={INPUT_FILE}")
    log(f"GEMINI={'ON' if USE_GEMINI else 'OFF'} / MODEL={GEMINI_MODEL}")
    log(f"GTI_SEND_EMAIL={os.getenv('GTI_SEND_EMAIL')} / SEND_EMAIL={SEND_EMAIL}")

    raw = pd.read_excel(INPUT_FILE)
    log(f"[LOAD] rows={len(raw)}")

    top = prepare_top(raw)
    log(f"[ANALYSIS COMPLETE] rows={len(top)}")

    save_excel(top, OUTPUT_XLSX)
    log(f"[SAVE] Excel: {OUTPUT_XLSX}")

    html_body = build_html(top)
    OUTPUT_HTML.write_text(html_body, encoding="utf-8")
    log(f"[SAVE] HTML: {OUTPUT_HTML}")

    update_cumulative(top)
    log(f"[SAVE] Cumulative: {MAIL_CUMULATIVE}")

    if SEND_EMAIL:
        send_email(html_body, [OUTPUT_XLSX])
    else:
        log("[MAIL SKIP] GTI_SEND_EMAIL=Y 설정 시 실제 발송")

    log("[DONE] GTI Mail Engine FINAL")


if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        log(f"[ERROR] {type(e).__name__}: {e}")
        log(traceback.format_exc())
        raise
